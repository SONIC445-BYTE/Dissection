from pathlib import Path
import csv,json,textwrap,html
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
OUT=Path('/home/user/regacore_research')
# feature xlsx
with (OUT/'feature_inventory.csv').open(encoding='utf-8') as f: rows=list(csv.reader(f))
wb=Workbook(); ws=wb.active; ws.title='Feature Inventory'
for r in rows: ws.append(r)
fill=PatternFill('solid',fgColor='111827'); font=Font(color='FFFFFF',bold=True); thin=Side(style='thin',color='D1D5DB')
for c in ws[1]: c.fill=fill; c.font=font; c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
for row in ws.iter_rows(min_row=2):
 for c in row: c.alignment=Alignment(vertical='top',wrap_text=True); c.border=Border(top=thin,bottom=thin,left=thin,right=thin)
for i,wid in enumerate([32,42,38,38,14,22,12,15,60],1): ws.column_dimensions[get_column_letter(i)].width=wid
ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
wb.save(OUT/'feature_inventory.xlsx')
# screenshot catalog
cap=json.loads((OUT/'capture_inventory.json').read_text())
fields=['Page name','URL','Viewport','Screenshot','Title','Visible components summary','Hosts observed','Observed/inferred','Confidence']
shot=[]
for r in cap:
 if r.get('screenshot'):
  shot.append({'Page name':r['name'],'URL':r['url'],'Viewport':f"{r['viewport'].get('width')}x{r['viewport'].get('height')}",'Screenshot':str(Path(r['screenshot']).relative_to(OUT)),'Title':r.get('title',''),'Visible components summary':r.get('body_text_sample','')[:300].replace('\n',' | '),'Hosts observed':', '.join(r.get('hosts',[])),'Observed/inferred':'Observed','Confidence':'High'})
with (OUT/'screenshot_catalog.csv').open('w',newline='',encoding='utf-8') as f:
 w=csv.DictWriter(f,fieldnames=fields); w.writeheader(); w.writerows(shot)
wb=Workbook(); ws=wb.active; ws.title='Screenshot Catalog'; ws.append(fields)
for row in shot: ws.append([row[f] for f in fields])
for c in ws[1]: c.fill=fill; c.font=font; c.alignment=Alignment(wrap_text=True,horizontal='center')
for row in ws.iter_rows(min_row=2):
 for c in row: c.alignment=Alignment(vertical='top',wrap_text=True); c.border=Border(top=thin,bottom=thin,left=thin,right=thin)
for i,wid in enumerate([24,48,14,42,30,80,50,18,12],1): ws.column_dimensions[get_column_letter(i)].width=wid
ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions; wb.save(OUT/'screenshot_catalog.xlsx')
# SVG helpers
def box(x,y,w,h,t,s='',fillc='#fff'):
 return f'<rect x="{x}" y="{y}" width="{w}" height="{h}" rx="14" fill="{fillc}" stroke="#111827" stroke-width="1.4"/><text x="{x+w/2}" y="{y+25}" text-anchor="middle" font-family="Arial" font-size="15" font-weight="700" fill="#111827">{html.escape(t)}</text>'+(f'<text x="{x+w/2}" y="{y+48}" text-anchor="middle" font-family="Arial" font-size="11" fill="#374151">{html.escape(s)}</text>' if s else '')
def arr(x1,y1,x2,y2,l=''):
 s=f'<line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" stroke="#4B5563" stroke-width="2" marker-end="url(#arrow)"/>'
 if l: s+=f'<text x="{(x1+x2)/2}" y="{(y1+y2)/2-6}" text-anchor="middle" font-family="Arial" font-size="10" fill="#374151">{html.escape(l)}</text>'
 return s
hdr='<defs><marker id="arrow" markerWidth="10" markerHeight="10" refX="7" refY="3" orient="auto" markerUnits="strokeWidth"><path d="M0,0 L0,6 L8,3 z" fill="#4B5563"/></marker></defs>'
arch=f'<svg xmlns="http://www.w3.org/2000/svg" width="1600" height="1000" viewBox="0 0 1600 1000">{hdr}<rect width="1600" height="1000" fill="#F8FAFC"/><text x="800" y="48" text-anchor="middle" font-family="Arial" font-size="30" font-weight="800">Regacore Product Architecture — Public + Inferred System</text><text x="800" y="78" text-anchor="middle" font-family="Arial" font-size="14" fill="#4B5563">Observed public pages, app demo routes, network fingerprints, and separated inferences</text>'
for args in [(60,130,210,80,'Visitor','SEO/referral/social','#E0F2FE'),(340,130,220,80,'Marketing Site','Landing/blog/FAQs','#fff'),(630,130,220,80,'Conversion','Checkout/waitlist/login','#FEF3C7'),(920,130,220,80,'Auth + Account','Supabase session','#FCE7F3'),(1210,130,250,80,'Member App Shell','Home/Data/Protocol/AI/Market','#EDE9FE'),(80,320,240,90,'Lab + Data Intake','Biomarkers/uploads/wearables','#DCFCE7'),(390,320,240,90,'Records Store','Biomarkers/categories/history','#fff'),(700,320,240,90,'Digital Twin/Scores','3D GLB, bio age, grades','#DBEAFE'),(1010,320,240,90,'AI Reasoning Layer','Risk reports/chat/insights','#FEE2E2'),(1320,320,220,90,'Human Care Layer','Care team/specialists','#ECFDF5'),(190,540,240,90,'Protocol Engine','Lifestyle/diet/supplements','#fff'),(500,540,240,90,'Marketplace','Tests/supplements/Rx','#FFEDD5'),(810,540,240,90,'Notifications','Prep/retest/follow-up','#F3F4F6'),(1120,540,240,90,'Retention Loop','Retest/progress/renewal','#E0E7FF'),(320,760,260,90,'Operational Partners','Labs/phlebotomy/clinicians','#fff'),(670,760,260,90,'Compliance/Safety','DPDP/SPI/HIPAA claims','#fff'),(1020,760,260,90,'Analytics','Cloudflare RUM/metrics','#fff')]: arch+=box(*args)
for a in [(270,170,340,170,''),(560,170,630,170,''),(850,170,920,170,''),(1140,170,1210,170,''),(1335,210,1335,320,'member use'),(320,365,390,365,''),(630,365,700,365,''),(940,365,1010,365,''),(1250,365,1320,365,''),(820,410,310,540,'actions'),(430,585,500,585,''),(740,585,810,585,''),(1050,585,1120,585,''),(1240,630,1240,740,'renew'),(450,760,250,630,'ops'),(800,760,1010,410,'guardrails'),(1150,760,1150,630,'performance')]: arch+=arr(*a)
arch+='<text x="70" y="925" font-family="Arial" font-size="13" fill="#374151"><tspan font-weight="700">Legend:</tspan> AI/back-end internals are inferred unless explicitly observed.</text></svg>'
(OUT/'product_architecture.svg').write_text(arch,encoding='utf-8')
steps=[('Visitor','SEO/referral/social'),('Landing','Hero + CTAs'),('Signup','Checkout or waitlist'),('Verification','Auth/payment unclear'),('Onboarding','History + consent'),('Assessment','Lab/upload/wearables'),('Reports','Dashboard/records'),('Recommendations','Protocol/market'),('Follow-up','AI/care team'),('Retention','Retest/progress'),('Subscription','Annual renewal'),('Referral','Invite reward'),('Long-term','Trends/twin')]
svg=f'<svg xmlns="http://www.w3.org/2000/svg" width="1200" height="700" viewBox="0 0 1200 700">{hdr}<rect width="1200" height="700" fill="#F8FAFC"/><text x="600" y="54" text-anchor="middle" font-family="Arial" font-size="28" font-weight="800">Regacore User Journey Map</text><text x="600" y="82" text-anchor="middle" font-family="Arial" font-size="13" fill="#4B5563">Observed public funnel + inferred paid lifecycle</text>'
x0,y0,w,h,g=60,170,150,74,70
pos=[]
for i,(t,s) in enumerate(steps):
 row=i//5; col=i%5; x=x0+col*(w+g); y=y0+row*180; pos.append((x,y)); svg+=box(x,y,w,h,t,s,'#ECFEFF' if i%2==0 else '#fff')
for i in range(len(steps)-1):
 x,y=pos[i]; nx,ny=pos[i+1]
 svg+=arr(x+w,y+h/2,nx,ny+h/2) if ny==y else arr(x+w/2,y+h,nx+w/2,ny)
svg+='<rect x="60" y="600" width="1080" height="60" rx="14" fill="#111827"/><text x="600" y="635" text-anchor="middle" font-family="Arial" font-size="15" fill="#fff">Ovexis takeaway: make each transition auditable — consent, lab order, clinician review, AI provenance, commerce disclosure, renewal value.</text></svg>'
(OUT/'user_journey_map.svg').write_text(svg,encoding='utf-8')
items=[('Key Partners','Labs/phlebotomy; clinicians; vendors; cloud; wearable platforms'),('Key Activities','Acquire; collect labs; interpret; generate protocols; care messaging; fulfillment'),('Value Proposition','100+ biomarker check + digital twin + AI/care team + action plan'),('Customer Relationships','AI concierge; care team; specialist ticketing; reminders; referrals'),('Customer Segments','Health-conscious adults 18+; longevity/performance users; not acute-care patients'),('Key Resources','Health data; interpretation engine; brand; lab network; clinical frameworks'),('Channels','Website; blog/SEO; waitlist; social; referral; partnerships'),('Cost Structure','Labs; phlebotomy; clinicians; AI/cloud; support; compliance; acquisition'),('Revenue Streams','Annual membership; add-on tests; supplements; prescriptions; specialists; corporate')]
pos=[(40,120),(330,120),(620,120),(910,120),(1200,120),(40,440),(330,440),(620,440),(910,440)]
canvas='<svg xmlns="http://www.w3.org/2000/svg" width="1500" height="1000" viewBox="0 0 1500 1000"><rect width="1500" height="1000" fill="#F8FAFC"/><text x="750" y="50" text-anchor="middle" font-family="Arial" font-size="30" font-weight="800">Regacore Business Model Canvas</text>'
for (t,txt),(x,y) in zip(items,pos):
 canvas+=f'<rect x="{x}" y="{y}" width="260" height="250" rx="16" fill="#FFFFFF" stroke="#CBD5E1"/><text x="{x+130}" y="{y+32}" text-anchor="middle" font-family="Arial" font-size="16" font-weight="800">{html.escape(t)}</text>'
 for j,line in enumerate(textwrap.wrap(txt,32)[:9]): canvas+=f'<text x="{x+20}" y="{y+68+j*22}" font-family="Arial" font-size="13" fill="#374151">{html.escape(line)}</text>'
canvas+='<text x="750" y="930" text-anchor="middle" font-family="Arial" font-size="14" fill="#374151">Canvas is reconstructed from public pages; unobserved contracts are inferred.</text></svg>'
(OUT/'business_model_canvas.svg').write_text(canvas,encoding='utf-8')
print('made support files',len(rows)-1,'features',len(shot),'screenshots')
