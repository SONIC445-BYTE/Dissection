from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Feature Inventory"

headers = [
    "Feature", "Purpose", "Evidence", "User Value", "Business Value",
    "Engineering Complexity", "Clinical Complexity", "Infrastructure Complexity", "Regulatory Complexity",
    "Estimated Team", "Estimated Months", "Priority", "Category", "Copy", "Improve", "Ignore", "Reinvent", "Moat", "Confidence"
]

features = [
    ["Google Health app rebrand from Fitbit", "Unify Fitbit app under Google Health brand and create health hub", "E04", "One destination for health and wellness data", "Brand consolidation; Premium upsell; Google ecosystem", "High", "Medium", "High", "Medium", "Mobile, backend, design, data migration, support", "12+", "High", "Consumer app", "Yes", "Yes", "No", "No", "Distribution/brand", "High"],
    ["Four-tab IA: Today/Fitness/Sleep/Health", "Simplify core navigation", "E04", "Find daily status, workouts, sleep, and health metrics faster", "Retention and discoverability", "Medium", "Low", "Low", "Low", "Product, design, mobile", "3-6", "High", "UX", "Yes", "Yes", "No", "No", "UX familiarity", "High"],
    ["Custom dashboards", "Let users pin favorite metrics", "E04", "Personalized view", "Engagement", "Medium", "Low", "Medium", "Low", "Mobile, backend", "3-6", "Medium", "UX", "Yes", "Yes", "No", "No", "Personalization", "High"],
    ["Google Health Coach", "AI coach across fitness, sleep, wellness", "E05", "Actionable guidance from data", "Premium subscription conversion/retention", "Very High", "Medium", "Very High", "Medium", "AI, mobile, backend, clinical, privacy, safety", "12+", "High", "AI coach", "Yes", "Yes", "No", "Yes", "AI/data", "High"],
    ["Coach onboarding conversation", "Capture goals, routines, equipment, injuries, lifestyle", "E05", "More relevant guidance", "Activation and personalization", "Medium", "Medium", "Medium", "Medium", "AI, UX, clinical safety", "3-6", "High", "AI coach", "Yes", "Yes", "No", "No", "Personalization", "High"],
    ["Ask Coach chat", "User-initiated health/wellness questions", "E05", "24/7 assistance", "Premium engagement", "High", "Medium", "High", "Medium", "AI, backend, safety", "6-12", "High", "AI interaction", "Yes", "Yes", "No", "Yes", "AI", "High"],
    ["Proactive Today insights", "Daily nudges and recommendations", "E05", "Timely action prompts", "DAU retention", "High", "Medium", "High", "Medium", "AI, mobile, notifications", "6-12", "High", "Retention", "Yes", "Yes", "No", "No", "Habit loop", "High"],
    ["Quick reply chips", "Reduce chat friction", "E05", "Faster coaching interactions", "Session depth", "Low", "Low", "Low", "Low", "UX, mobile", "1-2", "Medium", "AI UX", "Yes", "Yes", "No", "No", "UX", "High"],
    ["Weekly Plan", "Adaptive fitness plan", "E05", "Workout structure", "Premium differentiation", "High", "Medium", "Medium", "Medium", "AI, exercise science, mobile", "6-9", "High", "Fitness", "Yes", "Yes", "No", "Yes", "Personalization", "High"],
    ["Natural-language workout creation", "Create/save custom workouts from user text", "E05", "Less manual planning", "AI wow moment", "High", "Low", "Medium", "Low", "AI, mobile", "3-6", "Medium", "Fitness", "Yes", "Yes", "No", "No", "AI UX", "High"],
    ["Step-by-step workout guidance", "Guided workouts with visualizations and progress", "E05", "Execute plan correctly", "Engagement", "High", "Low", "Medium", "Low", "Mobile, content, design", "6-9", "Medium", "Fitness", "Yes", "Yes", "No", "No", "Content", "High"],
    ["Cardio Load", "Measure cardiovascular strain", "E05", "Balance training and recovery", "Wearable/Premium value", "Medium", "Medium", "Medium", "Low", "Data science, mobile", "3-6", "Medium", "Wearable metric", "Yes", "Yes", "No", "No", "Metric IP", "Medium"],
    ["HRV/recovery tracking", "Recovery insight from HRV and heart rate", "E05", "Know when to push/rest", "Engagement", "Medium", "Medium", "Medium", "Low", "Data science, wearable", "3-6", "Medium", "Wearable metric", "Yes", "Yes", "No", "No", "Data", "High"],
    ["Sleep stage tracking", "Light/deep/REM sleep breakdown", "E05", "Understand sleep quality", "Premium retention", "High", "Medium", "Medium", "Low", "ML, wearable, clinical validation", "6-12", "High", "Sleep", "Yes", "Yes", "No", "No", "Sensor/model", "High"],
    ["Sleep consistency coaching", "Weekly sleep feedback", "E05", "Improve rest habits", "Retention", "Medium", "Low", "Medium", "Low", "AI, sleep science", "3-6", "High", "Sleep", "Yes", "Yes", "No", "No", "Behavior loop", "High"],
    ["Cycle tracking + coach insights", "Connect cycle data to sleep/fitness/wellness", "E05", "Understand menstrual-health patterns", "Women’s health engagement", "High", "Medium", "Medium", "Medium", "Mobile, AI, clinical", "6-12", "High", "Women’s health", "Yes", "Yes", "No", "Yes", "Specialized dataset", "High"],
    ["Nutrition logging via voice/image", "Log meals with multimodal inputs", "E05", "Reduce food logging burden", "Premium value", "Very High", "Medium", "High", "Medium", "AI vision, nutrition DB, UX", "9-12", "High", "Nutrition", "Yes", "Yes", "No", "Yes", "Data loop", "High"],
    ["Mental wellbeing features", "Redesigned mental wellbeing in Coach experience", "E05", "Wellness support", "Broader engagement", "Medium", "Medium", "Medium", "Medium", "Clinical, UX, AI", "3-6", "Medium", "Wellbeing", "Yes", "Yes", "No", "No", "Engagement", "Medium"],
    ["Medical record sync/upload", "Import labs, meds, allergies, records", "E04/E05", "Central health record", "Differentiation and coach context", "Very High", "High", "High", "High", "FHIR, integrations, privacy, AI", "12+", "High", "PHR", "Yes", "Yes", "No", "Yes", "Data moat", "High"],
    ["Medical record summaries", "Simplify PHR and records", "E04/E05", "Understand complex medical data", "Premium value and trust", "High", "High", "High", "High", "AI, clinical, FHIR", "6-12", "High", "PHR/AI", "Yes", "Yes", "No", "Yes", "Clinical trust", "High"],
    ["Health data export/delete", "User control and portability", "E06", "Trust and control", "Regulatory/trust", "Medium", "Low", "Medium", "Medium", "Backend, privacy", "3-6", "High", "Privacy", "Yes", "Yes", "No", "No", "Trust", "High"],
    ["Optional feature toggles", "Turn sensitive features on/off", "E06", "Granular control", "Consent and trust", "Medium", "Low", "Medium", "Medium", "Product, backend", "3-6", "High", "Privacy", "Yes", "Yes", "No", "No", "Trust", "High"],
    ["No health data for Google Ads commitment", "Ads-use privacy covenant", "E06/E24", "Reassurance", "Regulatory approval and trust", "Low", "Low", "Medium", "High", "Legal, policy, data governance", "3-6", "High", "Trust", "Yes", "Yes", "No", "No", "Trust", "High"],
    ["Google Health Premium subscription", "Monetize Coach and advanced features", "E05", "Access advanced coaching", "Recurring revenue", "Medium", "Low", "Medium", "Low", "Payments, product", "3-6", "High", "Business model", "Yes", "Yes", "No", "No", "Monetization", "High"],
    ["AI Pro/Ultra bundling", "Bundle health Premium with AI subscriptions", "E05", "More value for existing subscribers", "Bundle retention", "Medium", "Low", "Medium", "Low", "Subscription/platform", "3-6", "Medium", "Pricing", "Yes", "Yes", "No", "No", "Distribution", "High"],
    ["Fitbit/Pixel wearable sync", "Continuous sensor data", "E03/E04", "Automatic tracking", "Hardware and data moat", "High", "Medium", "High", "Medium", "Device, app, backend", "12+", "High", "Wearables", "Yes", "Yes", "No", "No", "Data moat", "High"],
    ["Fitbit Air", "Screenless tracker optimized for Google Health", "E04/E05", "Comfortable 24/7 tracking", "Hardware + Premium attach", "Very High", "Medium", "High", "Medium", "Hardware, firmware, app", "18+", "Medium", "Hardware", "No", "Maybe", "No", "No", "Hardware distribution", "High"],
    ["Third-party app/device integrations", "Connect 100s of apps/devices", "E03/E46", "Holistic view", "Ecosystem lock-in", "High", "Low", "High", "Medium", "Integrations, API, partnerships", "12+", "High", "Integrations", "Yes", "Yes", "No", "Yes", "Network effects", "High"],
    ["Share/export via Health Connect or Google Health APIs", "Data portability to apps/coaches", "E46", "Use data elsewhere", "Platform adoption", "High", "Medium", "High", "High", "API, privacy, consent", "6-12", "High", "API", "Yes", "Yes", "No", "Yes", "Developer moat", "Medium"],
    ["Health Connect on-device encrypted store", "Android health data exchange", "E25", "Secure local sharing", "Platform leverage", "Very High", "Medium", "High", "High", "Android platform", "12+", "High", "Developer platform", "Yes", "Yes", "No", "No", "Platform moat", "High"],
    ["Health Connect granular permissions", "Per-data-type read/write access", "E07/E09", "Transparency/control", "Developer trust", "High", "Medium", "Medium", "High", "Android platform", "6-12", "High", "Permissions", "Yes", "Yes", "No", "No", "Trust", "High"],
    ["Health Connect data-source priority", "Prioritize sources when multiple apps write data", "E25", "Cleaner data", "Data quality", "Medium", "Low", "Medium", "Low", "Android platform", "3-6", "Medium", "Data quality", "Yes", "Yes", "No", "Yes", "Data quality", "High"],
    ["Health Connect Medical Records API", "FHIR-based medical record read/write", "E07/E08", "Unify clinical records with daily data", "Developer ecosystem expansion", "Very High", "High", "High", "Very High", "FHIR, Android, policy", "12+", "High", "Medical records", "Yes", "Yes", "No", "Yes", "Platform moat", "High"],
    ["Medical Records permissions screen", "Separate sensitive permissions UX", "E07", "Understand medical data access", "Trust/compliance", "Medium", "Medium", "Medium", "High", "Android UX/policy", "3-6", "High", "Permissions", "Yes", "Yes", "No", "No", "Trust", "High"],
    ["Medical Records browser", "Browse records stored in Health Connect", "E07", "View imported medical data", "User trust and utility", "Medium", "Medium", "Medium", "Medium", "Android UX", "3-6", "Medium", "Medical records", "Yes", "Yes", "No", "No", "Utility", "High"],
    ["FHIR R4/R4B medical categories", "Allergies, conditions, labs, meds, visits, vitals", "E08", "Structured medical data", "Developer adoption", "High", "High", "Medium", "High", "FHIR, terminology", "6-12", "High", "Data model", "Yes", "Yes", "No", "No", "Interoperability", "High"],
    ["Play health apps declaration", "Govern sensitive health access", "E09/E34", "Privacy assurance", "Platform governance", "Medium", "Medium", "Medium", "High", "Policy, console", "6-12", "High", "Governance", "Yes", "Yes", "No", "No", "Regulatory", "High"],
    ["Cloud Healthcare API datasets/stores", "Managed healthcare data stores", "E10", "Interoperable backend", "Cloud revenue and enterprise lock-in", "Very High", "High", "Very High", "High", "Cloud platform", "12+", "High", "Cloud", "Yes", "Yes", "No", "No", "Infrastructure moat", "High"],
    ["FHIR store", "Clinical resource storage/search", "E10", "FHIR interoperability", "Cloud adoption", "Very High", "High", "High", "High", "Cloud/FHIR", "12+", "High", "Cloud API", "Yes", "Yes", "No", "No", "Interoperability", "High"],
    ["HL7v2 store", "Clinical event message storage", "E10", "Legacy hospital integration", "Enterprise adoption", "High", "High", "High", "High", "Cloud/HL7", "9-12", "High", "Cloud API", "Yes", "Yes", "No", "No", "Enterprise moat", "High"],
    ["DICOM/DICOMweb store", "Medical imaging storage/exchange", "E10", "Imaging workflows", "Cloud imaging workloads", "Very High", "High", "Very High", "High", "Cloud/imaging", "12+", "High", "Cloud API", "Yes", "Yes", "No", "No", "Enterprise moat", "High"],
    ["Bulk import/export", "Move FHIR/DICOM data to/from Cloud Storage", "E10", "Migration and pipelines", "Cloud workload growth", "High", "Medium", "High", "Medium", "Cloud data engineering", "6-12", "High", "Cloud API", "Yes", "Yes", "No", "No", "Cloud stickiness", "High"],
    ["BigQuery export/streaming", "Analytics over healthcare data", "E10", "Population/analytics insights", "BigQuery/AI pull-through", "High", "Medium", "High", "High", "Cloud data", "6-12", "High", "Analytics", "Yes", "Yes", "No", "No", "Data platform", "High"],
    ["Pub/Sub notifications", "Event-driven workflows", "E13", "React to clinical events", "Cloud architecture adoption", "Medium", "Medium", "Medium", "Medium", "Cloud events", "3-6", "Medium", "Cloud API", "Yes", "Yes", "No", "No", "Integration", "High"],
    ["Data de-identification", "Redact sensitive data for research", "E10", "Research utility with privacy", "Research/analytics adoption", "High", "High", "High", "High", "Cloud, DLP, clinical", "6-12", "High", "Privacy", "Yes", "Yes", "No", "No", "Compliance", "High"],
    ["IAM access control", "Fine-grained enterprise permissions", "E10", "Secure PHI access", "Enterprise compliance", "High", "Medium", "High", "High", "Cloud IAM", "6-12", "High", "Security", "Yes", "Yes", "No", "No", "Trust", "High"],
    ["Cloud audit logs", "Audit admin/data access", "E13", "Compliance and investigation", "Enterprise trust", "Medium", "Medium", "Medium", "High", "Cloud logging", "3-6", "High", "Security", "Yes", "Yes", "No", "No", "Compliance", "High"],
    ["CMEK", "Customer-controlled encryption keys", "E12", "Key governance", "Regulated enterprise adoption", "High", "Low", "High", "High", "Cloud KMS", "3-6", "Medium", "Security", "Yes", "Yes", "No", "No", "Compliance", "High"],
    ["HIPAA BAA coverage", "Enable ePHI workloads", "E10/E14", "Enterprise compliance path", "Healthcare cloud sales", "Medium", "Medium", "High", "Very High", "Legal, cloud security", "6-12", "High", "Compliance", "Yes", "Yes", "No", "No", "Regulatory", "High"],
    ["Vertex AI Search for Healthcare", "Clinical search and Q&A", "E02/E17", "Find data faster", "Cloud AI revenue", "Very High", "High", "Very High", "High", "AI, search, cloud, clinical", "12+", "High", "Clinical AI", "Yes", "Yes", "No", "Yes", "AI + data", "High"],
    ["Grounded answers with citations", "Reduce hallucination risk", "E17/E18", "Clinician confidence", "Enterprise trust", "High", "High", "High", "High", "RAG, UX, audit", "6-12", "High", "Clinical AI", "Yes", "Yes", "No", "No", "Trust", "Medium"],
    ["Visual Q&A / multimodal clinical search", "Query charts/tables/images", "E18", "Comprehensive patient view", "Differentiation", "Very High", "High", "Very High", "High", "Multimodal AI", "12+", "High", "Clinical AI", "Yes", "Yes", "No", "Yes", "AI", "High"],
    ["MedLM", "Healthcare-tuned foundation models on Vertex AI", "E16", "Healthcare LLM tasks", "Cloud AI revenue", "Very High", "High", "Very High", "High", "AI platform", "12+", "High", "Clinical AI", "Yes", "Yes", "No", "No", "AI moat", "High"],
    ["MedGemma", "Open medical text/image model", "E19", "Build medical AI apps", "Developer ecosystem", "Very High", "High", "High", "Medium", "AI research/platform", "12+", "High", "Open model", "Yes", "Yes", "No", "No", "Developer moat", "High"],
    ["TxGemma", "Therapeutics open models", "E19/E20", "Drug discovery predictions", "Life-sciences ecosystem", "Very High", "High", "High", "Medium", "AI research", "12+", "Medium", "Open model", "Maybe", "Maybe", "No", "No", "AI moat", "High"],
    ["MedASR", "Medical audio ASR", "E19", "Dictation/notes", "Healthcare AI ecosystem", "High", "Medium", "Medium", "Medium", "AI speech", "6-12", "Medium", "Open model", "Yes", "Yes", "No", "No", "Developer", "High"],
    ["MedSigLIP", "Medical image encoder", "E19", "Image classification/retrieval", "Model ecosystem", "High", "High", "Medium", "Medium", "AI vision", "6-12", "Medium", "Open model", "Yes", "Yes", "No", "No", "AI", "High"],
    ["Open Health Stack Android FHIR SDK", "Offline-capable FHIR-native Android apps", "E41", "Mobile healthcare apps in low-resource contexts", "Developer goodwill/ecosystem", "High", "High", "Medium", "Medium", "Kotlin/FHIR", "12+", "Medium", "Developer platform", "Yes", "Yes", "No", "No", "Developer moat", "High"],
    ["Search AI Overviews / AI Mode for health", "Answer health questions in Search", "E01", "Health info access", "Search engagement/trust", "Very High", "High", "Very High", "High", "Search/AI", "12+", "High", "Search", "No", "Maybe", "No", "No", "Distribution", "High"],
    ["Lens skin-condition visual search", "Search visually similar skin conditions", "E01", "Visual health query", "Search utility", "High", "High", "High", "High", "Computer vision/Search", "12+", "Medium", "Search", "No", "Maybe", "No", "No", "Distribution", "High"],
    ["YouTube health content shelf", "Highlight authoritative health content", "E01", "Find credible videos", "YouTube trust/engagement", "Medium", "Medium", "Medium", "Medium", "Ranking, policy, partnerships", "6-12", "Medium", "Content", "No", "Maybe", "No", "No", "Distribution", "High"],
    ["YouTube personal stories shelf", "Surface lived-experience videos", "E01", "Support and stigma reduction", "Engagement", "Medium", "Medium", "Medium", "Medium", "Ranking/policy", "6-12", "Low", "Content", "No", "Maybe", "No", "No", "Community", "High"],
    ["First aid shelf", "Quick first aid videos", "E01", "Immediate help info", "Public-good trust", "Medium", "High", "Medium", "High", "Content partnerships", "6-12", "Medium", "Content", "No", "Maybe", "No", "No", "Trust", "High"],
    ["Google Wallet health passes", "Store insurance/vaccination/private passes", "E01", "Convenient access", "Wallet adoption", "Medium", "Low", "Medium", "Medium", "Wallet, privacy", "3-6", "Low", "Wallet", "Maybe", "Maybe", "No", "No", "Ecosystem", "High"],
    ["AMIE", "Conversational diagnostic AI research", "E02/E21", "Future clinical dialogue support", "Research leadership", "Very High", "Very High", "Very High", "Very High", "AI research, clinical eval", "24+", "Medium", "Research", "No", "Maybe", "No", "Yes", "AI moat", "High"],
    ["SensorFM/LSM", "Wearable foundation model", "E22", "Personalized health signals", "Future health AI moat", "Very High", "High", "Very High", "High", "AI research, sensors", "24+", "High", "Research", "No", "Maybe", "No", "Yes", "Data moat", "High"],
    ["PH-LLM", "Personal health LLM over sensor data", "E02", "Sleep/fitness insights", "Coach roadmap", "Very High", "Medium", "High", "Medium", "AI research", "12+", "High", "Research", "Yes", "Yes", "No", "Yes", "AI/data", "Medium"],
    ["AlphaFold", "Protein structure / drug discovery research", "E02", "Scientific discovery", "Research/life sciences leadership", "Very High", "Very High", "Very High", "Medium", "DeepMind research", "24+", "Low", "Research", "No", "Maybe", "No", "No", "AI/science", "High"],
    ["DeepVariant", "Genomic variant calling", "E28", "More accurate genomic analysis", "Research/developer goodwill", "High", "High", "High", "Medium", "AI genomics", "12+", "Low", "Research/open source", "Maybe", "Maybe", "No", "No", "AI science", "High"],
]

ws.append(headers)
for row in features:
    ws.append(row)

# Evidence register
sources = [
    ["E01", "Google for Health - Products", "https://health.google/products/", "Search AI, YouTube Health, Pixel/Fitbit, Health Connect, Medical Records API, Wallet, privacy messaging", "Official product page", "Public images on page", "High", "Observed"],
    ["E02", "Google for Health - AI Models", "https://health.google/ai-models/", "Gemini, MedGemma, TxGemma, AlphaFold, AMIE, LSM/PH-LLM, Vertex AI Search for Healthcare", "Official AI page", "Public images on page", "High", "Observed"],
    ["E03", "Google Health app marketing site", "https://healthapp.google/", "Google Health app brings wearables, apps, devices, and medical records into one place; Coach/Premium; privacy promises", "Official Google-owned domain", "Public images on page", "High", "Observed"],
    ["E04", "Introducing the Google Health app", "https://blog.google/products-and-platforms/products/google-health/google-health-app/", "Fitbit app becomes Google Health app; four tabs; data connections; medical records; app update; disclaimers", "Google blog", "Blog images", "High", "Observed"],
    ["E05", "Google Health Coach launch", "https://blog.google/products-and-platforms/products/google-health/google-health-coach/", "Gemini-built Coach; onboarding; proactive insights; data ecosystem; SHARP; pricing $9.99/mo $99/yr", "Google blog", "Blog images", "High", "Observed"],
    ["E06", "Google Health Privacy", "https://healthapp.google/privacy/", "No Fitbit health/wellness data for Google Ads; encryption; export/delete; optional features; 2FA", "Official privacy page", "Public image", "High", "Observed"],
    ["E07", "Health Connect Medical Records", "https://developer.android.com/health-and-fitness/health-connect/medical-records", "Medical Records APIs, FHIR, permissions screen, data browsing, experimental API", "Android docs", "Official permissions/browsing images", "High", "Observed"],
    ["E08", "Health Connect Medical Records data format", "https://developer.android.com/health-and-fitness/health-connect/medical-records/data-format", "FHIR R4/R4B, medical resource types and permission declarations", "Android docs", "Not captured—public URL provided", "High", "Observed"],
    ["E09", "Health Connect data types", "https://developer.android.com/health-and-fitness/health-connect/data-types", "Health/fitness and medical record data types; Play declaration requirement", "Android docs", "Not captured—public URL provided", "High", "Observed"],
    ["E10", "Cloud Healthcare API overview", "https://docs.cloud.google.com/healthcare-api/docs/introduction", "FHIR/HL7v2/DICOM, compliance, IAM, BigQuery, de-ID, audit, durability", "Google Cloud docs", "Not captured—public URL provided", "High", "Observed"],
    ["E11", "Cloud Healthcare API pricing", "https://cloud.google.com/healthcare-api/pricing", "Storage/request/notification/ETL/de-identification pricing", "Google Cloud pricing", "Not captured—public URL provided", "High", "Observed"],
    ["E12", "Cloud Healthcare API CMEK", "https://docs.cloud.google.com/healthcare-api/docs/cmek", "CMEK support and default Google-managed encryption", "Google Cloud docs", "Not captured—public URL provided", "High", "Observed"],
    ["E13", "Cloud Healthcare API audit logging", "https://docs.cloud.google.com/healthcare-api/docs/how-tos/audit-logging", "Cloud Audit Logs methods and custom headers", "Google Cloud docs", "Not captured—public URL provided", "High", "Observed"],
    ["E14", "Google Cloud HIPAA compliance", "https://cloud.google.com/security/compliance/hipaa-compliance", "BAA and HIPAA coverage for Google Cloud products", "Google Cloud compliance", "Not captured—public URL provided", "High", "Observed"],
    ["E15", "Healthcare Data Engine deprecation", "https://cloud.google.com/healthcare-data-engine/pricing", "Healthcare Data Engine deprecated and unavailable effective July 11, 2026", "Google Cloud docs", "Not captured—public URL provided", "High", "Observed"],
    ["E16", "Introducing MedLM", "https://cloud.google.com/blog/topics/healthcare-life-sciences/introducing-medlm-for-the-healthcare-industry", "MedLM models based on Med-PaLM 2 on Vertex AI", "Google Cloud blog", "Not captured—public URL provided", "High", "Observed"],
    ["E17", "Vertex AI Search healthcare coverage", "https://www.fiercehealthcare.com/ai-and-machine-learning/google-cloud-building-out-generative-ai-tools-lighten-load-healthcare", "GA/grounded clinical search, Gemini/MedLM integration, citations, FHIR/notes", "Press coverage", "Not captured—public URL provided", "Medium", "Observed"],
    ["E18", "Google Cloud HIMSS 2025 healthcare AI", "https://blog.google/innovation-and-ai/infrastructure-and-cloud/google-cloud/himss-2025/", "Counterpart, MEDITECH, Suki use Vertex AI Search for healthcare", "Google blog", "Not captured—public URL provided", "High", "Observed"],
    ["E19", "Health AI Developer Foundations", "https://developers.google.com/health-ai-developer-foundations", "MedGemma, MedASR, MedSigLIP, TxGemma, HeAR, Path Foundation; Vertex/Cloud Healthcare integration", "Google Developers", "Public images/GIF", "High", "Observed"],
    ["E20", "TxGemma/MedGemma blog", "https://developers.google.com/health-ai-developer-foundations/blog", "TxGemma and MedGemma HAI-DEF updates", "Google Developers", "Not captured—public URL provided", "High", "Observed"],
    ["E21", "AMIE papers/search results", "https://arxiv.org/abs/2401.05654", "AMIE conversational diagnostic AI, self-play, OSCE studies, limitations", "arXiv/Nature search results", "Not captured—public URL provided", "Medium", "Observed"],
    ["E22", "SensorFM Google Research", "https://research.google/blog/sensorfm-towards-a-general-intelligence-and-interface-for-wearable-health-data/", "SensorFM trained on over one trillion minutes from five million consented participants", "Google Research blog", "Not captured—public URL provided", "High", "Observed"],
    ["E24", "Fitbit acquisition", "https://techcrunch.com/2021/01/14/googles-fitbit-acquisition-is-official/", "$2.1B Fitbit acquisition and commitments not to use Fitbit data for ads", "Press coverage", "Not captured—public URL provided", "Medium", "Observed"],
    ["E25", "Health Connect introduction", "https://android-developers.googleblog.com/2022/05/introducing-health-connect.html", "On-device encrypted Health Connect; granular permissions; delete/prioritize data; Samsung/Fitbit/Fit adoption", "Android Developers blog", "Not captured—public URL provided", "High", "Observed"],
    ["E27", "Google mammography AI / Imperial NHS", "https://www.imperial.ac.uk/news/articles/global-health-innovation/2026/new-research-conducted-using-google-ai-can-match-or-exceed-radiologists-in-detecting-cancer-in-breast-scans-/", "AI mammography research with Imperial/NHS, 25% interval cancers, workload reductions", "Imperial public news", "Not captured—public URL provided", "Medium", "Observed"],
    ["E28", "Google genomics DeepVariant", "https://health.google/genomics/", "DeepVariant improves variant-calling accuracy and won PrecisionFDA challenge categories", "Google Health genomics page", "Not captured—public URL provided", "High", "Observed"],
    ["E29", "Google Health dismantled", "https://www.businessinsider.com/google-health-shutting-down-david-feinberg-leaves-2021-8", "Google Health division dismantled; teams spread across company", "Press coverage", "Not captured—public URL provided", "Medium", "Observed"],
    ["E30", "Fierce Healthcare division unwound", "https://www.fiercehealthcare.com/tech/google-dissolved-its-unified-health-division-what-s-next-for-its-health-tech-strategy", "Company-wide health effort, no single health division", "Press coverage", "Not captured—public URL provided", "Medium", "Observed"],
    ["E31", "Alphabet Q2 2026 SEC results", "https://www.sec.gov/Archives/edgar/data/1652044/000165204426000066/googexhibit991q22026.htm", "Alphabet segment revenues; Google Cloud Q2 2026 revenue", "SEC filing/exhibit", "Not captured—public URL provided", "High", "Observed"],
    ["E32", "Karen DeSalvo retirement / Michael Howell", "https://www.healthcaredive.com/news/karen-desalvo-google-chief-health-officer-retire/747201/", "Michael Howell to lead healthcare initiatives after DeSalvo retirement", "Healthcare Dive", "Not captured—public URL provided", "Medium", "Observed"],
    ["E33", "Original Google Health failure", "https://www.computerworld.com/article/1534713/why-google-health-failed-too-little-too-soon.html", "2008 Google Health PHR failed due low adoption and limited capabilities", "Press analysis", "Not captured—public URL provided", "Medium", "Observed"],
    ["E34", "Google Play Android Health Permissions", "https://support.google.com/googleplay/android-developer/answer/12991134?hl=en", "Forbidden uses: ads, sale/transfer, credit/insurance/employment/lending decisions", "Google Play policy", "Not captured—public URL provided", "High", "Observed"],
    ["E35", "Google Health shutdown reporting", "https://techcrunch.com/2011/06/24/google-shuts-down-medical-records-and-health-data-platform/", "Google Health shut down Jan 2012; data export through Jan 2013; broad adoption not achieved", "Press coverage", "Not captured—public URL provided", "Medium", "Observed"],
    ["E36", "David Feinberg hired", "https://www.cnbc.com/2018/11/08/google-hires-geisinger-ceo-david-feinberg-to-oversee-health.html", "Feinberg hired to coordinate Google health strategy", "CNBC", "Not captured—public URL provided", "Medium", "Observed"],
    ["E37", "DeepMind Health to Google Health", "https://www.moorfields.nhs.uk/research/google-deepmind", "DeepMind Health moved to Google Health; Moorfields partnership transferred", "Moorfields NHS", "Not captured—public URL provided", "High", "Observed"],
    ["E38", "Project Nightingale scrutiny", "https://www.healthcaredive.com/news/googles-project-nightingale-prompts-hhs-investigation/567078/", "HHS OCR inquiry into Google/Ascension Project Nightingale", "Healthcare Dive", "Not captured—public URL provided", "Medium", "Observed"],
    ["E39", "Karen DeSalvo public role", "https://blog.google/authors/karen-desalvo/", "Chief Health Officer profile and Google Health mission", "Google blog author page", "Not captured—public URL provided", "High", "Observed"],
    ["E40", "Rishi Chandra/wearables public interview", "https://www.wired.com/story/google-is-rebranding-the-fitbit-app-to-google-health/", "Rishi Chandra on Google Health/Fitbit and coach/hardware", "Wired", "Not captured—public URL provided", "Medium", "Observed"],
    ["E41", "Open Health Stack / Android FHIR SDK", "https://developers.google.com/open-health-stack/android-fhir", "Kotlin libraries for offline-capable FHIR-native Android apps; WHO SMART Guidelines", "Google Developers", "Not captured—public URL provided", "High", "Observed"],
    ["E42", "Google patent WO2019022779A1", "https://patents.google.com/patent/WO2019022779A1/en", "Google LLC patent: predicting/summarizing medical events from EHRs using standardized/FHIR data and deep learning", "Google Patents", "Patent figures public", "High", "Observed"],
    ["E43", "MEDITECH/Care Studio integration", "https://www.healthcaredive.com/news/google-meditech-ehr-hospital-himss/620393/", "Google clinical tools integrated into MEDITECH Expanse / Care Studio lineage", "Healthcare Dive", "Not captured—public URL provided", "Medium", "Observed"],
    ["E44", "The Check Up official", "https://health.google/the-check-up/", "Official event page with HCA, MEDITECH, WHO, Princess Maxima, Apollo, health AI updates", "Google Health event page", "Public video/page", "High", "Observed"],
    ["E45", "Fitbit AFib FDA clearance", "https://blog.google/products-and-platforms/devices/fitbit/irregular-heart-rhythm-notifications/", "FDA clearance of Fitbit PPG AFib algorithm/irregular rhythm notifications", "Google blog", "Not captured—public URL provided", "High", "Observed"],
    ["E46", "Google Health app data connectivity blog", "https://blog.google/products-and-platforms/products/google-health/connect-data-across-devices/", "Google Health APIs, Health Connect, Apple Health, TCX export, third-party data sharing", "Google blog", "Not captured—public URL provided", "High", "Observed"],
    ["E47", "Mashable Fitbit user backlash", "https://mashable.com/tech/fitbit-users-complain-about-new-google-health-app", "User backlash to forced Google Health app migration and AI coach", "Press/customer intelligence", "Not captured—public URL provided", "Medium", "Observed"],
    ["E48", "Reddit Fitbit complaints", "https://www.reddit.com/r/fitbit/comments/1tn2x4c/beyond_frustrated_with_the_forced_google_health/", "Complaints about AI coach, logging errors, forced update", "Public forum", "Not captured—public URL provided", "Low-Medium", "Observed"],
    ["E49", "Lifehacker AI Health Coach review", "https://lifehacker.com/health/google-health-app-is-replacing-the-fitbit-app-starting", "AI coach hallucination/inconsistency and Google Health app review", "Press review", "Not captured—public URL provided", "Medium", "Observed"],
    ["E50", "Healthcare providers in Search/Maps", "https://9to5google.com/2021/12/02/google-search-maps-healthcare-providers-insurance/", "Provider profiles show insurance/languages/booking-like info", "Press coverage", "Not captured—public URL provided", "Medium", "Observed"],
    ["E51", "FHIR consent/access control", "https://docs.cloud.google.com/healthcare-api/docs/fhir-consent", "Cloud Healthcare API FHIR consent/access concepts and audit logs", "Google Cloud docs", "Not captured—public URL provided", "High", "Observed"],
    ["E52", "Google Cloud Applied AI PM job", "https://www.google.com/about/careers/applications/jobs/results/133846090876101318-senior-product-manager-applied-ai-google-cloud", "FHIR/HL7/EHR/HIPAA/GxP applied AI PM skills", "Google Careers search result", "Not captured—public URL provided", "Medium", "Observed"],
    ["E53", "AI Research Health Clinical Specialist job", "https://jobs.anitab.org/companies/google-24698/jobs/55897887-ai-research-health-clinical-specialist", "Clinical AI health research role with MD/AI/generative AI/product/regulatory collaboration", "Job mirror", "Not captured—public URL provided", "Medium", "Observed"],
    ["E54", "Google health tech jobs pay coverage", "https://www.beckershospitalreview.com/disruptors/googles-pay-for-7-health-tech-jobs-2/", "AI evaluation, women’s health sensing, medically regulated Fitbit/product support jobs", "Becker's", "Not captured—public URL provided", "Medium", "Observed"],
    ["E55", "G2 Google Cloud Healthcare API reviews", "https://www.g2.com/products/google-cloud-healthcare-api/reviews", "Praise and complaints: standards, analytics, cost, complexity, cloud dependency", "G2 review snippets", "Not captured—public URL provided", "Low-Medium", "Observed"],
    ["E56", "Health Connect developer complaints", "https://www.reddit.com/r/androiddev/comments/1f6fyjn/where_is_the_new_health_connect_approval_process/", "Developer frustration with Health Connect approval process", "Reddit/StackOverflow snippets", "Not captured—public URL provided", "Low-Medium", "Observed"],
    ["E57", "OpenEvidence Reuters", "https://www.reuters.com/business/healthcare-pharmaceuticals/medical-ai-startup-openevidence-doubles-valuation-12-billion-latest-round-2026-01-21/", "$250M Series D at $12B; >40% US physicians daily; 10,000 hospitals/medical centers", "Reuters", "Not captured—public URL provided", "Medium", "Observed"],
    ["E58", "Function Health pricing sources", "https://crowncounseling.com/reviews/function-health-vs-superpower-vs-outlive/", "Function Health pricing/biomarker comparisons", "Review/market source", "Not captured—public URL provided", "Low-Medium", "Observed"],
    ["E59", "Superpower pricing sources", "https://crowncounseling.com/reviews/superpower-health-review/", "Superpower $199/yr, 100+ biomarkers, AI/care team", "Review/market source", "Not captured—public URL provided", "Low-Medium", "Observed"],
    ["E60", "Levels CGM review", "https://optimizebiomarkers.com/cgm-providers/levels", "Levels CGM/metabolic features and pricing tiers", "Review/market source", "Not captured—public URL provided", "Low", "Observed"],
    ["E61", "Atropos/Merck collaboration", "https://www.businesswire.com/news/home/20250110445683/en/Atropos-Health-Collaborates-with-Merck-for-Rapid-Evidence-Generation-to-Accelerate-Innovation-for-Life-Saving-Treatments", "Atropos GENEVA OS, Green Button, ChatRWD, RWE in minutes/under 48h", "BusinessWire", "Not captured—public URL provided", "Medium", "Observed"],
    ["E62", "UpToDate Expert AI", "https://www.wolterskluwer.com/en/expert-insights/reimagine-trust-for-todays-healthcare-with-uptodate-expert-ai", "GenAI grounded in UpToDate content with traceable reasoning/citations", "Wolters Kluwer", "Not captured—public URL provided", "High", "Observed"],
    ["E63", "AMBOSS clinician page", "https://www.amboss.com/us/clinicians/attendings-apps", "AI Mode Clinical Care, clinical decision support, clinician-curated library", "AMBOSS", "Not captured—public URL provided", "High", "Observed"],
    ["E64", "Apollo 24/7", "https://www.apollo247.com/", "Online pharmacy, doctor consultations, lab tests, digital vault", "Official Apollo", "Public site", "High", "Observed"],
    ["E65", "Tata 1mg AI/diagnostics reporting", "https://medicalbuyer.co.in/tata-1mg-diagnostics-biz-crosses-%E2%82%B9600cr-invests-in-ai-led-healthcare-tools/", "Tata 1mg diagnostics, AI Pulse, Health Insights Hub, Family Hub", "Press coverage", "Not captured—public URL provided", "Medium", "Observed"],
    ["E66", "Oura features", "https://ouraring.com/blog/the-oura-difference/", "Oura tracks sleep, stress, temp, glucose integration, Oura Advisor", "Oura official", "Public images", "High", "Observed"],
    ["E67", "WHOOP Coach", "https://support.whoop.com/s/article/How-to-Use-the-AI-Powered-WHOOP-Coach?language=en_US", "WHOOP AI coaching, recovery/sleep/strain/environmental data", "WHOOP support", "Not captured—public URL provided", "High", "Observed"],
    ["E68", "Ultrahuman AI/metabolic", "https://techgenyz.com/ultrahuman-m2-live-metabolic-health-platform-launch/", "Ultrahuman real-time AI, glucose, sleep, HRV, recovery", "Press coverage", "Not captured—public URL provided", "Low-Medium", "Observed"],
    ["E69", "PreventiveHealth.ai", "https://preventivehealth.ai/", "Healthspan membership, lifestyle/genetics/wearables/blood/microbiome inputs", "Official site", "Public site", "High", "Observed"],
]

ws2 = wb.create_sheet("Evidence Register")
ws2.append(["ID", "Source", "URL", "Evidence", "Source Type", "Screenshot/Public Image", "Confidence", "Observed vs Inferred"])
for row in sources:
    ws2.append(row)

# Decision Ledger sheet
ws3 = wb.create_sheet("Decision Ledger")
ws3.append(["Feature", "Why built", "Pain solved", "KPI", "Tradeoff", "Alternative", "Evidence", "Confidence"])
ledger_rows = [
    ["Google Health app rebrand", "Unify Fitbit under Google Health", "Brand/data fragmentation", "Premium attach; MAU", "Legacy backlash", "Keep Fitbit and Google Health separate", "E04/E47", "High"],
    ["Health Coach", "Interpret wearable/record data", "Users do not know what to do with data", "Subscription conversion; DAU", "Hallucination/overreliance", "Human coaching marketplace", "E05/E49", "High"],
    ["Medical record sync", "Give Coach clinical context", "Records split across portals", "Coach usefulness; differentiation", "Privacy/regulatory anxiety", "PDF-only upload", "E04/E05", "High"],
    ["Health Connect Medical Records", "Platform medical data exchange", "One-off integrations", "Developer adoption", "Experimental API instability", "Proprietary Google Health API", "E07/E08", "High"],
    ["Cloud Healthcare API", "Enterprise interoperability", "FHIR/HL7/DICOM fragmentation", "Cloud revenue; enterprise lock-in", "Complexity/cost", "Self-hosted FHIR server", "E10/E11/E55", "High"],
    ["Vertex AI Search for Healthcare", "Clinical search/Q&A", "Chart review burden", "Cloud AI usage; clinician workflow", "Clinical liability", "EHR-native search only", "E02/E17/E18", "High"],
    ["HAI-DEF open models", "Seed developer ecosystem", "Hard to build medical AI from scratch", "Model adoption; cloud pull-through", "Commoditizes base models", "Closed MedLM only", "E19", "High"],
    ["No-health-data-for-ads commitment", "Protect trust/regulatory approval", "User/regulator fear", "Consent; retention", "Limits ad monetization", "Ad monetization", "E06/E24", "High"],
    ["Proactive Today insights", "Daily habit loop", "Users forget to act", "DAU", "AI clutter", "Pull-only insights", "E05/E47", "High"],
    ["Nutrition image/voice logging", "Lower logging friction", "Manual food logging burden", "Logs/day", "Parsing errors", "Manual database only", "E05/E47", "Medium"],
]
for row in ledger_rows:
    ws3.append(row)

# Moat Matrix
ws4 = wb.create_sheet("Moat Matrix")
ws4.append(["Moat", "Strength", "Evidence", "Notes"])
moat_rows = [
    ["Data moat", "Strong/Future", "E03/E07/E22", "Fitbit/Pixel/Health Connect/Search/medical records/SensorFM"],
    ["AI moat", "Strong", "E02/E16/E19/E21/E22", "Gemini, MedLM, MedGemma, AMIE, SensorFM, AlphaFold"],
    ["Distribution moat", "Very Strong", "E01/E04/E05/E07/E31", "Search, YouTube, Android, Fitbit migration, Cloud"],
    ["Developer moat", "Strong", "E07/E10/E19/E41", "Health Connect, Cloud Healthcare API, HAI-DEF, Open Health Stack"],
    ["Clinical moat", "Medium", "E05/E18/E21/E27", "Partnerships and panels; not care delivery"],
    ["Trust moat", "Medium", "E06/E24/E38/E47", "Strong commitments but history/backlash"],
]
for row in moat_rows:
    ws4.append(row)

# Risk Register
ws5 = wb.create_sheet("Risk Register")
ws5.append(["Risk", "Severity", "Likelihood", "Evidence", "Google Mitigation", "Ovexis Countermove"])
risk_rows = [
    ["AI hallucination", "High", "High", "E49", "Disclaimers, SHARP, expert panel", "Citations, confidence, human review, undo"],
    ["Forced AI backlash", "High", "Medium", "E47/E48", "Roadmap/fixes", "AI optional/classic mode"],
    ["Privacy trust erosion", "High", "Medium", "E24/E38", "No ads, consent, encryption", "Independent audits; no employer/insurer covenant"],
    ["Developer approval friction", "Medium", "Medium", "E56", "Play declarations", "Concierge integration; clear docs"],
    ["Cloud complexity/cost", "Medium", "Medium", "E55", "Docs/pricing calculator", "Transparent packages"],
    ["Regulatory boundary", "High", "Medium", "E04/E05", "Not medical purpose disclaimers", "Feature-tier FDA/clinical strategy"],
]
for row in risk_rows:
    ws5.append(row)

# Competitive Landscape
ws6 = wb.create_sheet("Competitive Landscape")
ws6.append(["Competitor", "Category", "Confirmed evidence", "Google advantage", "Google weakness", "Ovexis opening"])
comp_rows = [
    ["Apple Health", "OS health app", "E46", "Google has cross-Android/Cloud/Search and shipped Coach", "Apple trust/hardware integration strong", "Cross-platform, doctor-ready records"],
    ["Oura", "Smart ring/recovery", "E66", "Google has app + Search + Android + Cloud", "Oura premium ring/sleep trust", "Deep longitudinal clinical layer"],
    ["Whoop", "Performance wearable/AI coach", "E67", "Google has broader ecosystem", "Whoop stronger athlete/performance identity", "Clinical-grade preventive intelligence"],
    ["Ultrahuman", "Ring + metabolic", "E68", "Google has broader platform", "Ultrahuman glucose/metabolic focus", "Lab/CGM/wearable fusion"],
    ["Function Health", "Lab membership", "E58", "Google has data/app/AI distribution", "No first-party lab loop", "Own biomarker-to-action loop"],
    ["Superpower", "Lab + AI health", "E59", "Google distribution and models", "No lab ordering; wellness disclaimers", "Affordable longitudinal preventive care"],
    ["Levels", "CGM metabolic", "E60", "Google can integrate CGM partners", "No first-party metabolic program", "Metabolic health OS"],
    ["OpenEvidence", "Physician AI search", "E57", "Google has Cloud/Gemini/EHR integrations", "OpenEvidence owns physician network/trust", "Patient+doctor shared intelligence"],
    ["UpToDate Expert AI", "Clinical reference", "E62", "Google has infrastructure/models", "UpToDate has trusted editorial content", "Personalized evidence + records"],
    ["Atropos", "RWE evidence", "E61", "Google has BigQuery/Cloud", "Atropos owns workflow/IP", "Personal cohort evidence layer"],
    ["Apollo 24/7", "India health super-app", "E64", "Google has platform/AI", "Apollo has care delivery/pharmacy/labs", "India care network partnerships"],
    ["Tata 1mg", "India pharmacy/diagnostics", "E65", "Google has AI/platform", "1mg has local healthcare commerce", "Health intelligence over Indian care stack"],
    ["PreventiveHealth.ai", "India preventive AI", "E69", "Google has scale", "Google not India-personalized healthspan", "India-first healthspan with labs/genomics"],
    ["Regacore", "Unknown", "No public verified source found", "Cannot assess", "Cannot assess", "Monitor only"],
]
for row in comp_rows:
    ws6.append(row)

# Format all sheets
for sheet in wb.worksheets:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col in range(1, sheet.max_column + 1):
        max_len = 0
        for row in range(1, min(sheet.max_row, 80) + 1):
            val = sheet.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        sheet.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 45)

wb.save("google_health_feature_inventory.xlsx")
print("Created google_health_feature_inventory.xlsx")
