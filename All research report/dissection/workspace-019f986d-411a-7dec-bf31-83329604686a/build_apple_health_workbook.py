from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

wb=Workbook(); wb.remove(wb.active)
headers={
'Feature Inventory':['Feature','Purpose','Evidence ID','User Value','Business Value','Engineering Complexity (1-5)','Clinical Complexity (1-5)','Infrastructure Complexity (1-5)','Regulatory Complexity (1-5)','Estimated Team','Estimated Months','Priority','Category','Copy','Improve','Ignore','Reinvent','Moat','Confidence','Observed vs inferred'],
'Evidence Register':['Evidence ID','Claim / evidence','Source','URL','Source type','Date / relevance','Confidence','Observed vs inferred','Screenshot','Limitations'],
'Decision Ledger':['Feature / decision','Why built / pain','KPI likely improved','Trade-off','Alternative architecture','Evidence','Status'],
'Risk Register':['Risk','Domain','Likelihood','Impact','Leading indicator','Mitigation','Owner recommendation','Evidence / basis','Status'],
'Business Model Canvas':['Block','Apple Health assessment','Label','Ovexis strategic implication','Evidence'],
'Roadmap Reconstruction':['Stage','Period','Visible scope','Evidence','Confidence','Unknowns'],
'SWOT + Five Forces':['Framework','Element','Assessment','Label','Evidence / rationale'],
'Competitive Landscape':['Competitor / cluster','Overlap','Apple advantage','Apple gap','Ovexis response','Confidence'],
}
for name, cols in headers.items():
    ws=wb.create_sheet(name); ws.append(cols)
    ws.freeze_panes='A2'; ws.auto_filter.ref=f'A1:{get_column_letter(len(cols))}1'
    for c in ws[1]:
        c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='17365D'); c.alignment=Alignment(wrap_text=True,vertical='top')

features=[
('Health app central view','Centralises user health information','E01','High','High',3,2,3,2,'Cross-functional platform',24,'P0','Consumer platform','Yes','Yes','No','Yes','Strong','High','🟢 Observed'),
('HealthKit repository','Permissioned repository/API for health and fitness data','E02; E05','High','Very High',5,3,5,4,'Platform org',36,'P0','Developer platform','Yes','Yes','No','No','Strong','High','🟢 Observed'),
('Granular read permission','User chooses app data types to read','E04; E05','High','High',4,2,4,4,'Privacy + mobile',12,'P0','Privacy','Yes','Yes','No','Yes','Strong','High','🟢 Observed'),
('Separate read/write permissions','Separates sharing data from app writes','E05','High','High',4,2,4,4,'Privacy + mobile',12,'P0','Privacy','Yes','Yes','No','Yes','Strong','High','🟢 Observed'),
('Revocation controls','Remove app access in Settings','E04; E05','High','High',3,1,3,4,'Privacy + mobile',8,'P0','Privacy','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('On-device health encryption','Protects local health data','E05','High','Very High',5,2,5,5,'Security platform',24,'P0','Security','No','Yes','No','Yes','Strong','High','🟢 Observed'),
('E2E encrypted health iCloud sync','Protects synced health data under requirements','E04; E05','High','Very High',5,2,5,5,'Security/cloud',24,'P0','Security','No','Yes','No','Yes','Strong','High','🟢 Observed'),
('Health Records','Download supported provider clinical records','E02; E03','High','High',5,5,5,5,'Interop + clinical',30,'P0','Clinical interoperability','Yes','Yes','No','Yes','Strong','High','🟢 Observed'),
('FHIR clinical record access','Developers query authorised FHIR resources','E02','High','High',4,5,4,5,'Interop SDK',12,'P0','Developer platform','Yes','Yes','No','Yes','Strong','High','🟢 Observed'),
('FHIRModels','Typed Swift models for FHIR releases','E02','Medium','Medium',3,4,2,2,'SDK team',8,'P1','Open source','Yes','Yes','No','No','Medium','High','🟢 Observed'),
('Multi-provider aggregation','Brings multiple institutions into HealthKit','E02','High','High',5,5,5,5,'Interop + identity',30,'P0','Longitudinal record','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('SMART Health Cards','Import/share verifiable clinical record','E06','Medium','Medium',4,5,3,5,'Interop + security',12,'P2','Credential','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('Per-record one-time sharing','User selects individual verifiable records','E06','High','High',4,3,3,4,'Privacy + mobile',9,'P1','Privacy','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('Trends and charts','Visualises health data over time','E01','High','Medium',3,2,3,2,'Mobile UX',9,'P0','Analytics UX','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('Medications','Organises medication information','E01','High','Medium',3,4,3,4,'Mobile + clinical',12,'P1','Patient self-management','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('Sleep / activity view','Shows device/app sourced lifestyle data','E01','High','High',3,2,3,2,'Mobile + data',9,'P1','Wellness','Yes','Yes','No','Yes','Strong','High','🟢 Observed'),
('Health sharing','User shares selected health information','E04','High','Medium',4,3,4,4,'Privacy + UX',12,'P1','Collaboration','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('Medical ID','Emergency-accessible user information','E05','High','Medium',3,4,3,5,'Mobile + clinical',10,'P1','Emergency','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('BLE device integration','Connects compatible BLE health devices','E05','Medium','High',4,2,3,3,'Device integration',10,'P1','Wearables','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('Background / observer queries','Notifies apps of matching store changes','E02','Medium','High',4,2,4,3,'SDK',10,'P1','Developer platform','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('ResearchKit','Research consent, survey and active-task framework','E14','Medium','Medium',4,5,3,5,'Research platform',14,'P2','Research','Yes','Yes','No','No','Medium','High','🟢 Observed'),
('CareKit','Patient-care task/persistence/chart framework','E14','Medium','Medium',4,4,3,4,'Care platform',14,'P2','Care workflow','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('Apple Research app / studies','Distributed research participation','E12; E13','Medium','High',5,5,5,5,'Research ops',24,'P2','Research','Yes','Yes','No','Yes','Strong','High','🟢 Observed'),
('ECG / irregular rhythm notifications','Bounded cardiac signal/notification feature','E19; E20','High','High',5,5,5,5,'Hardware + ML + regulatory',36,'P1','Regulated wearable','No','Yes','No','Yes','Strong','High','🟢 Observed'),
('Sleep apnea notification','OTC risk notification from Watch feature','E19','High','High',5,5,5,5,'Hardware + ML + regulatory',36,'P1','Regulated wearable','No','Yes','No','Yes','Strong','High','🟢 Observed'),
('Feature education / limitations','Explains bounds of health findings','E19','High','High',3,5,2,5,'Clinical UX',8,'P0','Safety','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('Source provenance model','Tracks source/resource identifiers','E02; E05','High','High',4,5,4,4,'Data platform',16,'P0','Data quality','Yes','Yes','No','Yes','Medium','High','🟢 Observed'),
('Health data export / sharing report','Patient mediated report/data sharing','E20','Medium','Medium',3,4,3,4,'UX + clinical',8,'P1','Interoperability','Yes','Yes','No','Yes','Medium','Medium','🟢 Observed'),
('General clinical AI copilot','No public Apple Health documentation','None','Potentially high','Unknown',5,5,5,5,'AI + clinical safety',24,'N/A','AI','No','Yes','Yes','Yes','Future','High','🟢 Negative finding'),
('Cross-platform longitudinal graph','Native Apple-only substrate; broader graph not documented','E01; E05','High','High',5,5,5,5,'Data platform + clinical',24,'P0 Ovexis','Market gap','No','Yes','No','Yes','Future','High','🟢 Negative finding'),
]
ws=wb['Feature Inventory']
for r in features: ws.append(r)

sources=[
('E01','Health app purpose, central secure store, charts/trends/privacy','Apple Health','https://www.apple.com/health/','First-party','Current landing page','High','🟢 Observed','None','Marketing page; exact feature availability varies'),
('E02','Direct provider API/FHIR download, HealthKit clinical records/FHIRModels','Apple WWDC 2020','https://developer.apple.com/videos/play/wwdc2020/10669/','First-party technical','2020, durable architecture','High','🟢 Observed','None','Version-specific SDK details may change'),
('E03','Health Records rollout','Apple Newsroom','https://www.apple.com/newsroom/2018/03/doctors-put-patients-in-charge-with-apples-health-records-feature/','First-party','2018 historical','High','🟢 Observed','None','Not current coverage'),
('E04','Privacy principles, permission/data-use limits','Apple Privacy PDF','https://www.apple.com/privacy/docs/Health_Privacy_White_Paper_May_2023.pdf','First-party','2023','High','🟢 Observed','None','Configuration conditions apply'),
('E05','Security model, data protection, permissions, BLE','Apple Platform Security','https://support.apple.com/guide/security/protecting-access-to-users-health-data-sec88be9900f/web','First-party','Current support documentation','High','🟢 Observed','None','Not a third-party compliance certification'),
('E06','SMART Health Cards, JWS, per-record sharing','Apple WWDC 2021','https://developer.apple.com/videos/play/wwdc2021/10089/','First-party technical','2021','High','🟢 Observed','None','Availability/entitlements apply'),
('E11','800 institutions/12000 locations in 2022','Apple Newsroom CA','https://www.apple.com/ca/newsroom/2022/07/how-apple-is-empowering-people-with-their-health-information/','First-party','2022 historical','High','🟢 Observed','None','Not current count'),
('E13','Apple Health Study and VP quote','Apple Newsroom','https://www.apple.com/newsroom/2025/02/new-holistic-apple-health-study-launches-today-in-the-research-app/','First-party','2025','High','🟢 Observed','None','Study does not establish every feature claim'),
('E14','CareKit and ResearchKit source repositories','GitHub','https://github.com/carekit-apple/CareKit','Primary project','Current repository','High','🟢 Observed','None','Open-source status not product adoption measure'),
('E15','Health Technology disciplines','Apple Careers','https://www.apple.com/careers/us/hardware.html','First-party','Current careers overview','High','🟢 Observed','None','Not a requisition/headcount list'),
('E16','Desai leadership / principles','Stanford Medicine','https://med.stanford.edu/news/all-news/2025/09/sumbul-desai-mgr.html','Academic interview','2025','Medium','🟢 Observed','None','Interview framing'),
('E17','Reported health-related acquisitions','Becker’s Hospital Review','https://www.beckershospitalreview.com/healthcare-information-technology/apple-s-health-it-acquisitions-a-timeline/','Trade press','2021','Medium','🟢 Reported','None','Feature attribution unavailable'),
('E18','Apple corporate reporting','Apple Investor Relations','https://investor.apple.com/','First-party','FY2025','High','🟢 Observed','None','No Health segment P&L'),
('E19','Sleep apnea feature FDA record','FDA','https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfpmn/pmn.cfm?ID=K240929','Regulator','2024','High','🟢 Observed','None','Specific feature only'),
('E20','Watch health feature/eligibility statements','Apple Newsroom','https://www.apple.com/newsroom/2025/09/apple-debuts-apple-watch-series-11-featuring-groundbreaking-health-insights/','First-party','2025','High','🟢 Observed','None','Region/device limitations'),
('E21','Anecdotal UX and tracking complaints','Reddit','https://www.reddit.com/r/apple/comments/146q482/bad_ui_in_apple_health/','User-generated','2023–25','Low','🟢 User report','None','Non-representative sentiment'),
]
for r in sources: wb['Evidence Register'].append(r)

decisions=[
('Local HealthKit repository','Fragmented app/device data','Integration and hardware utility','Less cloud coordination','Central cloud PHR','E02; E05','🟡 Strong inference'),
('Granular permissions','Health-data misuse concerns','Trust / consent acceptance','Partial data and onboarding friction','All-or-nothing consent','E04; E05','🟢 mechanism / 🟡 KPI'),
('Direct FHIR provider download','Portal fragmentation','Record utility','Coverage and semantic variation','Apple central record cloud','E02','🟢 mechanism'),
('On-device processing','Data minimisation','Trust differentiation','Less central analysis','Cloud analytics','E04','🟢 principle / 🟡 trade-off'),
('Bounded regulated features','Early-risk signals with credibility','Hardware differentiation','Cost/scope/alert risk','Unregulated general scoring','E19; E20','🟢 bounds / 🟡 KPI'),
('Open-source research/care frameworks','Developer/research adoption','Ecosystem reach','Lower direct control','Closed proprietary SDK','E14','🟢 mechanism / 🟡 KPI'),
]
for r in decisions: wb['Decision Ledger'].append(r)
risks=[
('False positive/negative or user over-reliance','Clinical','Medium','High','Unexpected support/complaint signal','Bound claims, education, prospective validation, escalation','Clinical safety lead','E19; E20','Open'),
('Incomplete/conflicting longitudinal data','Data','High','High','Low source coverage, duplicate/conflict rate','Provenance, reconciliation, missingness UX','Data platform lead','E02','Open'),
('Third-party overcollection/secondary use','Privacy','Medium','High','Permission scope / complaints','Minimum necessary, revocation, audits','Privacy lead','E04; E05','Open'),
('Provider FHIR onboarding friction','Operational','High','Medium','Connection completion rate','Partner toolkit, fallback documents','Interop lead','E02; E11','Open'),
('Regulatory claim expansion','Regulatory','Medium','High','Feature claim review failures','Feature risk classification and QMS','Regulatory lead','E19','Open'),
('Platform dependency','Business','High','High','API/OS policy changes','Multi-platform connectors, contractual review','CTO','E02; E05','Open'),
('AI hallucination/unsafe advice','AI','Medium','Very High','Unsupported output rate','Evidence grounding, abstention, clinician review','AI safety lead','Negative finding; Ovexis recommendation','Open'),
]
for r in risks: wb['Risk Register'].append(r)
canvas=[
('Customer segments','iPhone/Watch users; developers; provider partners; researchers','🟢','Choose a focused patient+clinician cohort first','E01; E02; E12'),
('Value proposition','Private data substrate and selected health insights','🟡','Evidence-backed cross-platform longitudinal intelligence','E01; E04'),
('Channels','Preinstalled OS/device, developer ecosystem, provider/research partnerships','🟢/🟡','B2B2C clinic/lab partners, app stores, trusted clinicians','E01; E12'),
('Customer relationships','User-controlled self-service permissions and sharing','🟢','High-touch onboarding for complex records plus self-service','E04; E05'),
('Revenue','No standalone Health price disclosed; indirect hardware/platform value','🟡','Per-engaged-member / clinic seats + premium review','E01; E18'),
('Key resources','Hardware, OS, HealthKit, brand, clinical studies','🟢/🟡','Provenance graph, evidence layer, local partners, clinical governance','E02; E13'),
('Key activities','Sensor/feature development, privacy controls, platform stewardship','🟡','Ingestion QA, synthesis, safety evaluation, workflow closure','E04; E20'),
('Key partners','Providers, researchers, developers, compatible device makers','🟢','EHR/lab/pharmacy/clinician partners','E02; E12'),
('Cost structure','Not disclosed by product','🟢','Clinical ops, integrations, security, model evaluation','E18'),
]
for r in canvas: wb['Business Model Canvas'].append(r)
roadmap=[
('MVP','2014–15','Health/HealthKit; ResearchKit follows','E02; E14','High','Detailed internal roadmap unknown'),
('V2','2016–20','Watch expansion; CareKit; FHIR Health Records','E02; E03; E14','High','Acquisition-to-feature mapping unknown'),
('V3','2021–24','Verifiable records; privacy articulation; sleep apnea notification','E04; E06; E19','High','Feature performance/coverage varies'),
('Visible current','2025–26','Apple Health Study; expanded Watch health surfaces','E13; E20','Medium','Precise future release schedule unknown'),
('Future','2026+','Broader signals/interoperability likely','Inference','Low','No confirmed roadmap'),
]
for r in roadmap: wb['Roadmap Reconstruction'].append(r)
swot=[
('SWOT','Strength','Integrated devices, platform permissions, privacy, studies','🟢/🟡','E01; E04; E13'),('SWOT','Weakness','Apple-only; provider coverage/semantic gaps; no published full care workflow','🟢/🟡','E02; E11'),('SWOT','Opportunity','Preventive and longitudinal intelligence','🟡','Strategic inference'),('SWOT','Threat','Regulation, trust failures, substitutes','🟡','E19; market structure'),
('Porter','Supplier power','Medium','🟡','Sensors/clinical partners vs Apple integration'),('Porter','Buyer power','Medium','🟡','Consumers/providers have alternatives'),('Porter','New entrants','Medium','🟡','Apps easy; trust/hardware/regulatory hard'),('Porter','Substitutes','High','🟡','Wearables, portals, clinics'),('Porter','Rivalry','High','🟡','Wellness/clinical intelligence crowded'),
]
for r in swot: wb['SWOT + Five Forces'].append(r)
comp=[
('WHOOP / Oura / Ultrahuman','Wearables + health insights','Integrated Apple OS/hardware/privacy','Focused coaching/cross-platform alternatives','Aggregate as inputs; win on longitudinal care','🟡'),
('Google Health / Health Connect','Health data platform','Apple device distribution','Cross-platform reach','Android parity','🟡'),
('Function Health / Levels / Superpower','Preventive consumer health','Data substrate and hardware','Labs/interpretation/programs','Reconcile labs with all records','🟡'),
('OpenEvidence / Glass Health / Atropos / AMBOSS / UpToDate','Clinical intelligence','Consumer sensor reach','Clinician evidence/workflow orientation','Citation-grade clinician layer','🟡'),
('Apollo 24/7 / Practo / Tata 1mg / Healthify','India care/service rails','Device/software integration','Local care/pharmacy distribution','Partner local rails','🟡'),
('Human API','Data aggregation API','Apple ecosystem position','Broader aggregation posture','Multi-source connector','🟡'),
('Regacore','Unverified in this pass','Not assessed','Not assessed','Perform separate diligence','🔴'),
]
for r in comp: wb['Competitive Landscape'].append(r)

# format all
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment=Alignment(wrap_text=True,vertical='top')
    for col in range(1,ws.max_column+1):
        maxlen=max(len(str(ws.cell(r,col).value or '')) for r in range(1,min(ws.max_row,100)+1))
        ws.column_dimensions[get_column_letter(col)].width=min(max(13,maxlen*0.85),42)
    ws.row_dimensions[1].height=34
    if ws.max_row>1:
        ref=f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
        tab=Table(displayName='T'+''.join(x for x in ws.title if x.isalnum())[:18], ref=ref)
        tab.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True,showColumnStripes=False)
        ws.add_table(tab)
# conditional complexity
ws=wb['Feature Inventory']
for col in ['F','G','H','I']:
    ws.conditional_formatting.add(f'{col}2:{col}{ws.max_row}',ColorScaleRule(start_type='min',start_color='63BE7B',mid_type='percentile',mid_value=50,mid_color='FFEB84',end_type='max',end_color='F8696B'))

wb.save('apple_health_feature_inventory.xlsx')
print('saved')
