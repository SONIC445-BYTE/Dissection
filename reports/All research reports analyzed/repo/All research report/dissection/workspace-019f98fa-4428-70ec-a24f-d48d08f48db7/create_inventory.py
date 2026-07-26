from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
rows = [
("Ring AIR passive sensing","Continuous sleep, movement, recovery and physiology capture","S1; public product page","Passive insight and daily habit","Hardware revenue and activation","High","Medium","High","Medium", "Embedded 4; mobile 3; data 3", "12-18", "P0", "Sensing", "Copy", "Medium", "Confirmed"),
("CGM / M1 integration","Expose glucose and metabolic response","S1; S10","Personalized metabolic feedback","Consumable/service attach","High","High","High","High", "Data 3; clinical 2; mobile 2", "9-15", "P0", "Metabolism", "Improve", "Medium", "Confirmed"),
("Blood biomarker ingestion","Connect episodic lab results to timeline","S4; S16","Understand lab trends in context","Paid testing / retention","Medium","High","Medium","High", "FHIR/data 3; clinical 2; OCR 2", "6-10", "P0", "Labs", "Improve", "Medium", "Confirmed"),
("Longitudinal health timeline","Unify observations by person and time","Ovexis proposal","Single source of truth","Retention and switching cost","Medium","High","High","High", "Platform 4; clinical informatics 2", "6-9", "P0", "Platform", "Reinvent", "Strong", "Strong Inference"),
("Consent ledger","Purpose-, source-, recipient- and time-bounded permissions","S2; Ovexis proposal","Trust and control","Enterprise readiness","Medium","High","High","High", "Security 2; backend 2; privacy 1", "4-7", "P0", "Governance", "Reinvent", "Strong", "Strong Inference"),
("Source provenance card","Show source, timestamp, device, algorithm version","S2; Ovexis proposal","Auditability and trust","Lower support and liability risk","Low","Medium","Medium","High", "Data 2; UX 1", "3-5", "P0", "Trust", "Reinvent", "Strong", "Strong Inference"),
("Missing-data aware insights","Abstain or lower confidence when data gaps exist","S6; S7; Ovexis proposal","Avoid false precision","Trust and safety","Medium","High","Medium","High", "ML 2; rules 2; UX 1", "4-7", "P0", "AI Safety", "Reinvent", "Strong", "Strong Inference"),
("Evidence retrieval","Cite research and guidelines in insights","S10; S16; Ovexis proposal","Know why recommendation exists","Differentiation and defensibility","Medium","High","Medium","High", "RAG 2; clinical 2; platform 1", "5-8", "P0", "AI", "Reinvent", "Strong", "Strong Inference"),
("Constrained health copilot","Natural language explanation over typed health tools","S2; public AI details not verified","Conversational understanding","Engagement and paid tier","Medium","High","Medium","High", "ML 3; backend 2; safety 2", "6-10", "P0", "AI", "Improve", "Future", "Strong Inference"),
("FHIR R4 export","Patient-authorized clinical interoperability","Not publicly verified for Ultrahuman; Ovexis proposal","Clinician use and portability","Enterprise/care access","Medium","High","Medium","High", "FHIR 2; backend 2; clinical 1", "4-7", "P0", "Interoperability", "Reinvent", "Strong", "Strong Inference"),
("Apple Health / HealthKit","Import on-device health data","S1 ecosystem context; integration not fully verified","Device neutrality","Top-of-funnel acquisition","Medium","Low","Medium","Medium", "iOS 2; data 1", "3-5", "P0", "Integration", "Copy", "Medium", "Strong Inference"),
("Google Health Connect","Import Android health data","Integration not publicly verified for Ultrahuman; Ovexis proposal","Android reach","Acquisition","Medium","Low","Medium","Medium", "Android 2; data 1", "3-5", "P0", "Integration", "Copy", "Medium", "Strong Inference"),
("Wearable adapter framework","Normalize Oura, WHOOP, Ultrahuman, Garmin and others","S21 third-party API evidence; Ovexis proposal","Avoid device lock-in","Distribution and retention","High","Low","High","Medium", "Integrations 4; data 2", "8-14", "P0", "Platform", "Reinvent", "Strong", "Strong Inference"),
("Lab PDF/OCR import","Extract values from legacy reports with verification","Ovexis proposal","Works with any lab","Fast activation","Medium","Medium","Medium","High", "OCR 2; clinical 2; UX 1", "5-8", "P1", "Labs", "Reinvent", "Medium", "Strong Inference"),
("Medication reconciliation","Track medication, dose, timing, adherence","Not publicly verified; Ovexis proposal","Contextualize physiology","Clinical value and retention","Medium","High","Medium","High", "Clinical 2; backend 2", "5-8", "P1", "Clinical", "Reinvent", "Strong", "Strong Inference"),
("Clinician summary","Generate evidence-linked patient report","S4; Ovexis proposal","Efficient appointment preparation","Provider distribution","Low","High","Medium","High", "Clinical UX 2; backend 2", "4-7", "P0", "Clinical", "Improve", "Medium", "Strong Inference"),
("Provider portal","Review trends, provenance, consent and flags","Not publicly verified for Ultrahuman; Ovexis proposal","Care coordination","B2B revenue and moat","High","High","High","High", "Full-stack 3; clinical 2; security 1", "9-15", "P1", "Clinical", "Reinvent", "Strong", "Strong Inference"),
("Intervention ledger","Record recommendation, adherence and outcome","Ovexis proposal","Measure what works for this person","Outcome moat","Medium","High","Medium","High", "Data 2; UX 2; clinical 1", "5-8", "P0", "AI", "Reinvent", "Strong", "Strong Inference"),
("Causal self-experiments","Baseline/control/intervention design","S1 metabolic action context; Ovexis proposal","Move beyond correlation","Retention and science","High","High","Medium","High", "Data science 3; clinical 2", "8-12", "P1", "Metabolism", "Reinvent", "Future", "Strong Inference"),
("Notification budget","Limit and prioritize nudges","S6; S20 complaint signals","Lower anxiety and fatigue","Retention and trust","Low","Medium","Low","Medium", "Backend 1; UX 1", "2-4", "P0", "UX", "Improve", "Medium", "Strong Inference"),
("User correction workflow","Correct sleep, lab, medication and identity errors","S6; S7 complaint signals","Data quality control","Trust and support efficiency","Medium","High","Medium","High", "Backend 2; UX 1", "3-6", "P0", "Data Quality", "Reinvent", "Strong", "Strong Inference"),
("Algorithm version diff","Explain changes after model updates","Ovexis proposal","Stable interpretation over time","Trust and reduced churn","Medium","Medium","Medium","High", "ML platform 2; UX 1", "4-6", "P1", "Trust", "Reinvent", "Strong", "Strong Inference"),
("Privacy center","Export, delete, revoke, view access history","S2; S17","User control","Compliance and trust","Medium","High","Medium","High", "Security 2; frontend 2", "4-7", "P0", "Governance", "Improve", "Strong", "Confirmed + proposal"),
("Open developer API","OAuth, webhooks, normalized schema, FHIR","S21 says native API exists; official details unverified","Developer adoption","Platform revenue and moat","High","Low","High","High", "Backend 3; docs 1; security 1", "6-10", "P1", "Developer", "Reinvent", "Strong", "Strong Inference"),
("Data quality dashboard","Show sync health, gaps, duplicates and stale sources","S6; S7; Ovexis proposal","Debug continuity","Lower support cost","Medium","Low","Medium","Medium", "Data 2; UX 1", "3-5", "P0", "Platform", "Reinvent", "Strong", "Strong Inference"),
("Clinically safe escalation","Triage red flags to care guidance, not diagnosis","S2 disclaimer; Ovexis proposal","Know when to seek care","Safety and clinical differentiation","Medium","High","Medium","High", "Clinical 3; safety 2", "6-10", "P1", "Clinical", "Reinvent", "Strong", "Strong Inference"),
("Community / referral","Share evidence cards and outcomes with consent","Public community positioning; Ovexis proposal","Social proof","Lower CAC","Low","Low","Medium","Medium", "Growth 2; privacy 1", "3-5", "P1", "Growth", "Copy", "Medium", "Strong Inference"),
("Support diagnostics","Remote device/integration troubleshooting","S1; S7 support evidence","Faster resolution","Lower returns and support cost","Medium","Low","Medium","Medium", "Support tools 2; data 1", "3-6", "P0", "Operations", "Improve", "Medium", "Strong Inference"),
("Regional data residency","Tenant/region policy and deletion controls","S2 cross-border policy; Ovexis proposal","Regulatory trust","Enterprise expansion","High","Low","High","High", "Platform 3; security 2; legal 1", "8-14", "P1", "Security", "Reinvent", "Strong", "Strong Inference"),
("Research consent exchange","Opt-in, revocable cohort/research data access","S10; S16; Ovexis proposal","Participate safely in research","Clinical/research partnerships","Medium","Medium","High","High", "Research 2; security 2; data 2", "6-10", "P2", "Research", "Reinvent", "Future", "Strong Inference"),
("Hardware manufacturing","Own sensor form factor","S3; S5 manufacturing evidence","Control signal quality and brand","Hardware revenue","Very High","High","Very High","High", "Hardware 6; supply chain 3; QA 3", "18-36", "P3", "Hardware", "Ignore", "Medium", "Confirmed market pattern"),
]
headers=["Feature","Purpose","Evidence","User Value","Business Value","Engineering Complexity","Clinical Complexity","Infrastructure Complexity","Regulatory Complexity","Estimated Team","Estimated Months","Priority","Category","Copy / Improve / Ignore / Reinvent","Moat","Confidence"]
wb=Workbook(); ws=wb.active; ws.title="Feature Inventory"
ws.append(headers)
for r in rows: ws.append(r)
ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
for c in ws[1]:
    c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid", fgColor="16324F"); c.alignment=Alignment(wrap_text=True,vertical="top")
for row in ws.iter_rows():
    for cell in row[1:]: cell.alignment=Alignment(wrap_text=True,vertical="top")
for i,w in enumerate([28,35,32,24,24,18,18,22,20,28,14,12,18,24,14,18],1): ws.column_dimensions[get_column_letter(i)].width=w
for row in ws.iter_rows(min_row=2):
    if row[11].value=="P0":
        for cell in row: cell.fill=PatternFill("solid", fgColor="EAF4EA")
# second sheet glossary
lg=wb.create_sheet("Methodology")
lg.append(["Field","Definition"])
for a,b in [("Evidence","Public source or clearly marked proposal."),("Confidence","Confidence in the evidence, not in a private implementation claim."),("Complexity","Ovexis directional estimate: Low/Medium/High/Very High."),("Estimated Team","Indicative cross-functional team count and roles."),("Estimated Months","Indicative build duration with a focused team; not a commitment."),("Copy/Improve/Ignore/Reinvent","Strategic recommendation for Ovexis."),("Confirmed","Directly observed in a cited public source."),("Strong Inference","Reasoned from multiple public signals, not directly observed."),("Speculation","Scenario or hypothesis; validate before using in decisions.")]: lg.append([a,b])
for c in lg[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid", fgColor="16324F")
lg.column_dimensions["A"].width=28; lg.column_dimensions["B"].width=100
for row in lg.iter_rows():
    for cell in row: cell.alignment=Alignment(wrap_text=True,vertical="top")
wb.save("/home/user/ultrahuman_feature_inventory.xlsx")
