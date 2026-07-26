from pathlib import Path
import csv,json,html,textwrap
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
OUT=Path('/home/user/humanapi_research'); OUT.mkdir(exist_ok=True)
features=[
('B2B enterprise product landing page','Explain Human API value and collect demo leads','risk.lexisnexis.com/products/humanapi','Helps buyers understand use cases','Enterprise lead gen','Low','1-2 weeks','No','Low','High','Yes','Yes','No','No','High','High'),
('Life-insurance EHR landing page','Position Health Intelligence for underwriting','risk.lexisnexis.com/products/health-intelligence-ehr','Clarifies underwriting ROI','Carrier pipeline','Low','1-2 weeks','No','Medium','High','Yes','Yes','No','No','High','High'),
('Contact/demo form','Capture enterprise leads','Product pages contact forms','Lets enterprise buyer request demo','Sales qualification','Medium','2-4 weeks','No','Low','High','Yes','Improve','No','No','High','High'),
('Global country selector','Collect geographic sales context','Contact form country list','Local routing','Sales ops','Low','1 week','No','Low','Medium','Yes','Improve','No','No','Medium','High'),
('Industry/subsector fields','Qualify use case','Contact form fields','Less back-and-forth','Lead scoring','Low','1 week','No','Low','Medium','Yes','Improve','No','No','Medium','High'),
('Consumer consent interface','Allow patient/applicant to authorize records','Human API docs: Connect widget, consumerLink','User controls sharing','Core data acquisition','High','3-6 months','No','High','Critical','Yes','Yes','No','No','P0','High'),
('Human API Connect JS client','Embed consent widget in customer apps','v2.3 web guide; humanapi-connect-client','Fast integration','Developer adoption','Medium','1-2 months','No','High','High','Yes','Improve','No','No','High','High'),
('NPM package','Install Connect in React/web apps','v2.3 web guide npm install humanapi-connect-client','Developer convenience','Integration velocity','Medium','1 month','No','Medium','Medium','Yes','Improve','No','No','Medium','High'),
('CDN script','Load Connect from cdn.humanapi.co','v2.3 web guide CDN snippets','Low setup friction','Adoption','Medium','2-4 weeks','No','Medium','Medium','Yes','Improve','No','No','Medium','High'),
('Lifecycle hooks','Emit connect/disconnect/close events','v2.3 web guide lifecycle hooks','Host app knows session outcome','Integration quality','Medium','2-4 weeks','No','Low','High','Yes','Improve','No','No','High','High'),
('Session token endpoint','Generate Connect session tokens','/v1/connect/token docs','Secure launch of consent flow','Access control','Medium','1-2 months','No','High','High','Yes','Improve','No','No','High','High'),
('Admin client token endpoint','Generate Admin API token','/v1/admin/token docs','Server-to-server auth','API monetization','Medium','1 month','No','High','High','Yes','Improve','No','No','High','High'),
('Create user/order API','Create applicant/user and order evidence','/api/v1/users docs','Starts data retrieval','Core transaction volume','High','2-4 months','No','High','Critical','Yes','Improve','No','No','P0','High'),
('Order type configuration','Bundle retrieval channels/timeouts/outputs','Configuring order types docs','Workflow flexibility','Enterprise customization','High','2-4 months','No','High','Critical','Yes','Improve','No','No','P0','High'),
('Order type API','List configured order types','/api/v1/order-types docs','Customer system knows options','Integration reliability','Medium','1 month','No','Medium','High','Yes','Improve','No','No','High','High'),
('Consumer-mediated retrieval channel','Portal/patient-mediated EHR access','Order types and product page','Patient-controlled data access','Higher hit rate','Very High','6-18 months','No','High','Critical','Yes','Improve','No','No','P0','High'),
('Digital HIPAA authorization channel','HIPAA-authorized EHR/HIE retrieval','Order types docs','No consumer portal dependency','Carrier automation','Very High','6-18 months','No','High','Critical','Yes','Improve','No','No','P0','High'),
('Traditional HIPAA/APS channel','Retrieve APS when EHR unavailable','Order types/docs','Fallback evidence','Completion revenue','High','6-12 months','Maybe','High','Critical','Yes','Improve','No','No','P0','High'),
('Smart evidence orchestration','Select retrieval path based on data/constraints','Official acquisition/product pages','Less friction','Cost/hit-rate optimization','Very High','9-18 months','Maybe','High','Critical','Yes','Reinvent','No','Yes','P0','High'),
('APS delay logic','Delay APS to allow digital channels first','Order type docs recommend 7 days','Lower unnecessary APS friction','Cost savings','Medium','1-2 months','No','Medium','High','Yes','Improve','No','No','High','High'),
('Order sufficiency logic','Keep searching until required data found','Order type docs','Data completeness','Better underwriting decisions','High','3-6 months','Maybe','High','Critical','Yes','Reinvent','No','Yes','High','High'),
('Custom outputs','Generate outputs by use case','Order type docs','Fit workflow needs','Enterprise upsell','High','3-6 months','Maybe','High','High','Yes','Improve','No','No','High','High'),
('Clinical History report','Shortened, readable EHR report','Reports docs; Health Intelligence page','Underwriters review faster','Differentiation','Very High','6-12 months','Maybe','High','Critical','Yes','Reinvent','No','Yes','P0','High'),
('Highlights Summary','Summary of profile, health measurements, social history, labs','Reports docs; product page','Fast triage','Adoption by underwriters','High','3-6 months','Maybe','High','Critical','Yes','Reinvent','No','Yes','P0','High'),
('HealthCheck report','Targeted labs/vitals alternative to labs','Reports docs/order types','Less invasive evidence','Faster decisions','High','3-6 months','Maybe','High','High','Yes','Improve','No','No','High','High'),
('FHIR R4 report','Structured zip ndjson output','Reports docs','Analytics ingestion','Developer/data product value','High','4-8 months','No','High','Critical','Yes','Improve','No','No','P0','High'),
('CCD/CCDA report','Legacy XML medical record payload','Epic docs and reports docs','Standards compatibility','Legacy integrations','Medium','2-4 months','No','High','Medium','Maybe','Improve','No','No','Medium','High'),
('API Data JSON report','Bulk export of clinical endpoints','Reports docs','Programmatic modeling','Decision automation','High','3-6 months','No','High','Critical','Yes','Improve','No','No','P0','High'),
('Report download API','Fetch available report list and content','/api/v1/user/reports docs','Automated retrieval','Integration value','Medium','1-2 months','No','Medium','High','Yes','Improve','No','No','High','High'),
('Automated report push','Push outputs automatically at completion','Reports delivery docs','No polling','Enterprise integration','Medium','2-3 months','No','Medium','High','Yes','Improve','No','No','High','High'),
('HTTP multipart delivery','Push report over HTTP','Reports delivery docs','Flexible delivery','Integration value','Medium','1-2 months','No','Medium','Medium','Yes','Improve','No','No','Medium','High'),
('SFTP delivery','Push reports to SFTP','Reports delivery docs','Legacy enterprise compatibility','Enterprise adoption','Medium','1-2 months','No','Medium','Medium','Yes','Improve','No','No','Medium','High'),
('Pre-signed URL delivery','Ask customer for URL then push file','Shipment docs','Handles large files/cloud storage','Modern integration','Medium','1-2 months','No','Medium','High','Yes','Improve','No','No','High','High'),
('Webhook order summary','Notify terminal order state and reports','Order summary notification docs','Automated case workflow','Integration reliability','Medium','1-2 months','No','Medium','High','Yes','Improve','No','No','High','High'),
('APS status note webhook','Notify vendor notes','APS status docs','Operational transparency','Case management value','Medium','1-2 months','No','Medium','High','Yes','Improve','No','No','High','High'),
('Pended task model','Consolidate blocking issues into tasks','Integration best practices docs','Clear next action','Reduced support friction','High','3-6 months','No','Medium','High','Yes','Improve','No','No','High','High'),
('Task Manager link','Hosted UI to resolve tasks','consumer-link docs','User/underwriter action path','Completion rate','Medium','1-3 months','No','Medium','High','Yes','Improve','No','No','High','High'),
('Order cancellation API','Abort/resync user/order','/api/v1/users/actions docs','Control workflow','Support reduction','Medium','1 month','No','Low','Medium','Yes','Improve','No','No','Medium','High'),
('Resync action','Trigger fresh data sync','users/actions enum resync','Current data','Retention/usefulness','Medium','1-2 months','No','Medium','High','Yes','Improve','No','No','High','High'),
('Provider search','Search/select provider/location','Product page: Provider Search','Less manual input','Better hit rate','High','3-6 months','No','High','High','Yes','Improve','No','No','High','High'),
('Suggested sources','Pass likely providers to API','Submitting orders docs','Faster selection','Hit-rate lift','Medium','1-2 months','No','Medium','High','Yes','Improve','No','No','High','High'),
('Provider object with NPI/MRN/fax','Supply provider details','Create user schema','Precise retrieval','APS success','Medium','1-2 months','No','High','High','Yes','Improve','No','No','High','High'),
('Demographic validation','Validate DOB, SSN, phone, address','Create user schema/order guidelines','Fewer failures','Retrieval quality','Medium','1-2 months','No','High','High','Yes','Improve','No','No','High','High'),
('Medical API: allergies','Normalized allergies endpoint','v2.3 allergies docs','Clinical context','Data completeness','Medium','1-2 months','No','High','Medium','Yes','Improve','No','No','Medium','High'),
('Medical API: problems/issues','Normalized problems endpoint','v2.3 problems docs','Condition history','Risk models','Medium','1-2 months','No','High','High','Yes','Improve','No','No','High','High'),
('Medical API: medications','Normalized meds endpoint','v2.3 medications docs','Medication history','Risk/clinical value','Medium','1-2 months','No','High','High','Yes','Improve','No','No','High','High'),
('Medical API: test results','Normalized lab/test endpoint','v2.3 test-results docs','Lab history','Risk/clinical value','High','2-4 months','No','High','Critical','Yes','Improve','No','No','P0','High'),
('Medical API: encounters','Clinical encounters','v2.3 endpoint list','Longitudinal care history','Completeness','Medium','1-2 months','No','High','High','Yes','Improve','No','No','High','Medium'),
('Medical API: procedures','Procedure history','v2.3 endpoint list','Clinical history','Risk value','Medium','1-2 months','No','High','High','Yes','Improve','No','No','High','Medium'),
('Medical API: immunizations','Vaccine history','v2.3 endpoint list','Preventive history','Use-case expansion','Medium','1-2 months','No','High','Medium','Yes','Improve','No','No','Medium','Medium'),
('Medical API: vitals','Vitals endpoint','v2.3 endpoint list','Objective measures','Underwriting/clinical value','Medium','1-2 months','No','High','High','Yes','Improve','No','No','High','Medium'),
('Wellness API: activities','Activity segments endpoint','v2.3 activities docs','Lifestyle data','Personalization/use cases','Medium','1-2 months','No','Low','Medium','Yes','Improve','No','No','Medium','High'),
('Wellness API: sleep','Sleep/sleep summary endpoint','v2.3 endpoint list','Recovery context','Personalization','Medium','1-2 months','No','Low','Medium','Yes','Improve','No','No','Medium','Medium'),
('Wellness API: heart rate','HR/HR summary endpoint','v2.3 endpoint list','Physiology context','Risk/engagement','Medium','1-2 months','No','Low','Medium','Yes','Improve','No','No','Medium','Medium'),
('Sources endpoint','List connected external accounts and sync status','v2.3 sources docs','Transparency','Debuggability','Medium','1-2 months','No','Medium','High','Yes','Improve','No','No','High','High'),
('Human summary endpoint','Latest normalized summary profile','v2.3 human-summary docs','Quick overview','App efficiency','Medium','1-2 months','No','Medium','Medium','Yes','Improve','No','No','Medium','High'),
('MyHumanAPI patient app','Patient aggregation/view/download','Epic docs','User self-service','Trust/compliance','High','3-6 months','No','High','High','Yes','Improve','No','No','High','High'),
('Enterprise Portal','Hosted portal to view/manage consumer data','Product page; v2.3 overview','Ops/admin visibility','Enterprise stickiness','High','4-8 months','Maybe','Medium','High','Yes','Improve','No','No','High','High'),
('Interactive timeline','Timeline view of health data sources','v2.3 overview portal description','Comprehension','Workflow adoption','High','3-6 months','Maybe','Medium','Medium','Maybe','Improve','No','No','Medium','High'),
('Subscriptions for agents/producers','Notify case stakeholders','Subscriptions docs','Distribution visibility','Carrier/broker adoption','Medium','2-4 months','No','Low','Medium','Yes','Improve','No','No','Medium','High'),
('Role-based producer subscription','Subscribe producer role to case','Subscriptions API docs','Agent experience','Distribution partner value','Medium','2-3 months','No','Low','Medium','Yes','Improve','No','No','Medium','High'),
('Medical Insights extraction','Targeted attributes from EHR','Health Intelligence page','Faster risk review','High-margin insights','Very High','6-12 months','Yes','High','Critical','Yes','Reinvent','No','Yes','P0','High'),
('Mortality management output','Select mortality predictors','Order type docs','Risk assessment','Carrier differentiation','Very High','6-12 months','Yes','High','Critical','Maybe','Reinvent','No','Yes','High','High'),
('APS summarization','Summarize long APS records','Order type docs','Less reading burden','Underwriter productivity','High','3-6 months','Yes','High','High','Yes','Reinvent','No','Yes','High','High'),
('Conditional summarization logic','Summarize based on pages/face amount','Order type docs','Cost-aware UX','Margin protection','High','2-4 months','Yes','Medium','Medium','Yes','Improve','No','No','Medium','High'),
('Advanced linking of consumers/providers','Link/cleanse identities and providers','2025 Lexis release','Higher hit rate','Lexis moat','Very High','12-24 months','Maybe','Medium','Critical','No','Reinvent','No','Yes','P0','High'),
('Identity resolution integration','Combine health data with Lexis identity data','Health Intelligence page and 2025 release','Better matching','Major moat','Very High','12-24 months','Maybe','Medium','Critical','No','Reinvent','No','Yes','P0','High'),
('Consent deletion/revocation','Patients can request deletion; connections active until revoked','Epic docs','Trust/control','Compliance','High','2-4 months','No','High','Critical','Yes','Improve','No','No','P0','High'),
('Audit logging','Log API calls/user/system events','Security/Epic docs','Accountability','Compliance','High','2-4 months','No','High','Critical','Yes','Improve','No','No','P0','High'),
('AES-256 at-rest encryption','Encrypt stored data','Security docs','Privacy/security','Enterprise trust','High','1-3 months','No','High','Critical','Yes','Improve','No','No','P0','High'),
('HTTPS/TLS everywhere','Encrypt transport','Security docs/headers','Security','Compliance','Medium','1 month','No','High','Critical','Yes','Improve','No','No','P0','High'),
('BAA support','Business Associate Agreements where appropriate','Security docs','Enterprise compliance','Sales enablement','High legal','1-3 months','No','High','Critical','Yes','Improve','No','No','P0','High'),
('ONC/Epic developer documentation','Show compliance with Epic guidelines/ONC criteria','Epic docs','Provider trust','Network access','Medium','1-2 months','No','High','High','Yes','Improve','No','No','High','High'),
('Consumer-mediated COVID credential use case','Support Health Pass/COVID results','CLEAR partnership public release','Safe access verification','Adjacent market expansion','High','3-6 months','No','High','Temporary','Maybe','Ignore','Yes','No','Low','High'),
('Sales resources and webinars','Educate enterprise buyers','Risk resources pages','Buyer education','Demand generation','Low','1-2 months','No','Low','Medium','Yes','Improve','No','No','Medium','High'),
('ROI stats/proof points','Quantify decisions, cycle time, placement','Health Intelligence page internal study','Buyer confidence','Conversion','Low','2-4 weeks','No','Low','High','Yes','Improve','No','No','High','High')
]
headers=['Feature','Purpose','Evidence','User Value','Business Value','Engineering Complexity','Estimated Build Time','AI Dependency','Clinical Dependency','Strategic Importance','Copy','Improve','Ignore','Reinvent','Priority','Confidence']
with (OUT/'feature_inventory.csv').open('w',newline='',encoding='utf-8') as f:
 w=csv.writer(f); w.writerow(headers); w.writerows(features)
wb=Workbook(); ws=wb.active; ws.title='Feature Inventory'; ws.append(headers)
for r in features: ws.append(r)
fill=PatternFill('solid',fgColor='111827'); font=Font(color='FFFFFF',bold=True); thin=Side(style='thin',color='D1D5DB')
for c in ws[1]: c.fill=fill; c.font=font; c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
for row in ws.iter_rows(min_row=2):
 for c in row: c.alignment=Alignment(vertical='top',wrap_text=True); c.border=Border(top=thin,bottom=thin,left=thin,right=thin)
widths=[32,38,46,34,34,20,18,14,16,18,10,10,10,10,12,12]
for i,wid in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=wid
ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions; wb.save(OUT/'feature_inventory.xlsx')
# Screenshot catalog
cap=json.loads((OUT/'capture_inventory.json').read_text())
fields=['Page name','URL','Final URL','Viewport','Screenshot','Title','Visible components summary','Hosts observed','Observed/inferred','Confidence']
rows=[]
for r in cap:
 rows.append({'Page name':r['name'],'URL':r['url'],'Final URL':r.get('final_url',''),'Viewport':f"{r.get('viewport',{}).get('width')}x{r.get('viewport',{}).get('height')}",'Screenshot':str(Path(r.get('screenshot','')).relative_to(OUT)) if r.get('screenshot') else '', 'Title':r.get('title',''),'Visible components summary':r.get('body_text_sample','')[:400].replace('\n',' | '),'Hosts observed':', '.join(r.get('hosts',[])),'Observed/inferred':'Observed','Confidence':'High'})
with (OUT/'screenshot_catalog.csv').open('w',newline='',encoding='utf-8') as f:
 w=csv.DictWriter(f,fieldnames=fields); w.writeheader(); w.writerows(rows)
wb2=Workbook(); ws2=wb2.active; ws2.title='Screenshot Catalog'; ws2.append(fields)
for r in rows: ws2.append([r[f] for f in fields])
for c in ws2[1]: c.fill=fill; c.font=font; c.alignment=Alignment(wrap_text=True,horizontal='center')
for row in ws2.iter_rows(min_row=2):
 for c in row: c.alignment=Alignment(vertical='top',wrap_text=True); c.border=Border(top=thin,bottom=thin,left=thin,right=thin)
for i,wid in enumerate([24,48,48,14,42,30,80,60,18,12],1): ws2.column_dimensions[get_column_letter(i)].width=wid
ws2.freeze_panes='A2'; ws2.auto_filter.ref=ws2.dimensions; wb2.save(OUT/'screenshot_catalog.xlsx')
# Diagrams
def box(x,y,w,h,t,s='',fillc='#fff'):
 return f'<rect x="{x}" y="{y}" width="{w}" height="{h}" rx="14" fill="{fillc}" stroke="#111827" stroke-width="1.4"/><text x="{x+w/2}" y="{y+25}" text-anchor="middle" font-family="Arial" font-size="15" font-weight="700" fill="#111827">{html.escape(t)}</text>'+(f'<text x="{x+w/2}" y="{y+47}" text-anchor="middle" font-family="Arial" font-size="11" fill="#374151">{html.escape(s)}</text>' if s else '')
def arr(x1,y1,x2,y2,l=''):
 s=f'<line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" stroke="#4B5563" stroke-width="2" marker-end="url(#arrow)"/>'
 if l: s+=f'<text x="{(x1+x2)/2}" y="{(y1+y2)/2-6}" text-anchor="middle" font-family="Arial" font-size="10" fill="#374151">{html.escape(l)}</text>'
 return s
hdr='<defs><marker id="arrow" markerWidth="10" markerHeight="10" refX="7" refY="3" orient="auto" markerUnits="strokeWidth"><path d="M0,0 L0,6 L8,3 z" fill="#4B5563"/></marker></defs>'
# product arch
svg=f'<svg xmlns="http://www.w3.org/2000/svg" width="1600" height="1050" viewBox="0 0 1600 1050">{hdr}<rect width="1600" height="1050" fill="#F8FAFC"/><text x="800" y="48" text-anchor="middle" font-family="Arial" font-size="30" font-weight="800">Human API / LexisNexis Health Intelligence Product Architecture</text><text x="800" y="78" text-anchor="middle" font-family="Arial" font-size="14" fill="#4B5563">Confirmed public architecture + clearly marked inferred analytics layers</text>'
for args in [(60,130,220,84,'Enterprise buyer','Carrier / life sciences / health org','#E0F2FE'),(350,130,220,84,'Sales site','LexisNexis product pages','#fff'),(640,130,220,84,'Contract + config','Order types / outputs / delivery','#FEF3C7'),(930,130,220,84,'Customer system','Application / underwriting workflow','#EDE9FE'),(1220,130,260,84,'Consumer / applicant','Consent, portal tasks, providers','#DCFCE7'),(80,330,250,90,'Connect / ConsumerLink','Hosted consent and task UI','#DCFCE7'),(390,330,250,90,'Admin API','Users/orders/actions/subscriptions','#fff'),(700,330,250,90,'Evidence Orchestration','Channels, timeouts, APS pivot','#FEE2E2'),(1010,330,250,90,'Data Network','EHR/HIE/portals/labs/Rx/wearables','#DBEAFE'),(1320,330,220,90,'Raw Records','CCD/CCDA/FHIR/source data','#fff'),(120,560,240,90,'Normalization','Parsing, dedupe, coding, linking','#F3F4F6'),(430,560,240,90,'Reports Engine','Clinical History, Highlights, FHIR, APS','#FFEDD5'),(740,560,240,90,'Insights Layer','Medical Insights, mortality predictors','#FCE7F3'),(1050,560,240,90,'Delivery Layer','API, webhooks, SFTP, presigned URL','#E0E7FF'),(1360,560,180,90,'Portal','View/manage/download','#fff'),(320,790,260,90,'Security/Compliance','HIPAA, BAA, AES-256, logs','#ECFDF5'),(670,790,260,90,'LexisNexis Assets','Identity resolution + risk data','#FDE68A'),(1020,790,260,90,'Ovexis Opportunity','AI longitudinal health intelligence','#E0F2FE')]: svg+=box(*args)
for a in [(280,172,350,172,''),(570,172,640,172,''),(860,172,930,172,''),(1150,172,1220,172,'embedded UX'),(1340,214,205,330,'consent/tasks'),(330,375,390,375,''),(640,375,700,375,''),(950,375,1010,375,''),(1260,375,1320,375,''),(1430,420,240,560,'raw+normalized'),(360,605,430,605,''),(670,605,740,605,''),(980,605,1050,605,''),(1290,605,1360,605,''),(800,650,800,790,'governance'),(800,790,800,650,'identity/risk enrichment')]: svg+=arr(*a)
svg+='<text x="70" y="990" font-family="Arial" font-size="13" fill="#374151">Confirmed: APIs, reports, channels, security claims. Inferred: internal scoring/insight algorithms and exact databases.</text></svg>'
(OUT/'product_architecture.svg').write_text(svg,encoding='utf-8')
# journey
steps=[('Visitor','Search/PR/referral'),('Marketing','Lexis pages/resources'),('Demo request','Enterprise form'),('Contract','BAA/config/order types'),('Integration','Admin API + Connect'),('Applicant','consent/provider search'),('Retrieval','EHR/HIE/APS/wearables'),('Processing','normalize/dedupe/link'),('Reports','Clinical History/FHIR/API'),('Decisioning','underwriter/model'),('Tasks','pended issues'),('Delivery','webhooks/API/SFTP'),('Renewal','usage/outcomes/expansion')]
svg=f'<svg xmlns="http://www.w3.org/2000/svg" width="1250" height="720" viewBox="0 0 1250 720">{hdr}<rect width="1250" height="720" fill="#F8FAFC"/><text x="625" y="54" text-anchor="middle" font-family="Arial" font-size="28" font-weight="800">Human API Enterprise + Consumer User Journey</text>'
x0,y0,w,h,g=60,150,160,78,60; pos=[]
for i,(t,s) in enumerate(steps):
 row=i//5; col=i%5; x=x0+col*(w+g); y=y0+row*180; pos.append((x,y)); svg+=box(x,y,w,h,t,s,'#ECFEFF' if i%2==0 else '#fff')
for i in range(len(steps)-1):
 x,y=pos[i]; nx,ny=pos[i+1]; svg+=arr(x+w,y+h/2,nx,ny+h/2) if y==ny else arr(x+w/2,y+h,nx+w/2,ny)
svg+='<rect x="70" y="630" width="1110" height="50" rx="12" fill="#111827"/><text x="625" y="662" text-anchor="middle" font-family="Arial" font-size="14" fill="#fff">Key: Human API has two users — the enterprise operator and the consumer/applicant whose data is retrieved with consent.</text></svg>'
(OUT/'user_journey_diagram.svg').write_text(svg,encoding='utf-8')
# data flow
nodes=[(40,150,210,80,'Consumer Consent','Connect / ConsumerLink','#DCFCE7'),(310,80,210,80,'Patient Portals','Consumer-mediated','#fff'),(310,200,210,80,'EHR/HIE Networks','HIPAA-authorized','#fff'),(310,320,210,80,'APS Vendors','Traditional retrieval','#fff'),(310,440,210,80,'Wearables/Labs/Rx','Apps, pharmacies, labs','#fff'),(600,230,230,90,'Human API Ingestion','Auth, retrieve, raw store','#DBEAFE'),(900,130,230,90,'Raw Source Records','CCD, CCDA, FHIR, PDFs','#F3F4F6'),(900,320,230,90,'Normalization Layer','Codes, dedupe, parse, link','#FFEDD5'),(1200,220,230,90,'Outputs','Reports, APIs, webhooks','#E0E7FF')]
svg=f'<svg xmlns="http://www.w3.org/2000/svg" width="1500" height="650" viewBox="0 0 1500 650">{hdr}<rect width="1500" height="650" fill="#F8FAFC"/><text x="750" y="50" text-anchor="middle" font-family="Arial" font-size="28" font-weight="800">Healthcare Data Flow Diagram</text>'
for n in nodes: svg+=box(*n)
for a in [(250,190,310,120,'select/search'),(250,190,310,240,'auth'),(250,190,310,360,'authorization'),(250,190,310,480,'device/app'),(520,120,600,260,''),(520,240,600,260,''),(520,360,600,260,''),(520,480,600,260,''),(830,260,900,175,'raw'),(830,285,900,365,'process'),(1130,175,1200,265,''),(1130,365,1200,265,'normalized')]: svg+=arr(*a)
svg+='</svg>'; (OUT/'healthcare_data_flow.svg').write_text(svg,encoding='utf-8')
# AI arch inferred
svg=f'<svg xmlns="http://www.w3.org/2000/svg" width="1400" height="760" viewBox="0 0 1400 760">{hdr}<rect width="1400" height="760" fill="#F8FAFC"/><text x="700" y="50" text-anchor="middle" font-family="Arial" font-size="28" font-weight="800">Inferred Analytics / AI Architecture</text><text x="700" y="78" text-anchor="middle" font-family="Arial" font-size="13" fill="#4B5563">No public evidence of a consumer LLM chatbot; confirmed analytics include parsing, normalization, summaries, Medical Insights and mortality predictors.</text>'
for args in [(80,140,230,85,'Records + APS','EHR/CCD/PDF/labs','#fff'),(380,140,230,85,'Parsing/OCR/NLP','Confirmed parsing; exact models unknown','#FEE2E2'),(680,140,230,85,'Normalization','Codes, units, dedupe, timelines','#FFEDD5'),(980,140,230,85,'Feature Extraction','Vitals, labs, impairments, social hx','#E0F2FE'),(230,360,230,85,'Rules/Guidelines','Sufficiency, order type, APS delay','#EDE9FE'),(530,360,230,85,'Summary Generation','Highlights, Clinical History, APS summary','#FCE7F3'),(830,360,230,85,'Risk/Decision Inputs','Mortality predictors, underwriting attrs','#DCFCE7'),(1130,360,210,85,'Human Review','Underwriter/case manager','#fff'),(530,580,280,85,'Evaluation + QA','Audit logs, request IDs, support triage','#F3F4F6')]: svg+=box(*args)
for a in [(310,182,380,182,''),(610,182,680,182,''),(910,182,980,182,''),(1095,225,345,360,'features'),(795,225,645,360,'summarize'),(1095,225,945,360,'risk attrs'),(1060,402,1130,402,''),(645,445,670,580,'monitor'),(945,445,670,580,'monitor')]: svg+=arr(*a)
svg+='</svg>'; (OUT/'ai_architecture_inferred.svg').write_text(svg,encoding='utf-8')
# business canvas
items=[('Key Partners','EHR networks, HIEs, patient portals, labs, pharmacies, APS vendors, life insurers, LexisNexis data assets'),('Key Activities','Consent, identity matching, retrieval, normalization, summarization, report delivery, enterprise integration'),('Value Proposition','Consumer-consented health data access and underwriting-ready intelligence at scale'),('Customer Relationships','Enterprise sales, account managers, workflow design, support, technical onboarding'),('Customer Segments','Life insurers, healthcare organizations, life sciences, digital health programs'),('Key Resources','30k+ connections, 270M lives claim, APIs, consent UX, reports engine, LexisNexis identity/risk assets'),('Channels','LexisNexis sales, PR, insurance conferences, developer docs, partner referrals'),('Cost Structure','Network integrations, compliance/security, data ops, support, APS vendor costs, cloud/storage'),('Revenue Streams','Enterprise contracts, per-order/per-report/API volume, professional services, add-on insights')]
pos=[(40,120),(330,120),(620,120),(910,120),(1200,120),(40,440),(330,440),(620,440),(910,440)]
svg='<svg xmlns="http://www.w3.org/2000/svg" width="1500" height="980" viewBox="0 0 1500 980"><rect width="1500" height="980" fill="#F8FAFC"/><text x="750" y="50" text-anchor="middle" font-family="Arial" font-size="30" font-weight="800">Human API Business Model Canvas</text>'
for (t,txt),(x,y) in zip(items,pos):
 svg+=f'<rect x="{x}" y="{y}" width="260" height="250" rx="16" fill="#FFFFFF" stroke="#CBD5E1"/><text x="{x+130}" y="{y+32}" text-anchor="middle" font-family="Arial" font-size="16" font-weight="800">{html.escape(t)}</text>'
 for j,line in enumerate(textwrap.wrap(txt,32)[:9]): svg+=f'<text x="{x+20}" y="{y+68+j*22}" font-family="Arial" font-size="13" fill="#374151">{html.escape(line)}</text>'
svg+='</svg>'; (OUT/'business_model_canvas.svg').write_text(svg,encoding='utf-8')
print('features',len(features),'screenshots',len(rows))
