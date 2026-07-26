from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
OUT=Path('/home/user/humanapi_research')
fill=PatternFill('solid',fgColor='111827'); font=Font(color='FFFFFF',bold=True); thin=Side(style='thin',color='D1D5DB')
def save_table(name, headers, rows, widths):
    wb=Workbook(); ws=wb.active; ws.title=name[:31]; ws.append(headers)
    for r in rows: ws.append(r)
    for c in ws[1]: c.fill=fill; c.font=font; c.alignment=Alignment(horizontal='center',wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for c in row: c.alignment=Alignment(vertical='top',wrap_text=True); c.border=Border(top=thin,bottom=thin,left=thin,right=thin)
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions; wb.save(OUT/f'{name}.xlsx')
comp_headers=['Company','Category','Primary customer','Data access','AI/analytics','Pricing/GTM','Where it beats Human API','Where Human API beats it','Implication for Ovexis']
comp_rows=[
['Human API / LexisNexis','Consumer-consented health data + underwriting intelligence','Life insurers, healthcare, life sciences','EHR/HIE/portals/labs/Rx/wearables/APS','Parsing, normalization, Medical Insights, summaries','Enterprise contract','Baseline','Baseline','Build AI-native and developer-first competitor'],
['Health Gorilla','QHIN/QHIO interoperability network','Payers/providers/health orgs','QHIN/QHIO, EHR, lab, ADT','Data exchange and security','Enterprise','Regulatory network position','Underwriting-specific workflow/reporting','Regulatory designation matters'],
['Particle Health','Medical record API and insights','Developers, VBC, payers/providers','National HIE networks, 320M+ claim','Product-ready insights','PLG + enterprise','Developer UX, API simplicity','Insurance evidence orchestration','Use PLG as wedge'],
['Apple Health','Consumer health records/device ecosystem','Consumers, providers, researchers','Device data, Health Records where supported','Device insights','Platform/ecosystem','Consumer trust/distribution','Enterprise retrieval/delivery','Own consumer trust layer'],
['Google Health/Fitbit','AI/search/cloud/wearable health','Consumers/orgs/startups','Fitbit, Search, cloud tools','Advanced AI research','Platform/enterprise','AI and distribution','Specific consented records workflow','AI + data access required'],
['Function Health','B2C biomarker membership','Consumers','Quest/Getlabs/labs','Doctor-reviewed insights','Subscription $365/yr public','Consumer UX and brand','Enterprise APIs/data network','Consumer product can sit on Ovexis data'],
['Superpower','B2C longevity membership','Consumers','Labs/wearables/upload','AI protocol and care team','Subscription $199/yr public','B2C AI health UX','Health-record data infra','Study retention loops'],
['OpenEvidence','Clinician medical AI','Clinicians','Medical literature/guidelines','Evidence-grounded clinical AI','Free/ad/enterprise','Clinical evidence and trust','Patient record data pipes','Combine RAG with records'],
['Regacore','Early longevity membership MVP','Consumers India/global','Labs/upload/wearables claimed','AI concierge/digital twin','₹10,800/yr public','Bold UX ambition','Operational network/proof','Avoid unproven claims'],
['Apollo 24/7','India healthcare marketplace','Indian consumers','Provider/pharmacy/lab ecosystem','Care navigation/AI emerging','B2C marketplace','Distribution in India','Standardized longitudinal data','Partner or compete by data layer'],
['Practo','Provider marketplace/booking','Indian consumers/providers','Provider network','Search/recommendation','B2C/B2B','Consumer demand and doctors','Health-data infrastructure','Distribution partner possibility'],
['Tata 1mg','Pharmacy/labs/telehealth','Indian consumers','Labs/pharmacy/consult data','Recommendations/commerce','B2C commerce','Scale and brand','EHR normalization APIs','Potential channel/competitor'],
['Healthify','AI nutrition/wellness','Consumers/employers','User-entered/wearable data','AI coaching','Subscription/employer','Behavior change','Verified clinical records','Layer verified data under coaching']]
save_table('competitive_comparison_matrix',comp_headers,comp_rows,[24,28,24,38,32,24,36,36,42])
risk_headers=['Risk','Type','Classification','Likelihood','Impact','Evidence','Mitigation for Ovexis']
risk_rows=[
['Wrong-patient match','Clinical/security','🟡 Strong inference','Medium','Very high','Matching uses demographics/SSN/address/provider details; LexID linking claimed','Use deterministic identity proof, source provenance, manual review for uncertain matches'],
['Incomplete retrieval','Product/clinical','🟡 Strong inference','High','High','Docs discuss timeouts, pended tasks, sufficiency logic','Gap detection, source coverage score, explicit “unknown” states'],
['Summary omission','AI/clinical','🟡 Strong inference','Medium','High','Reports are shortened/summarized; APS summarization exists','Citations, raw record access, QA sampling, adverse-impact checks'],
['Consent ambiguity','Regulatory','🟡 Strong inference','Medium','Very high','Consumer initiated consent is core; high-stakes insurance use','Consent ledger, purpose limitation, revocation UI, clear language'],
['Webhook misdelivery','Security','🟡 Strong inference','Medium','High','Reports delivered via HTTP/SFTP/presigned URL','Signed webhooks, mTLS, allowlists, delivery tests'],
['Connector breakage','Operational','🟡 Strong inference','High','Medium','30K+ connections imply high maintenance','Connector monitoring and fallback pipeline'],
['Docs staleness','Developer','🟢 Confirmed','High','Medium','Observed 404 pages and old links','Docs CI, llms.txt verification, changelog ownership'],
['Enterprise sales drag','Business','🟡 Strong inference','High','Medium','No public pricing, long forms, contract required','Self-serve sandbox and transparent tiers'],
['Data-source commoditization','Strategic','🟡 Strong inference','Medium','High','TEFCA/QHIN/FHIR competitors grow','AI insight layer, local network, workflow specialization'],
['Regulatory scrutiny of insurance data','Regulatory','🟡 Strong inference','Medium','High','Insurance underwriting uses health data','Fair-use policies, explainable decisions, compliance review']]
save_table('risk_register',risk_headers,risk_rows,[34,20,20,14,14,60,60])
road_headers=['Stage','Timeline','Probable / Recommended Components','Evidence or Rationale','Ovexis Countermove']
road_rows=[
['MVP','0-6 months','Consent widget, upload/OCR, labs/wearables, normalized profile, API, sandbox','Human API original core was Connect + API + portal','Build developer-first and consumer-trust-first'],
['V2','6-12 months','FHIR model, webhooks, source status, tasks, reports, portal, BAAs','Human API docs show order lifecycle/reports/delivery','Add AI provenance and source quality'],
['V3','12-24 months','EHR/HIE integrations, partner network, identity matching, custom outputs','Human API Health Intelligence evolved to order types/channels','Focus on India/ABHA + selected US niches'],
['V4','24+ months','AI summaries, risk APIs, evidence graph, outcome feedback, marketplace integrations','LexisNexis now emphasizes insights/analytics','Own AI-native longitudinal category']]
save_table('engineering_roadmap_reconstruction',road_headers,road_rows,[18,16,60,60,46])
print('extra tables written')
