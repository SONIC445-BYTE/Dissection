from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date

# Source shortcuts (URLs kept explicit for auditability)
SRC = {
    'about': 'https://www.amboss.com/int/about',
    'features': 'https://www.amboss.com/us/features',
    'students': 'https://www.amboss.com/us/students',
    'clinicians': 'https://www.amboss.com/us/clinicians',
    'attendings': 'https://www.amboss.com/us/clinicians/attendings-apps',
    'ai_clinical': 'https://www.amboss.com/us/clinical-ai-mode',
    'ai_learning': 'https://www.amboss.com/us/ai-mode-learning',
    'ai_innovation': 'https://www.amboss.com/us/ai-innovation',
    'ai_principles': 'https://www.amboss.com/us/ai-principles',
    'assistants': 'https://www.amboss.com/us/newsroom/amboss-assistants',
    'mcp': 'https://www.amboss.com/us/newsroom/amboss-mcp',
    'new_nav': 'https://www.amboss.com/us/newsroom/new-navigation',
    'pricing': 'https://www.amboss.com/us/pricing',
    'medical_schools': 'https://www.amboss.com/us/medical-schools',
    'educator': 'https://www.amboss.com/us/medical-schools/educator-tool',
    'residency': 'https://www.amboss.com/us/residency-programs',
    'support_platform': 'https://support.amboss.com/hc/en-us/articles/360034825692-Platform-overview',
    'support_qbank': 'https://support.amboss.com/hc/en-us/articles/360032477132-Creating-a-Qbank-session',
    'support_modes': 'https://support.amboss.com/hc/en-us/articles/360036038991-Using-Study-Mode-Exam-Mode',
    'support_feature_overview': 'https://support.amboss.com/hc/en-us/articles/360035199871-Feature-Overview',
    'support_search': 'https://support.amboss.com/hc/en-us/articles/4505593712148-Search-Function',
    'support_ai_learning': 'https://support.amboss.com/hc/en-us/articles/43601233276689-AI-Mode-Learning-FAQs',
    'support_ai_clinical': 'https://support.amboss.com/hc/en-us/articles/43829822392721-AI-Mode-Clinical-Care-FAQs',
    'support_educator': 'https://support.amboss.com/hc/en-us/articles/30006381173009-Educator-Tool-Overview',
    'support_assignments': 'https://support.amboss.com/hc/en-us/articles/30007809419665-Creating-Assignments',
    'support_analytics': 'https://support.amboss.com/hc/en-us/articles/30034537472657-Educator-Feature-Data-and-Analytics',
    'support_analysis': 'https://support.amboss.com/hc/en-us/articles/42885439769233-How-does-the-Analysis-work',
    'support_adaptive': 'https://support.amboss.com/hc/en-us/articles/45511273700369-Adaptive-Question-Sessions',
    'support_recs': 'https://support.amboss.com/hc/en-us/articles/360047938152-Study-Recommendations',
    'support_score': 'https://support.amboss.com/hc/en-us/articles/25051203159313-USMLE-Score-Predictor',
    'support_stats': 'https://support.amboss.com/hc/en-us/articles/360032830911-Show-Stats-Feature',
    'support_review': 'https://support.amboss.com/hc/en-us/articles/360035979311-Reviewing-a-Qbank-session',
    'support_clinical': 'https://support.amboss.com/hc/en-us/articles/360046945892-Clinical-Care-Features-Overview',
    'support_management': 'https://support.amboss.com/hc/en-us/articles/6203585104276-Management-Checklists',
    'support_ddx': 'https://support.amboss.com/hc/en-us/articles/4542048524308-Differential-Diagnoses',
    'support_drug': 'https://support.amboss.com/hc/en-us/articles/4410493072660-Clinical-Drug-Database',
    'support_dosing': 'https://support.amboss.com/hc/en-us/articles/6203840407956-Drug-Dosing',
    'support_calc': 'https://support.amboss.com/hc/en-us/articles/6204057842836-Clinical-Calculators',
    'content_policy': 'https://support.amboss.com/hc/en-us/articles/6365572887188-AMBOSS-Content-Policy',
    'cme_program': 'https://support.amboss.com/hc/en-us/articles/15744010801169-Program-Overview',
    'cme_faq': 'https://support.amboss.com/hc/en-us/articles/23670339860369-CME-MOC-FAQs',
    'cme_poc': 'https://support.amboss.com/hc/en-us/articles/23670009661713-Internet-Point-of-Care',
    'mobile': 'https://www.amboss.com/us/mobile-app',
    'anki': 'https://www.amboss.com/us/anki',
    'chrome': 'https://www.amboss.com/us/chrome',
    'gpt': 'https://support.amboss.com/hc/en-us/articles/33872741080081-AMBOSS-GPT',
    'account': 'https://support.amboss.com/hc/en-us/articles/360044684571-Set-up-your-AMBOSS-account',
    'free_trial': 'https://support.amboss.com/hc/en-us/articles/360045124131-Can-I-try-AMBOSS-for-free',
    'membership': 'https://support.amboss.com/hc/en-us/articles/360044237212-AMBOSS-Membership',
    'profile': 'https://support.amboss.com/hc/en-us/articles/360032858551-Edit-your-AMBOSS-profile',
    'delete': 'https://support.amboss.com/hc/en-us/articles/360032508792-Deleting-your-AMBOSS-account-and-profile',
    'sso': 'https://support.amboss.com/hc/en-us/articles/30228751806097-Activating-your-Institutional-License-Trial-Single-Sign-On',
    'email_license': 'https://support.amboss.com/hc/en-us/articles/4407238872084-Activating-your-Institutional-License-Trial-Email-Verification',
    'privacy': 'https://www.amboss.com/us/legal/privacy',
    'terms': 'https://www.amboss.com/us/legal/terms',
    'terms_clinics': 'https://www.amboss.com/us/legal/terms-clinics',
    'joint': 'https://www.amboss.com/us/legal/joint-data-processing',
    'assistants_privacy': 'https://www.amboss.com/us/legal/assistants',
    'accessibility': 'https://www.amboss.com/us/legal/accessibility',
    'funding': 'https://www.amboss.com/us/newsroom/financing-round-and-se-transition',
    'nejm_acq': 'https://www.amboss.com/us/newsroom/nejm-knowledge-plus-acquisition',
    'nejm_courses': 'https://www.amboss.com/us/newsroom/amboss-and-nejm-group-launch-master-classes-in-medicine',
    'novaheal_pr': 'https://www.prnewswire.com/news-releases/amboss-acquires-novaheal-to-strengthen-its-position-in-the-german-nursing-market-and-as-a-hub-for-healthcare-talent-and-knowledge-302265824.html',
    'careers': 'https://careers.amboss.com/',
    'careers_eng': 'https://careers.amboss.com/departments/engineering-product/',
    'shop_job': 'https://careers.amboss.com/jobs/job-09b09075-f834-4935-a4ec-8f5a878493f7/',
    'search_eval': 'https://careers.amboss.com/blog/enhancing-amboss-search-evaluation-with-chatgpt-generated-judgment-lists/',
    'noharm': 'https://arxiv.org/abs/2512.01241',
    'scores': 'https://www.amboss.com/us/usmle/scores',
    'reviews': 'https://www.amboss.com/us/reviews',
}

features = [
    ['Role-based navigation: Learning / Clinical Care / Teaching','Reduce context-switching by organizing the product around user intent and career role',SRC['new_nav'],'Faster entry to relevant tasks; clearer mental model','Supports expansion beyond students into clinicians, nurses, PAs and educators','Medium','Medium','Medium','Low','PM+UX+FE+BE+Data','4-8','High','Platform navigation','Yes','Add patient/family/coach roles for Ovexis','No','Reinvent as longitudinal role graph','Medium','🟢 Confirmed'],
    ['Knowledge Library','Provide evidence-based medical knowledge with article structure, references, images, flowcharts',SRC['features'],'Immediate trusted answers and deep learning context','Core subscription value and content moat','High','Very High','Medium','Medium','Medical editors+illustrators+CMS+Search','18+ ongoing','Critical','Medical content','Yes','Personalize to patient context if regulated','No','Reinvent as patient-specific evidence layer','Strong','🟢 Confirmed'],
    ['8-eyes content review / editorial policy','Quality-control content through multiple physician and copy-editing reviews',SRC['content_policy'],'Trust, safety, authority','Brand moat; lowers clinical-risk perception','Medium','Very High','Low','Medium','Clinical editorial operations','Ongoing','Critical','Content QA','Yes','Formalize evidence grades and update SLA','No','Use transparent provenance ledger','Strong','🟢 Confirmed'],
    ['Sources and references in articles','Expose source citations and DOI/link-outs at article end and throughout platform',SRC['content_policy'],'Verifiability and defensibility','Trust and clinician adoption','Medium','High','Medium','Low','Medical editors+CMS','6-12','Critical','Evidence','Yes','Add live guideline-diff alerts','No','Evidence graph with freshness/conflict scores','Strong','🟢 Confirmed'],
    ['High-Yield mode','Hide lower-yield article content relative to a selected study objective',SRC['support_feature_overview'],'Focus and reduced cognitive load during exam prep','Retention and exam-result differentiation','Medium','High','Medium','Low','Clinical tagging+FE','6-9','High','Learning UX','Yes','Make dynamic by exam date and goals','No','Personal yield mode for health goals','Medium','🟢 Confirmed'],
    ['Key Exam Info / Learning Radar','Highlight high-yield text and red underline content linked to previously missed questions',SRC['support_feature_overview'],'Turns mistakes into targeted remediation','Improves Qbank-library loop and retention','High','High','High','Low','Data science+Clinical tagging+FE','9-12','High','Adaptive learning','Yes','Use spaced repetition and mastery thresholds','No','Reinvent for longitudinal disease-risk gaps','Strong','🟢 Confirmed'],
    ['Personal highlighting','Let users highlight text inside AMBOSS articles',SRC['support_feature_overview'],'Personalized studying and review','Increases stickiness and user-generated context','Low','Low','Low','Low','FE+BE','1-2','Medium','Personalization','Yes','Add AI summarization of highlights','No','Use as memory substrate','Weak','🟢 Confirmed'],
    ['Notes in article sections','Add personal notes at section level; sync to Knowledge app',SRC['support_feature_overview'],'Capture context near the knowledge object','Retention and re-engagement','Medium','Low','Medium','Low','FE+BE','2-4','Medium','Personalization','Yes','Add note-to-task conversion','No','Collaborative care notes with governance','Medium','🟢 Confirmed'],
    ['Collections / Favorites','Save articles/media/substances to custom collections',SRC['features'],'Organize repeated workflows and curricula','Retention; group/educator workflows','Medium','Low','Medium','Low','FE+BE','2-4','Medium','Personalization','Yes','Add auto-curated collections','No','Care-plan collections in Ovexis','Medium','🟢 Confirmed'],
    ['Split View','Open related Library article next to current article or Qbank',SRC['support_feature_overview'],'No tab-hopping; efficient context building','Differentiates integrated platform from single-purpose Qbanks','Medium','Low','Medium','Low','FE','3-5','High','Workflow UX','Yes','Extend to patient timeline + guideline side-by-side','No','Multi-pane clinical reasoning workspace','Medium','🟢 Confirmed'],
    ['Qbank','Board-style questions with explanations, images, references and library linkage',SRC['features'],'Active recall and exam readiness','Major subscription/add-on revenue','Very High','Very High','High','Low','Medical writers+assessment psychometrics+Engineering','24+ ongoing','Critical','Assessment','Yes for education','Do not use generic quizzes for clinical advice','No','Reinvent as adaptive health literacy/problem-solving','Strong','🟢 Confirmed'],
    ['Custom Qbank filters','Filter by exam, discipline, system, topic, difficulty, status, saved folders, marked/images',SRC['support_qbank'],'Targeted study and deliberate practice','Increases usage depth; supports many exams','Medium','High','Medium','Low','FE+BE+taxonomy','4-8','High','Assessment','Yes','Add natural-language filter builder','No','Cohort-specific adaptive filters','Medium','🟢 Confirmed'],
    ['Adaptive Question Sessions','Select questions based on study objective, unanswered status, modeled weakness and spaced repetition',SRC['support_adaptive'],'Saves time and targets weak areas','Improves outcomes and retention loop','High','High','High','Low','Data science+BE+FE','9-12','High','Adaptive learning','Yes','Calibrate transparently to avoid frustration','No','Adaptive longitudinal risk-learning engine','Strong','🟢 Confirmed'],
    ['Study Mode','Qbank mode with Key Info, Attending Tip, notes, show answer, answer explanations, reset, rule-out, lab values',SRC['support_modes'],'Teaches while testing; scaffolds reasoning','Differentiated learning experience','Medium','High','Medium','Low','FE+assessment+content','6-10','High','Assessment UX','Yes','Add rubric explaining reasoning path','No','Simulation coach mode','Strong','🟢 Confirmed'],
    ['Exam Mode','USMLE/Shelf-like timed interface with mark, lab values, notes, calculator, reverse color, text zoom, pause/lock/end block',SRC['support_modes'],'Exam familiarity and lower test-day friction','Credibility for exam prep','Medium','Medium','Medium','Low','FE+Assessment','4-8','High','Assessment UX','Yes','Map to each board UI exactly where licensed','No','Credentialing simulations for health literacy','Medium','🟢 Confirmed'],
    ['Attending Tip','Provide hints that guide clinical reasoning without revealing answer',SRC['support_modes'],'Supports novice reasoning and confidence','Improves completion; reinforces brand as clinician-built','Medium','High','Low','Low','Clinical authors+FE','3-6','High','Learning aid','Yes','Make adaptive to learner level','No','Clinician mentor persona','Medium','🟢 Confirmed'],
    ['Key Info in Qbank','Highlight essential details in vignettes',SRC['support_modes'],'Trains signal detection in clinical vignettes','Improves perceived helpfulness for early learners','Medium','High','Low','Low','Clinical authors+FE','3-6','High','Learning aid','Yes','Fade hints over time','No','Progressive hinting engine','Medium','🟢 Confirmed'],
    ['Hammer difficulty rating','Tag questions from 1 to 5 hammers based on answer statistics',SRC['features'],'Controls challenge and prevents over/under-training','Segmentation and personalization','Medium','Medium','Medium','Low','Analytics+Assessment','3-6','High','Assessment metadata','Yes','Warn when 5-hammer may be overkill','No','Difficulty calibrated by clinical stakes','Medium','🟢 Confirmed'],
    ['Answer explanations with hyperlinks','Review answer choices with linked terms/articles and correct-answer article',SRC['support_review'],'Turns every question into content graph traversal','Strengthens integrated qbank-library moat','Medium','High','Medium','Low','Content graph+FE','6-12','Critical','Content graph','Yes','Add source-level evidence grade','No','Explain with patient-specific analogies','Strong','🟢 Confirmed'],
    ['Show Stats / peer answer distribution','Reveal percentage of users selecting each answer; first attempts >5 seconds; nightly update',SRC['support_stats'],'Normalizes difficulty and reveals distractor traps','Engagement; social proof','Medium','Low','Medium','Low','Analytics+FE','3-5','Medium','Analytics','Yes','Segment by cohort/exam date','No','Community benchmark with privacy guardrails','Medium','🟢 Confirmed'],
    ['Session Analysis','Post-session overview, question review, peer stats, explanations and article links',SRC['support_review'],'Immediate feedback and targeted learning','Re-engagement into next session','Medium','Medium','Medium','Low','Analytics+FE','4-8','High','Analytics','Yes','Prioritize next best action','No','After-action review for health plans','Medium','🟢 Confirmed'],
    ['Study Recommendations','Recommend topics/Qbank sessions ordered by individual need and exam importance',SRC['support_recs'],'Actionable next steps instead of raw analytics','Retention and outcome proof','High','High','High','Low','Data science+taxonomy+FE','9-12','Critical','Adaptive learning','Yes','Expose why a recommendation was made','No','Personalized health gap recommendations','Strong','🟢 Confirmed'],
    ['EPC / Peer Comparison / Expected Score','IRT-based performance modeling and peer percentile from Qbank activity',SRC['support_analysis'],'Objective readiness tracking','Strong retention and differentiation','Very High','High','High','Low','Data science+psychometrics','12+','Critical','Analytics','Yes','Validate externally and show uncertainty','No','Readiness score for health goals','Strong','🟢 Confirmed'],
    ['USMLE/COMLEX Score Predictor','Predict exam score/range/pass likelihood from practice exams using mixed-effects and Bayesian models',SRC['support_score'],'Reduces uncertainty and anxiety','Lead magnet and subscription value','High','Medium','Medium','Low','Data science+FE','6-12','High','Analytics','Yes','Avoid overclaiming; publish calibration','No','Predictive risk trajectory with confidence intervals','Strong','🟢 Confirmed'],
    ['Study Plans','Premade and custom plans by exam, end date, study days/hours, systems/disciplines',SRC['support_platform'],'Transforms broad curriculum into schedule','Increases daily habit retention','Medium','Medium','Medium','Low','PM+FE+BE+content','4-8','High','Planning','Yes','Add calendar integrations and streaks','No','Adaptive care-plan scheduler','Medium','🟢 Confirmed'],
    ['Self-Assessments','160 true-to-life questions over timed blocks with score report and recommendations',SRC['students'],'Baseline and exam simulation','Seasonal acquisition, paid upsell and retention','High','Very High','Medium','Low','Assessment+Data+FE','9-18','High','Assessment','Yes','Use as viral cohort event','No','Preventive health self-assessment with clinical governance','Strong','🟢 Confirmed'],
    ['Mobile Knowledge app offline','Clinical/library access on iOS/Android with offline mode',SRC['mobile'],'Use on wards without reliable Wi-Fi','Increases point-of-care usage and retention','High','Medium','High','Medium','Mobile+Content sync','9-18','Critical','Mobile','Yes','Add safe patient-context notes only if compliant','No','Offline longitudinal health vault','Medium','🟢 Confirmed'],
    ['Mobile Qbank app','Continue Qbank sessions and practice offline',SRC['mobile'],'Practice during downtime','Raises usage frequency','High','Medium','High','Low','Mobile+sync','6-12','High','Mobile assessment','Yes','Improve mobile highlighting and Anki support','No','Microlearning over health timeline','Medium','🟢 Confirmed'],
    ['Search Function','Search terms to articles/media/drugs with overview cards and AI-generated related questions',SRC['support_search'],'Fast retrieval and discovery','Core daily habit and CME entry','High','Medium','High','Medium','Search engineering+content graph','12+','Critical','Search','Yes','Unify with patient data and source provenance','No','Longitudinal semantic health search','Strong','🟢 Confirmed'],
    ['Semantic Search','Handle shorthand, symptoms, vague terms, spelling, trade names/abbreviations in multiple languages',SRC['features'],'Finds relevant content despite messy clinical queries','Differentiates from static reference','High','High','High','Medium','Search/ML+Clinical taxonomy','12+','Critical','AI search','Yes','Ground against structured patient timeline','No','Multi-modal health search','Strong','🟢 Confirmed'],
    ['AI Shortcuts / Related Questions','AI snippets and questions alongside search results',SRC['features'],'Fast answer preview without leaving flow','Improves search success and conversion','Medium','Medium','High','Medium','ML+Search+FE','6-12','High','AI search','Yes','Label reviewed vs unreviewed clearly','No','Next-question prompts for patients/clinicians','Medium','🟢 Confirmed'],
    ['AI Mode Clinical Care','Clinical AI search agent over AMBOSS articles, drug DB and selected guidelines; summary plus source links',SRC['ai_clinical'],'Natural-language evidence navigation','AI positioning and clinician expansion','Very High','Very High','Very High','High','AI+Search+Clinical+Safety','12-24','Critical','Clinical AI','Yes as evidence navigation','Do not copy without safety/regulatory model','No','Build longitudinal patient-context safe AI','Strong','🟢 Confirmed'],
    ['AI Mode Learning','AI study copilot for questions/uploads, explanations, Qbank/Anki recommendations and progress integration',SRC['ai_learning'],'Turns confusion into next steps','Defends against generic AI and improves retention','Very High','High','Very High','Medium','AI+Content+Data+FE','12-18','Critical','Learning AI','Yes','Add explainability and study-memory controls','No','Personal health tutor over individual data','Strong','🟢 Confirmed'],
    ['AMBOSS Assistants beta','Context-aware article assistants for learn/practice/teach workflows including DDx, dot phrases, consult prep, teaching tools',SRC['assistants'],'Guided AI use cases reduce prompt burden','Platformizes AI assistants; expands workflows','Very High','Very High','Very High','High','AI platform+Clinical+UX','12-24','High','AI assistants','Yes selectively','Avoid generating clinical docs without governance','No','Assistant marketplace with audited tools','Future strong','🟢 Confirmed'],
    ['AMBOSS GPT','Custom GPT that queries AMBOSS library and returns article links; free with quotas depending email',SRC['gpt'],'Meet users inside ChatGPT','Top-of-funnel and brand extension','Medium','Medium','Medium','Medium','Integration+RAG','2-4','Medium','Integration','Yes','Use as distribution but not core experience','No','External AI surface with verified handoff','Medium','🟢 Confirmed'],
    ['AMBOSS MCP','MCP server gives AI agents access to articles, drug monographs, flowcharts, calculators, clinical scores, patient cases; limited partner access',SRC['mcp'],'Developer/agent ecosystem access to trusted content','Potential platform/API moat','Very High','High','Very High','High','API+Auth+Search+BD','12+','High','Developer platform','Yes conceptually','Expose only safe, licensed scopes','No','Ovexis Health MCP over patient+evidence graph','Future strong','🟢 Confirmed'],
    ['Anki add-on','Popup definitions, side-by-side articles, Qbank sessions from cards, mobile popup support',SRC['anki'],'Fits entrenched med-student workflow','Distribution and retention loop','High','Medium','High','Low','Plugin+Content matching','6-12','Critical','Integration','Yes','Integrate bidirectionally with missed-question cards','No','Personal memory layer','Strong','🟢 Confirmed'],
    ['AnKing card matching','Get Anki cards linked to Qbank questions via AnKing deck',SRC['anki'],'Reduces workflow friction after missed questions','Increases switching cost','High','Medium','High','Low','Content mapping+integration','6-12','High','Integration','Yes','Automate deeper card creation carefully','No','Health education cards from personal data','Medium','🟢 Confirmed'],
    ['Chrome extension','Medical-term highlights/popups across websites, Qbanks and journals; free definitions; full articles require account',SRC['chrome'],'Ambient knowledge without leaving reading context','Top-of-funnel; embeds AMBOSS in competitor workflows','High','Medium','High','Medium','Browser extension+NLP','6-12','High','Integration','Yes','Add user-controlled sidebar/new-tab options','No','Health-context browser copilot','Strong','🟢 Confirmed'],
    ['AHFS drug database','Integrated AHFS clinical drug information, drug search, monographs and up-to-date source provider content',SRC['support_drug'],'Drug reference inside same workflow','Clinician credibility and subscription value','High','High','Medium','Medium','Licensing+integration+search','6-12','Critical','Drug reference','Yes if licensed','Add interaction checking with patient meds if compliant','No','Medication intelligence with patient-specific safety','Strong','🟢 Confirmed'],
    ['Drug dosing in articles','Dosage buttons from reliable treatment protocols in Clinical Care view',SRC['support_dosing'],'Faster prescribing reference','Point-of-care value','Medium','Very High','Medium','High','Clinical editors+drug data','6-12','Critical','Clinical workflow','Yes','Add renal/pregnancy personalization only with governance','No','Context-aware dosing guardrails','Strong','🟢 Confirmed'],
    ['Differential Diagnoses','Targeted differential diagnosis lists and cross-out workflow',SRC['support_ddx'],'Rapid diagnostic thinking','Clinical workflow engagement','Medium','Very High','Medium','High','Clinical editors+FE','6-12','High','Clinical workflow','Yes','Pair with probabilities/evidence cautiously','No','Patient-specific DDx with clinician review','Medium','🟢 Confirmed'],
    ['Management Checklists','Urgent-condition step checklists with task check-off',SRC['support_management'],'Prevents missed inpatient steps','Point-of-care retention and safety perception','Medium','Very High','Medium','High','Clinical editors+FE','6-12','High','Clinical workflow','Yes','Make local/SOP-specific and audit-aware','No','Care orchestration checklist engine','Strong','🟢 Confirmed'],
    ['Clinical calculators via QxMD','Embedded risk scores, dose adjustments, lab conversions',SRC['support_calc'],'Reduces calculator-hunting and errors','Clinician utility; third-party leverage','Medium','High','Medium','Medium','Integration+Clinical QA','3-6','High','Clinical tools','Yes','Track formula provenance and limitations','No','Patient-data auto-fill calculators','Medium','🟢 Confirmed'],
    ['Guidelines inside articles','Access current evidence-based protocols and external guideline links',SRC['features'],'Grounds decisions in standards','Trust and defensibility','High','Very High','Medium','Medium','Content ingestion+QA','12+','Critical','Evidence','Yes','Highlight conflicts and regional differences','No','Guideline-diff engine','Strong','🟢 Confirmed'],
    ['CME / Internet Point-of-Care','Award 0.5 AMA PRA Category 1 Credits per eligible clinical question researched; board review credits too',SRC['cme_faq'],'Turns daily search into professional credit','Clinician retention and differentiation','High','High','Medium','High','Accreditation+tracking+reporting','12+','High','CME','Yes','Add specialty-specific dashboards','No','Continuing-health education credits','Strong','🟢 Confirmed'],
    ['Educator Tool roles','Learner/Educator/Admin roles for assignments, analytics and institution groups',SRC['support_educator'],'Faculty management and learner support','B2B institutional revenue','High','Medium','High','Medium','RBAC+FE+BE','9-18','Critical','Institution','Yes','Add LTI/Canvas if not already','No','Care-team learning admin console','Strong','🟢 Confirmed'],
    ['Assignment creation','Questions/articles, filters by system/discipline/article/competency/difficulty, up to 1,000 questions, collaboration',SRC['support_assignments'],'Saves faculty time and standardizes teaching','Enterprise stickiness','High','High','High','Medium','Assessment+RBAC+FE','9-12','High','Institution','Yes','Use AI file matching with audit','No','Protocol-based patient education assignments','Strong','🟢 Confirmed'],
    ['Educator analytics dashboards','Assignment insights, institution insights, per-user performance, Qs completed, final/initial %, time per Q, recommendations',SRC['support_analytics'],'Find at-risk learners and remediate','Enterprise ROI and renewals','High','Medium','High','Medium','Analytics+RBAC+Data','9-18','Critical','Institution analytics','Yes','Add privacy-preserving cohort analytics','No','Clinical team performance+education analytics','Strong','🟢 Confirmed'],
    ['Institutional license activation and SSO','Email verification, SSO, access codes/direct invite/pre-set accounts; SSO re-verification every 12 months',SRC['sso'],'Frictionless school/clinic access','B2B distribution and lower CAC','Medium','Low','High','Medium','IAM+Support+SalesOps','6-12','Critical','Identity','Yes','Integrate modern SCIM/LTI where possible','No','Enterprise consent and identity fabric','Medium','🟢 Confirmed'],
    ['Profile / study objective settings','Current role, study objective, disability-related time accommodation and account tabs',SRC['profile'],'Personalizes content and accessibility','Data foundation for recommendations','Medium','Low','Medium','Medium','FE+BE+Privacy','3-6','High','Identity/personalization','Yes','Make settings transparent and exportable','No','Health/persona context with consent','Medium','🟢 Confirmed'],
    ['Account deletion','Users can delete account; data permanently erased; subscription cancellation separate',SRC['delete'],'Privacy control and trust','Compliance hygiene','Medium','Low','Medium','Medium','BE+Support+Legal','3-6','High','Privacy','Yes','Clarify subscription/account coupling','No','Consent/data-rights center','Medium','🟢 Confirmed'],
    ['Institution usage report','Institutions can request cohort usage reports based on AMBOSS activity/Qbank performance',SRC['medical_schools'],'Curriculum insight and at-risk identification','Sales wedge for institutional expansion','Medium','Medium','High','Medium','Data team+Sales','3-6','Medium','Institution analytics','Yes','Automate with privacy thresholds','No','Population insights with consent','Medium','🟢 Confirmed'],
    ['Group discounts','Students/residents can coordinate group discounts; bigger group = better deal',SRC['pricing'],'Lower price and peer coordination','Viral campus acquisition','Low','Low','Low','Low','Marketing+Ops','1-2','Medium','Growth','Yes','Add referral analytics and ambassador transparency','No','Community-led preventive-health cohorts','Medium','🟢 Confirmed'],
    ['Student Life membership','Unlimited platform through end of PGY-1',SRC['students'],'Longitudinal access across med school/residency','High upfront LTV and lock-in','Low','Low','Low','Low','Pricing+Billing','1-2','High','Pricing','Yes conceptually','Avoid overlong lock-in without ongoing value','No','Lifetime health intelligence membership','Medium','🟢 Confirmed'],
    ['30-day refund','No-questions-asked refund for direct purchases',SRC['pricing'],'Reduces purchase risk','Conversion lift and trust','Low','Low','Low','Low','Billing+Support','1','Medium','Commerce','Yes','Automate/refund UX','No','Trust guarantee','Weak','🟢 Confirmed'],
    ['AI data handling guardrails','No patient data/PHI intended; input stored for optimization; not used for LLM training; OpenAI/Anthropic/MongoDB processors',SRC['assistants_privacy'],'Privacy clarity and clinical-risk boundary','Legal risk control','High','High','High','Very High','Legal+Security+AI','6-12','Critical','Privacy/Safety','Yes policy rigor','Do not copy no-PHI if Ovexis needs patient intelligence; build compliant PHI stack','No','HIPAA-grade patient-consented AI architecture','Strong','🟢 Confirmed'],
    ['Accessibility statement and roadmap','WCAG 2.1 AA target for in-scope commerce; known exceptions and roadmap',SRC['accessibility'],'Improves inclusivity and legal compliance','Enterprise procurement trust','Medium','Low','Medium','Medium','UX+QA+Engineering+Legal','Ongoing','High','Accessibility','Yes','Make all core content in scope, not only commerce','No','Accessibility-by-default health OS','Medium','🟢 Confirmed'],
]

# Evidence register core claims
claims = [
    ['AMBOSS mission is to empower doctors to provide the best possible care','Official about page states this mission',SRC['about'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS was founded by doctors, for doctors, in 2012','Official about page states founded by doctors in 2012',SRC['about'],'Observed','High','🟢 Confirmed'],
    ['Founding problem was disconnected resources causing more time researching than mastering topics','Official about page states this frustration',SRC['about'],'Observed','High','🟢 Confirmed'],
    ['US launch occurred in 2017 with New York headquarters opening','Official about page states successful US launch in 2017 coincided with HQ opening in New York',SRC['about'],'Observed','High','🟢 Confirmed'],
    ['More than 1 million healthcare professionals in over 180 countries rely on AMBOSS','Official about and funding pages state this',SRC['about'],'Observed','High','🟢 Confirmed'],
    ['Over 50 medical schools adopted AMBOSS as an essential study resource','Official about page states over 50 medical schools around the world',SRC['about'],'Observed','High','🟢 Confirmed'],
    ['Helios rolled out AMBOSS as foundational medical reference for its doctors','Official about page states Helios rollout',SRC['about'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS closed a €240M financing round in March 2025','Official newsroom funding release',SRC['funding'],'Observed','High','🟢 Confirmed'],
    ['New 2025 investors included KIRKBI, M&G Investments and Lightrock','Official funding release',SRC['funding'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS converted to a European stock corporation (SE)','Official funding release',SRC['funding'],'Observed','High','🟢 Confirmed'],
    ['Funding supports international markets and expansion to nurses/other healthcare professionals','Official funding release',SRC['funding'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS acquired NEJM Knowledge+ in April 2024','Official AMBOSS NEJM acquisition release',SRC['nejm_acq'],'Observed','High','🟢 Confirmed'],
    ['NEJM Knowledge+ was trusted by over 500 residency programs','Official NEJM acquisition release',SRC['nejm_acq'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS acquired Novaheal in 2024','PRNewswire acquisition release and AMBOSS funding release',SRC['novaheal_pr'],'Observed','High','🟢 Confirmed'],
    ['Novaheal is a nursing education platform/start-up in Germany','PRNewswire Novaheal acquisition release',SRC['novaheal_pr'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS and NEJM launched Master Classes in Medicine in 2025','Official AMBOSS NEJM courses release',SRC['nejm_courses'],'Observed','High','🟢 Confirmed'],
    ['Knowledge Library has 1,400+ high-yield articles','Pricing/features pages state 1,400+ articles',SRC['pricing'],'Observed','High','🟢 Confirmed'],
    ['Student page states 1,500+ preclinical and clinical topics','Official students page FAQ',SRC['students'],'Observed','High','🟢 Confirmed'],
    ['Qbank includes 13,900+ questions on student page','Official students page',SRC['students'],'Observed','High','🟢 Confirmed'],
    ['Features page describes 13,000+ board-style questions','Official features page',SRC['features'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS offers Study Mode and Exam Mode','Support page Using Study Mode & Exam Mode',SRC['support_modes'],'Observed','High','🟢 Confirmed'],
    ['Exam Mode includes reverse color and text zoom accessibility features','Support Exam Mode page',SRC['support_modes'],'Observed','High','🟢 Confirmed'],
    ['Qbank difficulty hammers range from 1 to 5','Features/support assignments pages',SRC['features'],'Observed','High','🟢 Confirmed'],
    ['Adaptive sessions use performance data and estimate probability of correct answer','Support Adaptive Question Sessions',SRC['support_adaptive'],'Observed','High','🟢 Confirmed'],
    ['EPC uses IRT principles and real-time mapping to theta scores','Support Analysis page',SRC['support_analysis'],'Observed','High','🟢 Confirmed'],
    ['Score predictor uses mixed-effects model for Step 1/2 and Bayesian joint normal for Step 3','USMLE Score Predictor support FAQ',SRC['support_score'],'Observed','High','🟢 Confirmed'],
    ['AI Mode Clinical Care is an AI-powered search tool over AMBOSS articles, drug database and selected external guidelines','Clinical AI Mode page',SRC['ai_clinical'],'Observed','High','🟢 Confirmed'],
    ['AI Mode does not search the open web','Attendings FAQ states it never searches open web',SRC['attendings'],'Observed','High','🟢 Confirmed'],
    ['AI Mode should support but not replace clinical judgment and not be used in time-sensitive emergencies','Clinical AI Mode FAQ and terms',SRC['ai_clinical'],'Observed','High','🟢 Confirmed'],
    ['AI Mode Learning supports uploads of PDFs, DOCX, TXT, JPG/PNG/WEBP','Support AI Mode Learning FAQ',SRC['support_ai_learning'],'Observed','High','🟢 Confirmed'],
    ['AI Mode Learning not natively integrated into mobile apps yet; mobile web works','Support AI Mode Learning FAQ',SRC['support_ai_learning'],'Observed','High','🟢 Confirmed'],
    ['AI Mode Clinical Care is available in browser and iOS/Android apps to limited users','Support AI Clinical Care FAQ',SRC['support_ai_clinical'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS uses OpenAI and Anthropic for AMBOSS Assistants processing','AI assistants privacy page',SRC['assistants_privacy'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS says user data is not used to train the AI/LLMs','AI Mode Learning FAQ and terms',SRC['support_ai_learning'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS uses MongoDB Cloud Services for AMBOSS AI Features/Assistants','Privacy and assistants processing pages',SRC['privacy'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS terms prohibit patient data/PHI input into AI features','Terms Section 9 and clinics terms',SRC['terms'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS does not process PHI under clinics terms','Terms and Conditions for Clinics Section 12',SRC['terms_clinics'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS uses AWS hosting and CloudFront, Cloudflare, Cloudinary CDNs','Privacy Section 3.2',SRC['privacy'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS uses Auth0 for IAM','Privacy required technologies list',SRC['privacy'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS uses Braze for messages/notifications','Privacy required technologies list',SRC['privacy'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS uses Datadog and Sentry for monitoring/error analysis','Privacy required technologies list',SRC['privacy'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS uses Segment, Amplitude and Mixpanel/Looker-like stack from jobs?','Privacy confirms Segment/Amplitude; jobs mention Looker/Mixpanel','https://www.amboss.com/us/legal/privacy; https://careers.amboss.com/jobs/job-4a7634/','Observed and inferred','Medium','🟡 Strong Inference'],
    ['Marketing website is hosted through Webflow and CloudFront','HTTP headers show x-wf-page-id/x-wf-region and CloudFront; site HTML uses Webflow assets','HEAD request captured in workspace notes','Observed','Medium','🟢 Confirmed'],
    ['Core platform likely uses React/TypeScript/Go/Python/GraphQL/REST/AWS/Kubernetes','AMBOSS engineering job requires Go, React, GraphQL/REST, Docker, CI/CD, AWS/Kubernetes and Python in another job snippet','https://careers.amboss.com/jobs/job-09b09075-f834-4935-a4ec-8f5a878493f7/','Inferred from hiring','Medium','🟡 Strong Inference'],
    ['No public evidence of a FHIR/HL7/CCDA patient-data integration by AMBOSS was found','Terms prohibit PHI; searches did not reveal FHIR/HL7 API docs','Search across public web and legal terms','Observed absence','Medium','🟡 Strong Inference'],
    ['No public REST/OpenAPI/FHIR developer API docs were found for AMBOSS medical product','Search surfaced MCP and unrelated AmbossTech Bitcoin APIs only','Search across AMBOSS API/developer terms','Observed absence','Medium','🟡 Strong Inference'],
    ['AMBOSS MCP gives AI agents access to AMBOSS content through Model Context Protocol','Official MCP newsroom',SRC['mcp'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS content is not allowed to be crawled/scraped or used for AI training without consent','Terms rights of use Section 13.4/13.8',SRC['terms'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS is ACCME-accredited for CME','CME program overview',SRC['cme_program'],'Observed','High','🟢 Confirmed'],
    ['Internet Point-of-Care CME grants max 0.5 AMA PRA Category 1 Credit per researched clinical question','CME/MOC FAQ and Internet POC page',SRC['cme_faq'],'Observed','High','🟢 Confirmed'],
    ['Educator Tool includes Learner, Educator, Admin roles','Educator Tool Overview',SRC['support_educator'],'Observed','High','🟢 Confirmed'],
    ['Educator assignments can filter by system, discipline, article, competency and difficulty','Creating Assignments support page',SRC['support_assignments'],'Observed','High','🟢 Confirmed'],
    ['Institution dashboards share user assignment data and recommendations under role permissions','Educator Analytics support page',SRC['support_analytics'],'Observed','High','🟢 Confirmed'],
    ['Students in 350+ US medical schools and programs use AMBOSS','Official student/medical schools pages',SRC['students'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS claims 2 out of 3 US medical students use AMBOSS to prepare for exams','Medical schools page',SRC['medical_schools'],'Observed','High','🟢 Confirmed'],
    ['Residency page claims 98% renewals and 800+ partners','Residency programs page',SRC['residency'],'Observed','High','🟢 Confirmed'],
    ['Residency page claims residents save on average 23 minutes per shift and find answers in about 90 seconds','Residency page FAQ',SRC['residency'],'Observed vendor claim','Medium','🟢 Confirmed'],
    ['AMBOSS pricing: US student monthly $19.99, yearly $12.50/month, Qbank bundle $448, Student Life $1199','Pricing page',SRC['pricing'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS clinician pricing: monthly $29.99 and annual $21.58/month','Pricing page',SRC['pricing'],'Observed','High','🟢 Confirmed'],
    ['5-day free trial requires no payment information and includes full library plus 50 Qbank questions','Free-trial support page',SRC['free_trial'],'Observed','High','🟢 Confirmed'],
    ['Institutional license can activate through SSO, email verification and other methods','Support institutional license pages',SRC['sso'],'Observed','High','🟢 Confirmed'],
    ['User account deletion is immediate and irreversible, but subscription cancellation must be separate','Account deletion support page',SRC['delete'],'Observed','High','🟢 Confirmed'],
    ['AMBOSS has an Anki add-on code 1044112126','Anki page',SRC['anki'],'Observed','High','🟢 Confirmed'],
    ['Chrome extension highlights terms across websites and has settings for site toggling/theme','Chrome page',SRC['chrome'],'Observed','High','🟢 Confirmed'],
    ['Customer praise centers on integrated library/Qbank/Anki and clinical rotation utility','Official reviews plus Reddit search snippets',SRC['reviews'],'Observed sample','Medium','🟡 Strong Inference'],
    ['Customer complaints include 4-5 hammer questions being picky/low-yield and mobile/extension limitations','Reddit/Chrome/App Store review search snippets','Search results in browser session','Observed sample','Medium','🟡 Strong Inference'],
]

copy_ideas = [
'Role-based navigation around user intent, not feature silos','A trusted content core before AI features','Human expert review labels on every clinical output','Source links and inline citations everywhere','A single knowledge graph connecting education, search, AI, and workflows','High-yield mode for reducing cognitive load','Learning radar tied to mistakes and remediation','Adaptive sessions based on modeled weaknesses','Exam/performance predictors with uncertainty ranges','Peer benchmarking with privacy constraints','Side-by-side split view to avoid tab hopping','Browser extension that inserts knowledge into external workflows','Anki-like integration for spaced repetition','Mobile offline mode for unreliable clinical environments','Integrated drug database rather than standalone drug app','Checklists for high-risk workflows','Differential diagnosis aids with transparent uncertainty','Clinical calculators inside relevant content','CME/credit loops from everyday usage','Institution dashboards with role-based visibility','Faculty assignment authoring and analytics','Group discounts / cohort acquisition','Long-duration Student Life-type membership','30-day refund to lower purchase risk','Clear AI principles and safety page','No-open-web retrieval claim for clinical AI','AI output with transparent limits and refusal behavior','Separate learning AI from clinical-care AI','Assistant templates for common jobs-to-be-done','MCP/API strategy for trusted knowledge distribution','Human-generated content policy and evidence hierarchy','Content update changelog','Use of user profile/study goals for personalization','Specialty-aware responses','Highlighting conflicting recommendations','Explicit “support not replace judgment” stance','Trust badges from partnerships and accreditations','Search as the central daily habit','Free utility tools as top-of-funnel','Community testimonials and love notes','Institutional usage reports as a B2B wedge','Self-assessment weeks/events as acquisition moments','AI study copilot that converts answers into practice','Tooltips/popups instead of page-level help only','Embedded support channels (WhatsApp/contact/callback)','Data protection overview written for users','Accessibility roadmap with known gaps','Content-independent editorial stance/no pharma sponsorship','Use acquisitions to enter adjacent professions','Separate consumer/enterprise packaging without fragmenting core product']

improve_ideas = [
'Make evidence grading explicit, not just source-linked','Publish external validation, not only vendor claims','Convert every AI answer into an audit trail','Add regional guideline localization controls','Offer an explainer for why recommendations were made','Expose AI confidence and evidence sufficiency','Let users delete individual AI interactions','Make mobile feature parity a first-class requirement','Make browser extension sidebar behavior configurable','Add multi-provider health-data connections where allowed','Integrate wearable/lab/record timelines for Ovexis','Build a consent center richer than standard settings','Make institution dashboards privacy-preserving by default','Add patient-friendly translation with clinician review mode','Use structured clinical ontologies for all content objects','Support longitudinal goals beyond exams','Add outcome tracking after recommendations','Provide calibrated “unknown/needs clinician” states','Offer API tiers with sandbox data and compliance docs','Add developer documentation and schema governance','Add FHIR-native interfaces for Ovexis, unlike AMBOSS','Add family/caregiver/team roles','Add payer/lab/pharmacy workflows if building longitudinal care','Use retrieval over patient data and evidence simultaneously','Separate PHI and non-PHI AI infrastructure','Support multimodal records: labs, imaging, genomics, wearables','Support explainable normalization/deduplication','Publish security certifications and audit summaries','Offer BAA and HIPAA program if handling PHI','Build continuous safety evaluation set before launch','Add red-team output categories and model card','Support India localization for guidelines, drugs, labs','Make pricing transparent for institutions where possible','Instrument churn reasons by learner stage','Make learning content dynamic with micro-assessments','Add care-plan adherence loops','Add structured patient-reported outcomes','Integrate with provider EHR workflow via SMART on FHIR','Create “doctor review packet” exports','Build a marketplace of vetted clinical services carefully','Add data portability exports for users','Add explainable risk trajectories','Use physician+AI review for high-stakes insights','Build network effects around longitudinal records','Provide offline encrypted health vault','Add personalized screening guideline engine','Use local language support with medical QA','Add accessibility testing into CI/CD from day one','Make every recommendation actionable with next steps','Use a transparent evidence register in product UI']

ignore_ideas = [
'Do not copy “no PHI” if Ovexis wants longitudinal intelligence','Do not build only an education Qbank as the core wedge','Do not rely on static articles without user data loops','Do not make AI a chatbot bolted onto content','Do not hide evidence provenance behind generic citations','Do not overclaim hallucination-free guarantees','Do not make institutional dashboards invasive','Do not use exam anxiety as the only retention loop','Do not build hard/picky questions that create overthinking','Do not make mobile second-class','Do not design only for doctors and students','Do not ignore patients and caregivers','Do not use one-size-fits-all US guidelines globally','Do not bury deletion/export controls','Do not block all third-party workflows if ecosystem matters','Do not rely solely on Webflow-style marketing architecture for app scale','Do not use vague “AI-powered” labels without details','Do not expose output without confidence/uncertainty','Do not build hidden model memory without consent','Do not treat content licensing as a substitute for clinical workflow','Do not make paid AI premium before proving daily value','Do not overload users with dashboards','Do not use generic wellness recommendations without data','Do not skip FDA/HIPAA analysis if patient-specific','Do not make enterprise pricing a complete black box','Do not ignore India-specific affordability','Do not sell supplement/pharmacy marketplace too early','Do not let partners drive clinical bias','Do not create closed data silos','Do not require manual data entry as primary data source','Do not let PDFs become dead-end uploads','Do not use unstructured notes without ontology','Do not assume doctors will use another login','Do not ignore EHR context if clinical users are target','Do not build a tool that cannot explain conflicts','Do not use social proof without audited outcomes','Do not treat security page as afterthought','Do not ignore Chrome/Anki-like external workflow hooks','Do not build every profession at once without role UX','Do not use unvalidated biological-age gimmicks as core trust','Do not over-index on biohacker audience if mass health is goal','Do not promise diagnosis replacement','Do not bypass clinician review for high-risk insights','Do not store PHI with generic AI vendors without agreements','Do not train on user data by default','Do not make refund/cancellation difficult','Do not let accessibility exceptions remain undefined','Do not fragment user progress across products','Do not ignore support as part of trust','Do not make evidence static; freshness matters']

reinvent_ideas = [
'Reinvent AMBOSS content graph as a personal health knowledge graph','Reinvent Qbank as longitudinal health literacy and decision rehearsal','Reinvent Study Recommendations as health-action recommendations','Reinvent Score Predictor as risk trajectory predictor with uncertainty','Reinvent CME loop as patient/clinician co-learning loop','Reinvent browser extension as health-context copilot for lab portals and EHR portals','Reinvent checklists as personalized care orchestration plans','Reinvent differential diagnosis as “questions to ask your clinician” for consumers','Reinvent guidelines as patient-specific eligibility rules','Reinvent institution dashboards as clinician-patient-shared insight boards','Reinvent AI Mode Clinical Care as PHI-compliant, FHIR-grounded agent','Reinvent AI Mode Learning as a family/caregiver tutor','Reinvent Assistants as audited workflows with versioned prompts/tools','Reinvent MCP as consented health-data and evidence API','Reinvent role navigation into patient/doctor/caregiver/coach/payer roles','Reinvent mobile offline as encrypted longitudinal health vault','Reinvent AHFS-style drug data as medication safety graph with real meds','Reinvent Anki integration as habit formation/spaced action reminders','Reinvent self-assessment as annual preventive intelligence assessment','Reinvent group discounts as community preventive programs','Reinvent evidence register as in-product evidence ledger','Reinvent content updates as personalized guideline-diff notifications','Reinvent notes/highlights as clinician-reviewed memory objects','Reinvent search as multimodal search across labs, notes, imaging, wearables','Reinvent clinical calculators as auto-filled calculators with provenance','Reinvent educator assignments as remote care tasks and learning journeys','Reinvent board readiness as health readiness/goal readiness','Reinvent subscription as longitudinal membership spanning data custody','Reinvent review pages as after-visit intelligence summaries','Reinvent peer stats as privacy-safe population baselines','Reinvent Terms/Privacy into plain-language consent UX','Reinvent support chatbot as privacy-aware health-data navigator','Reinvent clinical safety benchmark as continuous product eval suite','Reinvent no-open-web as curated multimodal evidence garden','Reinvent specialty-aware into care-team-aware','Reinvent conflicting recommendations into shared-decision aids','Reinvent US-first content into India+global localization layer','Reinvent AI refusals into escalation pathways','Reinvent dots phrases as patient-friendly summaries and doctor packets','Reinvent institutional licenses as employer/preventive health programs','Reinvent cohort reports as public-health intelligence','Reinvent knowledge library as generated-but-reviewed microchapters','Reinvent patient cases as synthetic digital twin simulations','Reinvent retention loop around improvements in real biomarkers/outcomes','Reinvent marketplace only after trust layer is proven','Reinvent payment as insurance/HSA/employer hybrid','Reinvent data import as consented FHIR+wearable+lab ingestion','Reinvent article sections into computable recommendations','Reinvent user network as consented expert community','Reinvent content QA as human-in-the-loop governance operating system']

market_gaps = [
'Consumer-owned longitudinal health record with clinical-grade interpretation','PHI-compliant medical AI that can use patient data safely','India-localized longitudinal preventive health intelligence','FHIR/ABDM integration layer for consumer-controlled records','Wearable + lab + EHR + pharmacy + imaging normalization','Explainable deduplication and identity resolution for health data','Evidence-graded recommendations with source freshness','Doctor-ready exports that reduce appointment friction','Caregiver/family health intelligence roles','Clinician review marketplace separated from supplement bias','Affordable annual preventive assessment for emerging markets','Lab trend interpretation over years, not one-off panels','Medication interaction context using actual patient meds','Guideline-diff alerts personalized to patient risks','Patient-facing education that adapts to health literacy','Risk trajectory models with uncertainty and next-best-action','Trust ledger for every AI recommendation','Consent and data portability as product differentiator','Interoperability APIs for third-party health apps','Safety evaluation benchmark for patient-specific AI','Longitudinal retention loop beyond reminders','Offline encrypted mobile health vault','Doctor collaboration without EHR integration burden','Automated records acquisition and summarization','Multilingual medical UX with verified translation','Outcome tracking after interventions','Insurance/lab/pharmacy workflows under user control','Personalized screening and vaccination tracker','Genomics/microbiome integration with conservative guardrails','Mental-health and lifestyle context merged with medical data','Clinical-grade wearable interpretation beyond wellness scores','Human review escalation for critical results','Low-cost physician second opinion packets','Data-quality score for every imported source','Patient identity and household graph','No-ad, no-pharma-sponsored trust position','Developer ecosystem for consented health applications','Evidence marketplace for vetted protocols','Community cohorts for preventive health programs','Intervention ROI tracking','Adherence intelligence with context-aware coaching','Continuity across life stages, not school-to-residency','Patient-clinician shared planning workspace','Privacy-preserving population analytics','Regulatory-ready audit logs for AI actions','Secure upload and redaction pipeline for records','Personalized lab reference ranges and trends','Clinical trial eligibility intelligence','Health coach + physician + AI triage routing','Cross-border medical record translation and normalization']

blue_ocean = [
'PHI-compliant personal medical intelligence layer','India-first ABDM + FHIR health vault','Doctor-ready longitudinal summary generator','Evidence-graded personal guideline engine','AI + human review preventive health plan','Family/caregiver health intelligence workspace','Longitudinal medication safety graph','Wearable-to-clinical signal interpreter','Personal health MCP for consented agents','Clinic-light second-opinion workflow','Insurance-neutral preventive care membership','Lab trend and root-cause intelligence','Clinical-grade health literacy tutor','Population cohort insights with privacy','Post-visit plan adherence copilot','Critical-result escalation network','Chronic-condition digital twin with audit','Pharmacy/lab marketplace separated by trust rules','Low-bandwidth offline health vault for India','Regulatory-grade AI evaluation suite']

# Report markdown

def numbered(items, label='🟡'):
    return '\n'.join([f'{i+1}. {label} {x}.' for i,x in enumerate(items)])

report = f"""# AMBOSS Competitive Intelligence Report for Ovexis

**Target:** AMBOSS — Clinical Decision Support | Medical Education  
**Official website:** https://www.amboss.com/  
**Prepared:** 2026-07-25  
**Evidence standard:** 🟢 Confirmed = directly observed in public sources; 🟡 Strong Inference = evidence-backed inference; 🔴 Speculation = future-looking or unverifiable hypothesis.  
**Method constraint:** 🟢 Confirmed — This report uses public webpages, support docs, legal pages, public career pages, public search snippets, and publicly indexed third-party material only.  
**Screenshot constraint:** 🟢 Confirmed — This environment has no browser screenshot capture. The evidence register includes URLs, quoted evidence, confidence, and observed/inferred status; screenshot cells state “not captured; source URL available.”  
**Robots/ToS constraint:** 🟢 Confirmed — The investigation avoided disallowed/private account/API paths and did not attempt unauthorized access.

---

## 1. Executive Summary

### What they are building

- 🟢 Confirmed — AMBOSS is a global medical knowledge, medical education, Qbank, clinical decision support, educator, and AI medical intelligence platform for medical students, residents, clinicians, educators, PAs/NPs, and institutions. Evidence: {SRC['about']}; {SRC['features']}; {SRC['new_nav']}.
- 🟢 Confirmed — The company explicitly frames the platform as a connected continuum across Learning, Clinical Care, and Teaching rather than as separate products. Evidence: {SRC['new_nav']}.
- 🟢 Confirmed — Core product assets include a Knowledge Library, Qbank, study plans, analytics, mobile apps, Anki/Chrome/ChatGPT integrations, clinical tools, AI Mode Clinical Care, AI Mode Learning, AMBOSS Assistants, educator analytics, and institutional licensing. Evidence: {SRC['features']}; {SRC['ai_innovation']}; {SRC['support_platform']}.
- 🟢 Confirmed — The core clinical AI strategy is not open-web answer generation; AMBOSS says AI Mode draws from AMBOSS content, a drug database, and selected external guidelines/sources and links users back to sources. Evidence: {SRC['ai_clinical']}; {SRC['support_ai_clinical']}.

### Why it exists

- 🟢 Confirmed — AMBOSS says it was founded by doctors in 2012 after young residents became frustrated with disconnected resources that caused more time spent researching topics than mastering them. Evidence: {SRC['about']}.
- 🟢 Confirmed — AMBOSS’s stated mission is to “empower all doctors to provide the best possible care.” Evidence: {SRC['about']}.
- 🟡 Strong Inference — AMBOSS exists to compress the time and cognitive effort required to acquire, retrieve, apply, and teach medical knowledge across a clinician’s lifecycle; this inference follows from the official story, navigation redesign, clinician messaging, and product architecture. Evidence: {SRC['about']}; {SRC['clinicians']}; {SRC['new_nav']}.

### Customer problems solved

- 🟢 Confirmed — For students, AMBOSS addresses exam preparation, knowledge review, Qbank practice, Anki integration, score prediction, clinical rotations, and avoiding tool-switching. Evidence: {SRC['students']}.
- 🟢 Confirmed — For clinicians, AMBOSS addresses time-to-answer, clinical reference, drug lookup, differential diagnosis, diagnostic algorithms, checklists, offline mobile access, CME, and board/recertification prep. Evidence: {SRC['clinicians']}; {SRC['attendings']}.
- 🟢 Confirmed — For institutions, AMBOSS addresses assignments, assessments, learner progress, cohort analytics, at-risk learner identification, curriculum integration, and institutional licensing. Evidence: {SRC['medical_schools']}; {SRC['support_educator']}; {SRC['support_analytics']}.
- 🟢 Confirmed — AMBOSS explicitly says clinical tools are not intended for patients or laypeople. Evidence: {SRC['attendings']}; {SRC['terms']}.

### Emotional problem

- 🟡 Strong Inference — AMBOSS reduces anxiety from uncertainty: “What should I study?”, “Am I ready?”, “What am I missing?”, and “Can I defend this clinical answer?” This follows from score predictor, study recommendations, source links, and clinical AI disclaimers. Evidence: {SRC['support_score']}; {SRC['support_recs']}; {SRC['ai_clinical']}.
- 🟡 Strong Inference — AMBOSS offers identity reinforcement for medical learners and clinicians by making them feel faster, safer, more prepared, and less likely to look uninformed on wards. Evidence includes official reviews and clinician testimonials. Evidence: {SRC['reviews']}; {SRC['mobile']}; {SRC['residency']}.

### Operational problem

- 🟢 Confirmed — AMBOSS claims its integrated system reduces switching between resources and combines Qbank, library, Anki, content review, board prep, and clinical tools. Evidence: {SRC['students']}.
- 🟢 Confirmed — AMBOSS claims residents can save time by finding answers quickly and that unified resources reduce redundant tools. Evidence: {SRC['residency']}.
- 🟡 Strong Inference — Its operational win is not merely content; it is the conversion of content into workflow primitives: search, snippets, checklists, split view, highlighted knowledge gaps, Qbank sessions, and educator assignments. Evidence: {SRC['features']}; {SRC['support_feature_overview']}; {SRC['support_management']}.

### Who is the customer

- 🟢 Confirmed — Primary individual customers are medical students, residents, attending physicians, physician associates/assistants, nurse practitioners, PA/NP students, and other healthcare professionals. Evidence: {SRC['students']}; {SRC['attendings']}; {SRC['pa'] if 'pa' in SRC else 'https://www.amboss.com/us/pa'}; {SRC['np'] if 'np' in SRC else 'https://www.amboss.com/us/np'}.
- 🟢 Confirmed — Institutional customers include medical schools, PA programs, NP programs, residency programs, universities, clinics, and hospitals. Evidence: {SRC['medical_schools']}; {SRC['residency']}; {SRC['terms_clinics']}.
- 🟢 Confirmed — AMBOSS also targets AI development teams through its MCP access program. Evidence: {SRC['mcp']}.

### Who is NOT the customer

- 🟢 Confirmed — Patients and laypeople are not intended users of the clinical decision support tools. Evidence: {SRC['attendings']}; {SRC['terms']}.
- 🟢 Confirmed — AMBOSS legal terms state AI Mode Clinical Care is only for physicians and other healthcare professionals in the U.S., Canada, Australia, or Qatar, with different medical-research-only constraints elsewhere. Evidence: {SRC['terms']}.
- 🟢 Confirmed — AMBOSS clinics terms state the AMBOSS Program is not intended for storage, processing, or transmission of PHI. Evidence: {SRC['terms_clinics']}.

### Market category created / replaced

- 🟡 Strong Inference — AMBOSS is attempting to create the “medical intelligence platform” category: one longitudinal platform for medical learning, clinical reference, teaching, and AI evidence navigation. Evidence: {SRC['new_nav']}; {SRC['ai_innovation']}.
- 🟡 Strong Inference — It replaces or absorbs parts of textbook/reference search, Qbank-only subscriptions, point-of-care clinical reference, flashcard add-ons, board analytics, faculty assignment tools, and generic AI prompts. Evidence: {SRC['features']}; {SRC['support_platform']}; {SRC['ai_learning']}.
- 🟡 Strong Inference — Its strongest category wedge is “education-to-practice continuity,” not patient longitudinal health intelligence. Evidence: {SRC['new_nav']}; {SRC['terms_clinics']}.

### Jobs-To-Be-Done

| Status | User | Job | AMBOSS mechanism | Evidence |
|---|---|---|---|---|
| 🟢 Confirmed | Student | Prepare for boards and in-house exams | Qbank, study plans, self-assessments, score predictor, analytics | {SRC['students']} |
| 🟢 Confirmed | Clerkship student | Look up clinical topics quickly on wards | mobile app, clinician mode, treatment plans, flowcharts, drug dosages | {SRC['students']}; {SRC['mobile']} |
| 🟢 Confirmed | Resident | Confirm workup/treatment while learning for boards | Clinical Care, Qbank, board review, CME | {SRC['residency']}; {SRC['attendings']} |
| 🟢 Confirmed | Attending/NP/PA | Find evidence-based answers and earn CME | AI Mode Clinical Care, drug database, POC CME, guidelines | {SRC['attendings']}; {SRC['cme_faq']} |
| 🟢 Confirmed | Educator/Admin | Assign and track learner progress | Educator Tool, assignments, analytics, institution groups | {SRC['support_educator']} |
| 🟢 Confirmed | AI team | Ground agents in medical knowledge | AMBOSS MCP | {SRC['mcp']} |

### Value proposition

- 🟢 Confirmed — AMBOSS’s core value proposition is an all-in-one system combining knowledge, Qbank, analytics, clinical support, mobile/offline access, integrations, and AI, with source-linked and clinician-reviewed content. Evidence: {SRC['features']}; {SRC['content_policy']}.
- 🟡 Strong Inference — The board-level value proposition is “familiarity + trust + speed + continuity”: a tool learned in medical school becomes a trusted reference in residency and clinical care. Evidence: {SRC['new_nav']}; {SRC['about']}.

### Core philosophy

- 🟢 Confirmed — AMBOSS states that any separation between learning, clinical care, and teaching is artificial. Evidence: {SRC['new_nav']}.
- 🟢 Confirmed — AMBOSS states technology will not replace healthcare professionals, but those who embrace it will lead the way. Evidence: {SRC['funding']}.
- 🟢 Confirmed — AMBOSS’s AI principles emphasize medical-professional + technology-expert collaboration and rigorous validation. Evidence: {SRC['ai_principles']}.
- 🟡 Strong Inference — Their philosophy is clinician-centered augmentation, not autonomous diagnosis. Evidence: {SRC['terms']}; {SRC['ai_clinical']}.

---

## 2. Company Intelligence

### Timeline

| Status | Year | Event | Evidence |
|---|---:|---|---|
| 🟢 Confirmed | 2012 | Founded by doctors, for doctors | {SRC['about']} |
| 🟢 Confirmed | 2017 | US launch and New York headquarters opening | {SRC['about']} |
| 🟢 Confirmed | 2019 | €30M Series B led by Partech and Target Global per public press reporting | https://www.einnews.com/pr_news/497270609/medtech-company-amboss-raises-30m-in-series-b |
| 🟢 Confirmed | 2020 | Global Education Fund launched during COVID period; $2M initial budget later exceeded | {SRC['scholarships'] if 'scholarships' in SRC else 'https://www.amboss.com/us/scholarships-program'} |
| 🟢 Confirmed | 2024 | NEJM Knowledge+ acquired | {SRC['nejm_acq']} |
| 🟢 Confirmed | 2024 | Novaheal acquired | {SRC['novaheal_pr']} |
| 🟢 Confirmed | 2025 | AMBOSS converted to SE and closed €240M financing round | {SRC['funding']} |
| 🟢 Confirmed | 2025 | AMBOSS MCP announced | {SRC['mcp']} |
| 🟢 Confirmed | 2025 | AMBOSS Assistants beta announced | {SRC['assistants']} |
| 🟢 Confirmed | 2025 | AMBOSS and NEJM Group launched Master Classes in Medicine | {SRC['nejm_courses']} |
| 🟢 Confirmed | 2026 | Redesigned navigation announced as AMBOSS evolves as a medical intelligence platform | {SRC['new_nav']} |

### Founders and leadership

- 🟢 Confirmed — Legal/privacy pages list AMBOSS SE managing directors as Dr. med. Madjid Salimi, Dr. med. Nawid Salimi, and Benedikt Hochkirchen. Evidence: {SRC['privacy']}.
- 🟢 Confirmed — Official 2025 financing release quotes Benedikt Hochkirchen as Co-Founder and Co-CEO and Dr. Madjid Salimi as Co-Founder and Co-CEO. Evidence: {SRC['funding']}.
- 🟢 Confirmed — Official pages and public company profiles consistently indicate AMBOSS was founded by a team of physicians; public third-party sources vary on exact founder roster, so exact complete founder list should be treated carefully unless sourced directly from AMBOSS. Evidence: {SRC['about']}; https://cherry.vc/founders/amboss.
- 🟡 Strong Inference — The leadership operating model is physician-founder led with a substantial internal medical team and product/engineering leadership for AI/search/knowledge retrieval. Evidence: {SRC['funding']}; {SRC['careers_eng']}.

### Funding / investors / valuation

- 🟢 Confirmed — AMBOSS closed €240M with KIRKBI, M&G Investments, Lightrock, and existing shareholders. Evidence: {SRC['funding']}.
- 🟢 Confirmed — AMBOSS’s 2025 financing release says many new investors manage evergreen funds and plan to accompany AMBOSS until a possible IPO and beyond. Evidence: {SRC['funding']}.
- 🟢 Confirmed — AMBOSS says additional funds will go to technology, new market segments, international expansion, and selective acquisitions. Evidence: {SRC['funding']}.
- 🟢 Confirmed — Earlier public reports stated AMBOSS raised €30M Series B led by Partech Growth with Target Global as co-investor and Cherry Ventures, Wellington Partners, and Holtzbrinck Digital participating. Evidence: https://www.einnews.com/pr_news/497270609/medtech-company-amboss-raises-30m-in-series-b.
- 🟡 Strong Inference — AMBOSS is positioning for eventual public-market optionality, because the official release references possible IPO and SE conversion. Evidence: {SRC['funding']}.
- 🔴 Speculation — Exact current valuation is not publicly confirmed by AMBOSS in the sources reviewed; third-party estimates should not be treated as verified.

### Acquisitions and partnerships

- 🟢 Confirmed — AMBOSS acquired NEJM Knowledge+ from NEJM Group in April 2024. Evidence: {SRC['nejm_acq']}.
- 🟢 Confirmed — NEJM Knowledge+ integration supports AMBOSS’s goal of convergence between board prep and clinical practice. Evidence: {SRC['nejm_acq']}.
- 🟢 Confirmed — AMBOSS acquired Novaheal to strengthen German nursing-market presence and become a digital partner for physicians and nurses. Evidence: {SRC['novaheal_pr']}.
- 🟢 Confirmed — AMBOSS integrates AHFS clinical drug information. Evidence: {SRC['support_drug']}; {SRC['clinicians'] if 'clinicians' in SRC else SRC['features']}.
- 🟢 Confirmed — AMBOSS uses QxMD for clinical calculators. Evidence: {SRC['support_calc']}.
- 🟢 Confirmed — AMBOSS and NEJM Group launched two initial Master Classes in Medicine courses: From Evidence to Impact and Core Concepts in Clinical Research. Evidence: {SRC['nejm_courses']}.

### Patents, research, open source

- 🟡 Strong Inference — No public patent filings clearly attributable to AMBOSS medical education/CDS were identified in the searches run for this report; absence of evidence is not proof of absence.
- 🟢 Confirmed — AMBOSS publishes or references outcome/score reports and white papers, including Step 2 score association claims and score predictor modeling. Evidence: {SRC['scores']}; {SRC['score_predictor'] if 'score_predictor' in SRC else SRC['support_score']}.
- 🟢 Confirmed — AMBOSS’s careers blog describes internal search evaluation work using ChatGPT-generated judgment lists and compares GPT-generated judgments to medically trained raters. Evidence: {SRC['search_eval']}.
- 🟢 Confirmed — The Stanford–Harvard/ARISE NOHARM preprint evaluates AMBOSS LiSA 1.0 among clinical AI systems; AMBOSS cites a #1 clinical-care safety ranking on multiple official pages. Evidence: {SRC['noharm']}; {SRC['ai_innovation']}.
- 🟡 Strong Inference — AMBOSS has not made a broad public open-source portfolio visible under an obvious official AMBOSS medical GitHub identity; search results with “AmbossTech” refer to a separate Bitcoin Lightning company and should not be confused with medical AMBOSS.

### Geographic expansion

- 🟢 Confirmed — AMBOSS has offices/hubs in Berlin, Cologne, Cagliari, and New York. Evidence: {SRC['careers']}.
- 🟢 Confirmed — AMBOSS supports US/English, DE/Deutsch, INT/English, PL/Polish, and PT/Portuguese web localizations in its site footer. Evidence: {SRC['funding']}.
- 🟢 Confirmed — 2025 financing is intended to support additional international market exploration. Evidence: {SRC['funding']}.
- 🟡 Strong Inference — LATAM expansion is active or planned because public LinkedIn/job snippets include LATAM partnership/marketing roles; this should be validated with current job pages at the time of decision.

### Regulatory filings / legal identity

- 🟢 Confirmed — AMBOSS SE is listed as data controller at Torstrasse 19, 10119 Berlin, with Local Court Berlin (Charlottenburg), HRB 270315 B. Evidence: {SRC['privacy']}.
- 🟢 Confirmed — AMBOSS MD Inc. is listed as a subsidiary at 234 5th Avenue, 2nd Floor, New York, NY. Evidence: {SRC['privacy']}.
- 🟢 Confirmed — AMBOSS’s terms state Delaware law for user terms and New York law/forum for clinic terms where permissible. Evidence: {SRC['terms']}; {SRC['terms_clinics']}.

---

## 3. Founder Psychology Report

- 🟡 Strong Inference — Founder belief #1: medical professionals drown in fragmented knowledge, and the winning product is an integrated system that distills, connects, and operationalizes knowledge. Evidence: {SRC['about']}; {SRC['new_nav']}.
- 🟡 Strong Inference — Founder belief #2: physicians must remain at the center; technology and AI should augment, not replace, clinical judgment. Evidence: {SRC['funding']}; {SRC['ai_principles']}; {SRC['terms']}.
- 🟡 Strong Inference — Founder belief #3: trust is built by clinician-authored content, editorial independence, source linkage, and workflow fit. Evidence: {SRC['content_policy']}; {SRC['ai_clinical']}.
- 🟡 Strong Inference — Founder decision framework appears mission-first but commercially disciplined: prove wedge in medical education, expand to US, expand to clinicians/residents, add CME, acquire NEJM Knowledge+ and Novaheal, then raise long-duration capital. Evidence: {SRC['about']}; {SRC['nejm_acq']}; {SRC['novaheal_pr']}; {SRC['funding']}.
- 🟡 Strong Inference — Risk tolerance is moderate-to-high in product/category expansion but conservative in clinical claims; they launch AI but heavily constrain it as evidence navigation, exclude PHI, and require human judgment. Evidence: {SRC['ai_clinical']}; {SRC['terms']}.
- 🔴 Speculation — 10-year ambition is to become the default medical intelligence infrastructure layer for every healthcare professional from school through practice and, via MCP, for AI agents that require grounded medical knowledge.
- 🔴 Speculation — Likely internal strategy is to convert medical-student familiarity into resident/clinician/institutional contracts, then use AI/search to attack point-of-care reference incumbents while expanding into nurses and allied health.

---

## 4. Product Reverse Engineering

### Publicly visible product map

| Status | Surface | Key visible pages / actions | Evidence |
|---|---|---|---|
| 🟢 Confirmed | Marketing site | Students, clinicians, institutions, pricing, features, AI, careers, legal, help center | {SRC['features']} |
| 🟢 Confirmed | Registration | Email/password, SSO options, email verification, profile completion, 5-day free trial | {SRC['account']} |
| 🟢 Confirmed | Learning tab | Qbank, study plans & courses, Analysis, Library, AI Mode Learning | {SRC['support_platform']}; {SRC['ai_learning']} |
| 🟢 Confirmed | Clinical Care tab | Drugs, Clinical Resources, Quick Guides, Flowcharts, Calculators, Library, AI Mode entry points | {SRC['support_platform']}; {SRC['support_clinical']} |
| 🟢 Confirmed | Teaching tab | Assignments, analytics, groups, admin roles | {SRC['support_educator']} |
| 🟢 Confirmed | Account settings | Career/study profile, contact details, membership/licenses, payment, notes/network/settings, reset stats/password, deletion | {SRC['profile']} |
| 🟢 Confirmed | Mobile apps | Knowledge app and Qbank app, offline access, cross-device workflows | {SRC['mobile']}; {SRC['accessibility']} |
| 🟢 Confirmed | Integrations | Anki add-on, Chrome extension, AMBOSS GPT, MCP partner interface | {SRC['anki']}; {SRC['chrome']}; {SRC['gpt']}; {SRC['mcp']} |

### Feature-level reverse engineering summary

- 🟢 Confirmed — AMBOSS’s integrated retention loop is: search/read article → encounter highlighted concepts → start Qbank session from article → answer questions → get session analysis/study recommendations → return to relevant articles/questions → score predictor/peer comparison reinforces readiness. Evidence: {SRC['support_feature_overview']}; {SRC['support_platform']}; {SRC['support_analysis']}.
- 🟢 Confirmed — AMBOSS’s clinical workflow loop is: open Clinical Care/search → identify article/drug/algorithm/checklist/calculator → review sources/doses/guidelines → optionally earn CME → save notes/continue via mobile/offline. Evidence: {SRC['support_clinical']}; {SRC['cme_poc']}.
- 🟢 Confirmed — AMBOSS’s institution loop is: admin/educator creates groups → creates assignments → learners complete → educator views performance → recommendations/remediation → renewal/expansion. Evidence: {SRC['support_assignments']}; {SRC['support_analytics']}.
- 🟢 Confirmed — AMBOSS’s AI learning loop is: ask/upload → AI explanation → links to articles/Qbank/Anki → use recommended resources → activity counts toward progress. Evidence: {SRC['ai_learning']}; {SRC['support_ai_learning']}.
- 🟢 Confirmed — AMBOSS’s AI clinical loop is: natural-language query → trusted-source retrieval → short AI-generated summary → links to full source review → user remains responsible for judgment. Evidence: {SRC['ai_clinical']}; {SRC['support_ai_clinical']}.
- 🟡 Strong Inference — The hidden product dependency is a medical content graph joining articles, sections, terms, questions, answer choices, media, drugs, guidelines, calculators, competencies, exams, specialties, and user performance data. Evidence: feature behavior across search, split view, Qbank article links, study recommendations, AI retrieval. Evidence: {SRC['support_feature_overview']}; {SRC['support_recs']}; {SRC['gpt']}.

### Publicly visible “buttons / actions” inventory

- 🟢 Confirmed — Marketing CTAs include Start Free Trial, See Pricing, Buy Now, Full Access Bundle, Get Student Life, Book a Demo, Request Report, Request Consultation/Quote, Download apps, Install Chrome Extension, Download Anki add-on. Evidence: {SRC['pricing']}; {SRC['medical_schools']}; {SRC['residency']}; {SRC['chrome']}; {SRC['anki']}.
- 🟢 Confirmed — Learning actions include Create Qbank Session, Create study plan, Start session, Review session, Show Stats, Show Answer, Reset Question, Mark, Rule Out, Notes, Lab Values, Refine, Qbank from article, High-Yield toggle, Key Exam Info toggle, Add to Collections. Evidence: {SRC['support_qbank']}; {SRC['support_modes']}; {SRC['support_feature_overview']}.
- 🟢 Confirmed — Clinical actions include search condition/drug, jump to checklist/treatment, click DOSAGE, open QxMD calculator, cross out differential items, check management steps, open AI Mode entry point with prefilled prompt. Evidence: {SRC['support_management']}; {SRC['support_dosing']}; {SRC['support_calc']}; {SRC['support_ddx']}; {SRC['support_clinical']}.
- 🟢 Confirmed — Educator actions include Create Assignment, choose title/type/question pool, filter questions, assign to groups/learners, add articles, view analytics, manage groups/users/roles. Evidence: {SRC['support_assignments']}; {SRC['support_educator']}; {SRC['support_analytics']}.
- 🟢 Confirmed — Account/security actions include Manage Account, verify email, activate institutional license, connect/remove SSO, compare memberships, update payment, reset password, reset statistics, delete account. Evidence: {SRC['account']}; {SRC['profile']}; {SRC['sso']}; {SRC['delete']}.

---

## 5. Complete User Journeys

### Anonymous visitor → trial → subscription

```mermaid
flowchart TD
A[Anonymous visitor] --> B[Marketing page: students / clinicians / institutions]
B --> C[Start free trial CTA]
C --> D[Email/password or SSO registration]
D --> E[Email verification]
E --> F[Profile: role, study objective, optional settings]
F --> G[5-day trial: library + 50 Qbank questions]
G --> H[Learning / Clinical Care / Teaching intent selection]
H --> I[Search / Qbank / AI / mobile / integration activation]
I --> J[Analysis, recommendations, notes, saved content]
J --> K[Pricing page or Membership & licenses]
K --> L[Monthly/yearly/Student Life/Qbank add-on/clinician plan]
L --> M[Payment via supported methods]
M --> N[Renewal, cancellation, refund, support]
```

- 🟢 Confirmed — Free registration uses email/password or SSO options and requires email verification to keep using AMBOSS. Evidence: {SRC['account']}.
- 🟢 Confirmed — Free trial is 5 days, requires no payment info, and includes the platform plus 50 Qbank questions. Evidence: {SRC['free_trial']}.
- 🟢 Confirmed — Standard membership gives unlimited library plus 50 Qbank questions/month; unlimited Qbank requires add-on. Evidence: {SRC['membership']}.
- 🟢 Confirmed — Memberships auto-renew and can be canceled in Membership & Licenses. Evidence: {SRC['membership']}.
- 🟢 Confirmed — Direct purchases have a 30-day no-questions-asked refund policy. Evidence: {SRC['pricing']}.

### Institutional learner journey

```mermaid
flowchart TD
A[Student/resident/faculty eligible for license] --> B[Create AMBOSS account or login]
B --> C{{Activation path}}
C --> D[SSO via institution]
C --> E[Institution email verification]
C --> F[Access code / direct invite / pre-set account]
D --> G[License active in Membership & licenses]
E --> G
F --> G
G --> H[Complete assigned articles/Qbank]
H --> I[Educator/Admin analytics dashboard]
I --> J[Personalized recommendations / remediation]
```

- 🟢 Confirmed — AMBOSS supports institutional license activation through SSO, email verification, access codes, direct invites, and pre-set accounts. Evidence: {SRC['sso']}; {SRC['email_license']}.
- 🟢 Confirmed — SSO institutional license access must be refreshed every 12 months. Evidence: {SRC['sso']}.
- 🟢 Confirmed — Authorized educators/admins can view assignment performance data according to roles and groups. Evidence: {SRC['support_analytics']}.

### Clinical AI journey

```mermaid
flowchart TD
A[Clinician query] --> B[AI Mode Clinical Care]
B --> C[Search trusted sources: AMBOSS articles, drug database, selected guidelines]
C --> D[Generate brief structured summary]
D --> E[Inline links to original/full sources]
E --> F[Clinician independently reviews sources]
F --> G[Clinical judgment outside AMBOSS]
G --> H[Potential CME if eligible search/article workflow]
```

- 🟢 Confirmed — AI Mode is a search/research aid that supports but does not replace professional judgment. Evidence: {SRC['ai_clinical']}; {SRC['terms']}.
- 🟢 Confirmed — AI Mode refuses/indicates when it cannot find relevant information according to AMBOSS. Evidence: {SRC['ai_clinical']}.
- 🟢 Confirmed — AI Mode Clinical Care must not be used in emergencies/time-sensitive situations without time to review information. Evidence: {SRC['terms']}.

---

## 6. UX Research

- 🟢 Confirmed — AMBOSS’s redesigned navigation is explicitly based on role/current intent and allows switching among Learning, Clinical Care, and Teaching. Evidence: {SRC['new_nav']}.
- 🟢 Confirmed — The interface supports split view, article-section notes, collections, high-yield filtering, key-exam highlighting, personal highlighting, and search cards. Evidence: {SRC['support_feature_overview']}; {SRC['support_search']}.
- 🟢 Confirmed — Exam Mode includes reverse color and text zoom, while accessibility pages list broader WCAG efforts and known exceptions. Evidence: {SRC['support_modes']}; {SRC['accessibility']}.
- 🟢 Confirmed — The Chrome extension supports light/dark/system themes and site-level toggling. Evidence: {SRC['chrome']}.
- 🟢 Confirmed — AMBOSS’s accessibility statement says it strives to meet WCAG 2.1 AA for in-scope digital services and lists known issues in contrast, iFrames, signup/login, navigation, keyboard focus, form errors, and legacy purchase flows. Evidence: {SRC['accessibility']}.
- 🟡 Strong Inference — The UX philosophy is “progressive compression”: show enough for quick action, but let users drill into evidence, details, or full articles when needed. Evidence: AI Shortcuts, High-Yield mode, split view, references, source links. Evidence: {SRC['features']}; {SRC['ai_clinical']}.
- 🟡 Strong Inference — The primary UX risk is perceived density/text-heaviness for visual learners; this is inferred from product breadth and third-party/user feedback. Evidence: public review snippets and feature pages.
- 🟡 Strong Inference — A key conversion optimization pattern is risk removal: no-card 5-day trial, 30-day refund, student plans, institutional licenses, group discounts, and free utility integrations. Evidence: {SRC['free_trial']}; {SRC['pricing']}; {SRC['group_discounts'] if 'group_discounts' in SRC else 'https://www.amboss.com/us/group-discounts'}.

---

## 7. Healthcare Workflow Reverse Engineering

| Workflow | Status | AMBOSS coverage | Evidence |
|---|---|---|---|
| Clinical reference | 🟢 Confirmed | Search, articles, guidelines, AI Mode, drug DB, calculators, checklists, flowcharts | {SRC['clinicians']}; {SRC['support_clinical']} |
| Patient workflow | 🟢 Confirmed | Not intended for patients or laypeople | {SRC['attendings']}; {SRC['terms']} |
| Provider workflow | 🟢 Confirmed | Supports physicians/NPs/PAs with evidence lookup and CME | {SRC['attendings']} |
| Hospital workflow | 🟢 Confirmed | Supports management checklists, drug dosing, differential diagnosis, calculators and offline app; no PHI workflow | {SRC['support_clinical']}; {SRC['terms_clinics']} |
| Insurance workflow | 🟢 Confirmed | No public AMBOSS insurance workflow found | Public source review |
| Lab workflow | 🟢 Confirmed | Lab values in Qbank and clinical calculators; no lab data ingestion found | {SRC['support_modes']}; {SRC['support_calc']} |
| Pharmacy workflow | 🟢 Confirmed | Drug database/dosing reference; no pharmacy transaction integration found | {SRC['support_drug']} |
| Referral/consult workflow | 🟢 Confirmed | Prep for Consult assistant beta and consult guidance features; no referral network found | {SRC['assistants']}; {SRC['features']} |
| Medical records | 🟢 Confirmed | AMBOSS terms say not for PHI storage/processing/transmission | {SRC['terms_clinics']} |
| Clinical documentation | 🟢 Confirmed | Dot Phrase Generator beta and dot phrases feature; no confirmed EHR writeback | {SRC['assistants']}; {SRC['features']} |
| Care coordination | 🟡 Strong Inference | Supports knowledge coordination for learners/faculty; not care-team coordination over patients | {SRC['support_educator']}; {SRC['terms_clinics']} |

---

## 8. Healthcare Data Architecture

### AMBOSS actual public data architecture

```mermaid
flowchart LR
A[Medical content sources: guidelines, textbooks, primary literature] --> B[AMBOSS editorial process / content graph]
B --> C[Articles, sections, images, videos, flowcharts]
B --> D[Qbank questions, competencies, exams, answer explanations]
B --> E[Drug DB via AHFS]
B --> F[Calculators via QxMD]
B --> G[AI retrieval corpus]
H[User profile: role, study objective, accommodations] --> I[Personalization engine]
J[Qbank attempts, search activity, notes, highlights, collections] --> I
I --> K[Study recommendations, EPC, score predictor, peer comparison]
K --> L[User dashboards]
K --> M[Educator/Admin dashboards when institution-enabled]
N[AI prompts/uploads] --> O[AI processors and MongoDB storage under stated policies]
O --> G
G --> P[AI Mode Clinical Care / Learning / Assistants / GPT / MCP]
```

- 🟢 Confirmed — AMBOSS processes user account data, optional profile data, login/IP data, usage statistics, institutional usage data, AI inputs, score-predictor data, comments, and cookies according to privacy policy. Evidence: {SRC['privacy']}.
- 🟢 Confirmed — AMBOSS stores AI Feature input data for product optimization and does not allow users to delete individual interactions during beta. Evidence: {SRC['privacy']}; {SRC['assistants_privacy']}.
- 🟢 Confirmed — AMBOSS prohibits patient data/PHI in AI features and states clinics product is not intended to process PHI. Evidence: {SRC['terms']}; {SRC['terms_clinics']}.
- 🟡 Strong Inference — AMBOSS likely maintains a structured medical knowledge graph because semantic search, split view, term underlining, Qbank-to-article mapping, Anki matching, AI retrieval, and recommendations require entity/relationship metadata. Evidence: {SRC['support_feature_overview']}; {SRC['anki']}; {SRC['gpt']}.

### FHIR / HL7 / CCD / Apple Health / wearables / labs / genomics

- 🟢 Confirmed — No public source reviewed confirms AMBOSS medical product support for FHIR, HL7, CCD/CCDA, Apple Health, Google Health Connect, wearables, labs ingestion, imaging ingestion, genomics ingestion, insurance claims, pharmacy transaction data, or patient identity matching.
- 🟢 Confirmed — AMBOSS is explicitly not positioned as a patient longitudinal record or PHI-processing system in the reviewed legal/product materials. Evidence: {SRC['terms_clinics']}.
- 🟡 Strong Inference — This is AMBOSS’s largest gap relative to Ovexis: AMBOSS owns clinician knowledge workflows, not longitudinal personal health data workflows.

---

## 9. AI Reverse Engineering

### Confirmed AI surfaces

| Status | AI surface | What it does | Evidence |
|---|---|---|---|
| 🟢 Confirmed | AI Mode Clinical Care / LiSA | Natural-language clinical search over trusted sources with brief summary and source links | {SRC['ai_clinical']} |
| 🟢 Confirmed | AI Mode Learning | Study copilot for explanations, uploads, Qbank/Anki recommendations, progress | {SRC['ai_learning']} |
| 🟢 Confirmed | AMBOSS Assistants beta | Article-embedded tools for Learn, Practice, Teach use cases | {SRC['assistants']} |
| 🟢 Confirmed | AI Shortcuts / Related Questions | AI snippets and generated related questions in search results | {SRC['features']}; {SRC['support_search']} |
| 🟢 Confirmed | Study Recommendations | AI/ML-supported recommendations based on performance | {SRC['features']}; {SRC['support_recs']} |
| 🟢 Confirmed | AMBOSS GPT | Custom GPT queries AMBOSS library and returns links | {SRC['gpt']} |
| 🟢 Confirmed | MCP | Agent interface to AMBOSS content | {SRC['mcp']} |

### Inferred architecture

```mermaid
flowchart TD
A[User input: clinical query, study question, upload, article context] --> B[Intent and product-surface router]
B --> C[Profile context: role, specialty, study objective]
B --> D[Safety policy: no PHI, no emergency, output limits]
C --> E[Retrieval: AMBOSS articles, drug DB, selected guidelines, Qbank/media]
D --> E
E --> F[Ranker / context window builder]
F --> G[LLM via secure APIs: OpenAI and/or Anthropic for Assistants; providers undisclosed for all surfaces]
G --> H[Structured answer with citations, next steps, or assistant output]
H --> I[User verifies source / follows recommendation]
I --> J[Feedback / usage analytics / product optimization storage]
```

- 🟢 Confirmed — AMBOSS uses OpenAI and Anthropic for AMBOSS Assistants and MongoDB Cloud Services for storage/management of databases necessary for AMBOSS Assistants. Evidence: {SRC['assistants_privacy']}.
- 🟢 Confirmed — AI Mode Learning FAQ says requests are sent to large language models through secure APIs and are not used to train those models. Evidence: {SRC['support_ai_learning']}.
- 🟢 Confirmed — AI Mode Clinical Care FAQ says requests are sent to LLMs through secure APIs and are not used to train the models. Evidence: {SRC['support_ai_clinical']}.
- 🟡 Strong Inference — AMBOSS AI is RAG-like because it searches trusted sources, identifies relevant content, builds summaries, and links to source material. Evidence: {SRC['ai_clinical']}; {SRC['gpt']}.
- 🟡 Strong Inference — AMBOSS appears to use separate system prompts/policies per surface: Learning, Clinical Care, Teaching, Assistants, and GPT, because each surface has distinct roles, allowed inputs, and output formats. Evidence: {SRC['ai_learning']}; {SRC['terms']}.
- 🟢 Confirmed — AMBOSS says it regularly assesses output quality against curated questions using defined scoring criteria. Evidence: {SRC['ai_clinical']}.
- 🟢 Confirmed — AMBOSS AI principles say AMBOSS Intelligence-labeled content has medical professional review, clinical grounding, rigorous quality control, and continuous feedback/testing. Evidence: {SRC['ai_principles']}.
- 🔴 Speculation — AMBOSS likely maintains a retrieval evaluation harness with human-labeled queries, judgment lists, ranking metrics, and LLM-assisted evaluation due to its search evaluation blog and AI product claims, but exact production evaluation stack is not public.

### AI guardrails and safety

- 🟢 Confirmed — AMBOSS requires users to verify outputs and not substitute AMBOSS AI for professional judgment. Evidence: {SRC['terms']}.
- 🟢 Confirmed — AMBOSS AI Mode Clinical Care is not for emergencies/time-sensitive situations without time to independently review information. Evidence: {SRC['terms']}.
- 🟢 Confirmed — AMBOSS instructs users not to enter patient data/PHI into AI features. Evidence: {SRC['terms']}.
- 🟢 Confirmed — AMBOSS highlights differing recommendations when sources differ. Evidence: {SRC['ai_clinical']}.
- 🟢 Confirmed — AMBOSS says AI Mode indicates when it cannot find relevant information. Evidence: {SRC['ai_clinical']}.

---

## 10. Technical Reverse Engineering

| Layer | Status | Evidence-backed assessment | Evidence |
|---|---|---|---|
| Marketing frontend | 🟢 Confirmed | Webflow-hosted/served marketing pages with CloudFront/Cloudflare/CDN headers observed; Webflow asset URLs visible | HTTP header capture; {SRC['funding']} |
| Careers site | 🟢 Confirmed | Careers site returns nginx/PHP/Plesk headers; this applies to careers subdomain, not necessarily core app | HTTP header capture; {SRC['careers']} |
| Core app | 🟡 Strong Inference | likely SPA/web app on next.amboss.com with React/TypeScript front-end and backend APIs | Careers job requiring React/TypeScript and GraphQL/REST: {SRC['shop_job']} |
| Backend | 🟡 Strong Inference | Go and Python are used in production teams; backend services likely include Go/Python | {SRC['shop_job']} plus Data Engineer search snippets |
| APIs | 🟡 Strong Inference | Internal GraphQL and/or REST APIs are used; no public API docs found | {SRC['shop_job']} |
| Cloud/hosting | 🟢 Confirmed | AWS for website/registered area; CloudFront, Cloudflare, Cloudinary CDNs | {SRC['privacy']} |
| Identity | 🟢 Confirmed | Auth0 IAM; institutional SSO supported | {SRC['privacy']}; {SRC['sso']} |
| Data warehouse/analytics | 🟡 Strong Inference | BigQuery/Snowflake, Airflow, Airbyte, dbt, Looker, Mixpanel appear in public job snippets | Public career search snippets |
| Monitoring | 🟢 Confirmed | Datadog and Sentry used | {SRC['privacy']} |
| Product analytics | 🟢 Confirmed | Segment and Amplitude listed; Bunchbox also listed | {SRC['privacy']} |
| Messaging/CRM | 🟢 Confirmed | Braze for contract messages/notifications; HubSpot for forms/CRM/email marketing; Zendesk for support | {SRC['privacy']} |
| Payments | 🟢 Confirmed | Stripe Payments Europe and Shopify Shop Pay are listed; membership page mentions credit/debit cards, ApplePay, PayPal | {SRC['privacy']}; {SRC['membership']} |
| AI storage/providers | 🟢 Confirmed | OpenAI, Anthropic, MongoDB Cloud Services for Assistants; Zendesk AI Agent uses OpenAI via Zendesk | {SRC['assistants_privacy']}; {SRC['privacy']} |
| CI/CD/IaC | 🟡 Strong Inference | Docker, CI/CD pipelines, AWS/Kubernetes are required in engineering job; Terraform/GitHub Actions appear in data platform snippets | {SRC['shop_job']} |
| Security headers | 🟢 Confirmed | HSTS, CSP, X-Frame-Options, X-Content-Type-Options observed on public pages | HEAD capture |
| Feature flags | 🔴 Speculation | Likely used due to beta/limited-access AI and experiments, but no public explicit provider confirmed | product behavior only |

---

## 11. API Investigation

- 🟢 Confirmed — AMBOSS publicly announced an MCP server for AI agents, with access granted to a limited number of AI development teams. Evidence: {SRC['mcp']}.
- 🟢 Confirmed — MCP is described as a standardized open-source interface developed by Anthropic and adopted by OpenAI and others; AMBOSS’s MCP lets agents browse/interact with AMBOSS articles, drug monographs, flowcharts, calculators, clinical scores, patient cases and more. Evidence: {SRC['mcp']}.
- 🟡 Strong Inference — AMBOSS likely has internal APIs for web/mobile/platform functions; public job descriptions mention GraphQL and/or REST experience. Evidence: {SRC['shop_job']}.
- 🟢 Confirmed — No public OpenAPI specification, REST API docs, GraphQL schema, webhooks documentation, FHIR API, HL7 integration, SDK, or public rate-limit documentation for the AMBOSS medical platform was found in reviewed sources.
- 🟢 Confirmed — AMBOSS GPT has a quota detail: non-AMBOSS email users are limited to 50 prompts every 3 months, AMBOSS-email quota is lifted, and OpenAI rate limits still apply. Evidence: {SRC['gpt']}.
- 🟡 Strong Inference — AMBOSS is cautiously moving from closed SaaS to controlled developer ecosystem via MCP, not a fully open developer platform.

---

## 12. Security / Privacy / Compliance Investigation

| Topic | Status | Finding | Evidence |
|---|---|---|---|
| GDPR | 🟢 Confirmed | Privacy policy is GDPR-based with Art. 6 legal bases, DPO contact, SCC transfers, data subject rights context | {SRC['privacy']} |
| Data controller | 🟢 Confirmed | AMBOSS SE, Torstrasse 19, Berlin; privacy contact listed | {SRC['privacy']} |
| PHI/HIPAA posture | 🟢 Confirmed | Clinics terms state AMBOSS does not process PHI and program is not intended for PHI storage/processing/transmission | {SRC['terms_clinics']} |
| BAA | 🟢 Confirmed | No public evidence of a BAA offering found in reviewed sources; PHI is prohibited | {SRC['terms_clinics']} |
| SOC 2 / ISO 27001 | 🟢 Confirmed | No public AMBOSS SOC 2/ISO certification evidence found; AWS certifications are cited for AWS, not AMBOSS itself | {SRC['joint']} |
| Encryption | 🟢 Confirmed | Joint data processing measures list HTTPS/SFTP, encrypted connections, encrypted notebooks/tablets/data carriers | {SRC['joint']} |
| Access control | 🟢 Confirmed | Measures include user/password auth, firewalls, VPN, central password rules, 2FA, authorization concept and low admin count | {SRC['joint']} |
| Audit logs | 🟢 Confirmed | Measures include logging access and data entry/modification/deletion | {SRC['joint']} |
| Incident response | 🟢 Confirmed | Measures include breach reporting/notification processes and DPO involvement | {SRC['joint']} |
| AI data training | 🟢 Confirmed | Terms state AMBOSS does not include Input in datasets for LLM training | {SRC['terms']} |
| AI data deletion | 🟢 Confirmed | AI input stored for product optimization; individual interaction deletion unavailable during beta | {SRC['privacy']} |
| FDA / medical device | 🟢 Confirmed | AMBOSS terms repeatedly disclaim diagnostic/treatment tool status and responsibility remains with clinician | {SRC['terms']} |
| ONC/FHIR | 🟢 Confirmed | No public evidence found; AMBOSS is not a patient record/EHR tool in reviewed sources | public source review |

### Threat model

- 🟡 Strong Inference — Major risks are medical misinformation, hallucination/misinterpretation, stale guidelines, overreliance by trainees, institutional data privacy leakage, account sharing, AI prompt data exposure, content scraping, and extension/plugin attack surface.
- 🟢 Confirmed — AMBOSS mitigates some risks through no-PHI rules, source grounding, medical review, legal disclaimers, privacy policies, access controls, and content-scraping prohibitions. Evidence: {SRC['content_policy']}; {SRC['terms']}; {SRC['joint']}.
- 🔴 Speculation — Enterprise procurement may increasingly demand SOC 2 Type II, BAA, pen-test summaries, model cards, and safety evaluations if AMBOSS moves deeper into EHR/patient-context workflows.

---

## 13. Business Model

### Pricing and revenue streams

- 🟢 Confirmed — US student Monthly Access is $19.99/month billed monthly and Yearly Access is $12.50/month billed yearly. Evidence: {SRC['pricing']}.
- 🟢 Confirmed — US student 12-month Qbank Bundle is listed at $448 and Student Life at $1199. Evidence: {SRC['pricing']}; {SRC['students']}.
- 🟢 Confirmed — Clinician Monthly Access is $29.99/month and Yearly Access is $21.58/month billed yearly. Evidence: {SRC['pricing']}.
- 🟢 Confirmed — AMBOSS sells institutional licenses with flexible coverage plans and implementation solutions; pricing is not public. Evidence: {SRC['pricing']}; {SRC['medical_schools']}; {SRC['residency']}.
- 🟢 Confirmed — AMBOSS sells/activates AMBOSS Courses separately from membership according to terms. Evidence: {SRC['terms']}.
- 🟢 Confirmed — Self-Assessments are included in Student Life; otherwise official page lists $49.99. Evidence: {SRC['self_assessment'] if 'self_assessment' in SRC else 'https://www.amboss.com/us/usmle/self-assessment'}.

### Unit economics and sales motion

- 🟡 Strong Inference — Gross margins are likely software-like but lower than pure SaaS due to large medical editorial team, content licensing, AI inference costs, mobile/offline infrastructure, and support. Evidence: {SRC['content_policy']}; {SRC['funding']}; {SRC['privacy']}.
- 🟡 Strong Inference — B2C CAC is reduced by SEO, free trial, group discounts, Anki/Chrome integrations, self-assessments, institutional familiarity, and word-of-mouth. Evidence: {SRC['chrome']}; {SRC['anki']}; {SRC['group_discounts'] if 'group_discounts' in SRC else 'https://www.amboss.com/us/group-discounts'}; {SRC['self_assessment'] if 'self_assessment' in SRC else 'https://www.amboss.com/us/usmle/self-assessment'}.
- 🟡 Strong Inference — B2B motion is consultative sales to medical schools/residency programs/clinics with demos, quotes, implementation, usage reports, and renewals. Evidence: {SRC['medical_schools']}; {SRC['residency']}.
- 🟡 Strong Inference — LTV is lifted by career-stage continuity: preclinical → clerkships → boards → residency → attending CME/recertification. Evidence: {SRC['new_nav']}; {SRC['students']}; {SRC['residency']}.
- 🔴 Speculation — AI Mode may become a premium subscription layer later; AMBOSS explicitly reserves the right to offer AI Mode in a Premium subscription. Evidence: {SRC['ai_clinical']}; {SRC['ai_learning']}.

---

## 14. Growth Strategy

- 🟢 Confirmed — AMBOSS uses a 5-day no-card free trial. Evidence: {SRC['free_trial']}.
- 🟢 Confirmed — AMBOSS uses group discounts where larger groups get higher discounts. Evidence: https://www.amboss.com/us/group-discounts.
- 🟢 Confirmed — AMBOSS runs or has run scholarship/global education initiatives. Evidence: https://www.amboss.com/us/scholarships-program.
- 🟢 Confirmed — AMBOSS has free/low-friction tools and integrations including Chrome extension, Anki add-on, AMBOSS GPT, score predictor pages, self-assessment events, and content pages. Evidence: {SRC['chrome']}; {SRC['anki']}; {SRC['gpt']}; {SRC['support_score']}.
- 🟢 Confirmed — AMBOSS uses institutional logos/social proof and testimonials across student, clinician, residency and reviews pages. Evidence: {SRC['students']}; {SRC['clinicians']}; {SRC['residency']}; {SRC['reviews']}.
- 🟢 Confirmed — AMBOSS has official social channels linked in footers: YouTube, Facebook, Instagram, X, LinkedIn, TikTok. Evidence: {SRC['funding']}.
- 🟢 Confirmed — AMBOSS participates in conferences for residency/GME audiences. Evidence: {SRC['residency']}.
- 🟢 Confirmed — AMBOSS highlights NOHARM ranking for AI Mode in clinician/residency/AI pages. Evidence: {SRC['ai_innovation']}; {SRC['residency']}.
- 🟡 Strong Inference — AMBOSS’s best growth loop is “student familiarity becomes institutional clinician adoption,” because 2/3 US med students claim and resident/clinic products build on existing use. Evidence: {SRC['medical_schools']}; {SRC['residency']}.

---

## 15. Hiring Intelligence

- 🟢 Confirmed — AMBOSS careers site says offices are Berlin, Cologne, Cagliari, New York, and remote. Evidence: {SRC['careers']}.
- 🟢 Confirmed — Careers site organizes departments into Medical, Product & Engineering, Commercial, General & Administration. Evidence: {SRC['careers']}.
- 🟢 Confirmed — Product & Engineering page lists leadership roles including Director of AI & Knowledge Retrieval, Director of Research, Director of Medical Product, Director of User Experience Design. Evidence: {SRC['careers_eng']}.
- 🟢 Confirmed — An open Senior Backend Engineer, Shop & Payments role requires Go, React willingness, GraphQL/REST API design, Docker, CI/CD, AWS/Kubernetes, monitoring, database management, and payment/subscription domain expertise. Evidence: {SRC['shop_job']}.
- 🟡 Strong Inference — Hiring implies active work on commerce, localization, subscriptions, international tax compliance, payment reliability, and global access funnels. Evidence: {SRC['shop_job']}.
- 🟡 Strong Inference — Search/AI/data roles and blog posts imply continued investment in AI retrieval, evaluation, data platforms, personalization, and product analytics. Evidence: {SRC['careers_eng']}; {SRC['search_eval']}.
- 🟡 Strong Inference — Medical Project Manager / course role snippets imply ongoing CME/course content expansion, particularly in German-speaking markets.

---

## 16. Customer Intelligence

### Praise themes

- 🟢 Confirmed — Official reviews and public snippets praise AMBOSS as organized, efficient, useful for clinical rotations, and integrated with library/Qbank/Anki. Evidence: {SRC['reviews']}; public Reddit search results.
- 🟡 Strong Inference — Users value AMBOSS most when they need foundations, context, and integrated explanations rather than only a final answer.
- 🟡 Strong Inference — Anki and Chrome integrations create disproportionate love because they fit existing study workflows rather than asking users to change habits.
- 🟡 Strong Inference — Clinician/resident praise centers on speed, confidence, and not breaking workflow on wards/rounds. Evidence: {SRC['residency']}.

### Complaint themes

- 🟡 Strong Inference — Recurrent public complaints include AMBOSS Qbank being hard, picky, low-yield, or 4–5 hammer questions causing overthinking compared with UWorld/NBME style.
- 🟡 Strong Inference — Some users view UWorld as the stronger or gold-standard pure Qbank, while AMBOSS is often praised as better integrated learning/reference.
- 🟡 Strong Inference — Mobile/extension pain points include inability to highlight passages on mobile, iPad/Anki limitations, Chrome extension sidebar preferences, bugs, and desire for Edge support.
- 🟡 Strong Inference — Student price sensitivity appears in Reddit discussions about whether Student Life is worth it, especially when schools provide UWorld.

### Churn / expansion signals

- 🟡 Strong Inference — Churn risk is highest where schools provide UWorld/free alternatives and users frame AMBOSS as optional supplement.
- 🟡 Strong Inference — Expansion potential is highest in residency programs, clinicians, PA/NP students, nurses, CME, and AI-enabled point-of-care search.

---

## 17. Decision Ledger

| Status | Feature | Why built | Pain solved | KPI improved | Trade-off |
|---|---|---|---|---|---|
| 🟡 Strong Inference | Integrated Library+Qbank | Link passive learning to active recall | Fragmented study resources | Retention, exam outcomes, LTV | Large editorial/assessment cost |
| 🟡 Strong Inference | High-Yield mode | Reduce curriculum overload | Too much information near exams | Session duration, completion | Risk of oversimplifying |
| 🟡 Strong Inference | Key Exam Info/Learning Radar | Convert mistakes to remediation | Learners do not know what to revisit | Qbank reuse, mastery | Requires precise mapping |
| 🟡 Strong Inference | Adaptive sessions | Save time and target weak areas | Manual planning fatigue | Qbank completion, EPC improvement | Opaque algorithm frustration |
| 🟡 Strong Inference | Score Predictor | Reduce exam uncertainty | Anxiety and poor readiness estimation | Trial conversion, retention | Prediction liability/perception |
| 🟡 Strong Inference | Anki add-on | Meet students in dominant workflow | Flashcards lack context | Acquisition, engagement | Dependency on external app/deck |
| 🟡 Strong Inference | Chrome extension | Ambient knowledge outside AMBOSS | Constant Googling/tab switching | Top-of-funnel, daily active use | Extension bugs/privacy perception |
| 🟡 Strong Inference | Clinical Mode | Extend beyond exams into wards | Residents need fast clinical answers | Career-stage retention | Higher clinical risk |
| 🟡 Strong Inference | AHFS drug DB | Avoid separate drug app | Dosing/drug lookup fragmentation | Clinician conversion | Licensing cost |
| 🟡 Strong Inference | Management Checklists | Prevent missed urgent-care steps | Cognitive load in admissions | Clinical usage frequency | Must keep current and safe |
| 🟡 Strong Inference | CME | Turn search into required credit | Clinicians need CME anyway | Clinician retention | Accreditation overhead |
| 🟡 Strong Inference | Educator Tool | Sell to institutions and improve learning oversight | Faculty need assignments/analytics | B2B ACV, renewals | Privacy/RBAC complexity |
| 🟡 Strong Inference | AI Mode Clinical Care | Defend against generic AI and attack clinical reference | Natural-language evidence search | Clinician adoption, AI brand | AI safety/regulatory risk |
| 🟡 Strong Inference | AI Mode Learning | Convert generic AI usage into AMBOSS-native study flow | Students ask ChatGPT but lack trusted next steps | Engagement, retention | LLM cost and reliability |
| 🟡 Strong Inference | MCP | Become infrastructure for medical AI agents | AI teams need grounded medical sources | Platform partnerships | Content licensing/safety exposure |

---

## 18. Feature Dependency Graph

```mermaid
flowchart TD
Identity[Identity + role + study objective] --> Consent[Terms + privacy + AI no-PHI consent]
Consent --> Content[Clinician-reviewed content graph]
Content --> Search[Search + semantic retrieval]
Content --> Qbank[Qbank + assessment metadata]
Content --> ClinicalTools[Drugs + calculators + checklists + flowcharts]
Qbank --> Analytics[Attempts + EPC + peer comparison]
Analytics --> Recommendations[Study recommendations + adaptive sessions]
Search --> AI[AI Mode / GPT / Assistants / MCP]
ClinicalTools --> AI
Recommendations --> Retention[Daily study habit + progress]
AI --> Verification[Sources + citations + clinician judgment]
Identity --> Institution[Institution license + roles]
Institution --> Dashboards[Educator/admin analytics]
Dashboards --> Remediation[Assignments + remediation]
```

- 🟢 Confirmed — AMBOSS dependencies begin with identity/profile/study objective and membership/license state. Evidence: {SRC['profile']}; {SRC['membership']}.
- 🟢 Confirmed — Content graph and Qbank drive search, study recommendations, AI Mode Learning, Anki, and GPT. Evidence: {SRC['support_feature_overview']}; {SRC['ai_learning']}; {SRC['gpt']}.
- 🟢 Confirmed — Institution dashboards depend on institutional license activation, user roles, groups, assignments, and usage data. Evidence: {SRC['support_analytics']}.

---

## 19. Engineering Roadmap Reconstruction

| Status | Phase | Likely scope | Evidence / rationale |
|---|---|---|---|
| 🟡 Strong Inference | MVP | German medical exam Qbank + library + high-yield content | Founding story and early German exam frustration |
| 🟡 Strong Inference | V2 | USMLE/NBME localization, English library/Qbank, New York launch | Official 2017 US launch and USMLE/NBME materials |
| 🟡 Strong Inference | V3 | Mobile apps, Anki, Chrome, analytics, score predictor, study plans | Current support/product docs show mature learning ecosystem |
| 🟡 Strong Inference | V4 | Clinical Care: drug DB, calculators, differential, checklists, CME, resident/clinician expansion | Clinician/residency pages and CME docs |
| 🟡 Strong Inference | V5 | Institutional educator suite: assignments, analytics, group/RBAC, usage reports | Medical school/residency and support docs |
| 🟢 Confirmed | Current | AI Mode Clinical Care, AI Mode Learning, Assistants beta, MCP, redesigned navigation | 2025/2026 official product updates |
| 🔴 Speculation | Next | deeper AI assistants, nursing/allied health, international markets, controlled agent/API ecosystem, possible EHR-context experiments only if legal posture changes | Financing release and product direction |

### Technical debt hypotheses

- 🟡 Strong Inference — Accessibility statement reveals legacy purchase flows, legacy pages, iOS login/paywall issues, navigation inconsistencies, and commerce redesign/migration work. Evidence: {SRC['accessibility']}.
- 🟡 Strong Inference — Payment/shop job indicates commerce stack complexity around subscriptions, international tax, global payment processing, and potential monolith decomposition. Evidence: {SRC['shop_job']}.
- 🟡 Strong Inference — Multiple product surfaces and acquisitions likely create content model, identity, permission, and taxonomy integration debt.

---

## 20. Competitive Landscape

| Status | Company | Category | Compared with AMBOSS | Evidence |
|---|---|---|---|---|
| 🟢 Confirmed | UpToDate | Incumbent clinical reference/CDS | Stronger hospital enterprise penetration and expert-authored reference; AMBOSS stronger learner/Qbank lifecycle | https://www.wolterskluwer.com/en/news/uptodate-expert-ai-genai-clinical-decision-support |
| 🟢 Confirmed | OpenEvidence | AI medical search for verified clinicians | Stronger free AI search adoption claims; AMBOSS stronger education/Qbank/institution learning stack | https://www.prnewswire.com/news-releases/openevidence-achieves-historic-milestone-1-million-clinical-consultations-between-verified-doctors-and-an-artificial-intelligence-system-in-a-single-day-302712459.html |
| 🟡 Strong Inference | Glass Health | AI clinical workflow/DDx/docs | Glass closer to encounter workflow; AMBOSS closer to evidence/reference/education | https://glass.health/compare/uptodate |
| 🟢 Confirmed | Atropos Health | Real-world evidence generation | Atropos uses deidentified patient/RWD evidence; AMBOSS uses curated medical knowledge | https://www.atroposhealth.com/atropos-health-launches-new-geneva-os-and-chatrwd-application-for-rapid-real-world-evidence-with-generative-ai/ |
| 🟢 Confirmed | Function Health | Consumer biomarker testing | Function owns longitudinal consumer labs; AMBOSS owns clinician education/reference | https://www.functionhealth.com/article/function365 |
| 🟢 Confirmed | Superpower | Preventive health super-app | Superpower owns consumer lab/AI concierge; AMBOSS excludes patients/PHI | https://www.forbes.com.au/news/innovation/superpower-raises-30-million-to-launch-worlds-first-health-super-app/ |
| 🟢 Confirmed | Levels | CGM/metabolic health | Levels owns glucose/lifestyle data; AMBOSS owns medical knowledge workflows | https://apps.apple.com/us/app/levels-metabolic-health/id1481511675 |
| 🟢 Confirmed | Apple Health | Consumer health records/wearables | Apple owns device/record aggregation; AMBOSS has no public patient-data integration | https://www.healthcareitnews.com/news/apple-launch-health-records-app-hl7s-fhir-specifications-12-hospitals |
| 🟢 Confirmed | Google Health Connect | Android health-data API | Google provides on-device health data exchange; AMBOSS not in this data layer | https://developer.android.com/health-and-fitness/health-services/health-platform |
| 🟢 Confirmed | Human API | Health-data connectivity | Human API normalizes EHR/lab/wearable data for insurance/health workflows; AMBOSS is content/CDS | https://www.prnewswire.com/news-releases/human-api-launches-health-intelligence-platform-to-modernize-life-insurance-underwriting-and-customer-experience-301320064.html |
| 🟢 Confirmed | Practo / Apollo 24/7 / Tata 1mg | India consumer digital health | These own consultation/pharmacy/lab commerce; AMBOSS owns professional education/reference | https://www.apollo247.com/; https://apps.apple.com/us/app/tata-1mg-healthcare-app/id554578419 |
| 🟢 Confirmed | Healthify | AI nutrition/fitness coaching | Healthify owns consumer coaching; AMBOSS owns professional knowledge | https://www.business-standard.com/companies/news/ai-powered-fitness-app-healthify-secures-45-mn-to-drive-us-expansion-124102500232_1.html |
| 🟢 Confirmed | WHOOP / Oura / Ultrahuman | Wearable biometrics | Wearables own continuous biometric data; AMBOSS does not process user biometrics publicly | https://athletechnews.com/oura-whoop-push-deeper-into-healthcare/ |
| 🟢 Confirmed | PreventiveHealth.ai | Preventive personalized health | PreventiveHealth.ai appears consumer personalized wellness; AMBOSS is medical professional education/CDS | https://preventivehealth.ai/pages/about-us |
| 🟢 Confirmed | RegCore.AI | Regulatory intelligence, not health | Search found RegCore.AI as financial regulatory intelligence, not healthcare competitor; user may mean another Regacore | https://regcore.ai/ |

### Competitive advantages

- 🟢 Confirmed — AMBOSS’s unique advantage is integration across Qbank, library, clinical tools, mobile/offline, educators, and AI. Evidence: {SRC['new_nav']}; {SRC['features']}.
- 🟡 Strong Inference — AMBOSS’s main weakness versus Ovexis-like platforms is lack of longitudinal personal health data ingestion and patient-facing workflows.
- 🟡 Strong Inference — AMBOSS’s main weakness versus OpenEvidence is potential friction/paywall and lower “free physician AI search” virality, though AMBOSS has deeper education/learning loops.
- 🟡 Strong Inference — AMBOSS’s main weakness versus UpToDate is incumbent hospital procurement and enterprise trust; its strength is modern UX and education-to-practice continuity.

---

## 21. Moat Analysis

| Moat | Status | Strength | Rationale |
|---|---|---|---|
| Content/data moat | 🟢 Confirmed | Strong | Clinician-reviewed library, Qbank, questions, analytics, sources, updates |
| AI moat | 🟡 Strong Inference | Medium → Future Strong | RAG over proprietary content + LiSA/NOHARM + MCP, but model providers are not unique |
| Clinical moat | 🟢 Confirmed | Strong | 150+ physicians in 2025 release; editorial policy; ACCME; AHFS/NEJM |
| Brand moat | 🟢 Confirmed | Strong in students, medium in attendings | 2/3 US med student claim; 1M+ users; clinician expansion ongoing |
| Distribution moat | 🟢 Confirmed | Strong | Schools/residencies, apps, Anki, Chrome, GPT, group discounts |
| Developer moat | 🟡 Strong Inference | Future | MCP limited access, not yet broad public platform |
| Marketplace moat | 🟢 Confirmed | Weak | No visible marketplace business beyond courses and content partnerships |
| Regulatory moat | 🟢 Confirmed | Medium | ACCME, GDPR posture, no-PHI constraints; not a PHI/EHR regulated moat |
| Network effects | 🟡 Strong Inference | Medium | Peer stats, institutional cohorts, educator assignments, Anki/community loops |
| Switching costs | 🟡 Strong Inference | Medium | Notes, progress, analytics, institution license, study plans; content can be substituted by UWorld/UpToDate |
| Trust moat | 🟢 Confirmed | Strong | No ads/pharma sponsorship, content policy, source citations, clinician review |

---

## 22. Failure Analysis

- 🟡 Strong Inference — Technical failure mode: AI answers can miss/misinterpret rare/new evidence, increasing clinician distrust or legal risk. Evidence: AMBOSS itself acknowledges AI limitations. {SRC['ai_clinical']}.
- 🟡 Strong Inference — Business failure mode: UWorld remains dominant in high-stakes Qbank prep while OpenEvidence/UpToDate Expert AI erode clinician AI search adoption.
- 🟡 Strong Inference — Clinical failure mode: a widely publicized AI or content error could damage trust because clinical accuracy is existential.
- 🟡 Strong Inference — Regulatory failure mode: if AMBOSS moves into patient-specific recommendations or EHR integration without changing legal/compliance posture, it may trigger HIPAA/FDA/medical device obligations.
- 🟡 Strong Inference — Operational failure mode: content maintenance, AI evaluation, multiple professions, multiple countries, acquisitions, and institutional dashboards create complexity and cost.
- 🟡 Strong Inference — Distribution failure mode: institutions may prefer one incumbent enterprise tool; students may get UWorld/other resources free through schools.
- 🟡 Strong Inference — AI economics failure mode: included AI features may pressure gross margins if usage grows faster than monetization.
- 🔴 Speculation — IPO ambition may force margin discipline that slows editorial depth or free/included AI access.

---

## 23. Competitive Attack Plan

- 🟡 Recommendation — Do not attack AMBOSS as “another Qbank”; attack the category boundary by building the longitudinal patient-data intelligence layer AMBOSS intentionally does not cover.
- 🟡 Recommendation — Build PHI-compliant, consent-native FHIR/ABDM/wearable/lab/pharmacy ingestion first, then add evidence-grounded AI; AMBOSS built content first and excludes PHI.
- 🟡 Recommendation — Use clinicians for review, but let users own the data graph; AMBOSS owns professional content and user learning data.
- 🟡 Recommendation — Differentiate on personal longitudinal records, trends, medication safety, lab interpretation, screening gaps, and doctor-ready packets.
- 🟡 Recommendation — Match AMBOSS trust primitives: sources, evidence grades, provenance, human review, disclaimers, no ads/pharma influence.
- 🟡 Recommendation — Beat AMBOSS in India by integrating ABDM, WhatsApp, local labs, multilingual UX, family/caregiver roles, low-cost pricing, and doctor export workflows.
- 🟡 Recommendation — Use free utility wedges: lab report explainer, medication interaction checker, doctor visit summary, health vault import, screening checklist.
- 🟡 Recommendation — Build an Ovexis MCP/API around consented health data and evidence, but with strict scopes, audit logs, and de-identification.
- 🟡 Recommendation — Avoid competing on medical-school exam prep unless it is a GTM wedge; AMBOSS’s entrenched Qbank/content ecosystem is expensive to replicate.

---

## 24. Future Prediction

- 🔴 Speculation — Next 12 months: AMBOSS expands AI Mode availability, refines Assistants, uses NOHARM ranking in clinician/institutional sales, and continues navigation/purchase-flow/accessibility modernization.
- 🔴 Speculation — Next 12 months: AMBOSS pushes nursing/PA/NP and residency segments using Novaheal/NEJM Knowledge+ assets and the €240M financing.
- 🔴 Speculation — Next 3 years: AMBOSS becomes a broader medical intelligence platform with controlled agent access, more courses/CME, more AI assistant templates, and more institutional analytics.
- 🔴 Speculation — Next 3 years: AMBOSS may partner with ambient documentation/EHR vendors only if it can preserve no-PHI/search-aid posture or create separate compliant product lines.
- 🔴 Speculation — Next 5 years: AMBOSS’s likely strategic paths are IPO, acquisition by a major health information company, or durable private compounding through evergreen investors.
- 🔴 Speculation — Likely acquisitions include nursing/allied health content tools, specialty-board assets, guideline localization assets, workflow/AI evaluation tools, and institutional analytics products.

---

## 25. Ovexis Strategy Memo

### Top 50 ideas to copy
{numbered(copy_ideas, '🟡')}

### Top 50 ideas to improve
{numbered(improve_ideas, '🟡')}

### Top 50 ideas to ignore
{numbered(ignore_ideas, '🟡')}

### Top 50 ideas to reinvent
{numbered(reinvent_ideas, '🟡')}

### Top 50 market gaps
{numbered(market_gaps, '🟡')}

### Top 20 blue-ocean opportunities
{numbered(blue_ocean, '🔴')}

### Recommended Ovexis MVP

- 🟡 Recommendation — MVP should be a consented longitudinal health vault plus AI explainer, not a medical-school Qbank.
- 🟡 Recommendation — MVP data ingestion should support PDF lab uploads, manual medications, conditions, family history, wearable imports, and at least one standards path such as FHIR/ABDM where feasible.
- 🟡 Recommendation — MVP outputs should be doctor-ready summaries, lab trend explanations, medication safety checks, screening/prevention gaps, and questions-to-ask-your-doctor.
- 🟡 Recommendation — MVP guardrails should classify every output as education, suggestion for clinician discussion, or urgent escalation.
- 🟡 Recommendation — MVP trust layer should include evidence grade, source links, confidence/uncertainty, freshness, and human review option.

### Recommended GTM

- 🟡 Recommendation — Start with high-anxiety, high-data users: chronic-condition families, executives/biohackers, health-conscious Indian urban consumers, and caregivers managing parents.
- 🟡 Recommendation — Use free tools as wedges: lab report parser, medication list checker, vaccination/screening gap checker, and doctor-visit summary generator.
- 🟡 Recommendation — Build clinician partnership channels with independent doctors and clinics, not hospitals first.
- 🟡 Recommendation — Price India differently from US/EU and avoid luxury-only positioning.

### Recommended moat

- 🟡 Recommendation — Build a data moat around normalized, consented longitudinal health records, not around generic medical content.
- 🟡 Recommendation — Build a trust moat with human review, evidence ledger, no-ad/no-supplement-bias policy, and explicit clinical escalation.
- 🟡 Recommendation — Build an integration moat with labs, ABDM/FHIR, wearables, pharmacy, and clinician export workflows.

### Recommended AI architecture

```mermaid
flowchart TD
A[User consent + identity] --> B[Data connectors: PDF/labs/FHIR/ABDM/wearables/pharmacy]
B --> C[Normalization + deduplication + provenance ledger]
C --> D[Longitudinal health graph]
E[Curated medical evidence + guidelines + drug data] --> F[Evidence graph]
D --> G[Context builder with minimum necessary data]
F --> G
G --> H[Policy router: education vs clinician-review vs urgent escalation]
H --> I[LLM/RAG agents with tool-use]
I --> J[Evidence-graded output + uncertainty + sources]
J --> K[Doctor packet / user action plan / follow-up task]
K --> L[Outcome feedback + audit log]
```

- 🟡 Recommendation — Use separate environments for PHI and non-PHI content; AMBOSS’s no-PHI stance is not sufficient for Ovexis.
- 🟡 Recommendation — Implement data minimization, encryption, audit logs, user-controlled memory, deletion/export, and BAA-ready infrastructure from day one.
- 🟡 Recommendation — Use RAG over both personal data and curated evidence, with a policy layer that blocks diagnosis claims and escalates critical findings.

### Recommended pricing

- 🟡 Recommendation — Offer a freemium vault/import tier, a low-cost annual personal intelligence tier, a family plan, and a clinician-reviewed premium tier.
- 🟡 Recommendation — Avoid tying core safety features to upsells; monetize convenience, review, integrations, and advanced longitudinal insights.

### Recommended roadmap

| Status | Period | Roadmap |
|---|---|---|
| 🟡 Recommendation | 0-3 months | PDF lab parser, health profile, meds, conditions, evidence-based explainer, doctor packet |
| 🟡 Recommendation | 3-6 months | Longitudinal trends, wearable import, screening gaps, family/caregiver roles, consent center |
| 🟡 Recommendation | 6-12 months | FHIR/ABDM, clinician review network, medication safety graph, multilingual India UX |
| 🟡 Recommendation | 12-24 months | Health MCP/API, employer/family plans, provider integrations, outcomes tracking |
| 🟡 Recommendation | 24+ months | Digital twin simulations, personalized guideline engine, population insights, regulated clinical pathways if chosen |

---

## 26. Master Feature Inventory

- 🟢 Confirmed — The complete feature inventory spreadsheet is provided as `AMBOSS_Feature_Inventory.xlsx` with requested columns.
- 🟢 Confirmed — Spreadsheet rows include feature, purpose, evidence, user/business value, engineering/clinical/infrastructure/regulatory complexity, estimated team/months, priority, category, copy/improve/ignore/reinvent guidance, moat, and confidence.

---

## 27. Evidence Register

- 🟢 Confirmed — The evidence register is provided as `AMBOSS_Evidence_Register.xlsx`.
- 🟢 Confirmed — Each row contains claim, evidence quote/summary, source URL, confidence, observed vs inferred, and screenshot status.
- 🟢 Confirmed — Screenshot status is “not captured; source URL available” because this environment lacks browser screenshot capture.

---

## SWOT

| Status | Strengths | Weaknesses |
|---|---|---|
| 🟢/🟡 | 🟢 Clinician-authored content; 🟢 integrated Qbank/library; 🟢 strong student distribution; 🟢 AI grounded in curated sources; 🟢 ACCME; 🟢 institutional tools; 🟢 NEJM/AHFS/QxMD assets; 🟡 emerging MCP platform | 🟢 Not patient-facing; 🟢 no PHI/longitudinal health data posture; 🟡 Qbank perceived as hard/picky by some; 🟡 clinician AI faces OpenEvidence/UpToDate; 🟡 broad scope creates complexity; 🟢 no public FHIR/EHR API found |

| Status | Opportunities | Threats |
|---|---|---|
| 🟡/🔴 | 🟡 Nurses/allied health expansion; 🟡 AI assistants; 🟡 institutional residency growth; 🟡 MCP/developer ecosystem; 🟡 CME/courses; 🔴 potential EHR/ambient partnerships | 🟡 OpenEvidence free clinician AI; 🟡 UpToDate enterprise AI; 🟡 UWorld exam prep incumbency; 🟡 AI safety incident; 🟡 regulatory changes; 🟡 content licensing cost; 🟡 generic AI improving |

---

## Porter’s Five Forces

- 🟡 Strong Inference — Rivalry is high because AMBOSS competes with UWorld, UpToDate, OpenEvidence, Glass, Lecturio/Osmosis, and institution-provided tools.
- 🟡 Strong Inference — Buyer power is medium-high for institutions because multiple education/CDS vendors exist and pricing is negotiated.
- 🟡 Strong Inference — Supplier power is medium because AMBOSS depends on clinical experts, licensed drug/content assets, AI providers, cloud providers, and guideline availability.
- 🟡 Strong Inference — Threat of substitutes is high because users can combine UWorld, UpToDate, PubMed, Google, ChatGPT, Anki, and institutional resources.
- 🟡 Strong Inference — Threat of new entrants is medium because AI lowers interface cost but trusted medical content, assessment data, institutional distribution, and clinical QA are difficult.

---

## Value Chain

```mermaid
flowchart LR
A[Medical evidence + exam blueprints + guidelines] --> B[Clinician editorial distillation]
B --> C[Content graph + Qbank + media + drug/calculator integrations]
C --> D[Search + recommendations + AI retrieval]
D --> E[Learning / Clinical Care / Teaching workflows]
E --> F[User outcomes: faster answers, exam readiness, CME, faculty insight]
F --> G[Usage data + feedback + content updates]
G --> B
```

- 🟢 Confirmed — AMBOSS’s value chain explicitly includes physician-authored content, sources/guidelines, platform features, AI, and user feedback. Evidence: {SRC['content_policy']}; {SRC['ai_principles']}.

---

## Risk Register

| Status | Risk | Likelihood | Impact | Mitigation |
|---|---|---:|---:|---|
| 🟡 Strong Inference | AI hallucination/misinterpretation | Medium | Very High | RAG, citations, refusal, eval, human review |
| 🟡 Strong Inference | Stale clinical content | Medium | High | Editorial update process and changelog |
| 🟡 Strong Inference | PHI accidentally entered into AI | Medium | High | Terms, UI warnings, redaction, technical PHI detection |
| 🟡 Strong Inference | Student churn to UWorld/free school resources | Medium | Medium | Integrated value, pricing, institution licenses |
| 🟡 Strong Inference | OpenEvidence captures clinician AI search | High | High | Unique education+CDS continuity and source trust |
| 🟡 Strong Inference | Accessibility/legal issues in legacy flows | Medium | Medium | WCAG roadmap and redesign |
| 🟡 Strong Inference | AI cost pressure | Medium | Medium | Premium packaging, retrieval efficiency, caching |
| 🟡 Strong Inference | Content scraping/AI training misuse | High | Medium | Legal restrictions, MCP licensing, bot controls |
| 🟡 Strong Inference | Institution privacy concerns | Medium | High | RBAC, transparency, joint-controller agreements |
| 🔴 Speculation | IPO pressure | Medium | Medium | Long-term evergreen investors and margin discipline |

---

## References

- 🟢 Confirmed — Official AMBOSS About: {SRC['about']}
- 🟢 Confirmed — Official Product Features: {SRC['features']}
- 🟢 Confirmed — Official Students: {SRC['students']}
- 🟢 Confirmed — Official Clinicians: {SRC['clinicians']}
- 🟢 Confirmed — Official AI Mode Clinical Care: {SRC['ai_clinical']}
- 🟢 Confirmed — Official AI Mode Learning: {SRC['ai_learning']}
- 🟢 Confirmed — Official AI Innovation: {SRC['ai_innovation']}
- 🟢 Confirmed — Official AI Principles: {SRC['ai_principles']}
- 🟢 Confirmed — Official New Navigation: {SRC['new_nav']}
- 🟢 Confirmed — Official Pricing: {SRC['pricing']}
- 🟢 Confirmed — Official Financing: {SRC['funding']}
- 🟢 Confirmed — Official NEJM Knowledge+ acquisition: {SRC['nejm_acq']}
- 🟢 Confirmed — Official MCP: {SRC['mcp']}
- 🟢 Confirmed — Official Assistants beta: {SRC['assistants']}
- 🟢 Confirmed — Support platform overview: {SRC['support_platform']}
- 🟢 Confirmed — Support Qbank: {SRC['support_qbank']}
- 🟢 Confirmed — Support Study/Exam Mode: {SRC['support_modes']}
- 🟢 Confirmed — Support Feature Overview: {SRC['support_feature_overview']}
- 🟢 Confirmed — Privacy: {SRC['privacy']}
- 🟢 Confirmed — Terms: {SRC['terms']}
- 🟢 Confirmed — Clinics terms: {SRC['terms_clinics']}
- 🟢 Confirmed — Joint Data Processing TOMs: {SRC['joint']}
- 🟢 Confirmed — Careers: {SRC['careers']}
- 🟢 Confirmed — Product & Engineering careers: {SRC['careers_eng']}
- 🟢 Confirmed — NOHARM arXiv: {SRC['noharm']}
"""

# Patch any missing SRC used in f-strings not defined
report = report.replace("{SRC['scholarships'] if 'scholarships' in SRC else 'https://www.amboss.com/us/scholarships-program'}", "https://www.amboss.com/us/scholarships-program")
report = report.replace("{SRC['group_discounts'] if 'group_discounts' in SRC else 'https://www.amboss.com/us/group-discounts'}", "https://www.amboss.com/us/group-discounts")
report = report.replace("{SRC['self_assessment'] if 'self_assessment' in SRC else 'https://www.amboss.com/us/usmle/self-assessment'}", "https://www.amboss.com/us/usmle/self-assessment")
report = report.replace("{SRC['pa'] if 'pa' in SRC else 'https://www.amboss.com/us/pa'}", "https://www.amboss.com/us/pa")
report = report.replace("{SRC['np'] if 'np' in SRC else 'https://www.amboss.com/us/np'}", "https://www.amboss.com/us/np")
report = report.replace("{SRC['score_predictor'] if 'score_predictor' in SRC else SRC['support_score']}", SRC['support_score'])

with open('/home/user/AMBOSS_Competitive_Intelligence_Report.md','w',encoding='utf-8') as f:
    f.write(report)

# Diagrams file
with open('/home/user/AMBOSS_Diagrams.md','w',encoding='utf-8') as f:
    f.write("""# AMBOSS Diagrams

## Product Architecture Diagram
```mermaid
flowchart TB
Visitor[Visitor] --> Marketing[Marketing Website]
Marketing --> Registration[Registration / Free Trial]
Registration --> Identity[Identity, role, study objective, membership]
Identity --> Learning[Learning]
Identity --> Clinical[Clinical Care]
Identity --> Teaching[Teaching]
Learning --> Library[Knowledge Library]
Learning --> Qbank[Qbank / Study Plans]
Learning --> Analysis[Analysis / Score Predictor]
Learning --> AILearning[AI Mode Learning]
Clinical --> Search[Clinical Search]
Clinical --> Drug[AHFS Drug DB]
Clinical --> Tools[DDx / Checklists / Flowcharts / QxMD Calculators]
Clinical --> AIClinical[AI Mode Clinical Care]
Teaching --> Assign[Assignments]
Teaching --> Dash[Analytics / Groups / Roles]
Library --> ContentGraph[Medical Content Graph]
Qbank --> ContentGraph
Drug --> ContentGraph
Tools --> ContentGraph
ContentGraph --> Retrieval[Semantic Retrieval]
Retrieval --> AI[AI/GPT/MCP/Assistants]
```

## AI Architecture Diagram
```mermaid
flowchart LR
Input[User query/upload/article context] --> Policy[Policy checks: no PHI, no emergency, role]
Policy --> Retrieval[Retrieve trusted AMBOSS/drug/guideline/Qbank content]
Retrieval --> Context[Context builder + ranker]
Context --> LLM[LLM provider via secure API]
LLM --> Output[Structured answer + citations + limits]
Output --> Verify[User verifies source]
Verify --> Feedback[Usage/quality feedback]
```

## Healthcare Data Flow Diagram
```mermaid
flowchart TD
Guidelines[Guidelines / literature / exam blueprints] --> Editorial[AMBOSS editorial process]
Editorial --> Knowledge[Articles, media, sources]
Editorial --> Questions[Qbank questions]
AHFS[AHFS Drug DB] --> Drug[Drug monographs]
QxMD[QxMD calculators] --> Calc[Calculators]
User[User profile + usage + attempts + notes] --> Analytics[Analytics and personalization]
Knowledge --> SearchAI[Search and AI]
Questions --> Analytics
Drug --> SearchAI
Calc --> SearchAI
Analytics --> Recs[Study recommendations / EPC / score]
Recs --> User
SearchAI --> User
Institution[Institution admins/educators] --> Dashboards[Dashboards if license enabled]
Analytics --> Dashboards
```

## User Journey Diagram
```mermaid
flowchart LR
Anon[Anonymous visitor] --> Trial[5-day free trial]
Trial --> Profile[Role/profile]
Profile --> Mode{{Choose intent}}
Mode --> Learn[Learning: Qbank, Library, AI Learning]
Mode --> Care[Clinical Care: Search, AI, Drugs, Tools]
Mode --> Teach[Teaching: Assignments, Analytics]
Learn --> Retain[Progress, recommendations, notes]
Care --> Verify[Source review + CME]
Teach --> Remediate[Assignments + remediation]
Retain --> Subscribe[Subscription / institutional license]
Verify --> Subscribe
Remediate --> Renew[Institution renewal]
Subscribe --> Renew
```

## Feature Dependency Graph
```mermaid
flowchart TD
Consent[Consent/Terms/Privacy] --> Identity[Identity + Roles]
Identity --> ContentGraph[Content Graph]
ContentGraph --> Search[Search]
ContentGraph --> Qbank[Qbank]
ContentGraph --> ClinicalTools[Clinical Tools]
Qbank --> Attempts[Attempts]
Attempts --> Analytics[Analytics]
Analytics --> Recommendations[Recommendations]
Search --> AI[AI Modes]
ClinicalTools --> AI
Identity --> Institution[Institution/RBAC]
Institution --> Dashboards[Educator Dashboards]
AI --> Citations[Source Citations]
Citations --> HumanJudgment[Human Judgment]
```
""")

# Create feature inventory XLSX
headers = ['Feature','Purpose','Evidence','User Value','Business Value','Engineering Complexity','Clinical Complexity','Infrastructure Complexity','Regulatory Complexity','Estimated Team','Estimated Months','Priority','Category','Copy','Improve','Ignore','Reinvent','Moat','Confidence']
wb = Workbook()
ws = wb.active
ws.title = 'Feature Inventory'
ws.append(headers)
for row in features:
    ws.append(row)

# Style
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
    cell.alignment = Alignment(wrap_text=True, vertical='top')
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
widths = [28,45,38,36,36,16,16,18,18,24,14,12,18,10,40,10,40,16,16]
for i,w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
wb.save('/home/user/AMBOSS_Feature_Inventory.xlsx')

# Evidence register XLSX
wb2 = Workbook()
ws2 = wb2.active
ws2.title = 'Evidence Register'
headers2 = ['Claim','Evidence / Quote Summary','Source','Screenshot','Confidence','Observed vs Inferred','Status Label']
ws2.append(headers2)
for claim, ev, source, obs, conf, status in claims:
    ws2.append([claim, ev, source, 'Not captured in this environment; source URL available', conf, obs, status])
for cell in ws2[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='7030A0')
    cell.alignment = Alignment(wrap_text=True, vertical='top')
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
for i,w in enumerate([55,70,55,35,14,20,16], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = ws2.dimensions
wb2.save('/home/user/AMBOSS_Evidence_Register.xlsx')

# Also save CSV versions for quick diff/readability
import csv
with open('/home/user/AMBOSS_Feature_Inventory.csv','w',newline='',encoding='utf-8') as f:
    writer=csv.writer(f); writer.writerow(headers); writer.writerows(features)
with open('/home/user/AMBOSS_Evidence_Register.csv','w',newline='',encoding='utf-8') as f:
    writer=csv.writer(f); writer.writerow(headers2)
    for claim, ev, source, obs, conf, status in claims:
        writer.writerow([claim, ev, source, 'Not captured in this environment; source URL available', conf, obs, status])

print('Created report and spreadsheets.')
