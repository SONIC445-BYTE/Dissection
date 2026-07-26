from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from datetime import date

OUT = 'OpenEvidence_Master_Feature_Inventory_2026-07-25.xlsx'

# Exact requested columns. All estimates are explicitly marked inference in Evidence/Confidence.
features = [
('Professional credential gate','Restrict sensitive clinical tool to verified professionals','🟢 E13,E14','High','High','M','M','M','H','3–5','2–4','P0','Trust & identity','Copy','Improve','','','Medium','High'),
('Free verified-clinician access','Remove price/procurement friction and build habitual adoption','🟢 E01,E05,E13','High','High','H','L','H','M','5–8','3–6','P0','Distribution','Copy','Improve','','','Strong','High'),
('Clinician profile/specialty data','Personalise services and verify audience','🟢 E13','Medium','High','M','L','M','H','3–5','2–4','P1','Identity & personalisation','Copy','Improve','','','Medium','High'),
('Natural-language medical search','Translate clinician question into evidence retrieval','🟢 E05','High','High','H','H','H','H','8–12','6–12','P0','Evidence retrieval','Copy','Improve','','','Strong','High'),
('Cited answer generation','Give concise answer with references','🟢 E05,E12','High','High','H','H','H','H','8–12','6–12','P0','Clinical AI','Copy','Improve','','','Strong','High'),
('Source drill-down','Let clinician inspect underlying evidence','🟢 E05,E12','High','High','M','H','M','M','4–6','3–5','P0','Trust & citations','Copy','Improve','','','Medium','High'),
('Follow-up questions','Refine query against prior conversational context','🟢 E05','High','Medium','M','M','M','M','4–6','3–5','P1','Clinical AI','Copy','Improve','','','Medium','High'),
('Licensed journal corpus','Use licensed peer-reviewed content for retrieval','🟢 E12,E18,E19','High','High','H','H','H','H','8–15 + legal','9–18','P0','Content moat','Copy','Improve','','','Strong','High'),
('Figures/tables/multimedia retrieval','Use full-text visual and structured clinical findings','🟢 E01,E18,E19','Medium','High','H','H','H','H','6–10','6–12','P1','Content experience','Copy','Improve','','','Medium','High'),
('Society guideline integrations','Surface authoritative specialty guidance/algorithms','🟢 E01,E34','High','High','H','H','M','H','5–8 + legal','4–9','P0','Content moat','Copy','Improve','','','Strong','High'),
('Specialty routing / conductor','Route query to medically specialised model(s)','🟢 E26','High','High','H','H','H','H','8–15','9–18','P0','Model orchestration','Copy','Improve','','','Medium','High'),
('DeepConsult agentic research','Run parallel long-form cross-study synthesis','🟢 E05','High','High','H','H','H','H','8–15','9–18','P1','Agentic AI','Copy','Improve','','','Medium','High'),
('EvidenceGrade','Grade retrieved evidence with GRADE-inspired method','🟢 E09','High','High','H','H','M','H','6–10','6–12','P0','Clinical safety','Copy','Improve','','','Medium','High'),
('Claim-level evidence entailment','Verify each answer claim against source span','🟡 Recommended; no public OE disclosure','High','High','H','H','H','H','8–14','9–18','P0','Ovexis differentiator','','Improve','','Reinvent','Strong','Medium'),
('Conflict/contradiction detector','Show contradictory guidelines/studies and why','🟡 Recommended; no public OE disclosure','High','High','H','H','H','H','7–12','9–15','P0','Ovexis differentiator','','Improve','','Reinvent','Strong','Medium'),
('Patient–evidence applicability scoring','Score fit of study population to patient/state','🟡 Recommended; no public OE disclosure','High','High','H','H','H','H','8–14','12–18','P0','Ovexis differentiator','','Improve','','Reinvent','Strong','Medium'),
('High-risk abstention/escalation','Stop unsafe output and route to review/next action','🟡 Recommended; terms only establish human responsibility E14','High','High','H','H','M','H','6–10','6–12','P0','Clinical safety','','Improve','','Reinvent','Strong','Medium'),
('Patient case context','Use case details in clinical questions','🟢 E05','High','High','M','H','M','H','5–8','4–8','P0','Clinical AI','Copy','Improve','','','Medium','High'),
('Document upload','Collect patient documents in Visit workflow','🟢 E06','High','High','H','H','H','H','6–10','6–12','P0','Patient context','Copy','Improve','','','Medium','High'),
('Document repository / cross-document query','Organise and query histories and past treatments','🟢 E06','High','High','H','H','H','H','7–12','9–15','P0','Patient context','Copy','Improve','','Reinvent','Medium','High'),
('Longitudinal provenance timeline','Reconcile events with source/time/conflict state','🟡 Recommended; no public OE full longitudinal record','High','High','H','H','H','H','10–18','12–24','P0','Ovexis differentiator','','Improve','','Reinvent','Strong','Medium'),
('Visits audio transcription','Transcribe encounter into note','🟢 E06','High','High','H','H','H','H','6–10','6–12','P1','Documentation','Copy','Improve','','','Medium','High'),
('Custom note templates','Adapt generated notes to clinician style','🟢 E06','High','Medium','M','M','L','M','3–5','2–4','P1','Documentation','Copy','Improve','','','Weak','High'),
('Evidence-enriched assessment/plan','Surface evidence inside documentation flow','🟢 E06','High','High','H','H','M','H','6–10','6–12','P0','Documentation & CDS','Copy','Improve','','','Medium','High'),
('AI note editing/literature search','Edit note/refine reasoning/search inside Visit','🟢 E06','High','Medium','H','H','M','H','6–10','6–12','P1','Documentation & CDS','Copy','Improve','','','Medium','High'),
('Patient-consented action ledger','Assign owner/due date/closure for each plan','🟡 Recommended; no public OE disclosure','High','High','H','H','H','H','7–12','9–15','P0','Ovexis differentiator','','Improve','','Reinvent','Strong','Medium'),
('Conversation private by default','Limit unintended disclosure; support controlled sharing','🟢 E30','High','High','M','M','M','H','3–5','2–4','P0','Privacy & collaboration','Copy','Improve','','','Medium','High'),
('Invite sharing / non-PHI public links','Support collaboration with explicit PHI boundary','🟢 E30','Medium','Medium','M','M','M','H','3–5','2–4','P1','Collaboration','Copy','Improve','','','Weak','High'),
('Prior authorisation letters','Draft evidence-backed coverage request','🟢 E10','High','High','M','H','M','H','4–7','4–8','P1','Administrative workflow','Copy','Improve','','','Medium','High'),
('Patient handouts/home instructions','Draft patient-facing education material','🟢 E10','High','Medium','M','H','L','M','3–5','3–5','P1','Patient communication','Copy','Improve','','','Weak','High'),
('Clinical calculators','Calculate widely used risk/clinical scores','🟢 E10','High','Medium','M','H','L','H','3–6','3–6','P1','Clinical workflow','Copy','Improve','','','Medium','High'),
('Drug monograph modules','Provide drug reference content','🟢 E10','High','Medium','H','H','M','H','5–8 + data','6–12','P1','Clinical reference','Copy','Improve','','','Medium','High'),
('Guideline-based modules','Offer structured guideline information','🟢 E10','High','High','H','H','M','H','5–8 + legal','6–12','P0','Clinical reference','Copy','Improve','','','Strong','High'),
('Clinical Trial Matching','Match patient context to active/recruiting trials','🟢 E08','High','High','H','H','H','H','7–12','9–15','P1','Research & referral','Copy','Improve','','Reinvent','Medium','High'),
('Trial location/site filtering','Use eligibility, status, site/location/contact','🟢 E08','High','High','M','H','M','H','4–7','4–8','P1','Research & referral','Copy','Improve','','','Weak','High'),
('Trial navigator / referral closure','Ensure match becomes screened referral and access support','🟡 Recommended; no public OE closure workflow','High','High','H','H','H','H','6–10','6–12','P1','Ovexis differentiator','','Improve','','Reinvent','Medium','Medium'),
('Dialer with hospital caller ID','Protect personal number/increase pickup rate','🟢 E07','High','High','H','M','H','H','5–8','5–9','P1','Communications','Copy','Improve','','','Medium','High'),
('Secure two-way messaging','Communicate with optional patient replies','🟢 E07','High','High','H','M','H','H','5–8','5–9','P1','Communications','Copy','Improve','','','Medium','High'),
('Fax / scan workflow','Send/receive faxes, upload/scan docs','🟢 E07','Medium','Medium','H','L','H','H','5–8','5–9','P2','Communications','Copy','Improve','','','Weak','High'),
('Straight-to-voicemail outreach','Send nonurgent patient messages without ringing','🟢 E07','Medium','Medium','M','L','H','H','3–5','3–6','P2','Communications','Copy','Improve','','','Weak','High'),
('Create Visit from call','Turn call into structured evidence-integrated note','🟢 E07','High','High','H','H','H','H','7–11','9–15','P1','Communications & documentation','Copy','Improve','','','Medium','High'),
('EHR workflow launch','Surface evidence inside Epic workflow','🟢 E33','High','High','H','H','H','H','8–15','9–18','P0','Enterprise integration','Copy','Improve','','Reinvent','Strong','High'),
('Patient-aware Epic context','Use procedures/comorbidities/meds/allergies/longitudinal data in session','🟢 E35','High','High','H','H','H','H','10–18','12–24','P0','Enterprise integration','Copy','Improve','','Reinvent','Strong','High'),
('FHIR/SMART interoperability layer','Standards-based read-only patient context with audit','🟡 Recommended; OE specific standard not public','High','High','H','H','H','H','8–14','9–18','P0','Ovexis differentiator','','Improve','','Reinvent','Strong','Medium'),
('Dragon Copilot channel','Embed evidence in ambient clinical application','🟢 Planned integration E11','High','High','H','H','H','H','8–15','9–18','P2','Platform partnership','Copy','Improve','','','Medium','High'),
('iOS mobile app','Point-of-care native mobile access','🟢 E29,E38','High','High','H','M','H','H','5–9','6–12','P1','Mobile','Copy','Improve','','','Medium','High'),
('Android mobile app','Point-of-care native mobile access','🟢 E29,E39','High','High','H','M','H','H','5–9','6–12','P1','Mobile','Copy','Improve','','','Medium','High'),
('Discover/news/learning feed','Curate medical updates/engagement in app','🟢 E38 app-store update description','Medium','Medium','M','M','M','M','3–5','3–5','P3','Engagement','','','Ignore','Reinvent','Weak','Medium'),
('Medical content partnerships','Negotiate rights and distribution with publishers/societies','🟢 E12,E18,E19,E34','High','High','H','H','H','H','6–12 + legal','9–24','P0','Content moat','Copy','Improve','','','Strong','High'),
('Medical advisor network','Bring expert review/credibility to product','🟢 E01','High','High','M','H','L','M','3–6','3–6','P0','Clinical governance','Copy','Improve','','Reinvent','Medium','High'),
('Formal clinical safety case','Document intended use, hazards, controls and evaluation','🟡 Recommended; no public full safety case','High','High','H','H','M','H','5–9','6–12','P0','Ovexis differentiator','','Improve','','Reinvent','Strong','Medium'),
('HIPAA BAA','Contractual PHI rules for covered entities','🟢 E15,E40','High','High','M','M','M','H','3–6','3–6','P0','Security & compliance','Copy','Improve','','','Medium','High'),
('SOC 2 Type II security posture','Demonstrate operating security controls','🟢 E15','Medium','High','M','L','M','H','3–6','6–12','P0','Security & compliance','Copy','Improve','','','Medium','High'),
('Encryption transit/rest','Protect data in transport/storage','🟢 E15','High','High','M','L','M','H','3–5','2–4','P0','Security & compliance','Copy','Improve','','','Table stakes','High'),
('Responsible vulnerability disclosure','Receive/report security issues','🟢 E15','Medium','Medium','L','L','M','M','1–2','1–2','P2','Security & compliance','Copy','Improve','','','Weak','High'),
('Audience extension advertising','Create US off-platform advertiser audiences from on-platform activity','🟢 E13','Low','High','H','L','H','H','5–9','6–12','P3','Monetisation','','','Ignore','Reinvent','Medium','High'),
('Advertising data firewall','Prevent any commercial influence on evidence/ranking','🟡 Recommended; OE policy says questions not shared for ad purposes E13','High','High','H','H','M','H','4–7','4–8','P0','Ovexis differentiator','','Improve','','Reinvent','Strong','Medium'),
('Patient consent graph','Purpose/role/time-bound consent for data, sharing and recording','🟡 Recommended; OE terms assign consent duty to user E14','High','High','H','H','H','H','7–12','9–15','P0','Ovexis differentiator','','Improve','','Reinvent','Strong','Medium'),
('Outcome/error feedback loop','Capture error reports and downstream outcomes for safety improvement','🟡 Feedback contact public; no public end-to-end loop E06','High','High','H','H','H','H','6–10','6–12','P0','Clinical safety','','Improve','','Reinvent','Strong','Medium'),
]

sources = [
('E01','OpenEvidence About','https://www.openevidence.com/about','A','Team, mission, content partners, advisors','Public page / logos','High'),
('E02','OpenEvidence Home','https://www.openevidence.com/','A','Public marketing, free access, mobile/investor links','Public page','High'),
('E03','Announcements index','https://www.openevidence.com/announcements','A','Chronology of releases/partnerships','Announcement thumbnails','High'),
('E04','Security and Compliance','https://www.openevidence.com/security','A','HIPAA/SOC2 claims, GCP/Vercel, encryption, testing','Text-only public page','High self-report'),
('E05','Series B and DeepConsult','https://www.openevidence.com/announcements/openevidence-the-fastest-growing-application-for-physicians-in-history-announces-dollar210-million-round-at-dollar35-billion-valuation','A','Funding, point-of-care, DeepConsult, adoption claims','Public release','High self-report'),
('E06','Visits','https://www.openevidence.com/announcements/visits-real-time-medical-intelligence','A','Visit transcription, templates, docs, query','Public product images','High'),
('E07','Dialer release','https://www.openevidence.com/announcements/messaging-faxing-and-voicemail-are-now-live-in-the-openevidence-dialer','A','Calls/text/fax/voicemail/Create Visit','Public product image','High'),
('E08','Clinical Trial Matching','https://www.openevidence.com/announcements/new-feature-clinical-trials-matching-in-openevidence','A','Matching/filter/location details','Public product images','High'),
('E09','EvidenceGrade technical post','https://www.openevidence.com/blog/introducing-evidencegrade-grading-the-strength-of-medical-evidence-in-real-time','A','Grading/retrieval method','Public diagrams','High self-report'),
('E10','OpenEvidence 2.0','https://www.openevidence.com/announcements/openevidence-20','A','Prior auth, handouts, calculators, modules','Public product image','High'),
('E11','Microsoft/Dragon Copilot','https://www.openevidence.com/announcements/openevidence-collaborates-with-microsoft-to-expand-ai-leadership-in-healthcare-bringing-clinical-evidence-and-guidelines-to-enterprise-clinician-workflows','A','Planned integration','Public release image','High announcement'),
('E12','Wiley and OpenEvidence','https://www.openevidence.com/announcements/wiley-and-openevidence-partner-to-deliver-trusted-research-to-physicians-at-the-point-of-care','A','Wiley/Cochrane/400+ journals','Public release image','High'),
('E13','Privacy policy','https://www.openevidence.com/policies/privacy','A','Registration, ads/profiling, questions/PHI, transfers','Public legal text','High'),
('E14','Terms of Use','https://www.openevidence.com/policies/terms','A','Professional use, disclaimer, user data/content','Public legal text','High'),
('E15','Trust Center','https://trust.openevidence.com/','A','SOC2/HIPAA claims, controls','Public trust portal','High self-report'),
('E16','PRNewswire Amaro acquisition','https://www.prnewswire.com/news-releases/openevidence-acquires-google-ventures-backed-ai-startup-amaro-302547047.html','A/B','Advertising acquisition/rationale','Public press release','High'),
('E17','Mergr Amaro transaction','https://mergr.com/transaction/openevidence-acquires-amaro','C','Reported transaction date','Text listing','Medium'),
('E18','NEJM agreement','https://www.openevidence.com/announcements/openevidence-and-nejm','A','NEJM content from 1990','Public release','High'),
('E19','JAMA agreement','https://www.openevidence.com/announcements/openevidence-and-the-jama-network-sign-strategic-content-agreement','A','JAMA content agreement','Public release','High'),
('E20','medRxiv complex-question evaluation','https://www.medrxiv.org/content/10.64898/2025.11.29.25341091v1.full.pdf','C','Preprint performance/repeatability caution','PDF','Medium; preprint'),
('E21','Nature Medicine evaluation','https://www.nature.com/articles/s41591-026-04431-5','B','Independent benchmark comparison','Article figures','High'),
('E22','r/medicine discussion','https://www.reddit.com/r/medicine/comments/1mslx0z/openevidence_not_quite_as_accurate_as_id_have/','C','Anecdotal clinician praise/caution','Community comments','Low incidence'),
('E23','CNBC Series D','https://www.cnbc.com/2026/01/21/openevidence-chatgpt-for-doctors-doubles-valuation-to-12-billion.html','B','Funding, revenue/adoption statements','News page','High'),
('E24','Reuters Series D','https://www.reuters.com/business/healthcare-pharmaceuticals/medical-ai-startup-openevidence-doubles-valuation-12-billion-latest-round-2026-01-21/','B','Funding/valuation confirmation','News page','High'),
('E25','Fierce Series C','https://www.fiercehealthcare.com/ai-and-machine-learning/open-evidence-raises-200m-6b-valuation-rapid-adoption-doctors-continues','B','Series C/adoption reporting','News page','Medium–High'),
('E26','BusinessWire Series D','https://www.businesswire.com/news/home/20260121029132/en/OpenEvidence-Raises-$250-Million-to-Build-Medical-Superintelligence-for-Doctors','A','Specialised models/conductor claim','Press release','High self-report'),
('E27','Reported $20B fundraising','https://www.digitalhealthnews.com/openevidence-reportedly-seeks-200m-funding-at-20b-valuation-amid-rapid-ai-healthcare-growth','C','Unconfirmed rumour','News page','Low'),
('E28','CHIL Best Paper','https://www.openevidence.com/announcements/openevidence-wins-best-paper-award-at-chil-2023','A','Research paper/award','Public release','High'),
('E29','iOS/Android launch','https://www.openevidence.com/announcements/openevidence-is-now-available-for-ios-and-android','A','Mobile availability','Public release','High'),
('E30','HIPAA announcement','https://www.openevidence.com/announcements/openevidence-is-now-hipaa-compliant','A','BAA/PHI/conversation share','Public release','High self-report'),
('E31','Veeva/Open Vista','https://www.openevidence.com/announcements/openevidence-and-veeva-announce-open-vista-partnership','A','Life-sciences partnership','Public release','High'),
('E32','Initial Dialer','https://www.openevidence.com/announcements/openevidence-hipaa-secure-dialer-now-available','A','Initial dialer','Public release','High'),
('E33','Sutter/Epic','https://www.openevidence.com/announcements/sutter-health-collaborates-with-openevidence-to-bring-evidence-based-ai-powered-insights-into-physician-workflows','A','Epic workflow launch','Public release image','High'),
('E34','NCCN collaboration','https://www.openevidence.com/announcements/openevidence-collaborates-with-nccn-to-integrate-canonical-oncology-treatment-algorithms-at-the-point-of-care','A','NCCN algorithms','Public release','High'),
('E35','Cedars-Sinai','https://www.openevidence.com/announcements/openevidence-partners-with-cedars-sinai-to-create-patient-aware-clinical-intelligence-with-agentic-clinical-ai','A','Epic context/session non-storage claim','Public release image','High announcement'),
('E36','NYP/Columbia/Weill Cornell','https://www.openevidence.com/announcements/openevidence-and-newyork-presbyterian-columbia-university-and-weill-cornell-medicine-expand-clinical-ai-tools-across-new-york-city-and-westchester','A','NYC collaboration','Announcement','High'),
('E37','Contrary research','https://research.contrary.com/company/openevidence','C','Founder/market analysis','Analyst page','Medium'),
('E38','Apple App Store','https://apps.apple.com/us/app/openevidence/id6612007783','B','Mobile description/features/rating snapshot','Store images','Medium–High'),
('E39','Google Play','https://play.google.com/store/apps/details?id=com.openevidence','B','Android/data safety/rating snapshot','Store images','Medium–High'),
('E40','Business Associate Agreement','https://www.openevidence.com/policies/baa','A','PHI obligations/subcontractors','Public legal text','High'),
('E41','100% USMLE release','https://www.openevidence.com/announcements/openevidence-creates-the-first-ai-in-history-to-score-a-perfect-100percent-on-the-united-states-medical-licensing-examination-usmle','A','Benchmark claim','Public release','High self-report'),
('E42','r/hospitalist discussion','https://www.reddit.com/r/hospitalist/comments/1je4ria/open_evidence/','C','Citation mismatch anecdote/company reply','Community comments','Low incidence'),
('E43','r/Residency discussion','https://www.reddit.com/r/Residency/comments/1nofa70/open_evidence_examples_of_ai_hallucination/','C','Rare/specialty caution','Community comments','Low incidence'),
('E44','Lancet Europe geoblocking note','https://www.thelancet.com/journals/lanepe/article/PIIS2666-7762(26)00130-4/fulltext','B','EU/UK availability/regulatory uncertainty','Article page','Medium–High'),
('E45','NBC coverage','https://www.nbcnews.com/tech/tech-news/openevidence-ai-doctor-medical-physician-login-app-what-npi-uptodate-rcna341064','B','Free/ad-supported model','News page','High'),
('E46','Research Scientist job','https://jobs.ashbyhq.com/openevidence/80ca886f-2c07-43b2-8978-07c37542a207','A','Team/culture/evaluation hiring signal','Job post','High employer claim'),
('E47','Data Infrastructure job','https://jobs.thrivecap.com/companies/openevidence-2/jobs/81122458-software-engineer-data-infrastructure','A/B','Infrastructure/culture hiring signal','Job post','High employer claim'),
('E48','Superpower','https://superpower.com/','A','Comparator reference','Vendor site','Low–Medium'),
('E49','Function Health','https://www.functionhealth.com/','A','Comparator reference','Vendor site','Low–Medium'),
('E50','Levels','https://www.levels.com/','A','Comparator reference','Vendor site','Low–Medium'),
('E51','Glass Health','https://glass.health/','A','Comparator reference','Vendor site','Low–Medium'),
('E52','Atropos Health','https://www.atroposhealth.com/','A','Comparator reference','Vendor site','Low–Medium'),
('E53','AMBOSS','https://www.amboss.com/','A','Comparator reference','Vendor site','Low–Medium'),
('E54','UpToDate','https://www.wolterskluwer.com/en/solutions/uptodate','A','Comparator reference','Vendor site','Low–Medium'),
('E55','Apollo 24|7','https://www.apollo247.com/','A','Comparator reference','Vendor site','Low–Medium'),
('E56','Practo','https://www.practo.com/','A','Comparator reference','Vendor site','Low–Medium'),
('E57','Tata 1mg','https://www.1mg.com/','A','Comparator reference','Vendor site','Low–Medium'),
('E58','Apple Health','https://www.apple.com/health/','A','Comparator reference','Vendor site','Low–Medium'),
('E59','Android Health Connect','https://developer.android.com/health-and-fitness/guides/health-connect','A','Comparator/interoperability reference','Developer site','Low–Medium'),
('E60','Research boundary','N/A','No public first-party documentation located for asserted absent items','No screenshot','N/A'),
('E61','Human API','https://www.humanapi.co/','A','Comparator reference','Vendor site','Low–Medium'),
('E62','WHOOP','https://www.whoop.com/','A','Comparator reference','Vendor site','Low–Medium'),
('E63','Oura','https://ouraring.com/','A','Comparator reference','Vendor site','Low–Medium'),
('E64','Ultrahuman','https://www.ultrahuman.com/','A','Comparator reference','Vendor site','Low–Medium'),
]

claims = [
('C01','OpenEvidence is a cited medical information / clinical decision-support platform for verified healthcare professionals.','🟢 Confirmed','E01,E05,E13','Company describes AI copilot/medical information platform; terms/professional verification.','Public About / release text','Observed','High'),
('C02','Core answer workflow uses natural-language questions and source-cited research synthesis.','🟢 Confirmed','E05,E12','Company release and Wiley agreement describe cited synthesis/retrieval.','Public release text','Observed','High'),
('C03','OpenEvidence offers a free verified-U.S.-clinician access model.','🟢 Confirmed','E01,E05,E13','Website/releases/policy state free verified professional access.','Public page','Observed','High'),
('C04','The core economic model includes advertising and partnership revenue.','🟢 Confirmed','E13,E45','Privacy policy explicitly describes advertising/partnership support; NBC reports CEO statement.','Public policy/news','Observed','High'),
('C05','OpenEvidence has licensed content relationships with NEJM, JAMA, Wiley/Cochrane and NCCN.','🟢 Confirmed','E12,E18,E19,E34','Named agreements/announcement.','Public releases','Observed','High'),
('C06','OpenEvidence announced DeepConsult as agentic deep research across hundreds of studies.','🟢 Confirmed','E05','Company announcement.','Public release','Observed','High'),
('C07','OpenEvidence announced Visits with transcription, templates, patient documents and evidence integration.','🟢 Confirmed','E06','Company announcement.','Public product images/release','Observed','High'),
('C08','OpenEvidence announced Dialer with call, SMS, fax, voicemail and Create Visit.','🟢 Confirmed','E07','Company announcement.','Public product image/release','Observed','High'),
('C09','OpenEvidence announced trial matching against patient characteristics/history/location.','🟢 Confirmed','E08','Company announcement.','Public product images/release','Observed','High'),
('C10','EvidenceGrade is a GRADE-inspired real-time evidence-strength system, not stated to be a formal systematic review.','🟢 Confirmed','E09','Method post describes both method and limitations.','Public diagrams/text','Observed','High'),
('C11','OpenEvidence says specialised models are coordinated by a central conductor.','🟢 Confirmed','E26','Series D release.','Public release','Observed','High self-report'),
('C12','OpenEvidence has announced Sutter/Epic workflow deployment and Cedars Epic patient context.','🟢 Confirmed','E33,E35','Named partner announcements.','Public release images','Observed','High'),
('C13','Cedars announcement says patient EHR data will not be stored after the clinical session for that integration.','🟢 Confirmed','E35','Specific partnership statement.','Public release text','Observed','High for stated integration'),
('C14','OpenEvidence says services are primarily hosted on GCP and Vercel.','🟢 Confirmed','E04','Security page.','Public security text','Observed','High self-report'),
('C15','OpenEvidence claims HIPAA compliance, BAA availability and SOC 2 Type II Security certification.','🟢 Confirmed','E04,E15,E40','Company security page/trust/BAA.','Public legal/trust text','Observed','High self-report'),
('C16','OpenEvidence uses profile/activity data for personalised advertising but says question/conversation text is not shared for that purpose.','🟢 Confirmed','E13','Privacy policy makes both statements.','Public policy text','Observed','High'),
('C17','Public sources do not confirm OpenEvidence base model providers, vector DB, public API, FHIR resource mapping, or broad EHR write-back.','🟢 Research boundary','E60','No reviewed first-party documentation located; not proof of absence.','No screenshot','Observed limitation','Medium'),
('C18','Citations alone are insufficient to prove clinical safety because retrieval, entailment, applicability and contradiction can fail separately.','🟡 Strong Inference','E05,E09,E20,E21,E22','Public design + independent/user caveats support analytical conclusion.','Textual sources','Inferred','High'),
('C19','OpenEvidence should be treated as evidence-grounded decision support, not hallucination-proof clinical reasoning.','🟡 Strong Inference','E09,E14,E20,E21,E22','Terms disclaim treatment advice; studies/community show performance/citation caution.','Textual sources','Inferred','High'),
('C20','OpenEvidence’s strongest moat is the combination of content rights, clinician distribution and workflow integration.','🟡 Strong Inference','E05,E06,E07,E12,E16,E33','Multiple direct evidence sources support combined strategic conclusion.','Public releases','Inferred','High'),
('C21','OpenEvidence’s largest strategic whitespace is a patient-owned reconciled longitudinal intelligence layer.','🟡 Strong Inference','E06,E35,E60','Documents/session context are public; no full patient-owned longitudinal platform documented.','Public/research boundary','Inferred','Medium'),
('C22','A clinically safer Ovexis architecture should provide claim-to-source-span verification, applicability and conflict signals.','🟡 Strong Inference','E09,E20,E21,E22','Identified safety gaps inform design recommendation.','Analytical','Inferred','High'),
('C23','Series B was $210M at $3.5B valuation.','🟢 Confirmed','E05','Company release.','Public release','Observed','High'),
('C24','Series D was $250M at $12B valuation, co-led by Thrive and DST.','🟢 Confirmed','E23,E24,E26','Company and independent reporting.','News/release','Observed','High'),
('C25','The reported $20B July 2026 financing was unconfirmed at cutoff.','🔴 Speculation / rumour','E27','Reporting says company did not confirm.','News page','Reported rumour','Low'),
('C26','Amaro acquisition was strategic advertising infrastructure, not a clinical-content acquisition.','🟢 Confirmed','E16','Acquisition release explains target/rationale.','Public PR release','Observed','High'),
('C27','Clinician community reports include both paper-discovery praise and summary/citation caution.','🟢 Confirmed as anecdotal reports','E22,E42,E43','Public Reddit threads.','Community comments','Observed anecdote','Low incidence'),
('C28','Independent studies should be interpreted by version/benchmark and do not alone determine real-world product utility.','🟡 Strong Inference','E20,E21','Study scope limitations.','Academic sources','Inferred','High'),
('C29','No public API should be assumed from named enterprise integrations.','🟡 Strong Inference','E11,E33,E35,E60','Named integrations do not document a general developer API.','Public/research boundary','Inferred','High'),
('C30','Advertising creates a structural trust risk even if question text is not used for ads.','🟡 Strong Inference','E13,E16,E45','Data/monetisation design supports risk conclusion.','Policy/release/news','Inferred','High'),
]

risks = [
('R01','Citation–claim mismatch or unsupported synthesis','Clinical AI','🟡 Medium','🟡 Critical','E20,E22,E42','Claim-to-passage verifier; high-risk review; visible correction history','Clinical AI lead','Open'),
('R02','Weak/old/irrelevant evidence appears authoritative','Clinical','🟡 High','🟡 Critical','E09,E22','Evidence hierarchy, recency, directness and contradiction display','Clinical governance','Open'),
('R03','Stale/missing/misattributed patient context','Data/clinical','🟡 Medium','🟡 Critical','E35','Reconciliation, provenance, user confirmation','Interoperability lead','Open'),
('R04','EHR/foundation model incumbents bundle competing capability','Business','🟡 High','🟡 High','E11,E33,E35','Longitudinal cross-system differentiation','Strategy','Open'),
('R05','Content licensing concentration/cost','Business','🟡 Medium','🟡 High','E12,E18,E19','Rights diversification and versioned open evidence','Content lead','Open'),
('R06','Advertising undermines clinician trust','Brand/regulatory','🟡 Medium','🟡 High','E13,E16,E45','No clinical-path ads; transparent firewall','CEO/Trust','Open'),
('R07','Deep-agent compute economics fail','Economic','🟡 Medium','🟡 High','E05,E26','Risk-tier routing, budgeting, asynchronous deep reviews','Platform lead','Open'),
('R08','PHI/telecom/document breach','Security','🟡 Low–Medium','🟡 Critical','E07,E15,E40','Data minimisation, feature BAAs, audit, incident drills','CISO','Open'),
('R09','Regulation/reclassification/geoblocking','Regulatory','🟡 Medium','🟡 High','E14,E44','Safety case, regional architecture, intended-use controls','Regulatory lead','Open'),
('R10','High-profile patient harm/reputational event','Clinical/brand','🟡 Medium','🟡 Critical','E14,E20,E21','Hazard testing, abstention, escalation, transparency','CMO','Open'),
]

competitors = [
('OpenEvidence','🟢 Clinical AI evidence / workflow','🟢 E01,E05','Direct benchmark','🟡 Licensed evidence + clinician PLG + workflow'),
('Regacore','🟢 Public positioning insufficiently verified','🟢 E60','Unknown','🟡 Disambiguate before comparison'),
('Superpower','🟡 Consumer preventive health intelligence','🟢 E48','Adjacent longitudinal health','🟡 Benchmark engagement'),
('Function Health','🟡 Consumer lab health insights','🟢 E49','Adjacent labs','🟡 Benchmark lab logistics'),
('Levels','🟡 Metabolic-health/sensor interpretation','🟢 E50','Adjacent continuous data','🟡 Benchmark biomarker education'),
('PreventiveHealth.ai','🟢 Public positioning insufficiently verified','🟢 E60','Unknown','🟡 Validate separately'),
('Glass Health','🟡 Clinician AI/workflow','🟢 E51','Direct/adjacent','🟡 Compare evidence transparency'),
('Atropos Health','🟡 Real-world evidence','🟢 E52','Adjacent evidence','🟡 Local RWE opportunity'),
('AMBOSS','🟡 Medical knowledge/education','🟢 E53','Direct reference','🟡 Structured content incumbent'),
('UpToDate','🟡 Curated clinical reference','🟢 E54','Direct reference','🟡 Editorial accountability benchmark'),
('Apollo 24|7','🟡 India care/telehealth','🟢 E55','Low direct overlap','🟡 India care-navigation benchmark'),
('Practo','🟡 India provider discovery','🟢 E56','Low direct overlap','🟡 Provider network benchmark'),
('Tata 1mg','🟡 India pharmacy/diagnostics','🟢 E57','Low direct overlap','🟡 Fulfilment benchmark'),
('Healthify','🟢 Ambiguous entity','🟢 E60','Unknown','🟡 Disambiguate'),
('Apple Health','🟡 Consumer longitudinal health data','🟢 E58','Complementary','🟡 Consent/data UX benchmark'),
('Google Health Connect','🟡 Device health-data platform','🟢 E59','Complementary','🟡 Interoperability benchmark'),
('Human API','🟡 Health data connectivity','🟢 E61','Complementary','🟡 Record acquisition archetype'),
('WHOOP','🟡 Wearable/recovery platform','🟢 E62','Adjacent','🟡 Engagement benchmark'),
('Oura','🟡 Wearable/readiness platform','🟢 E63','Adjacent','🟡 Trend UX benchmark'),
('Ultrahuman','🟡 Wearable/metabolic platform','🟢 E64','Adjacent','🟡 Hardware/data UX benchmark'),
]

roadmap = [
('2023–2024','🟢 MVP','Medical search, cited answers, credentialed access, early research, mobile','E28,E29','High'),
('2024-12–2025-07','🟢 V2','Admin workflows, calculators, licences, HIPAA/BAA, DeepConsult','E05,E10,E18,E30','High'),
('2025-08–2026-07','🟢 V3/current','Visits, Dialer, trial matching, EHR context, EvidenceGrade, societies','E06,E07,E08,E09,E12,E33,E35','High'),
('2026–2027','🟡 Near roadmap','More speciality models/guidelines, EHR embedding, evaluation/governance','E09,E26,E33,E35','Medium'),
('Ovexis 0–6 months','🟡 Recommended MVP','Timeline + evidence packets + applicability + action ledger','C21,C22','Medium'),
('Ovexis 6–12 months','🟡 Recommended V2','Local policy, tasks, FHIR, prospective evaluation','C22','Medium'),
('Ovexis 12–24 months','🟡 Recommended V3','Pharmacy/labs/claims, referral/trial, multilingual regional expansion','C21,C22','Medium'),
]

wb = Workbook()
ws = wb.active
ws.title = 'README'
readme = [
['OpenEvidence — Master Feature Inventory & Evidence Workbook'],
['Snapshot date','25 July 2026 (Asia/Kolkata)'],
['Purpose','Companion structured workbook to the board report. Public information only; no authenticated product access.'],
['Status legend','🟢 Confirmed = direct source; 🟡 Strong Inference = explicitly analytical; 🔴 Speculation = unconfirmed scenario.'],
['How to use','Filter Master Feature Inventory by priority/category/recommendation. Use Evidence Register and Claim Ledger to trace material statements.'],
['Important limitation','“Not publicly confirmed” or E60 does not prove absence; it records the reviewed public-source boundary.'],
['Source quality','A first-party/legal; B reputable independent or app store; C community/preprint/analyst; D research boundary.'],
['Commercial caution','Company-reported adoption/revenue/security metrics are not audited unless separately labelled.'],
]
for r in readme: ws.append(r)
ws.column_dimensions['A'].width=28; ws.column_dimensions['B'].width=130
ws['A1'].font=Font(bold=True,size=18,color='FFFFFF'); ws['A1'].fill=PatternFill('solid',fgColor='153B5B'); ws.merge_cells('A1:B1')
for row in range(2,len(readme)+1):
    ws.cell(row,1).font=Font(bold=True,color='153B5B')
    ws.cell(row,2).alignment=Alignment(wrap_text=True,vertical='top')
    ws.row_dimensions[row].height=35

# helper styles
header_fill=PatternFill('solid',fgColor='153B5B')
sub_fill=PatternFill('solid',fgColor='D9EAF7')
confirmed_fill=PatternFill('solid',fgColor='D9EAD3')
inference_fill=PatternFill('solid',fgColor='FFF2CC')
spec_fill=PatternFill('solid',fgColor='F4CCCC')
thin=Side(style='thin',color='D9E2F3')

def setup_sheet(ws, headers, widths, table_name=None, freeze='A2'):
    ws.append(headers)
    for c, h in enumerate(headers,1):
        cell=ws.cell(1,c); cell.font=Font(bold=True,color='FFFFFF'); cell.fill=header_fill; cell.alignment=Alignment(wrap_text=True,vertical='center')
        ws.column_dimensions[get_column_letter(c)].width=widths[c-1] if c-1<len(widths) else 18
    ws.freeze_panes=freeze
    ws.auto_filter.ref=f'A1:{get_column_letter(len(headers))}1'
    ws.row_dimensions[1].height=34

def add_rows(ws, rows):
    for row in rows:
        ws.append(list(row))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment=Alignment(wrap_text=True,vertical='top')
            cell.border=Border(bottom=thin)
        # status marker wherever present
        text=' '.join(str(c.value) for c in row if c.value is not None)
        if '🟢' in text: row[0].fill=confirmed_fill
        elif '🟡' in text: row[0].fill=inference_fill
        elif '🔴' in text: row[0].fill=spec_fill
    for r in range(2,ws.max_row+1): ws.row_dimensions[r].height=45

def table(ws,name):
    ref=f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
    tab=Table(displayName=name, ref=ref)
    tab.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=False)
    ws.add_table(tab)

# Feature Inventory
ws=wb.create_sheet('Master Feature Inventory')
headers=['Feature','Purpose','Evidence','User Value','Business Value','Engineering Complexity','Clinical Complexity','Infrastructure Complexity','Regulatory Complexity','Estimated Team','Estimated Months','Priority','Category','Copy','Improve','Ignore','Reinvent','Moat','Confidence']
widths=[30,43,30,14,15,16,16,18,17,14,14,10,24,12,12,12,12,16,14]
setup_sheet(ws,headers,widths)
add_rows(ws,features); table(ws,'FeatureInventory')
# Conditional colours for priority
for r in range(2,ws.max_row+1):
    c=ws.cell(r,12)
    if c.value=='P0': c.fill=PatternFill('solid',fgColor='F4CCCC')
    elif c.value=='P1': c.fill=PatternFill('solid',fgColor='FFF2CC')
    elif c.value=='P2': c.fill=PatternFill('solid',fgColor='D9EAD3')
    else: c.fill=PatternFill('solid',fgColor='EDEDED')

# Evidence
ws=wb.create_sheet('Evidence Register')
headers=['Evidence ID','Source','URL','Source Grade','Evidence / observation','Screenshot / visual evidence','Confidence']
setup_sheet(ws,headers,[12,32,70,15,54,30,18])
add_rows(ws,sources); table(ws,'EvidenceRegister')
for row in range(2,ws.max_row+1):
    ws.cell(row,3).hyperlink=ws.cell(row,3).value if str(ws.cell(row,3).value).startswith('http') else None
    ws.cell(row,3).style='Hyperlink' if ws.cell(row,3).hyperlink else 'Normal'

# Claim ledger
ws=wb.create_sheet('Claim Ledger')
headers=['Claim ID','Claim','Status','Evidence ID(s)','Evidence summary','Screenshot / observation','Observed vs inferred','Confidence']
setup_sheet(ws,headers,[12,58,22,18,62,32,20,16])
add_rows(ws,claims); table(ws,'ClaimLedger')

# Decision ledger
ws=wb.create_sheet('Decision Ledger')
headers=['Feature decision','Why built / pain','KPI likely improved','Trade-off','Alternative architecture','Evidence / confidence']
decisions = [
('Free verified access','🟡 Remove price/procurement friction; make audience valuable','🟡 Registration/DAU/ad inventory','🟡 Ad-trust tension and subsidised compute','🟡 Paid individual or employer licence','🟢 E01,E05,E13; High'),
('Credential gate','🟢 Verify professional access','🟡 Audience quality/trust','🟡 Excludes non-NPI/international users','🟡 Open access + risk-tiered mode','🟢 E13; High'),
('Licensed content','🟢 Better authoritative retrieval','🟡 Trust/coverage','🟡 Rights cost/dependency','🟡 Open corpus/editorial summaries','🟢 E12,E18,E19; High'),
('Cited chat','🟢 Fast answer with sources','🟡 Activation/repeat use','🟡 False confidence if citation mismatch','🟡 Search-first evidence cards','🟢 E05; High'),
('DeepConsult','🟢 Complex cross-study research','🟡 Complex-query satisfaction','🟡 Compute/latency/agent failure','🟡 Async analyst queue','🟢 E05; High'),
('EvidenceGrade','🟢 Evidence strength context','🟡 Trust calibration','🟡 Single grade masks nuance','🟡 Claim-level evidence graph','🟢 E09; High'),
('Visits','🟢 Documentation in care flow','🟡 Daily use/stickiness','🟡 PHI/consent/liability','🟡 Standalone scribe','🟢 E06; High'),
('Dialer','🟢 Privacy + communications + note capture','🟡 Frequency/patient reach','🟡 Telecom compliance','🟡 CPaaS integration only','🟢 E07; High'),
('Trial matching','🟢 Find relevant active trials','🟡 Specialty value/life-science relevance','🟡 Eligibility and site-data error','🟡 Navigator service','🟢 E08; High'),
('Epic context','🟢 Patient-aware answers','🟡 Enterprise value','🟡 Integration/PHI/local governance','🟡 User-entered structured context','🟢 E33,E35; High'),
('Amaro acquisition','🟢 Advertising infrastructure','🟡 Monetisation/ad operations','🟡 Clinical brand conflict','🟡 External ad tech/subscription','🟢 E16; High'),
]
setup_sheet(ws,headers,[25,43,28,38,38,26]); add_rows(ws,decisions); table(ws,'DecisionLedger')

# Risk register
ws=wb.create_sheet('Risk Register')
headers=['Risk ID','Risk','Type','Likelihood','Impact','Evidence','Mitigation / Ovexis lesson','Owner','Status']
setup_sheet(ws,headers,[10,38,20,18,18,20,62,22,14]); add_rows(ws,risks); table(ws,'RiskRegister')

# Competitors
ws=wb.create_sheet('Competitor Matrix')
headers=['Company','Public category lens','Evidence','Overlap','Ovexis strategic read']
setup_sheet(ws,headers,[28,43,18,25,50]); add_rows(ws,competitors); table(ws,'CompetitorMatrix')

# Roadmap
ws=wb.create_sheet('Roadmap Reconstruction')
headers=['Horizon','Status','Scope','Evidence','Confidence']
setup_sheet(ws,headers,[24,18,90,30,18]); add_rows(ws,roadmap); table(ws,'Roadmap')

# Business model canvas
ws=wb.create_sheet('Business Model Canvas')
headers=['Block','Assessment','Evidence / status']
canvas = [
('Customer segments','Clinicians/users; health systems; pharma/med-device advertisers; publishers/societies; platform partners','🟡 E01,E13,E16,E33'),
('Value proposition','Fast cited evidence at point of care, increasingly documentation/communication','🟢 E05,E06,E07'),
('Channels','Web, iOS, Android, society/publisher, EHR/platform, word-of-mouth claim','🟢 E01,E05,E29'),
('Relationships','Self-serve clinician; contracted enterprise/content/advertiser relations','🟡 E13,E14'),
('Key resources','Content rights, verified audience, specialised models, clinical team, trust brand','🟡 E01,E12,E26'),
('Key activities','Content ingestion, retrieval/evaluation, clinical safety, product, ads, enterprise integration','🟡 E05,E13,E16'),
('Key partners','Publishers/societies, health systems, Microsoft, Veeva, GCP/Vercel','🟢 E01,E11,E12,E15,E33'),
('Cost structure','Content rights, model compute, cloud, staff, security, telecom, sales/integration','🟡 E05,E07,E12,E15'),
('Revenue','Advertising/partnerships confirmed; enterprise/API economics not publicly disclosed','🟢 E13,E45; 🟡 enterprise inference'),
]
setup_sheet(ws,headers,[28,90,30]); add_rows(ws,canvas); table(ws,'BusinessModelCanvas')

# apply page settings all sheets
for ws in wb.worksheets:
    ws.sheet_view.showGridLines=False
    ws.page_setup.orientation='landscape'
    ws.page_setup.fitToWidth=1
    ws.page_margins.left=0.25; ws.page_margins.right=0.25; ws.page_margins.top=0.5; ws.page_margins.bottom=0.5
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment=Alignment(wrap_text=True,vertical='top')

wb.save(OUT)
print(f'Wrote {OUT} with {len(features)} feature rows, {len(sources)} evidence sources, and {len(claims)} material claims.')
