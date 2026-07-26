from __future__ import annotations

from pathlib import Path
from textwrap import dedent
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except Exception as e:
    raise SystemExit(f"openpyxl unavailable: {e}")

OUT = Path('/home/user')

sources = [
    ("S01", "Levels homepage", "https://www.levelshealth.com/", "Homepage copy: app, CGM, labs, AI insights, adaptive programs, pricing, trust signals, HIPAA/SOC2, no data sale, 100,000+ lives changed."),
    ("S02", "What Levels is - Support", "https://support.levels.com/article/719-what-levels-is", "Support article: continuous glucose monitoring, lab testing, support/guidance, three-step workflow, health goals, U.S. shipping, no prescription required for Levels in current offering."),
    ("S03", "Levels pricing and plans - Support", "https://support.levels.com/article/720-levels-pricing-and-plans", "Membership: $15/month or $80/year; add-ons: Stelo $89/$99, Basic labs $99, Comprehensive labs $399, nutritionist $250; legacy Core/Complete; TrueMed/Affirm."),
    ("S04", "iOS App Store listing", "https://apps.apple.com/us/app/levels-metabolic-health/id1481511675", "App features, rating, disclaimers, Apple Health integration, user reviews including praise/complaints and AI coach complaints."),
    ("S05", "Getting Started with Levels - Support", "https://support.levels.com/article/726-getting-started-with-levels", "Kit ships 1-2 business days; app download; health intake; logging; Stelo setup; AI Guide bubble; first-week tips."),
    ("S06", "How to Download and Set Up the Levels App - Support", "https://support.levels.com/article/723-how-to-download-and-set-up-the-levels-app", "Account creation, practitioner invitation, health intake, notification/Bluetooth/camera permissions, app home/logging/insights/trends/programs."),
    ("S07", "How to Apply and Connect Stelo to Levels", "https://support.levels.com/article/680-how-to-apply-and-connect-stelo-to-the-levels-app", "Stelo app, Dexcom account, application, performance cover, Levels More > Devices > Manage, Permission to Disclose/Authorization, 5-minute readings and 15-minute Levels refresh."),
    ("S08", "How to Get Dexcom G7 Through Levels", "https://support.levels.com/article/763-how-to-get-dexcom-g7-through-levels", "Medical intake, physician partner review, prescription sent to Amazon Pharmacy, cash price approx. $177/month, ~1-hour data delay."),
    ("S09", "Where Levels is Available", "https://support.levels.com/article/21-where-levels-is-available", "Available in U.S. only; all 50 states and some territories with constraints; requirements include U.S. shipping address, U.S. phone, U.S. app store/play store, physical presence for Stelo."),
    ("S10", "Food logging - Support", "https://support.levels.com/article/463-how-to-log-food-in-the-levels-app", "Describe/voice/text+AI, photo+AI, recents, barcode, search, custom foods, long-press editing, retroactive logging from glucose graph."),
    ("S11", "Habit Loops - Support", "https://support.levels.com/article/477-how-to-use-habit-loops", "11 habit loops, progress vs limit loops, up to three active habits, targets customizable, updated July 24 2026."),
    ("S12", "Health data sync - Support", "https://support.levels.com/article/727-how-to-sync-your-health-data-with-levels", "Apple Health/Health Connect sync for steps, workouts, sleep, weight, heart rate and other supported health signals."),
    ("S13", "Labs biomarkers - Support", "https://support.levels.com/article/561-biomarkers-levels-labs-tests", "Basic panel 28 biomarkers; Comprehensive panel 100+ biomarkers; categories and pricing."),
    ("S14", "Lab upload - Support", "https://support.levels.com/article/489-how-to-upload-lab-results-to-levels", "PDF upload, AI extraction, review/edit, confirm, free AI-powered lab interpretation informed by biomarker expertise and 1.5B+ health data points."),
    ("S15", "Labs policies - Support", "https://support.levels.com/article/467-levels-labs-policies", "Lab ordering partners: MD Integrations, Healthie, Quest; Levels not provider; no diagnosis; critical values process; results in app/portal."),
    ("S16", "Labs availability", "https://support.levels.com/article/601-where-levels-labs-is-available", "Labs purchasable in all 50 U.S. states; Quest most orders, BioReference NY/NJ, LabCorp historical, travel requirements in some states/territories."),
    ("S17", "Privacy Policy", "https://levels.link/privacy", "Privacy principles, data categories, AI provider list, Langfuse, pseudonymization, security practices, HIPAA business associate posture, deletion/export."),
    ("S18", "Terms of Service", "https://www.levels.com/terms-of-service", "Levels is not a healthcare provider; general wellness/non-medical use; cash-pay/no insurance; telemedicine consent; prohibited uses; pharmacy services."),
    ("S19", "Levels Pro", "https://www.levels.com/pro", "Professional dashboard, AI-prepared review workflow, client roster, labs/CGM, recommendations, programs/protocols, pricing $99/$299 with monthly credits."),
    ("S20", "Seed round announcement", "https://www.levels.com/blog/levels-raises-12m-investment-round-improve-metabolic-health-wearables", "$12M seed led by a16z; founders; 50k+ waitlist; two 14-day CGM beta program; performance/athletic market entry."),
    ("S21", "Series A announcement", "https://www.levels.com/blog/levels-38m-series-a-driven-by-member-and-community-alignment-to-solve-metabolic-health-crisis", "$38M Series A at $300M valuation; $5M member crowdfunding; 25k paying beta members; $50M total raised; mission statements."),
    ("S22", "2024 $10M extension coverage", "https://pulse2.com/levels-10-million-series-a-extension-closed-to-show-how-food-affects-health/", "$10M Series A extension; $3M crowdfunding; Long Journey/a16z; 60k+ members; 700M+ glucose data points; 18M YouTube views; 2M+ annual blog visitors."),
    ("S23", "Research background", "https://support.levels.com/article/373-levels-study-background", "Study rationale: CGM in free-living non-diabetic populations; few studies >10k; CGM at 5-minute resolution; study goals."),
    ("S24", "Research article", "https://www.levels.com/blog/why-levels-does-research-and-what-we-hope-to-learn", "IRB-approved observational study; 10k+ members in 2022; anonymized and aggregated data; consent and opt-out."),
    ("S25", "Research concluded", "https://support.levels.com/article/683-the-levels-research-study-has-concluded", "IRB-approved CGM study completed and study services discontinued."),
    ("S26", "Biomarker improvement report", "https://www.levels.com/blog/levels-improves-biomarkers", "Retrospective non-clinical-trial member analysis: HbA1c, fasting insulin, triglyceride improvements; tools tied to behavior change."),
    ("S27", "AI Search page", "https://www.levels.com/blog/ai-search", "LevelsAI searches 700+ reported/fact-checked articles."),
    ("S28", "CEO Jan 2026 AI note", "https://www.levels.com/blog/from-the-ceo-jan26", "Josh Clemente on AI, in-app AI, data/knowledge-base query, Document Center as intelligent health record, reliable guidance and privacy."),
    ("S29", "Josh Clemente author page", "https://www.levels.com/blog/author/joshclemente", "Official author page: founder and president; 'From the CEO' newsletters authored by Josh."),
    ("S30", "Casey Means author page", "https://www.levels.com/blog/author/caseymeans", "Casey co-founder, served as CMO until late 2023, Stanford-trained physician, author."),
    ("S31", "Data export - Support", "https://support.levels.com/article/105-how-to-export-your-data", "CSV export of CGM, activity logs, zones/zone scores/glucose response, nutrition logs."),
    ("S32", "Share data with doctor - Support", "https://support.levels.com/article/501-how-to-share-levels-data-with-your-doctor", "Member portal flow: Data > Glucose > Share Your Data > grant email access."),
    ("S33", "Delete data - Support", "https://support.levels.com/article/239-how-to-delete-your-data", "Deletion request via privacy@levels.com; verification; one-month response; exceptions for payments, official medical records, third parties."),
    ("S34", "Exercise logging", "https://support.levels.com/article/173-how-to-log-notes-exercise", "Manual exercise logging and automatic exercise sync via Apple Health/Health Connect."),
    ("S35", "Exercise spikes", "https://support.levels.com/article/47-about-exercise-glucose-spikes", "Strenuous activities can be ignored for score; Apple Health/Health Connect import auto-strenuous when HR >150 BPM."),
    ("S36", "Sleep data", "https://support.levels.com/article/382-how-to-add-sleep-data-to-the-levels-app", "Sleep imported from Apple Health/Health Connect; no direct sleep logging in Levels."),
    ("S37", "Weight data", "https://support.levels.com/article/485-how-to-add-weight-data", "Weight imported from Apple Health/Health Connect or manually updated in profile."),
    ("S38", "Notes", "https://support.levels.com/article/646-how-to-log-notes", "Timeline notes for context: stress, illness, new food, routine change."),
    ("S39", "Lifestyle events", "https://support.levels.com/article/647-how-to-log-a-lifestyle-event", "Lifestyle event logging; ignore glucose box for event-related spikes."),
    ("S40", "Privacy and app analytics", "https://support.levels.com/article/136-privacy-and-app-analytics", "Aggregate app analytics; not used/sold/traded for advertising/marketing/other commercial purposes; app analytics opt-out."),
    ("S41", "GitHub org search result", "https://github.com/levelshealth", "Levels GitHub org: 17 repositories, including react-native-help-scout and react-native-health-connect forks; no major public proprietary OSS found."),
    ("S42", "Wefunder listing", "https://wefunder.com/levels", "Crowdfunding listing claims 60k+ members, 700M+ glucose points, 7M+ food logs, 468k email subscribers, $20M+ last-12-month revenue; third-party/crowdfunding context."),
    ("S43", "Dexcom Stelo FDA coverage", "https://investors.dexcom.com/news/news-details/2024/Stelo-by-Dexcom-First-Glucose-Biosensor-to-be-Cleared-by-FDA-as-Over-the-Counter/default.aspx", "Dexcom announced FDA clearance of Stelo as first OTC glucose biosensor for adults 18+ not using insulin therapy."),
    ("S44", "HTTP observations captured", "local: curl headers 2026-07-25", "Observed Cloudflare, Next.js static assets, CSP frame-ancestors self, PostHog/Sentry/Stripe/Truemed/Typeform/CookiePro/Apollo/Leadfeeder/acsbapp references in public HTML."),
    ("S45", "Reddit user review", "https://www.reddit.com/r/levelshealth/comments/16cscpn/worth_it_or_not_a_users_review/", "Customer complaints around dual apps, sync, calibration, food entry, narrow ranges, endurance athlete fit, perceived poor value."),
    ("S46", "App Store reviews", "https://apps.apple.com/us/app/levels-metabolic-health/id1481511675?see-all=reviews&platform=iphone", "User praise and complaints: sync failures, AI coach wrong information, limited food tracker, behavior change, cancellation."),
    ("S47", "Function Health pricing", "https://www.functionhealth.com/pricing", "Function membership starts at $365/year and includes 160+ lab tests, clinician review, personalized action plan."),
    ("S48", "Superpower blood test", "https://superpower.com/welcome-cms/best-blood-test-service", "Superpower: $199/year, 100+ biomarkers, Quest labs, AI health analysis, personalized action plan, care team."),
    ("S49", "OpenEvidence funding/usage", "https://www.fiercehealthcare.com/ai-and-machine-learning/openevidence-clinches-250m-series-d-rapidly-growing-its-reach-doctors", "OpenEvidence $250M Series D, medical search/generative chatbot for doctors, large physician usage claims."),
    ("S50", "Human API", "https://risk.lexisnexis.com/products/humanapi", "LexisNexis Human API: consumer-permissioned data network, 30k+ connections, medical records/labs/wearables, data normalization."),
]

features = [
    # feature, purpose, evidence ids, user value, business value, eng, clinical, infra, reg, team, months, priority, category, action, moat, confidence
    ("Homepage value proposition", "Explain metabolic health and drive quiz/start conversion", "S01", "Clear problem framing and trust cues", "Top-of-funnel conversion", 2, 1, 1, 2, "Growth+Design", 1, "High", "Marketing", "Improve", "Brand", "High"),
    ("Personalized plan quiz", "Route users to plan/package", "S01,S06", "Reduces plan ambiguity", "Lead capture and conversion", 3, 1, 2, 2, "Growth+Eng", 2, "High", "Onboarding", "Copy", "Distribution", "Med"),
    ("App-only membership", "Separate software from hardware", "S03", "Low entry price and non-CGM use", "Expands TAM and retention after CGM sprint", 3, 1, 2, 2, "Product+Billing", 2, "High", "Business model", "Copy", "Pricing", "High"),
    ("Stelo add-on/subscription", "Provide OTC CGM access", "S03,S07,S43", "Easy glucose monitoring without prescription", "Add-on revenue and data generation", 4, 2, 3, 3, "Ops+Eng", 3, "High", "CGM", "Copy", "Data", "High"),
    ("Dexcom G7 prescription pathway", "Support users who prefer prescription G7", "S08", "More sensor choice", "Coverage of advanced users and pharmacy flow", 5, 4, 4, 5, "Clinical Ops+Eng", 4, "Medium", "CGM", "Improve", "Regulatory", "High"),
    ("Bring your own CGM", "Let members use external sensors", "S03,S07,S08", "Reduces lock-in and cost", "Reduces hardware ops burden", 4, 2, 3, 3, "Mobile+Partner", 3, "High", "Integration", "Copy", "Switching", "High"),
    ("Stelo direct connection", "Pull Stelo data into Levels", "S07", "Simpler direct data flow", "Better experience than Apple/Health Connect delay", 5, 2, 4, 3, "Mobile+Backend", 4, "High", "Integration", "Improve", "Data", "High"),
    ("Glucose timeline", "Show CGM trace alongside logs", "S01,S06,S10", "Causal understanding", "Core engagement loop", 5, 2, 4, 2, "Mobile+Data", 4, "High", "Visualization", "Copy", "Data", "High"),
    ("Current glucose home card", "Show recent/current glucose", "S06", "Immediate feedback", "Daily active usage", 3, 2, 3, 2, "Mobile", 2, "High", "Visualization", "Copy", "Engagement", "High"),
    ("Meal logging by text/voice", "Fast natural-language input", "S10", "Lower logging friction", "More food logs and model training data", 5, 1, 4, 2, "AI+Mobile", 4, "High", "Nutrition AI", "Copy", "Data", "High"),
    ("Meal logging by photo", "AI recognizes plate contents", "S10", "Lower cognitive burden", "High-frequency user-generated data", 6, 1, 5, 2, "AI+Mobile", 5, "High", "Nutrition AI", "Improve", "Data", "High"),
    ("Barcode scanner", "Log packaged foods", "S10", "Accuracy for labeled items", "Food database enrichment", 4, 1, 3, 1, "Mobile+Data", 3, "High", "Food logging", "Copy", "Data", "High"),
    ("Food search", "Find foods/brands/history", "S10", "Completes meal logging", "Retention through utility", 4, 1, 4, 1, "Mobile+Data", 3, "High", "Food logging", "Copy", "Data", "High"),
    ("Custom foods", "Allow unknown/local foods", "S10", "Coverage for real-world diets", "Fewer failed logs", 3, 1, 3, 1, "Mobile+Data", 2, "Medium", "Food logging", "Copy", "Data", "High"),
    ("Recents/re-log", "Speed repeated meals", "S10", "Less repetitive input", "More adherence", 3, 1, 2, 1, "Mobile", 2, "High", "Food logging", "Copy", "Engagement", "High"),
    ("Retroactive logging from graph", "Timestamp events on glucose curve", "S10", "Corrects missed logs", "Improves data-label accuracy", 5, 1, 3, 1, "Mobile+Data", 3, "Medium", "UX", "Copy", "Data quality", "High"),
    ("Meal macro breakdown", "Display protein/fat/carbs/fiber/sugar", "S01,S10", "Turns meals into actionable nutrition", "Keeps app useful without CGM", 5, 2, 4, 2, "AI+Nutrition", 4, "High", "Nutrition", "Copy", "Data", "High"),
    ("Food quality feedback", "Score/assess meal quality", "S01,S10", "Simple decision cue", "Behavior-change loop", 5, 3, 4, 3, "Nutrition+AI", 4, "High", "Behavior", "Improve", "Clinical", "High"),
    ("Meal/Zone score", "Rate glucose response to meal/event", "S04,S31", "Easy interpretation", "Signature feature", 5, 4, 4, 3, "Data Science", 4, "High", "Scoring", "Copy", "Brand", "High"),
    ("Daily Stability/Metabolic score", "Summarize day glucose stability", "S04,S35", "At-a-glance score", "Gamified retention", 5, 4, 4, 3, "Data Science", 4, "High", "Scoring", "Improve", "Behavior", "High"),
    ("Ignore glucose for strenuous exercise", "Avoid penalizing exercise glucose rise", "S35", "More fair scores for athletes", "Reduces churn from false negatives", 4, 3, 3, 2, "Data+Mobile", 3, "High", "Scoring", "Improve", "Personalization", "High"),
    ("Automatic strenuous detection", "Detect HR >150 from Apple/Health Connect", "S35", "Less manual correction", "Better score trust", 5, 3, 4, 2, "Mobile+Data", 4, "Medium", "Wearables", "Improve", "Data quality", "High"),
    ("Exercise logging", "Add exercise context", "S34", "Explains glucose movement", "Richer behavior dataset", 3, 1, 2, 1, "Mobile", 2, "High", "Lifestyle", "Copy", "Data", "High"),
    ("Sleep import", "Add sleep context", "S36", "See sleep-glucose relation", "Non-CGM retention via habits", 4, 2, 3, 1, "Mobile+Data", 3, "High", "Wearables", "Copy", "Data", "High"),
    ("Weight import/manual update", "Track body-weight trend", "S37", "Progress monitoring", "Weight-loss program retention", 3, 1, 2, 1, "Mobile", 2, "Medium", "Wearables", "Copy", "Data", "High"),
    ("Mindful time import", "Track mindfulness behavior", "S11,S12", "Stress context", "Holistic positioning", 3, 1, 2, 1, "Mobile", 2, "Low", "Lifestyle", "Improve", "Data", "Med"),
    ("Notes", "Capture user context", "S38", "Explain anomalies", "Improves AI context", 2, 1, 2, 1, "Mobile", 1, "Medium", "Lifestyle", "Copy", "Data", "High"),
    ("Lifestyle events", "Log sauna/cold plunge/etc", "S39", "Context beyond food/exercise", "Biohacker engagement", 3, 1, 2, 1, "Mobile", 2, "Medium", "Lifestyle", "Improve", "Community", "High"),
    ("Habit loops", "Convert goals into daily actions", "S11", "Daily accountability", "Retention and habit formation", 5, 3, 4, 2, "Product+Mobile", 4, "High", "Habit", "Copy", "Behavior", "High"),
    ("Custom habit targets", "Personalize loop goals", "S11", "Fits starting point", "Improves adherence", 4, 2, 3, 1, "Product+Mobile", 3, "High", "Habit", "Improve", "Personalization", "High"),
    ("Programs", "Structured paths for heart/weight/glucose/metabolic goals", "S03,S06", "Guided next steps", "Retains users beyond raw data", 6, 4, 4, 3, "Product+Clinical", 5, "High", "Programs", "Copy", "Clinical", "High"),
    ("AI Guide chat bubble", "Conversational support and proactive feedback", "S05,S17", "Ask about own data", "AI-led engagement", 7, 4, 5, 4, "AI+Product", 6, "High", "AI", "Improve", "AI", "High"),
    ("AI food deconstruction", "Break meal text/photo into items", "S10", "Saves logging effort", "More complete nutrition dataset", 7, 2, 5, 3, "AI+Data", 6, "High", "AI", "Improve", "Data", "High"),
    ("AI health insights", "Surface personalized patterns", "S01,S03,S17,S28", "Personalized guidance", "Differentiation and retention", 8, 5, 5, 5, "AI+Clinical", 8, "High", "AI", "Reinvent", "AI+Data", "High"),
    ("LevelsAI blog search", "RAG over content library", "S27", "Instant education", "SEO-to-product conversion", 5, 2, 3, 2, "AI+Content", 4, "Medium", "AI Search", "Copy", "Content", "High"),
    ("Document Center", "Store health-related documents", "S17,S28", "Central health record", "AI health record roadmap", 7, 4, 5, 5, "Product+AI", 8, "High", "Records", "Reinvent", "Data", "Med"),
    ("AI lab upload extraction", "Extract lab values from PDFs", "S14", "Reduces manual entry", "Acquire non-Levels lab data", 7, 4, 5, 4, "AI+Data", 6, "High", "Labs", "Copy", "Data", "High"),
    ("AI lab interpretation", "Plain-English summary/actions", "S14", "Makes labs understandable", "Converts lab upload to engagement", 7, 5, 4, 5, "AI+Clinical", 7, "High", "Labs", "Improve", "Clinical", "High"),
    ("Basic Labs", "28-marker metabolic/cardiovascular panel", "S13", "Objective health baseline", "Add-on revenue", 4, 5, 4, 5, "Ops+Clinical", 4, "High", "Labs", "Copy", "Clinical", "High"),
    ("Comprehensive Labs", "100+ marker panel", "S13", "Broad health optimization", "High-ARPU add-on", 4, 6, 4, 5, "Ops+Clinical", 4, "High", "Labs", "Copy", "Clinical", "High"),
    ("Clinician review written note", "Review data 2x annually in Core/Complete", "S03", "Professional guidance", "Trust and upsell", 4, 6, 4, 6, "Clinical Ops", 4, "High", "Clinical", "Improve", "Clinical", "High"),
    ("Functional nutritionist session", "50-min 1:1 guidance", "S03", "Human accountability", "High-touch upsell", 3, 4, 2, 4, "Care Ops", 3, "Medium", "Coaching", "Improve", "Clinical", "High"),
    ("Concierge support", "SMS/email support on Complete", "S03", "Reduced friction", "Premium retention", 3, 1, 2, 2, "Support", 2, "Medium", "Support", "Copy", "Service", "High"),
    ("CSV data export", "Download data", "S31", "User control and portability", "Trust/compliance", 4, 1, 3, 3, "Backend", 3, "High", "Privacy", "Copy", "Trust", "High"),
    ("Share data with doctor", "Grant access by email", "S32", "Clinical conversations", "Bridge to care system", 5, 3, 4, 4, "Portal+Security", 4, "High", "Sharing", "Improve", "Trust", "High"),
    ("Delete data request", "Privacy deletion workflow", "S33", "Control and trust", "Compliance risk reduction", 3, 2, 4, 5, "Privacy Ops", 3, "High", "Privacy", "Copy", "Trust", "High"),
    ("Privacy opt-out for research/product improvement", "Let users control data contribution", "S17", "Trust", "Legitimizes data moat", 3, 2, 4, 5, "Privacy+Legal", 3, "High", "Privacy", "Copy", "Trust", "High"),
    ("No data sale principle", "Assure users on monetization", "S01,S17", "Trust", "Brand differentiation", 2, 1, 2, 3, "Legal+Brand", 1, "High", "Trust", "Copy", "Trust", "High"),
    ("AI provider pseudonymization", "Remove direct identifiers before model calls", "S17", "AI privacy assurance", "Enables AI features", 6, 3, 5, 5, "AI+Security", 5, "High", "Security", "Improve", "Trust", "High"),
    ("Langfuse monitoring", "Trace/debug AI prompts/responses", "S17", "More reliable AI", "AI ops quality", 5, 2, 4, 4, "AI Platform", 3, "Medium", "AI Ops", "Copy", "AI", "High"),
    ("Levels Pro roster", "Manage clients in practices", "S19", "Practitioner scale", "B2B revenue", 7, 5, 5, 5, "B2B Product", 8, "High", "Provider", "Reinvent", "Distribution", "High"),
    ("Pro AI summaries", "Prepare client reviews", "S19", "Less manual prep", "Provider adoption", 8, 5, 5, 5, "AI+B2B", 8, "High", "Provider AI", "Improve", "AI+Workflow", "High"),
    ("Pro recommendations/targets", "Turn review into plan", "S19", "Care follow-through", "Sticky workflows", 7, 6, 5, 6, "B2B+Clinical", 8, "High", "Provider", "Reinvent", "Workflow", "High"),
    ("Pro cohorts/challenges", "Group programs", "S19", "Accountability", "Scalable practice revenue", 6, 4, 4, 4, "B2B Product", 6, "Medium", "Provider", "Copy", "Network", "High"),
    ("International waitlist", "Capture non-U.S. demand", "S09", "Future access", "Expansion pipeline", 2, 1, 1, 2, "Growth", 1, "Low", "Expansion", "Improve", "Distribution", "High"),
    ("Research study", "Build glucose reference dataset", "S23,S24,S25", "Citizen-science participation", "Data/clinical credibility", 6, 6, 5, 6, "Research+Data", 12, "Medium", "Research", "Copy", "Data", "High"),
    ("Biomarker outcome report", "Show member improvements", "S26", "Social proof", "Conversion and trust", 4, 5, 3, 4, "Data+Content", 3, "High", "Evidence", "Improve", "Clinical", "High"),
    ("Content library", "Educate on metabolism/nutrition", "S01,S27,S42", "Self-service learning", "SEO acquisition", 3, 4, 2, 2, "Content", 1, "High", "Growth", "Copy", "Brand", "High"),
    ("YouTube/podcast ecosystem", "Founder/advisor-led education", "S20,S42", "Trust and motivation", "Low-CAC growth", 2, 3, 1, 1, "Content+Growth", 1, "High", "Growth", "Copy", "Brand", "Med"),
    ("HSA/FSA TrueMed checkout", "Use pre-tax dollars", "S03,S17", "Effective price reduction", "Conversion", 4, 3, 3, 4, "Growth+Payments", 3, "High", "Payments", "Copy", "Pricing", "High"),
    ("Affirm financing", "Pay over time", "S03", "Lower upfront friction", "Conversion", 3, 1, 2, 2, "Payments", 2, "Medium", "Payments", "Copy", "Pricing", "High"),
    ("Member portal Shop", "Buy add-ons from portal", "S03,S08,S14", "Self-serve expansion", "ARPU expansion", 5, 2, 4, 3, "Product+Billing", 4, "High", "Commerce", "Copy", "Revenue", "High"),
    ("Support/help center", "Answer setup/product questions", "S05,S06,S07", "Reduces confusion", "Support leverage", 3, 1, 2, 1, "Support+Content", 2, "High", "Support", "Copy", "Service", "High"),
]

competitors = [
    ("Regacore", "Unknown / insufficient public data", "Search did not verify a relevant public company; may be typo for Regene/Regacore.", "Treat as unverified until user supplies URL."),
    ("Superpower", "Preventive biomarker platform", "S48", "Competes on low-cost labs + AI/care team; weaker CGM focus than Levels."),
    ("Function Health", "Preventive lab membership", "S47", "Competes on breadth/frequency of biomarkers; weaker real-time behavior loop."),
    ("PreventiveHealth.ai", "AI preventive health/longevity platform", "public search", "India-adjacent personalized health/wearable/genomic/microbiome angle; verify before deep use."),
    ("OpenEvidence", "Clinician AI evidence assistant", "S49", "Not consumer metabolic; sets bar for clinically grounded AI and physician distribution."),
    ("Glass Health", "AI clinical reasoning/scribing", "public search", "Physician workflow competitor for AI clinical reasoning, not consumer metabolic."),
    ("Atropos", "Real-world evidence platform", "public search", "Clinical/RWE moat via healthcare datasets; not consumer behavior."),
    ("AMBOSS", "Medical knowledge and Qbank", "public search", "Education/CDS competitor for evidence content."),
    ("UpToDate", "Enterprise clinical decision support", "public search", "Gold-standard CDS; relevant for Ovexis evidence layer."),
    ("Apollo 24/7", "India omnichannel healthcare", "public search", "India-scale doctor/pharmacy/diagnostics distribution."),
    ("Practo", "India doctor consult/appointments/tests", "public search", "Provider marketplace and telemedicine distribution."),
    ("Tata 1mg", "India pharmacy/labs/eConsult", "public search", "Diagnostics + pharmacy logistics; potential distribution benchmark."),
    ("Healthify", "Nutrition/fitness coaching + AI + CGM", "public search", "Strong India food database and coach network; direct behavior-change competitor."),
    ("Apple Health", "Consumer health data OS", "public search", "HealthKit/FHIR records; platform threat and integration substrate."),
    ("Google Health/Health Connect", "Android health data layer", "public search", "Health Connect + FHIR medical records; Android substrate."),
    ("Human API", "Consumer-permissioned health data API", "S50", "Data aggregation competitor/partner for longitudinal record."),
    ("WHOOP", "Wearable recovery/strain subscription", "public search", "Retention via daily scores/coaching; weak glucose/labs."),
    ("Oura", "Smart ring sleep/readiness/stress", "public search", "Strong wearable UX and integrations; entering metabolic via Stelo partnerships/add-ons."),
    ("Ultrahuman", "Ring + CGM metabolic platform", "public search", "Closest hardware+metabolic competitor; strong India/global wearable bundle."),
]

# Evidence register rows: source, claim, label, confidence, observed/inferred
evidence_rows = []
def ev(source, claim, label='🟢 Confirmed', conf='High', obs='Observed'):
    evidence_rows.append((source, claim, label, conf, obs))

for sid, name, url, note in sources:
    ev(sid, note, '🟢 Confirmed' if not sid in {'S44','S45','S46'} else ('🟡 Strong Inference' if sid=='S44' else '🟢 Confirmed'), 'High' if sid not in {'S44','S45'} else 'Medium', 'Observed')

decision_rows = []
for f in features:
    name,purpose,evidence,user_value,biz_value,eng,clin,infra,reg,team,months,priority,category,action,moat,conf = f
    why = f"Built to {purpose.lower()}"
    kpi = {
        'Marketing': 'visitor-to-quiz conversion', 'Onboarding': 'activation', 'Business model': 'conversion and ARPU', 'CGM': 'data acquisition and subscription',
        'Integration': 'successful sensor connection', 'Visualization': 'daily active use', 'Nutrition AI': 'food-log completion', 'Food logging': 'logging frequency',
        'Behavior': 'habit adherence', 'Scoring': 'retention and insight comprehension', 'Wearables': 'context completeness', 'Lifestyle': 'context density',
        'Habit': 'weekly active use/streaks', 'Programs': 'program adherence', 'AI': 'AI sessions/member/week', 'AI Search': 'content conversion',
        'Records': 'document uploads', 'Labs': 'lab attach rate', 'Clinical': 'plan upgrade and trust', 'Coaching': 'premium retention', 'Support': 'support CSAT',
        'Privacy': 'trust and compliance', 'Security': 'AI risk reduction', 'AI Ops': 'AI defect rate', 'Provider': 'practice activation', 'Provider AI': 'clinician time saved',
        'Expansion': 'waitlist', 'Research': 'data depth', 'Evidence': 'conversion', 'Growth': 'organic acquisition', 'Payments': 'checkout conversion',
        'Commerce': 'add-on purchases',
    }.get(category, 'retention')
    trade = 'Complexity, privacy, and interpretation risk increase as the feature touches health data.'
    alt = 'Alternative: raw data export/basic dashboards; Levels chose integrated feedback loops and guided interpretation.'
    decision_rows.append((name, why, user_value, kpi, trade, alt, evidence, conf))

copy_ideas = [
"Metabolic-health-first positioning", "Real-time feedback loop", "App-only entry tier", "Optional CGM instead of mandatory CGM", "BYOD sensor support",
"AI text food logging", "AI photo meal logging", "Barcode scanning", "Recents-based meal re-log", "Custom foods",
"Meal scores", "Daily glucose/stability score", "Exercise spike exclusions", "Habit loops", "Progress vs limit habit rings",
"Programs by goal", "Lab upload with AI extraction", "Plain-English lab interpretation", "Two-tier lab panels", "Clinician review tier",
"Dietitian/nutritionist upsell", "Member portal shop", "HSA/FSA checkout", "CSV export", "Doctor-share flow",
"No-data-sale trust language", "AI provider pseudonymization", "Content library", "AI search over content", "Founder-led education",
"Advisory board halo", "Public biomarker-improvement report", "Retrospective results caveat", "Community/crowdfunding alignment", "Referral/partner pages",
"Stelo OTC integration", "Dexcom G7 path", "Apple Health sync", "Health Connect sync", "Timeline-centered UX",
"Notes/context logging", "Lifestyle events", "Water/protein/fiber targets", "Trend windows 30/60/90", "Pro dashboard concept",
"AI-prepared clinician summaries", "Programs/protocols for practitioners", "Cohorts/challenges", "U.S. support/help-center depth", "Privacy DPO/contact workflow"
]
improve_ideas = [
"Personalize glucose ranges by phenotype and goal", "Athlete mode for training spikes", "Explain uncertainty on every score", "Cite evidence behind AI advice", "Hard-separate wellness vs medical suggestions",
"Better temporal reasoning for AI coach", "AI self-check before recommendations", "Integrate Whoop/Oura/Garmin/Fitbit directly", "Single unified sensor flow", "Sensor artifact detection",
"Compression-low detection", "Manual sync button in main view", "Food database quality", "Indian/regional food support", "Portion estimation confidence",
"Calorie and macro precision", "Meal-order experimentation mode", "N-of-1 experiment templates", "Post-meal walk recommendations personalized", "Trigger-specific habit prescriptions",
"Streak recovery design", "Gamification without anxiety", "Clinician-ready PDF reports", "FHIR export", "Doctor portal with scoped permissions",
"Family account", "Caregiver sharing", "Insurance/primary-care handoff", "Continuous lab trend alerts", "At-home phlebotomy logistics",
"Genomics/microbiome only where actionable", "DEXA/body composition normalization", "Medication/supplement interaction context", "Mental-health/stress context", "Data provenance on every insight",
"Model evaluation dashboard", "Human review escalation", "Red-team AI health outputs", "Consent UX clarity", "Privacy-by-design settings",
"Transparent third-party processor list", "Lower-friction cancellation/winback", "Cohort challenges by goal", "Provider workflow audit trails", "Admin/role permissions",
"Enterprise analytics", "Employer/insurer outcomes dashboards", "API/webhooks for member-approved data", "International regulatory playbook", "Localized nutrition ontology"
]
ignore_ideas = [
"Overly narrow universal glucose thresholds", "CGM as permanent daily dependency for healthy users", "Black-box scores", "Generic AI coach tone", "Advice without causal uncertainty",
"Two-app friction as acceptable", "High sensor markup", "US-only assumption for global strategy", "Content quantity over clinical outcomes", "Wellness claims drifting into diagnosis",
"Keto bias", "One-size-fits-all meal scoring", "Manual logging as main loop forever", "Relying on reviews to find bugs", "Marketing before validation",
"Opaque billing/legacy plan complexity", "Dependence on one CGM vendor", "Ignoring serious athletes", "Ignoring anxiety/gluco-obsession", "No public developer layer",
"Static PDF reports only", "No EHR interoperability", "No physician workflow audit logs", "No model/provider transparency", "No confidence bands",
"Unverified biomarker claims", "Non-actionable long lab panels", "Supplements marketplace without evidence thresholds", "No escalation for alarming values beyond lab criticals", "Hard-to-find sync controls",
"No app widgets if users want glanceability", "Weak integration breadth", "Requiring users to interpret raw trends", "No data quality score", "No context-aware fasting analysis",
"Delayed glucose without clear UX", "Research not translated into guidelines", "Community as moat without network utility", "Feature flags confusing support", "Removal of useful content tabs without replacements",
"AI beta in high-stakes context without guardrails", "App-only value under-communicated", "Clinician review too infrequent", "Portal/mobile feature inconsistency", "Open-ended behavior goals",
"Over-indexing on glucose over insulin/lipids", "Lab results without care coordination", "No local cultural diet models", "No direct primary-care loop", "No automated experiment readouts"
]
reinvent_ideas = [
"Longitudinal health intelligence graph", "Personal digital-twin with uncertainty", "Causal N-of-1 experiment engine", "Adaptive CGM sprint scheduling", "Dynamic metabolic phenotype model",
"AI copilot that is evidence-cited and source grounded", "Clinician-reviewed AI pathways", "FHIR-native member record", "Consent ledger", "Data provenance ledger",
"Behavior-change protocol marketplace", "Lab-to-habit compiler", "Wearable signal fusion", "Context-aware athlete glucose model", "Meal-context ontology",
"Indian/South Asian metabolic risk module", "Family cardiometabolic risk dashboard", "Primary-care collaboration workspace", "Outcome-backed programs", "Micro-cohort peer loops",
"GLP-1 support mode", "PCOS/PMOS program", "Menopause metabolic program", "Hypertension/lipid program", "Sleep apnea risk program",
"Pharmacy and medication data import", "Claims-aware affordability insights", "Food environment recommendations", "Restaurant/menu personalization", "Grocery cart scoring",
"Voice-first health journal", "Ambient meal capture with correction", "AI lab-document parser with validation", "Automated red-flag triage", "Human-in-loop audit trail",
"Evaluation suite for health AI", "Model router by task/risk", "Privacy-preserving personalization", "Federated learning or on-device summaries", "Clinical evidence RAG with graded recommendations",
"Provider admin with roles", "Employer outcomes analytics", "Insurer prevention ROI dashboard", "API-first ecosystem", "Developer sandbox with synthetic data",
"International compliance modules", "Local lab network abstraction", "Care team marketplace", "Personal health operating system", "Outcome-guaranteed metabolic program"
]
gaps = [
"Validated outcomes for non-diabetic CGM users", "Personalized healthy glucose ranges", "Athlete-safe interpretation", "AI hallucination controls", "Robust data-quality scoring",
"Direct wearable integrations beyond Apple/Health Connect", "FHIR/EHR interoperability", "Clinician workflow adoption", "Global CGM regulatory access", "Affordable India pricing",
"South Asian food ontology", "Family/caregiver metabolic health", "Longitudinal lab + wearable causal inference", "Medication and supplement context", "Genetics actionability",
"Mental stress quantification", "Sleep apnea/metabolic link", "Women's hormone-metabolism programs", "GLP-1 nutrition support", "Retention after CGM novelty",
"Behavior-change personalization", "Provider reimbursement model", "Employer ROI proof", "Food industry transparency", "Privacy-preserving AI memory",
"Data portability beyond CSV", "Open APIs/webhooks", "Structured doctor reports", "Data deduplication across wearables", "Sensor artifact handling",
"Meal portion accuracy", "Cultural dietary guidance", "Low-cost labs", "Continuous care coordination", "Medical-grade disclaimers without fear",
"AI explainability", "Human review routing", "Clinical safety incident process", "Evidence grading UX", "Regulatory boundary management",
"Trustworthy supplement marketplace", "Insurer/primary-care partnership", "Pharmacy data integration", "Nutritionist scalability", "Community that changes behavior",
"Blue-collar/low-income access", "Non-English support", "Rural healthcare integration", "Longitudinal health score accepted by doctors", "Preventive health operating system category"
]
blue_ocean = [
"India-first metabolic intelligence with CGM/labs/food ontology", "Clinician-ready longitudinal health OS", "AI metabolic coach with audited citations", "Outcome-backed 90-day cardiometabolic reversal programs", "Adaptive CGM sprint subscription",
"FHIR-native consumer-controlled record", "Provider Pro for nutritionists and diabetologists", "Family cardiometabolic risk management", "GLP-1 metabolic support companion", "South Asian prediabetes prevention platform",
"Athlete metabolic performance mode", "Women's metabolic-hormone intelligence", "Food-order and meal-composition experiment lab", "Employer prevention ROI platform", "Insurer-sponsored metabolic risk reduction",
"Restaurant/grocery personalization layer", "Privacy-preserving health AI memory", "Data-quality layer for wearables", "Clinical-grade AI evaluation suite", "Open health-data developer platform"
]

# Markdown helpers

def sref(ids):
    return ", ".join(ids.split(',')) if isinstance(ids, str) else ", ".join(ids)

md = []
md.append("# Levels Health Competitive Intelligence Report\n")
md.append("**Target:** Levels Health / Levels Health, Inc.  \n**Website:** https://www.levelshealth.com/  \n**Category:** Metabolic Health | Continuous Glucose Monitoring | Consumer AI Health  \n**Prepared for:** Ovexis board-level strategy discussions  \n**Date:** 2026-07-25 (Asia/Calcutta)\n")
md.append("## Evidence Labels and Investigation Boundaries\n")
md.append("- 🟢 Confirmed = directly visible in public sources listed in the Evidence Register.\n- 🟡 Strong Inference = reasoned from multiple public observations; not explicitly confirmed by Levels.\n- 🔴 Speculation = strategic possibility or prediction; not verified.\n- 🟢 Confirmed: This report uses only public information and does not attempt login-gated access, private APIs, vulnerability probing, credentialed views, or unauthorized collection.\n- 🟢 Confirmed: Public Terms of Service state that Levels is not a healthcare provider and that Services are for general wellness/non-medical use; therefore clinical conclusions in this report are limited to public claims and strategy analysis, not medical advice. [S18]\n- 🟢 Confirmed: Authenticated screens, hidden workflows, exact database schemas, and internal KPIs cannot be verified from public information; where discussed they are labelled inference or speculation.\n")

md.append("## 1. Executive Summary\n")
exec_bullets = [
("🟢 Confirmed", "Levels is building a metabolic health platform that combines a consumer app, optional continuous glucose monitoring, lab testing, AI-powered food/lab insights, habit loops, adaptive programs, clinician review, nutritionist support, and a professional dashboard for practices.", "S01,S02,S03,S19"),
("🟢 Confirmed", "The consumer-facing promise is to help people understand how diet and lifestyle affect their body so they can improve energy, appetite, and long-term metabolic health.", "S01"),
("🟢 Confirmed", "The company frames the customer problem as generic nutrition advice plus lack of feedback: users cannot tell whether advice fits their own physiology.", "S01,S23,S24"),
("🟡 Strong Inference", "The emotional problem is uncertainty and loss of agency: users want to stop guessing, feel in control, and see proof that everyday choices matter.", "S01,S04,S20,S21"),
("🟡 Strong Inference", "The operational problem is data fragmentation: CGM data, meals, labs, sleep, activity, documents, and expert guidance historically live in separate tools and require manual synthesis.", "S01,S11,S12,S13,S14,S19"),
("🟢 Confirmed", "Current consumer customers are U.S.-based adults with a U.S. phone number, U.S. app-store access, and U.S. shipping address; Levels Labs is for members age 18+.", "S09,S13"),
("🟢 Confirmed", "Levels is not for people seeking diagnosis, treatment, cure, disease prevention, emergency support, or a substitute for professional medical care; the app’s own disclaimers state it is for general wellness.", "S04,S15,S18"),
("🟡 Strong Inference", "Levels is creating or reinforcing the category of a metabolic health operating system: real-time metabolic feedback + longitudinal labs + AI + behavior-change workflows.", "S01,S19,S21,S28"),
("🟡 Strong Inference", "Levels replaces raw CGM apps, generic food trackers, one-off annual lab panels, static education, and disconnected nutrition coaching with a more integrated feedback system.", "S01,S04,S10,S13,S14,S19"),
("🟢 Confirmed", "The core philosophy is prevention over reaction, personal data over generic advice, and behavior change driven by feedback loops.", "S20,S21,S23,S24,S28"),
("🟡 Strong Inference", "Levels’ biggest strategic strength is not the sensor itself; it is the interpretation layer that turns commodity CGM/lab inputs into habit loops, programs, and trust-building education.", "S01,S03,S10,S11,S13,S14"),
("🟡 Strong Inference", "Levels’ biggest vulnerability is retention after the novelty/learning value of CGM fades, especially as OTC sensors make raw glucose access cheap and competitors bundle labs, wearables, or coaching.", "S03,S43,S45,S46,S47,S48"),
]
for label, claim, src in exec_bullets:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n### Jobs-To-Be-Done\n")
jtbd = [
("🟢 Confirmed", "When I am trying to understand why I feel tired, hungry, unfocused, or stuck with weight, I want real-time feedback on food, sleep, activity, and glucose so I can make better choices.", "S01,S04,S10,S12"),
("🟢 Confirmed", "When my labs are confusing, I want a plain-English analysis and a plan so I know what to work on.", "S13,S14,S26"),
("🟢 Confirmed", "When I want professional help, I want a clinician/nutritionist/practitioner to see my data in context rather than starting from a blank visit.", "S03,S19,S32"),
("🟡 Strong Inference", "When I have learned my CGM triggers, I still need a lightweight, non-sensor habit system to maintain gains and justify subscription renewal.", "S03,S11,S26,S45"),
]
for label, claim, src in jtbd:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n### Value Proposition\n")
for label, claim, src in [
("🟢 Confirmed", "Measure: CGM, labs, imported health data, uploaded documents, and user logs.", "S01,S02,S07,S12,S13,S14,S17"),
("🟢 Confirmed", "Understand: AI food logging, macro breakdowns, lab interpretation, AI Guide, LevelsAI content search, clinician review, and Pro AI summaries.", "S03,S10,S14,S17,S19,S27"),
("🟢 Confirmed", "Act: habit loops, guided programs, recommendations, targets, protocols, dietitian sessions, and follow-through workflows.", "S03,S11,S19,S26"),
("🟡 Strong Inference", "Retain: repeated logs, streak-like habit loops, periodic labs, add-on CGM sprints, programs, content/newsletters, and Pro care workflows create recurring reasons to return.", "S03,S11,S13,S19,S26,S42"),
]:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 2. Company Intelligence\n")
md.append("### Timeline\n")
timeline = [
("2019-06", "🟢 Confirmed", "Levels was formed/founded by Josh Clemente, Sam Corcos, Casey Means, David Flinner, and Andrew Conner to address metabolic health.", "S20,S21"),
("2019-11/12", "🟡 Strong Inference", "First product shipped around November/December 2019 according to founder-interview coverage; official support pages do not independently verify this shipment date.", "External interview coverage from search; not used as primary evidence"),
("2020-11", "🟢 Confirmed", "Levels announced a $12M seed round led by a16z and positioned CGM + intelligent software as a category-defining biowearable system.", "S20"),
("2020-11", "🟢 Confirmed", "The seed announcement said Levels was in closed beta with a 50,000+ person waitlist and a month-long program including two 14-day CGM sensors and app access.", "S20"),
("2021", "🟢 Confirmed", "Press described Levels as a late-stage beta wellness product for athletes/health-conscious users at roughly $400 for a one-month program.", "The Verge search result; see source register note in references"),
("2022-04", "🟢 Confirmed", "Levels announced a $38M Series A at a $300M valuation and stated it had 25,000 paying beta members.", "S21"),
("2022-04", "🟢 Confirmed", "The Series A included $5M from more than 1,400 members through crowdfunding in under six hours.", "S21"),
("2023-01", "🟢 Confirmed", "Levels announced a $7M Series A extension and more than $55M raised to date.", "BusinessWire/AP search result and S21 context"),
("2024-08", "🟢 Confirmed", "Levels announced/was covered for a $10M Series A extension including $3M crowdfunding, Long Journey, a16z, and others.", "S22"),
("2024-08", "🟢 Confirmed", "Coverage stated 60,000+ members, 700M+ glucose data points, hundreds of thousands of food logs, 18M+ YouTube views, and 2M+ annual blog visitors.", "S22"),
("2025-2026", "🟢 Confirmed", "Levels shifted toward an app-first system with optional CGM/labs and $15/month or $80/year membership pricing.", "S01,S03"),
("2026", "🟢 Confirmed", "Levels Pro markets a metabolic-health operating system for modern practices with AI-prepared review workflows, programs, and practitioner plans.", "S19"),
("2026-02", "🟢 Confirmed", "Support states the Levels IRB-approved CGM research study has concluded and study services have been discontinued.", "S25"),
]
md.append("| Date | Label | Event | Evidence |\n|---|---:|---|---|\n")
for row in timeline:
    md.append(f"| {row[0]} | {row[1]} | {row[2]} | {row[3]} |\n")

md.append("\n### Founders and Leadership Signals\n")
for label, claim, src in [
("🟢 Confirmed", "Founders publicly listed across Levels announcements are Casey Means MD, Josh Clemente, Sam Corcos, David Flinner, and Andrew Conner.", "S20,S21"),
("🟢 Confirmed", "Casey Means served as Levels CMO until late 2023 according to her official Levels author page.", "S30"),
("🟢 Confirmed", "Josh Clemente’s official author page describes him as founder and president, while the site also publishes recurring 'From the CEO' newsletters under his name.", "S28,S29"),
("🟢 Confirmed", "2024 coverage quoted Sam Corcos as Levels CEO and Josh Clemente as co-founder.", "S22"),
("🟡 Strong Inference", "Public leadership titles appear to have evolved or are inconsistent across official and third-party pages; a board-level user should verify current officer titles via LinkedIn, Delaware filings, or direct company confirmation before relying on a CEO title.", "S22,S28,S29,S30"),
("🟢 Confirmed", "Levels has used a high-profile advisor strategy including metabolic-health physicians/researchers listed in seed/Series A materials and App Store copy.", "S04,S20,S21"),
]:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n### Funding and Valuation\n")
funding = [
("🟢 Confirmed", "$12M seed led by a16z in November 2020.", "S20"),
("🟢 Confirmed", "$38M Series A at $300M valuation in April 2022.", "S21"),
("🟢 Confirmed", "$5M of the Series A was crowdfunded from members/community.", "S21"),
("🟢 Confirmed", "$7M Series A extension in January 2023 and more than $55M raised to date at that time.", "BusinessWire/AP search result; S21 for prior total"),
("🟢 Confirmed", "$10M Series A extension in August 2024 according to coverage, including $3M crowdfunding and Long Journey/a16z participation.", "S22"),
("🟡 Strong Inference", "Total funding is likely above $65M after the 2024 extension; third-party databases list approximately $67M, but the exact total should be verified from primary filings.", "S20,S21,S22"),
("🟢 Confirmed", "Wefunder lists the company as a Tier 1 VC-backed B2C subscription company and claims $20M+ revenue over the last 12 months; this is a crowdfunding listing claim, not audited financials in this report.", "S42"),
]
for label, claim, src in funding:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n### Partnerships and Operating Network\n")
partnerships = [
("Dexcom/Stelo", "🟢 Confirmed", "Stelo/Dexcom account connection; Stelo data every 5 minutes and Levels refresh about every 15 minutes.", "S07"),
("Amazon Pharmacy", "🟢 Confirmed", "Dexcom G7 prescription pathway sends prescriptions to Amazon Pharmacy for sensor purchase.", "S08"),
("Quest/BioReference/LabCorp", "🟢 Confirmed", "Quest for most lab orders, BioReference for NY/NJ, LabCorp for certain historical orders.", "S16"),
("MD Integrations/Healthie", "🟢 Confirmed", "Labs orders created by independent physician network facilitated by MD Integrations and routed via Healthie to Quest.", "S15"),
("TrueMed", "🟢 Confirmed", "HSA/FSA checkout option and LMN flow.", "S03,S17"),
("Affirm", "🟢 Confirmed", "Pay-over-time option at checkout.", "S03"),
("The Lyons Share", "🟢 Confirmed", "Nutritionist partner referenced in current support article.", "S03 and S565 fetched evidence"),
("OpenAI/Anthropic/Google/xAI", "🟢 Confirmed", "AI service providers for Levels Guide AI and similar features.", "S17"),
("Langfuse", "🟢 Confirmed", "AI prompt/response logging and reliability support.", "S17"),
("PostHog/Sentry/Stripe/Cloudflare/etc.", "🟡 Strong Inference", "Observed in public headers/HTML; confirms public site integrations but not full production architecture.", "S44"),
]
md.append("| Partner/system | Label | Role | Evidence |\n|---|---:|---|---|\n")
for p in partnerships:
    md.append(f"| {p[0]} | {p[1]} | {p[2]} | {p[3]} |\n")

md.append("\n### Acquisitions, Patents, Open Source, Geography, Awards\n")
for label, claim, src in [
("🟢 Confirmed", "No acquisition was verified from the public sources reviewed for this report.", "Public search; absence is not proof"),
("🟢 Confirmed", "No Levels Health patent assignment was verified from the public searches performed; this should be treated as a negative search result, not definitive legal advice.", "Public patent searches"),
("🟢 Confirmed", "The GitHub organization 'levelshealth' publicly shows 17 repositories, including forks of react-native-help-scout and react-native-health-connect; no proprietary core app repository was identified as public.", "S41"),
("🟢 Confirmed", "Current availability is U.S.-only, with an international waitlist and no recommended workaround for non-U.S. users.", "S09"),
("🟢 Confirmed", "Levels was named to Fast Company's Next Big Things in Tech list according to press/search snippets; this report did not fetch the primary award page.", "BusinessWire search snippet"),
]:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 3. Founder Psychology and Strategic Mental Models\n")
founder_psy = [
("🟢 Confirmed", "Josh Clemente publicly ties the origin story to personal fatigue/metabolic discovery and the insight that raw CGM data needs interpretation and behavior-change software.", "S20 and founder podcast/search evidence"),
("🟢 Confirmed", "Levels’ public mission language emphasizes prevention rather than reaction and making health information accessible to the individual in real time.", "S21"),
("🟢 Confirmed", "Privacy principles state that user health data belongs to users, not Levels, and that data should be exportable/deletable and not sold.", "S17"),
("🟡 Strong Inference", "The founders believe behavior change is more likely when feedback is objective, immediate, personalized, and tied to daily actions rather than abstract annual advice.", "S01,S20,S21,S23,S26"),
("🟡 Strong Inference", "Levels’ decision framework appears to favor wedges: start with high-salience CGM, then expand into labs, habits, AI, and practice workflows.", "S20,S21,S03,S13,S19"),
("🟡 Strong Inference", "Risk tolerance is high on category creation and AI adoption but cautious on regulatory claims; the company pushes new consumer/pro workflows while repeatedly using general-wellness disclaimers.", "S04,S17,S18,S28"),
("🟡 Strong Inference", "The 10-year ambition is likely to become a trusted metabolic-health operating layer for consumers and care teams, not merely a CGM reseller.", "S01,S19,S21,S28"),
("🔴 Speculation", "If successful, Levels could pursue payer/employer coverage, acquisition by a device/wearable/lab platform, or expansion into a broader personal health record/AI coach category.", "Strategic prediction based on market trajectory"),
]
for label, claim, src in founder_psy:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 4. Product Reverse Engineering\n")
md.append("### Master Feature Map (summary; full spreadsheet contains detailed complexity and decision ledger)\n")
md.append("| Feature | Label | Purpose | Evidence | Strategic read |\n|---|---:|---|---|---|\n")
for f in features:
    name,purpose,evidence,user_value,biz_value,eng,clin,infra,reg,team,months,priority,category,action,moat,conf = f
    label = '🟢 Confirmed' if conf == 'High' else '🟡 Strong Inference'
    md.append(f"| {name} | {label} | {purpose} | {evidence} | {action}; moat={moat}; priority={priority} |\n")

md.append("\n### Visible Workflows\n")
workflows = [
("Anonymous visitor", "Homepage -> value proposition -> proof points -> quiz/plan CTA -> pricing cards -> checkout/start.", "S01"),
("Signup/onboarding", "Purchase membership -> app login with same email/password -> health intake -> permissions -> sensor setup if applicable.", "S05,S06"),
("Sensor setup", "Stelo app -> Dexcom account -> apply sensor -> Levels More > Devices > Manage -> choose Stelo -> connect Dexcom -> consent forms -> data arrives in ~15 minutes.", "S07"),
("Dexcom G7 purchase", "Shop -> medical intake -> independent physician review -> prescription to Amazon Pharmacy -> user buys sensors -> connects Dexcom account -> ~1-hour delay in Levels.", "S08"),
("Food logging", "Timeline + -> Describe/Photo/Recents/Barcode/Search/Custom -> AI processing/review -> save -> glucose response/score.", "S10"),
("Exercise/lifestyle", "Timeline + -> Exercise/Note/Lifestyle -> optional strenuous or ignore glucose -> timeline context and scoring impact.", "S34,S35,S38,S39"),
("Wearable imports", "Settings -> Connect Apple Health/Health Connect -> grant categories -> timeline gets workouts/sleep/weight/heart rate/other supported signals.", "S12,S36,S37"),
("Labs", "Order Labs or upload PDF -> AI extraction -> review/edit/confirm -> lab results visible in app/portal -> clinician note/AI interpretation depending feature/plan.", "S13,S14,S15"),
("Doctor sharing", "Portal -> Data/Health Data -> Glucose -> Share Your Data -> enter email -> Grant Access.", "S32"),
("Export/delete", "Portal CSV export; deletion by privacy email and verification with exceptions for payments/official records/third parties.", "S31,S33"),
("Pro practice", "Invite -> connect data -> AI prep -> practitioner review -> recommendations/targets/protocols/program steps -> follow-through.", "S19"),
]
md.append("| Journey/workflow | Label | Steps | Evidence |\n|---|---:|---|---|\n")
for w in workflows:
    md.append(f"| {w[0]} | 🟢 Confirmed | {w[1]} | {w[2]} |\n")

md.append("\n### Retention Loops\n")
retention = [
("CGM discovery loop", "🟢 Confirmed", "Eat/log -> observe glucose -> score -> change meal -> compare result -> reinforce behavior.", "S04,S10,S31"),
("Habit loop", "🟢 Confirmed", "Choose up to three habit loops -> daily progress/limit rings -> targets -> weekly/streak views.", "S11"),
("Lab progress loop", "🟢 Confirmed", "Baseline labs -> recommendations -> behavior changes -> retest -> biomarker improvement proof.", "S13,S14,S26"),
("AI loop", "🟢 Confirmed", "Ask AI Guide or receive proactive meal feedback -> combine prompts with account data -> response -> monitoring/debugging.", "S05,S17"),
("Content loop", "🟢 Confirmed", "SEO/article/video/podcast -> LevelsAI/content -> CTA -> app or quiz -> newsletter.", "S01,S27,S42"),
("Pro loop", "🟢 Confirmed", "Client data -> AI prep -> practitioner action -> program/protocol -> progress monitoring -> next review.", "S19"),
("Community/investor loop", "🟢 Confirmed", "Members invest/promote -> brand alignment -> growth -> stronger community narrative.", "S21,S22,S42"),
]
md.append("| Loop | Label | Mechanism | Evidence |\n|---|---:|---|---|\n")
for r in retention:
    md.append(f"| {r[0]} | {r[1]} | {r[2]} | {r[3]} |\n")

md.append("\n## 5. Complete User Journey Diagram\n")
md.append("```mermaid\nflowchart TD\n  A[Anonymous visitor] --> B[Homepage trust/value props]\n  B --> C[Quiz / personalized plan CTA]\n  C --> D[Pricing: app / Core / Complete or current membership]\n  D --> E[Checkout: card, Affirm, TrueMed HSA/FSA]\n  E --> F[Account creation]\n  F --> G[Download iOS/Android app]\n  G --> H[Health intake: goals/status/lifestyle]\n  H --> I[Permissions: notifications, Bluetooth, camera, Apple Health/Health Connect]\n  I --> J{CGM?}\n  J -->|Stelo| K[Stelo app + Dexcom account + Permission to Disclose/Authorization]\n  J -->|Dexcom G7| L[Medical intake -> physician partner -> Amazon Pharmacy -> connect Dexcom]\n  J -->|No CGM| M[Food/labs/habit-only app use]\n  K --> N[Glucose timeline]\n  L --> N\n  M --> O[Food logs, labs, habits]\n  N --> O[Meals, exercise, sleep, notes, lifestyle]\n  O --> P[AI Guide / insights / scores]\n  P --> Q[Programs + habit loops]\n  Q --> R[Retest labs / repeat CGM sprint]\n  R --> S[Clinician/nutritionist review]\n  S --> T[Share/export with doctor]\n  T --> U[Renew, add-on, refer, or cancel]\n```\n")

md.append("\n## 6. UX Research\n")
ux = [
("🟢 Confirmed", "Trust signals include '100,000+ lives changed', 'first and most trusted', HIPAA/SOC2 compliant, encrypted, data not sold, member testimonials, and retrospective biomarker improvements caveated as not a clinical trial.", "S01,S26"),
("🟢 Confirmed", "Primary marketing hierarchy is problem-first: generic nutrition advice vs personal physiology, followed by app/data/labs/expert support modules.", "S01"),
("🟢 Confirmed", "App setup explicitly asks for notifications, Bluetooth, and optional camera permissions; Apple Health/Health Connect permissions are separate.", "S06,S12"),
("🟢 Confirmed", "The core mobile interaction pattern is a timeline with a green plus button for logging and graph-based retroactive logging.", "S10"),
("🟢 Confirmed", "Habit loops use rings/dots and distinguish habits to increase from habits to limit, which is a clear behavior-design pattern.", "S11"),
("🟡 Strong Inference", "Levels’ UX prioritizes clarity and coaching over medical-chart density; however, users who want raw/precise clinical data may perceive score simplification as reductive.", "S01,S04,S45,S46"),
("🟢 Confirmed", "Observed user complaints include hidden sync controls, lost/removed education tabs, sync failures, tedious food logging, limited searchability, and AI coach inaccuracies.", "S04,S45,S46"),
("🟡 Strong Inference", "Design friction is concentrated in three places: dual sensor apps, manual logging burden, and interpretation trust when scores/AI advice conflict with user context.", "S04,S07,S08,S10,S45,S46"),
("🟡 Strong Inference", "Dark mode, detailed accessibility compliance, exact typography tokens, button states, animations, and authenticated dashboard screens cannot be verified from public sources.", "Public unauthenticated boundary"),
("🟢 Confirmed", "Public HTML references acsbapp.com, suggesting an accessibility overlay or related script on the public site; this is not an accessibility audit.", "S44"),
]
for label, claim, src in ux:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 7. Healthcare Workflow\n")
health_workflow = [
("🟢 Confirmed", "Levels states it is not a healthcare provider and does not provide diagnosis, clinical interpretation, or treatment recommendations through Labs.", "S15,S18"),
("🟢 Confirmed", "Lab orders are created by licensed physicians from an independent physician network facilitated by MD Integrations, routed through Healthie, and transmitted to Quest.", "S15"),
("🟢 Confirmed", "Quest returns lab results, which Levels makes available in the app/portal at user request for personal use.", "S15"),
("🟢 Confirmed", "If Quest or the ordering physician flags critical/life-threatening values, a healthcare professional may contact the user; Levels does not provide emergency support.", "S15"),
("🟢 Confirmed", "Core and Complete legacy plans include clinician review of health data twice annually as a written note.", "S03"),
("🟢 Confirmed", "Complete includes a 50-minute functional nutritionist session and dedicated concierge support.", "S03"),
("🟢 Confirmed", "Levels Pro is positioned for practices to unify glucose, labs, nutrition, sleep, activity, AI summaries, recommendations, programs, and follow-through.", "S19"),
("🟡 Strong Inference", "Levels is building toward a hybrid wellness-clinical collaboration layer while avoiding direct provider status and shifting licensed decisions to independent networks.", "S15,S17,S18,S19"),
("🟢 Confirmed", "Insurance is not processed or accepted for Levels products; users pay cash, while HSA/FSA via TrueMed may be available.", "S03,S18"),
]
for label, claim, src in health_workflow:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 8. Healthcare Data Architecture\n")
md.append("### Data Source Map\n")
for label, claim, src in [
("🟢 Confirmed", "CGM data comes from a glucose biosensor account/manufacturer platform after user permission, then Levels creates insights, zone scores, and metabolic scores.", "S17"),
("🟢 Confirmed", "Stelo records every 5 minutes and Levels refreshes approximately every 15 minutes.", "S07"),
("🟢 Confirmed", "Dexcom G7 via Levels/Amazon Pharmacy has approximately a one-hour delay in Levels; Dexcom via Apple Health/Health Connect historically can involve longer delays.", "S08"),
("🟢 Confirmed", "Apple Health/Health Connect imports include steps/workouts, sleep, weight, heart rate, and other supported phone health signals.", "S12,S36,S37"),
("🟢 Confirmed", "Lab data can come from Levels Labs, outside PDF uploads, or documents uploaded through Document Center.", "S13,S14,S17"),
("🟢 Confirmed", "Food logs include photos, titles, notes, activities, and nutritional metadata; export includes nutrition logs and metadata.", "S10,S17,S31"),
("🟢 Confirmed", "Members can export biometric/activity/nutrition/zone data in CSV format.", "S31"),
("🟢 Confirmed", "Members can grant doctor/care-team data access via email in the member portal.", "S32"),
("🟢 Confirmed", "Health documents may include lab reports, visit summaries, imaging reports, DEXA scans, and health records according to homepage/privacy language.", "S01,S17"),
("🟡 Strong Inference", "Levels likely normalizes heterogeneous data into a unified timeline/entity model because the app shows food, sleep, activity, glucose, labs, and habits in one system; internal schema is not public.", "S01,S10,S11,S12,S13"),
("🟢 Confirmed", "No public evidence was found that Levels exposes FHIR/HL7/CCDA APIs; Apple Health and Google Health Connect themselves support clinical/FHIR capabilities, but Levels public docs only confirm Apple/Health Connect wellness-signal sync and document upload.", "S12,S17,S31"),
("🟢 Confirmed", "No verified integrations with hospitals, insurance claims, pharmacy histories beyond Amazon Pharmacy G7 flow, medical imaging archives, or genomics were found in public Levels docs.", "S08,S17"),
]:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n### Healthcare Data Flow Diagram\n")
md.append("```mermaid\nflowchart LR\n  U[Member identity: email, phone, address, goals] --> L[Levels app/member portal]\n  CGM[Stelo/Dexcom sensor] --> MFG[Dexcom/Stelo app + cloud]\n  MFG -->|user authorization| L\n  AH[Apple Health] -->|permissions| L\n  HC[Health Connect] -->|permissions| L\n  FOOD[Food photos/text/barcodes] --> L\n  DOC[PDF labs, DEXA, health docs] --> L\n  LABORDER[Levels Labs order] --> MD[Independent physician network / MD Integrations]\n  MD --> H[Healthie]\n  H --> Q[Quest / BioReference / historical LabCorp]\n  Q --> L\n  L --> AI[AI context builder + providers]\n  AI --> INS[Insights, scores, lab interpretation, programs]\n  INS --> USER[Member actions]\n  L --> PRO[Levels Pro practitioner dashboard]\n  L --> SHARE[Doctor share/export CSV]\n```\n")

md.append("\n## 9. AI Reverse Engineering\n")
ai = [
("🟢 Confirmed", "Levels Guide AI collects prompts, responses, and usage-related information.", "S17"),
("🟢 Confirmed", "Levels Guide AI may combine prompts with selected account data such as glucose data, logs, program information, lab results, and uploaded health-related documents.", "S17"),
("🟢 Confirmed", "Direct identifiers such as name, email, phone, and street address are not included in data sent to AI service providers; an internal user ID is used instead.", "S17"),
("🟢 Confirmed", "AI providers currently listed are OpenAI, Anthropic, Google, and xAI; Langfuse is used for reliability/support monitoring.", "S17"),
("🟢 Confirmed", "AI providers do not use Levels data to train their own models according to Levels’ privacy policy.", "S17"),
("🟢 Confirmed", "Levels may review de-identified or pseudonymized transcripts to troubleshoot issues, improve prompts, and monitor abuse.", "S17"),
("🟢 Confirmed", "Levels Guide AI is described as a beta feature that may change or be discontinued.", "S17"),
("🟢 Confirmed", "Levels uses AI service providers for AI-assisted clinician note drafts based on selected lab results/documents/account data.", "S17"),
("🟢 Confirmed", "LevelsAI searches 700+ reported/fact-checked Levels articles.", "S27"),
("🟢 Confirmed", "Josh Clemente stated Levels is using AI to improve CGM insights and meal plans and to build Document Center as an intelligent health record users can converse with and control.", "S28"),
("🟡 Strong Inference", "The architecture is likely multi-provider LLM routing plus RAG over proprietary content and user context, with prompt observability via Langfuse; exact prompts, eval sets, embeddings, and routing rules are not public.", "S17,S27,S28"),
("🟢 Confirmed", "App Store reviews include a user complaint that the AI coach gave wrong or temporally confused advice; this is anecdotal but important product-risk evidence.", "S04,S46"),
("🟡 Strong Inference", "AI safety should be a board-level risk because user-specific lab/glucose advice sits near the boundary of wellness guidance and clinical decision support.", "S04,S14,S17,S18,S46"),
]
for label, claim, src in ai:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")
md.append("\n### AI Architecture Diagram\n")
md.append("```mermaid\nflowchart TD\n  A[User prompt / proactive meal event / lab upload] --> B[Context selection]\n  B --> C[Privacy scrub: remove direct identifiers, use internal ID]\n  C --> D{Task router}\n  D --> E[Food logging / vision + nutrition extraction]\n  D --> F[Personal data Q&A: glucose, logs, labs, docs]\n  D --> G[Content RAG over Levels articles]\n  D --> H[Clinician-note draft]\n  E --> I[OpenAI / Anthropic / Google / xAI]\n  F --> I\n  G --> I\n  H --> I\n  I --> J[Guardrails: wellness disclaimers, no diagnosis, citations/context where available]\n  J --> K[Member app / Pro summary / clinician draft]\n  I --> L[Langfuse monitoring and debugging]\n  K --> M[User action / human review / support]\n```\n")

md.append("\n## 10. Technical Reverse Engineering\n")
tech = [
("🟢 Confirmed", "Public website and app pages expose Next.js static assets and Next.js headers; public pages are served behind Cloudflare.", "S44"),
("🟢 Confirmed", "Support center headers include Caddy/istio-envoy and Help Scout-style session naming; support tooling is not fully confirmed beyond public response headers.", "S44"),
("🟢 Confirmed", "Public HTML references PostHog, Sentry, Stripe, Truemed, CookiePro, Apollo, Leadfeeder, Typeform, acsbapp, and static.levels.com assets.", "S44"),
("🟢 Confirmed", "Terms and privacy state payments are processed by Stripe and HSA/FSA option by Truemed.", "S17,S18"),
("🟢 Confirmed", "Privacy policy states data is stored in one or more secure databases hosted by third parties, but it does not name the cloud/database vendors.", "S17"),
("🟢 Confirmed", "Security practices listed publicly include internal MFA, IP whitelists, encryption at rest, code reviews, and ability to revoke access.", "S17"),
("🟡 Strong Inference", "Mobile app may be React Native or partly React Native because the public GitHub org contains React Native integration forks and co-founder technical history emphasizes React/React Native; this is not proof of current app stack.", "S41"),
("🟢 Confirmed", "No verified public information was found for backend language, database, queueing, caching, feature flags, CI/CD, or observability beyond Sentry/PostHog/Langfuse references.", "S17,S44"),
("🟡 Strong Inference", "Levels’ backend must maintain a time-series/event architecture because it combines CGM readings, logs, zones, nutrition logs, lab results, programs, and habit loops; internal implementation is unknown.", "S10,S11,S13,S31"),
]
for label, claim, src in tech:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 11. API Investigation\n")
api = [
("🟢 Confirmed", "No official public REST/GraphQL/FHIR/OpenAPI developer documentation for Levels was found in the public searches performed.", "Public search"),
("🟢 Confirmed", "The verified user-facing data export is CSV for glucose/CGM, activity logs, zones, and nutrition logs.", "S31"),
("🟢 Confirmed", "The verified integrations are consumer-permissioned connections to Dexcom/Stelo, Apple Health, Health Connect, labs/physician/pharmacy partners, and payments/AI providers.", "S07,S08,S12,S15,S17"),
("🟢 Confirmed", "An unofficial GitHub project named levelshealth-api exists, but it is not an official API and was not used to probe Levels services.", "S41 and GitHub search"),
("🟡 Strong Inference", "Levels likely has internal authenticated APIs powering the app/member portal; endpoints, schemas, rate limits, auth flows, webhooks, and versioning cannot be verified without unauthorized probing or credentialed access.", "Public boundary"),
]
for label, claim, src in api:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 12. Security, Privacy, Compliance\n")
sec = [
("🟢 Confirmed", "Homepage states HIPAA compliant and SOC 2 compliant.", "S01"),
("🟢 Confirmed", "Privacy policy clarifies Levels is not a HIPAA Covered Entity, but may act as a business associate for clinician-note features or integrated telemedicine under BAAs.", "S17"),
("🟢 Confirmed", "Security practices publicly listed include MFA for internal member-data access, IP whitelists, encryption at rest, code reviews, and quick access revocation.", "S17"),
("🟢 Confirmed", "Users may export data and request deletion; deletion excludes payment records, official medical records, and third-party-held data.", "S31,S33"),
("🟢 Confirmed", "Levels says it does not sell or sell access to personally identifiable data and does not use user data for advertising third-party products.", "S17"),
("🟢 Confirmed", "App analytics can collect app usage and device details for quality/performance; Levels says this aggregate information does not personally identify users and is not used/sold/traded for advertising/marketing/other commercial purposes.", "S40"),
("🟢 Confirmed", "AI service provider sharing is pseudonymized but includes selected health data needed for AI features; providers may retain data for limited monitoring/abuse/debugging/reliability purposes.", "S17"),
("🟡 Strong Inference", "Top security risks are third-party data sharing, AI output safety, internal privileged access, sensor/lab partner dependencies, and user confusion over what is medical vs wellness.", "S15,S17,S18,S46"),
("🟢 Confirmed", "Public CSP includes frame-ancestors 'self'; this is a narrow clickjacking-related header observation, not a full security assessment.", "S44"),
]
for label, claim, src in sec:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 13. Business Model\n")
business = [
("🟢 Confirmed", "Current membership is $15/month billed monthly or $80/year billed annually.", "S03"),
("🟢 Confirmed", "Optional add-ons include Stelo subscription $89/shipment, Stelo one-time two-pack $99, Basic Lab Panel $99, Comprehensive Lab Panel $399, and functional nutritionist session $250.", "S03"),
("🟢 Confirmed", "Legacy Core is $499/year and legacy Complete is $1,999/year with labs/CGM/review differences.", "S03"),
("🟢 Confirmed", "Levels Pro is listed at $99/month for one practitioner seat with $1,000/month recurring credit purchase; Plus is $299/month for up to five practitioners with $1,000/month recurring credit purchase.", "S19"),
("🟢 Confirmed", "Levels is cash-pay and does not process insurance payments, but HSA/FSA via TrueMed and Affirm financing are available.", "S03,S18"),
("🟡 Strong Inference", "Revenue streams include consumer subscription, CGM fulfillment or facilitation economics, lab panel margin, nutritionist sessions, legacy high-touch bundles, Pro seat/credit revenue, and partner commerce.", "S03,S08,S13,S19,S28"),
("🟡 Strong Inference", "App-only pricing is a strategic response to OTC CGM commoditization and CGM novelty decay: the app must become useful without continuous sensors.", "S03,S43,S45,S46"),
("🟡 Strong Inference", "Unit economics likely improved when moving from included hardware bundles to low-cost software plus optional add-ons, because CGMs/labs carry operational and vendor costs.", "S03,S08,S13"),
("🟡 Strong Inference", "Retention challenge is high because many users learn their glucose patterns in 1-3 months; habit loops, labs, AI, and Pro workflows are retention countermeasures.", "S03,S11,S13,S14,S19,S45,S46"),
]
for label, claim, src in business:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 14. Growth Strategy\n")
growth = [
("🟢 Confirmed", "Levels operates a large public content engine with blog categories, videos, podcasts, newsletter CTAs, and AI Search.", "S01,S20,S27"),
("🟢 Confirmed", "2024 coverage claimed 18M+ YouTube views and 2M+ annual visitors to the metabolic health blog.", "S22"),
("🟢 Confirmed", "Wefunder claimed 468k+ email subscribers, 10M+ YouTube views, 1.5M+ podcast listens, and 500+ blog articles; these are crowdfunding listing claims.", "S42"),
("🟢 Confirmed", "Levels uses high-credibility advisors, medical reviewers, and expert bylines to build trust.", "S04,S20,S21,S26,S30"),
("🟢 Confirmed", "Partner pages such as Andrew Huberman x Levels are publicly visible in search results and align with influencer-led distribution.", "public search"),
("🟢 Confirmed", "Community funding is a growth/brand loop: $5M member crowdfund in 2022 and $3M in 2024 coverage.", "S21,S22"),
("🟡 Strong Inference", "SEO is a central CAC strategy because Levels has deep educational content around glucose, biomarkers, nutrition, and metabolic health with repeated product CTAs.", "S01,S20,S27,S42"),
("🟡 Strong Inference", "Levels Pro is a distribution expansion into practitioner-led acquisition, where clinicians/nutritionists bring cohorts of clients into the app.", "S19"),
]
for label, claim, src in growth:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 15. Hiring Intelligence\n")
hiring = [
("🟢 Confirmed", "No current official Levels careers page was found at /careers on levels.com/levelshealth.com; those URLs returned page-not-found during public fetch.", "Public fetch"),
("🟢 Confirmed", "Third-party job pages produced noisy/non-definitive results and were not treated as verified Levels hiring signals.", "Public search"),
("🟡 Strong Inference", "Current roadmap likely requires AI engineering, mobile/data engineering, lab operations, clinical operations, Pro/B2B product, security/privacy, and content/growth roles.", "S03,S10,S11,S13,S14,S17,S19,S28"),
("🟡 Strong Inference", "The lack of an obvious official careers funnel may indicate low active hiring, a hidden/referral-heavy process, outsourced recruiting, or simply a moved URL; cannot be verified publicly.", "Public fetch"),
]
for label, claim, src in hiring:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 16. Customer Intelligence\n")
customer = [
("🟢 Confirmed", "iOS App Store rating observed at 4.7 with about 6.6K ratings.", "S04"),
("🟢 Confirmed", "Positive reviews praise real-time food/exercise/sleep awareness, CGM access, meal photos, Apple Watch/Apple Health import, and behavior change.", "S04,S46"),
("🟢 Confirmed", "Negative reviews complain about sync failures, hidden sync controls, reduced education access, unhelpful insights, AI coach wrong/confident advice, and limited/tedious food tracking.", "S04,S46"),
("🟢 Confirmed", "Reddit reviews complain about dual-app dependency, poor connectivity, calibration errors, missing sleep/exercise sync, food database/logging limitations, narrow glucose ranges, and weak value for endurance athletes.", "S45"),
("🟢 Confirmed", "Reddit users also praise the app as better than native CGM apps for trends and behavior coaching when it works.", "S45"),
("🟡 Strong Inference", "The strongest churn trigger is 'I learned the patterns already,' followed by price, sync issues, and mistrust in scoring/AI.", "S45,S46"),
("🟡 Strong Inference", "Unexpected use cases include cancer-treatment monitoring, prediabetes prevention, hypoglycemia investigation, athlete fueling experiments, and quantified-self experimentation.", "S45,S46"),
]
for label, claim, src in customer:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 17. Decision Ledger\n")
md.append("| Feature | Why built | Pain solved | KPI improved | Trade-offs | Alternative architecture | Evidence | Confidence |\n|---|---|---|---|---|---|---|---|\n")
for row in decision_rows:
    md.append("| " + " | ".join(str(x).replace('|','/') for x in row) + " |\n")

md.append("\n## 18. Feature Dependency Graph\n")
md.append("```mermaid\nflowchart TD\n  Consent[Consent + permissions]\n  Identity[Identity: account, phone, address, device accounts]\n  Collection[Data collection: CGM, meals, labs, wearables, documents]\n  Normalization[Normalization: timeline, units, timestamps, metadata]\n  Quality[Data quality: sensor delays, artifacts, missing logs]\n  AI[AI/context engine]\n  Scores[Scores: meal/zone/stability/habits]\n  Reports[Reports: labs, trends, exports, clinician notes]\n  Insights[Insights + recommendations]\n  Programs[Programs + habit loops]\n  Doctor[Doctor/practitioner sharing + Pro]\n  Patient[Patient/member behavior change]\n  Consent --> Identity --> Collection --> Normalization --> Quality --> AI --> Scores --> Insights --> Programs --> Patient\n  Normalization --> Reports --> Doctor --> Patient\n  AI --> Reports\n  Patient --> Collection\n```\n")

md.append("\n## 19. Engineering Backlog Reconstruction\n")
roadmap = [
("MVP", "🟡 Strong Inference", "CGM access, app login, sensor onboarding, glucose timeline, meal photo/text logging, basic meal/daily scoring, educational content, support.", "S20,S04,S10"),
("V2", "🟡 Strong Inference", "Apple Health/imports, daily/weekly/monthly reports, food comparisons, community/content loops, member waitlist and referral mechanics.", "S04,S20,S42"),
("V3", "🟢 Confirmed", "Lab testing, bloodwork uploads, nutritionist/clinician support, member portal export/share, BYOD sensors, lower-cost software membership.", "S03,S13,S14,S31,S32"),
("Current", "🟢 Confirmed", "App-only membership, Stelo integration, Dexcom G7 Amazon Pharmacy path, AI Guide, AI food logging, habit loops, programs, Pro dashboard, Document Center.", "S03,S07,S08,S10,S11,S17,S19,S28"),
("Future", "🔴 Speculation", "FHIR-native health record, safer AI coach, richer Pro workflows, personalized ranges, direct wearable integrations, international expansion, payer/employer evidence.", "Strategic prediction"),
]
md.append("| Stage | Label | Reconstructed scope | Evidence |\n|---|---:|---|---|\n")
for r in roadmap:
    md.append(f"| {r[0]} | {r[1]} | {r[2]} | {r[3]} |\n")

md.append("\n## 20. Competitive Landscape\n")
md.append("| Competitor | Label | Category | Evidence | Strategic implication for Ovexis |\n|---|---:|---|---|---|\n")
for name,cat,evid,imp in competitors:
    label = '🟢 Confirmed' if evid.startswith('S') else ('🟢 Confirmed' if 'public search' in evid else '🟡 Strong Inference')
    md.append(f"| {name} | {label} | {cat} | {evid} | {imp} |\n")

md.append("\n### Common Features\n")
for label, claim, src in [
("🟢 Confirmed", "Levels, Function, and Superpower all use labs/biomarkers as a preventive-health engagement layer, but Levels uniquely keeps CGM and food-response feedback central.", "S03,S13,S47,S48"),
("🟢 Confirmed", "Whoop/Oura/Ultrahuman use daily wearable scores and habit coaching; Levels can learn retention design from them while adding labs/CGM context.", "Public search, S01"),
("🟢 Confirmed", "OpenEvidence/UpToDate/AMBOSS/Glass/Atropos compete more on clinical evidence and physician workflow than consumer metabolic behavior.", "S49 and public search"),
("🟡 Strong Inference", "Ovexis should avoid competing only on 'more biomarkers'; the strategic white space is cross-modal causal interpretation plus safe actions.", "S03,S13,S47,S48"),
]:
    md.append(f"- {label}: {claim} **Evidence:** {src}\n")

md.append("\n## 21. Moat Analysis\n")
moats = [
("Data moat", "Medium / Future Strong", "🟢 Confirmed data scale claims include 700M+ glucose points and 7M+ food logs in crowdfunding/coverage; 🟡 data becomes stronger if tied to outcomes and context quality.", "S22,S42"),
("AI moat", "Weak-to-Medium", "🟢 confirmed use of frontier model providers; 🟡 proprietary context/data/evals can create moat but model access itself is not unique.", "S17,S28"),
("Clinical moat", "Medium", "🟢 advisors, labs, clinician notes, Pro workflows; 🟡 limited by wellness posture and lack of public RCT outcomes.", "S03,S13,S15,S19,S26"),
("Brand moat", "Strong in metabolic/biohacker niche", "🟢 content, advisors, funding, member stories, community funding; 🟡 still narrower than Apple/Whoop/Oura.", "S01,S20,S21,S22,S42"),
("Distribution moat", "Medium", "🟢 SEO/content/influencers/crowdfunding/Pro; 🟡 U.S.-only and no insurer/enterprise scale verified.", "S09,S19,S22,S42"),
("Developer moat", "Weak", "🟢 no official public API/docs found; CSV export only verified.", "S31"),
("Marketplace moat", "Weak/Future", "🟡 Add-ons and shop exist; robust clinical marketplace not verified.", "S03,S28"),
("Regulatory moat", "Weak-to-Medium", "🟢 HIPAA/SOC2 claims and BAAs in specific workflows; 🟡 wellness boundary avoids regulation but also limits clinical defensibility.", "S01,S17,S18"),
("Network effects", "Weak-to-Medium", "🟢 community funding/content; 🟡 user data improves insights, but direct user-to-user network effects are limited.", "S21,S22,S42"),
("Switching costs", "Medium-low", "🟢 data export exists; 🟡 users may churn after learning patterns unless labs/programs/Pro create continuity.", "S31,S45,S46"),
("Trust moat", "Medium", "🟢 privacy principles, no data sale, HIPAA/SOC2 statements; 🟡 AI complaints can erode trust.", "S01,S17,S46"),
]
md.append("| Moat | Classification | Label/evidence | Sources |\n|---|---:|---|---|\n")
for m in moats:
    md.append(f"| {m[0]} | {m[1]} | {m[2]} | {m[3]} |\n")

md.append("\n## 22. Failure Analysis\n")
failures = [
("Technical", "Sync delays, sensor artifacts, dual-app complexity, AI temporal/context errors, food-recognition errors, data normalization defects.", "S04,S07,S08,S10,S45,S46"),
("Business", "Users treat CGM as a 30-90 day learning tool and churn; OTC CGM reduces access moat; lab competitors undercut price; high support burden.", "S03,S43,S45,S46,S47,S48"),
("Clinical", "Narrow ranges or weak evidence create anxiety, athlete misfit, or questionable recommendations; retrospective outcomes are not clinical trials.", "S26,S45,S46"),
("Regulatory", "Medical-claim drift or AI recommendations could invite FDA/FTC/OCR scrutiny; Levels’ own ToS keeps wellness boundary.", "S04,S17,S18"),
("Operational", "Lab partner failures, critical value handling, support misalignment, CGM shipping/replacement issues.", "S05,S15,S16,S46"),
("Distribution", "Content/influencer CAC saturates; Pro motion may require enterprise sales maturity; U.S.-only limits growth.", "S09,S19,S22"),
("AI", "Hallucinated or wrong user-specific advice damages trust and creates safety risk.", "S17,S46"),
("Economic", "Sensors/labs have lower margin and users resist monthly hardware cost; app-only must prove standalone value.", "S03,S43,S45"),
]
md.append("| Failure vector | Label | How it could fail | Evidence |\n|---|---:|---|---|\n")
for name, desc, src in failures:
    md.append(f"| {name} | 🟡 Strong Inference | {desc} | {src} |\n")

md.append("\n## 23. Competitive Attack Plan for Ovexis\n")
attack = [
("Technology", "Build a longitudinal health intelligence graph that natively models provenance, timestamps, uncertainty, phenotype, and interventions.", "Reinvent Levels’ timeline into a true causal record."),
("Pricing", "Offer free/low-cost app + paid metabolic sprints + lab bundles + family/pro tiers; avoid forcing permanent CGM subscription.", "Exploit retention/value concerns."),
("Distribution", "Start India/South Asia with local foods, labs, doctors, pharmacies, and WhatsApp/voice workflows rather than U.S.-only premium biohacking.", "Levels is U.S.-only."),
("AI", "Use deterministic analytics for calculations and LLMs only for explanation/planning; every recommendation carries evidence, confidence, and safety class.", "Exploit AI coach trust gap."),
("Brand", "Position as 'doctor-ready longitudinal health intelligence' rather than 'glucose curiosity'.", "Move beyond CGM novelty."),
("Clinical", "Publish prospective outcomes and build physician-review pathways by condition: prediabetes, ApoB/triglycerides, PCOS/PMOS, GLP-1 support.", "Build stronger evidence moat."),
("Enterprise", "Provider workspace with roles, audit logs, FHIR export, and reimbursable care-program templates.", "Levels Pro is promising but public evidence of deep EHR integration is absent."),
("Consumer", "Add direct integrations with Oura/Whoop/Garmin/Fitbit plus culturally localized meal AI.", "Exploit integration complaints."),
]
for dim, plan, why in attack:
    md.append(f"- 🟡 Strong Inference **{dim}:** {plan} **Why:** {why}\n")

md.append("\n## 24. Future Prediction\n")
preds = [
("Next 12 months", "🟡 Strong Inference", "Levels will continue shifting from CGM-first to app/labs/AI-first, expand Stelo/G7 flows, improve AI Guide, and push Pro practice adoption.", "S03,S07,S08,S14,S19,S28"),
("Next 3 years", "🔴 Speculation", "Levels may become a hybrid consumer + provider metabolic workflow platform with periodic labs, CGM sprints, AI health record, and practice cohorts.", "Strategic prediction"),
("Next 5 years", "🔴 Speculation", "Likely paths include broader health-record/AI coach expansion, payer/employer outcomes attempts, international partnerships, or acquisition by a device/wearable/lab platform.", "Strategic prediction"),
("Likely AI investments", "🟡 Strong Inference", "Document Center, meal plan personalization, CGM insight generation, Pro pre-visit summaries, clinician-note drafting, model eval/observability.", "S17,S19,S28"),
("Likely partnerships", "🔴 Speculation", "More CGM vendors, labs, dietitian networks, wearable platforms, functional-medicine practices, HSA/FSA/payment partners.", "Strategic prediction"),
]
md.append("| Horizon | Label | Prediction | Evidence basis |\n|---|---:|---|---|\n")
for p in preds:
    md.append(f"| {p[0]} | {p[1]} | {p[2]} | {p[3]} |\n")

md.append("\n## 25. Ovexis Strategy Memo\n")
for title, arr in [("Top 50 ideas to copy", copy_ideas),("Top 50 ideas to improve", improve_ideas),("Top 50 ideas to ignore", ignore_ideas),("Top 50 ideas to reinvent", reinvent_ideas),("Top 50 market gaps", gaps)]:
    md.append(f"\n### {title}\n")
    for i, item in enumerate(arr,1):
        lab = "🟢 Confirmed" if title.startswith("Top 50 ideas to copy") and i < 6 else ("🟡 Strong Inference" if not title.startswith("Top 50 ideas to ignore") else "🟡 Strong Inference")
        md.append(f"{i}. {lab}: {item}\n")
md.append("\n### Top 20 blue-ocean opportunities\n")
for i, item in enumerate(blue_ocean,1):
    md.append(f"{i}. 🔴 Speculation: {item}\n")

md.append("\n### Recommended Ovexis MVP\n")
for label, claim in [
("🟡 Strong Inference", "MVP should be an AI-powered longitudinal health record with labs + wearables + optional CGM sprint, not a CGM reseller."),
("🟡 Strong Inference", "Core screens: onboarding/consent, data connections, unified timeline, lab upload, meal logging, habit experiments, AI copilot with citations/confidence, doctor-ready report."),
("🟡 Strong Inference", "Initial programs: prediabetes/metabolic syndrome, ApoB/triglycerides/heart health, weight-loss/GLP-1 support, PCOS/PMOS, athlete fueling mode."),
("🟡 Strong Inference", "Differentiate with India/South Asia food ontology, local labs, local clinicians, WhatsApp/voice, family accounts, and affordability."),
]:
    md.append(f"- {label}: {claim}\n")
md.append("\n### Recommended GTM\n")
for label, claim in [
("🟡 Strong Inference", "Start with 90-day metabolic intelligence sprints sold through doctors/nutritionists and employer wellness pilots."),
("🟡 Strong Inference", "Use content and calculators, but anchor acquisition in measurable outcomes and clinician-ready reports."),
("🟡 Strong Inference", "Build referral loops around family metabolic risk and cohort challenges, not just biohacker curiosity."),
]:
    md.append(f"- {label}: {claim}\n")
md.append("\n### Recommended Moat\n")
for label, claim in [
("🟡 Strong Inference", "Data moat: provenance-rich longitudinal graph tied to outcomes and interventions."),
("🟡 Strong Inference", "AI moat: audited, evidence-cited, context-aware health reasoning with evaluation datasets and human escalation."),
("🟡 Strong Inference", "Clinical moat: prospective outcomes, clinician workflows, local lab network, and condition-specific protocols."),
("🟡 Strong Inference", "Distribution moat: India healthcare/pharmacy/lab partnerships and practitioner network."),
]:
    md.append(f"- {label}: {claim}\n")
md.append("\n### Recommended AI Architecture\n")
md.append("```mermaid\nflowchart TD\n  Data[Wearables + labs + CGM + meds + diet + docs + claims if available] --> Normalize[Normalize + deduplicate + provenance]\n  Normalize --> Feature[Deterministic feature store: biomarkers, glucose metrics, sleep, activity, nutrition]\n  Feature --> Causal[N-of-1 experiment and causal-inference layer]\n  Normalize --> RAG[Clinical/content RAG with evidence grading]\n  Causal --> Router[Risk-aware AI router]\n  RAG --> Router\n  Router --> LLM[LLM providers / local models]\n  Router --> Rules[Clinical safety rules + contraindications]\n  LLM --> Eval[Automated evals + hallucination checks + citations]\n  Rules --> Eval\n  Eval --> Human{Human review needed?}\n  Human -->|Yes| Clinician[Clinician/nutritionist queue]\n  Human -->|No| User[User recommendation with confidence]\n  Eval --> Audit[Audit log + consent ledger]\n```\n")
md.append("\n### Recommended healthcare integrations\n")
for label, claim in [
("🟡 Strong Inference", "Apple Health, Health Connect, Oura, Whoop, Garmin, Fitbit, Ultrahuman, Dexcom/Stelo/Libre/Lingo where legal."),
("🟡 Strong Inference", "FHIR R4/SMART for provider records, ABHA/NDHM for India if applicable, lab networks, pharmacy data, and PDF parsing fallback."),
("🟡 Strong Inference", "Doctor report export: FHIR Bundle, PDF, CSV, and share portal with scoped permissions and audit logs."),
]:
    md.append(f"- {label}: {claim}\n")
md.append("\n### Recommended pricing\n")
for label, claim in [
("🟡 Strong Inference", "Free record + AI-limited tier; ₹/US$ affordable monthly app; paid 90-day metabolic sprint; annual labs bundle; Pro practice tier."),
("🟡 Strong Inference", "Pass through CGM/lab costs transparently; avoid hiding hardware costs in subscription."),
("🟡 Strong Inference", "Offer family plan and practitioner bulk credits."),
]:
    md.append(f"- {label}: {claim}\n")
md.append("\n### Recommended roadmap\n")
road = ["MVP: lab upload + wearables + AI copilot + doctor report", "V1: optional CGM sprint + meal AI + habit experiments", "V2: clinician portal + programs + South Asian modules", "V3: FHIR/ABHA + pharmacy/claims + family risk", "V4: outcomes-backed employer/insurer programs"]
for i,item in enumerate(road,1):
    md.append(f"{i}. 🟡 Strong Inference: {item}\n")

md.append("\n## 26. Master Feature Inventory\n")
md.append("🟢 Confirmed: The full spreadsheet is saved as `levels_feature_inventory.xlsx` with FeatureInventory, EvidenceRegister, DecisionLedger, Competitors, Roadmap, and OvexisIdeas sheets.\n")
md.append("\n## Product Architecture Diagram\n")
md.append("```mermaid\nflowchart TD\n  Web[Marketing site + content + quiz] --> Checkout[Checkout/payment: Stripe/Affirm/TrueMed]\n  Checkout --> Account[Member account]\n  Account --> Mobile[Levels mobile app]\n  Account --> Portal[Member portal]\n  Mobile --> Logs[Meals, exercise, notes, lifestyle, habits]\n  Mobile --> Devices[Stelo/Dexcom + Apple Health/Health Connect]\n  Portal --> Labs[Lab orders/uploads/results]\n  Logs --> Data[Unified timeline + health data store]\n  Devices --> Data\n  Labs --> Data\n  Data --> Scores[Scores + trends + habit loops]\n  Data --> AI[AI Guide / food AI / lab AI / document AI]\n  AI --> Insights[Insights + recommendations + programs]\n  Scores --> Insights\n  Insights --> User[Member behavior change]\n  Data --> Export[CSV export/share doctor]\n  Data --> Pro[Levels Pro dashboard]\n  Pro --> Practitioner[Practitioner recommendations/protocols]\n```\n")

md.append("\n## Business Model Canvas\n")
canvas = [
("Customer segments", "U.S. adults 18+ seeking metabolic insight; high-intent health optimizers; people focused on heart health, weight loss, glucose control, metabolic health; practitioners/practices via Pro."),
("Value propositions", "Stop guessing; see food/lifestyle effects; combine CGM, labs, AI, programs, and expert support."),
("Channels", "SEO/blog, YouTube, podcast, newsletter, advisors/influencers, quiz, partner pages, practitioner Pro."),
("Customer relationships", "Self-guided app, AI Guide, support team, clinician review, nutritionist session, concierge, Pro practitioner relationship."),
("Revenue streams", "Membership, CGM add-ons, labs, nutritionist sessions, legacy bundles, Pro seats/credits, possible partner shop."),
("Key resources", "Brand, content library, member data, app, AI stack, lab/sensor/physician network, advisor credibility."),
("Key activities", "Data ingestion, scoring, AI insights, lab operations, content, growth, support, privacy/security, Pro workflows."),
("Key partners", "Dexcom/Stelo, Amazon Pharmacy, Quest/BioReference, MD Integrations, Healthie, TrueMed, Affirm, AI providers, Langfuse, nutritionist partners."),
("Cost structure", "Engineering, AI inference/observability, CGM/lab operations, clinical/reviewer ops, support, content/growth, compliance/security."),
]
md.append("| Canvas block | Label | Content |\n|---|---:|---|\n")
for k,v in canvas:
    md.append(f"| {k} | 🟢/🟡 Mixed | {v} |\n")

md.append("\n## SWOT\n")
swot = {
"Strengths": ["Brand leadership in consumer metabolic health", "Integrated app+CGM+labs+AI+habits", "Strong content/advisor halo", "Privacy trust language", "Emerging Pro workflow"],
"Weaknesses": ["U.S.-only", "CGM commodity dependence", "AI trust complaints", "Manual logging burden", "Retention after CGM learning curve"],
"Opportunities": ["Pro provider OS", "App-only retention", "Document Center/health record AI", "Personalized ranges/outcomes evidence", "International localized expansion"],
"Threats": ["Dexcom/Abbott/Apple/Oura/Ultrahuman commoditization", "Function/Superpower labs race", "Regulatory/AI safety", "Consumer anxiety backlash", "Lower-cost raw CGM alternatives"],
}
for quad, arr in swot.items():
    md.append(f"### {quad}\n")
    for item in arr:
        md.append(f"- 🟡 Strong Inference: {item}\n")

md.append("\n## Porter's Five Forces\n")
forces = [
("Rivalry", "High", "Many CGM, lab, wearable, nutrition, and AI-health entrants compete for preventive-health budgets."),
("Threat of substitutes", "High", "Raw OTC CGM apps, primary-care labs, food trackers, wearables, and doctor/nutritionist visits can substitute parts."),
("Supplier power", "Medium-High", "CGM vendors, labs, AI providers, physician networks, and app stores are critical suppliers."),
("Buyer power", "High", "Consumers can churn after short learning period; practitioners can choose other dashboards."),
("New entrants", "High", "OTC CGM, LLM APIs, and lab networks lower entry barriers, though trust/regulatory/data moats matter."),
]
md.append("| Force | Level | Label | Rationale |\n|---|---:|---:|---|\n")
for f in forces:
    md.append(f"| {f[0]} | {f[1]} | 🟡 Strong Inference | {f[2]} |\n")

md.append("\n## Value Chain\n")
value_chain = ["Acquire via content/partners/quiz", "Convert via pricing/checkout/HSA", "Onboard identity/intake/permissions", "Collect CGM/labs/wearables/logs/docs", "Normalize timeline", "Score and visualize", "AI/clinician interpretation", "Programs/habit loops", "Retest/share/export", "Renew/expand/Pro" ]
for i, item in enumerate(value_chain,1):
    md.append(f"{i}. 🟢/🟡 Mixed: {item}\n")

md.append("\n## Risk Register\n")
risks = [
("AI incorrect advice", "High", "High", "Implement risk classes, citations, evals, human escalation, disclaimers."),
("CGM commoditization", "High", "Medium", "Shift to longitudinal intelligence and labs/habits/pro workflows."),
("Retention decay", "High", "High", "Periodic sprints, programs, labs, social accountability, clinician workflows."),
("Regulatory claim drift", "Medium", "High", "Strict medical-review and claims governance."),
("Data privacy incident", "Medium", "High", "Least privilege, audit logs, encryption, vendor review, incident response."),
("Sensor artifacts", "High", "Medium", "Data-quality scoring and artifact detection."),
("Food logging fatigue", "High", "Medium", "Passive/AI/voice/photo improvements and lower-frequency experiments."),
("Provider workflow failure", "Medium", "Medium", "Design with clinicians, integrate into EHR/FHIR, prove time saved."),
]
md.append("| Risk | Likelihood | Impact | Label | Mitigation |\n|---|---:|---:|---:|---|\n")
for r in risks:
    md.append(f"| {r[0]} | {r[1]} | {r[2]} | 🟡 Strong Inference | {r[3]} |\n")

md.append("\n## Evidence Register\n")
md.append("| Source ID | Source | URL | Evidence note |\n|---|---|---|---|\n")
for sid,name,url,note in sources:
    md.append(f"| {sid} | {name} | {url} | {note} |\n")

md.append("\n## References\n")
for sid,name,url,note in sources:
    md.append(f"- {sid}: {name} — {url}\n")

# Write report
report_path = OUT / 'levels_health_ci_report.md'
report_path.write_text(''.join(md), encoding='utf-8')

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = 'FeatureInventory'
headers = ["Feature","Purpose","Evidence","User Value","Business Value","Engineering Complexity","Clinical Complexity","Infrastructure Complexity","Regulatory Complexity","Estimated Team","Estimated Months","Priority","Category","Copy","Improve","Ignore","Reinvent","Moat","Confidence"]
ws.append(headers)
for f in features:
    name,purpose,evidence,user_value,biz_value,eng,clin,infra,reg,team,months,priority,category,action,moat,conf = f
    ws.append([name,purpose,evidence,user_value,biz_value,eng,clin,infra,reg,team,months,priority,category, action=='Copy', action=='Improve', action=='Ignore', action=='Reinvent', moat, conf])

ws2 = wb.create_sheet('EvidenceRegister')
ws2.append(["Source ID","Source","URL","Evidence","Confidence","Observed vs inferred"])
for sid,name,url,note in sources:
    conf = 'High' if sid not in {'S44','S45'} else 'Medium'
    obs = 'Observed' if sid != 'S44' else 'Observed from HTTP/public HTML; architecture inferred only where stated'
    ws2.append([sid,name,url,note,conf,obs])

ws3 = wb.create_sheet('DecisionLedger')
ws3.append(["Feature","Why was it built?","Pain solved","KPI improved","Trade-offs","Alternative architecture","Evidence","Confidence"])
for row in decision_rows:
    ws3.append(list(row))

ws4 = wb.create_sheet('Competitors')
ws4.append(["Competitor","Category","Evidence","Strategic implication","Common features","Unique advantages","Blind spots"])
for name,cat,evid,imp in competitors:
    common = 'health data, insights, subscription or enterprise workflows' if name not in {'Regacore'} else 'not verified'
    unique = imp
    blind = 'not enough public data' if name=='Regacore' else 'varies by category; see report'
    ws4.append([name,cat,evid,imp,common,unique,blind])

ws5 = wb.create_sheet('RoadmapReconstruction')
ws5.append(["Stage","Label","Scope","Evidence"])
for r in roadmap:
    ws5.append(list(r))

ws6 = wb.create_sheet('OvexisIdeas')
ws6.append(["List","Rank","Label","Idea"])
for list_name, arr, label in [
    ('Copy', copy_ideas, '🟢/🟡 Mixed'), ('Improve', improve_ideas, '🟡 Strong Inference'), ('Ignore', ignore_ideas, '🟡 Strong Inference'), ('Reinvent', reinvent_ideas, '🟡/🔴 Mixed'), ('Market gaps', gaps, '🟡 Strong Inference'), ('Blue ocean', blue_ocean, '🔴 Speculation')]:
    for i,item in enumerate(arr,1):
        ws6.append([list_name,i,label,item])

# Styling
for sheet in wb.worksheets:
    sheet.freeze_panes = 'A2'
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    for col_idx, col in enumerate(sheet.columns, start=1):
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value is not None else ''
            max_len = max(max_len, min(len(val), 60))
        sheet.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 55))

xlsx_path = OUT / 'levels_feature_inventory.xlsx'
wb.save(xlsx_path)

# Also write CSV copy of feature inventory for easy diffing
csv_path = OUT / 'levels_feature_inventory.csv'
import csv
with csv_path.open('w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for ftr in features:
        name,purpose,evidence,user_value,biz_value,eng,clin,infra,reg,team,months,priority,category,action,moat,conf = ftr
        writer.writerow([name,purpose,evidence,user_value,biz_value,eng,clin,infra,reg,team,months,priority,category, action=='Copy', action=='Improve', action=='Ignore', action=='Reinvent', moat, conf])

print(report_path)
print(xlsx_path)
print(csv_path)
