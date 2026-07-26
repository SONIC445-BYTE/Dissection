#!/usr/bin/env python3
"""Build Practo competitive-intelligence companion workbook for Ovexis."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- styles ----
HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(color="FFFFFF", bold=True, size=11)
SUB = PatternFill("solid", fgColor="D9E1F2")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YEL = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
WRAP = Alignment(wrap_text=True, vertical="top")
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, ncols, row=1):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDRF; cell.alignment = CEN; cell.border = BORD

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_table(ws, headers, rows, widths, start=1):
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start, column=j, value=h)
    style_header(ws, len(headers), row=start)
    for i, row in enumerate(rows, start=start+1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = WRAP; cell.border = BORD
            if isinstance(val, str):
                if val.strip().startswith("🟢"): cell.fill = GREEN
                elif val.strip().startswith("🟡"): cell.fill = YEL
                elif val.strip().startswith("🔴"): cell.fill = RED
    autosize(ws, widths)
    ws.freeze_panes = ws.cell(row=start+1, column=1)

# =====================================================================
# SHEET 1: Feature Inventory (60 rows)
# =====================================================================
ws = wb.active
ws.title = "Feature Inventory"
fi_headers = ["#","Feature","Purpose","Evidence","User Value","Business Value",
              "Eng Complexity","Clinical Complexity","Infra Complexity","Regulatory Complexity",
              "Est. Team","Est. Months","Priority","Category","Copy/Improve/Ignore/Reinvent","Moat","Confidence"]
fi = [
 ["1","Free doctor discovery","Find verified doctors","Practo.com homepage","High","High (funnel)","Low","Low","Low","Low","3","2","P0","Marketplace","Copy","Marketplace","🟢"],
 ["2","Verified doctor badge","Trust signal","Homepage 'verified'","High","High","Low","Low","Low","Med","2","1","P0","Trust","Copy","Trust","🟢"],
 ["3","Instant video consult (<60s)","24/7 access","Homepage CTA","High","High (GMV)","Med","Med","Med","High","8","4","P0","Teleconsult","Improve","Retention","🟢"],
 ["4","Symptom->specialty shortcuts","Triage entry","Homepage icons","Med","Med","Low","Med","Low","Med","3","2","P1","UX","Copy","Conversion","🟢"],
 ["5","In-clinic appointment booking","Scheduling","Doctors page","High","High","Med","Low","Med","Med","6","3","P0","Marketplace","Copy","Marketplace","🟢"],
 ["6","Digital prescriptions","Rx delivery","App feature","High","Med","Med","High","Med","High","5","3","P1","Clinical","Copy","Clinical","🟢"],
 ["7","Medicine ordering","Pharmacy","Homepage","Med","Med (GMV)","Med","Low","Med","Med","6","4","P1","Commerce","Copy","Distribution","🟢"],
 ["8","Lab test booking","Diagnostics","Homepage","Med","Med","Med","Med","Med","Med","6","4","P1","Commerce","Copy","Distribution","🟢"],
 ["9","Surgery coordination","Secondary care","Care page","Med","Med","High","High","High","High","10","6","P2","Care","Copy","Clinical","🟢"],
 ["10","Practo Ray (clinic PMS)","Practice mgmt","Founders/Forbes","High","High (ARR)","High","Med","High","High","20","12","P0","Enterprise","Copy","Switching","🟢"],
 ["11","Practo Pro (doctor app)","Doctor workflow","App Store","Med","High","Med","Med","Med","Med","8","6","P1","Enterprise","Copy","Switching","🟢"],
 ["12","Insta HMS (hospital HIS)","Hospital mgmt","Capterra/Techjockey","High","High (ARR)","High","High","High","High","30","18","P0","Enterprise","Copy","Enterprise","🟢"],
 ["13","HL7 / API integration","Hospital interop","Insta docs","High","High","High","Med","High","High","10","8","P1","Enterprise","Copy","Enterprise","🟢"],
 ["14","Practo Reach (sponsored)","Monetize listings","LiveMint","Low","High","Low","Low","Low","Med","4","2","P1","Monetization","Copy","Revenue","🟢"],
 ["15","Practo Prime (guarantee)","Quality promise","Help center","High","Med","Low","Low","Low","Med","3","2","P1","Trust","Copy","Trust","🟢"],
 ["16","Practo Plus (subscription)","Recurring revenue","Help center","Med","High","Med","Low","Low","Med","6","4","P1","Monetization","Copy","Retention","🟢"],
 ["17","Practo Assured (curation)","Quality network","FE May-2025","Med","Med","Med","Med","Med","Med","8","6","P2","Trust","Copy","Trust","🟢"],
 ["18","Health articles / SEO","Top-of-funnel","Homepage","Med","Med","Low","Low","Low","Low","3","2","P2","Growth","Copy","SEO","🟢"],
 ["19","App-download SMS capture","Acquisition","Homepage +91","Med","Med","Low","Low","Low","Low","2","1","P2","Growth","Copy","Distribution","🟢"],
 ["20","Patient history storage","Records per visit","App reviews","Med","Med","Med","Med","Med","High","5","4","P1","Records","Improve","Data","🟢"],
 ["21","Appointment reminders (SMS)","No-show reduction","Ray docs","Med","Med","Low","Low","Low","Low","3","2","P1","Workflow","Copy","Retention","🟢"],
 ["22","2FA for clinicians","Security","Security page","Med","Med","Low","Low","Low","High","3","2","P1","Security","Copy","Security","🟢"],
 ["23","Access zones (geo)","Security","Security page","Med","Med","Med","Low","Med","High","4","3","P1","Security","Copy","Security","🟢"],
 ["24","Role-based staff profiles","Access control","Security page","Med","Med","Med","Low","Med","High","4","3","P1","Security","Copy","Security","🟢"],
 ["25","256-bit encryption","Security","Security page","High","High","Low","Low","Low","High","2","1","P0","Security","Copy","Security","🟢"],
 ["26","ISO 27001 cert","Compliance","Security blog","High","High","Low","Low","Low","High","2","6","P1","Security","Copy","Regulatory","🟢"],
 ["27","HIPAA-compliant servers","Compliance","Security page","High","High","Low","Low","Low","High","2","3","P1","Security","Copy","Regulatory","🟢"],
 ["28","Responsible disclosure","Security","Security page","Med","Med","Low","Low","Low","Med","2","1","P2","Security","Copy","Security","🟢"],
 ["29","Agentic AI brain","Differentiation","CIOL 2026","High","High","High","High","High","High","25","12","P0","AI","Improve/Reinvent","AI","🟢"],
 ["30","Clinical Intelligence layer","Data sense-making","CPTO quote","High","High","High","High","High","High","15","10","P0","AI","Reinvent","AI","🟢"],
 ["31","Consultation Intelligence","Context to doctor","CPTO quote","High","Med","High","Med","Med","High","10","8","P1","AI","Reinvent","AI","🟢"],
 ["32","Care Companion","End-to-end guide","CPTO quote","High","Med","High","Med","Med","Med","10","8","P1","AI","Reinvent","AI","🟢"],
 ["33","20k AI calls/chats/day","Engagement","CIOL 2026","Med","High","Med","Low","Med","Med","5","4","P1","AI","Copy","AI","🟢"],
 ["34","40M structured data points","AI fuel","FY25 letter","High","High","Low","Low","Med","Med","3","2","P1","AI","Copy","Data","🟢"],
 ["35","PROMs publishing","Outcomes credibility","FY25 letter","Med","High","Med","High","Low","High","6","4","P1","Evidence","Copy","Trust","🟢"],
 ["36","UAE consumer launch","International","BS 2025","Med","High","Med","Low","Med","High","10","6","P1","Expansion","Copy","Distribution","🟢"],
 ["37","US care-navigation launch","International/GMV","BS/CIOL 2026","High","High","High","Med","High","High","20","12","P0","Expansion","Copy","Distribution","🟢"],
 ["38","Tata AIA insurance tie-up","B2B2C","Press","Med","High","Med","Low","Med","High","6","4","P1","Partnership","Copy","Distribution","🟢"],
 ["39","Corporate wellness (B2B)","Recurring ARR","Competitor analog","High","High","Med","Low","Med","Med","8","6","P2","Enterprise","Improve","Retention","🟡"],
 ["40","Patient-owned longitudinal PHR","Lifelong record","GAP","High","High","High","High","High","High","20","12","P0","Records","Reinvent","Data","🔴"],
 ["41","Wearables ingestion","Continuous data","GAP","High","High","Med","Med","Med","High","8","6","P1","Data","Reinvent","Data","🔴"],
 ["42","Genomics intelligence","Risk model","GAP","High","High","High","High","Med","High","12","10","P2","Data","Reinvent","Data","🔴"],
 ["43","Open FHIR API","Developer ecosystem","GAP (no API)","High","High","High","Med","High","High","15","10","P0","Platform","Reinvent","Developer","🔴"],
 ["44","Doctor choice on instant consult","Trust fix","Complaint theme","High","Med","Low","Low","Low","Med","3","2","P1","Trust","Reinvent","Trust","🔴"],
 ["45","Transparent pricing","Trust fix","Complaint theme","High","Med","Low","Low","Low","Med","2","1","P1","Trust","Reinvent","Trust","🟢"],
 ["46","Refund SLA","Accountability","Complaint theme","Med","Med","Low","Low","Low","Med","3","2","P1","Support","Improve","Trust","🟢"],
 ["47","Confirmation sync to provider","Reliability","Complaint theme","High","High","Med","Low","Med","Med","5","3","P1","Reliability","Improve","Trust","🟢"],
 ["48","Data export/delete (DPDP)","Rights/compliance","Reddit 2026","High","Med","Med","Low","Med","High","5","3","P1","Compliance","Improve","Trust/Reg","🟢"],
 ["49","Clinician EBM copilot","Decision support","GAP","High","High","High","High","Med","High","15","12","P2","AI","Reinvent","Clinical","🔴"],
 ["50","Explainable AI + citations","Trust/ safety","GAP","High","Med","High","High","Med","High","10","8","P1","AI","Reinvent","Trust","🔴"],
 ["51","Confidence display","Calibration","GAP","Med","Med","Med","Med","Low","Med","4","3","P2","AI","Reinvent","Trust","🔴"],
 ["52","Human-in-the-loop review","Safety","GAP","High","Med","Med","High","Low","High","6","4","P1","AI","Reinvent","Clinical","🔴"],
 ["53","Consent-as-a-graph","Granular control","GAP","High","Med","High","Low","High","High","10","8","P1","Privacy","Reinvent","Trust/Reg","🔴"],
 ["54","Patient-visible audit log","Transparency","GAP","Med","Med","Med","Low","Med","High","5","4","P1","Privacy","Reinvent","Trust","🔴"],
 ["55","ABHA / US Core FHIR","India/US interop","GAP","High","High","High","Med","High","High","12","10","P1","Interop","Reinvent","Regulatory","🔴"],
 ["56","Family/household graph","Household health","GAP","Med","Med","Med","Med","Med","Med","8","6","P2","Data","Reinvent","Data","🔴"],
 ["57","Proactive risk forecasting","Prevention","GAP","High","High","High","High","Med","High","15","12","P2","AI","Reinvent","Clinical","🔴"],
 ["58","Integration marketplace","3rd-party apps","GAP","High","High","High","Med","High","High","15","10","P2","Platform","Reinvent","Developer","🔴"],
 ["59","Outcome-based pricing","Alignment","GAP","Med","High","Med","High","Med","High","8","6","P2","Monetization","Reinvent","Revenue","🔴"],
 ["60","Patient advocacy concierge","Support trust","GAP","Med","Med","Low","Low","Low","Low","5","3","P2","Support","Reinvent","Trust","🔴"],
]
write_table(ws, fi_headers, fi, [4,26,30,22,14,16,12,12,12,14,9,10,8,14,16,14,12])
ws.row_dimensions[1].height = 30

# =====================================================================
# SHEET 2: Decision Ledger (40 rows)
# =====================================================================
ws2 = wb.create_sheet("Decision Ledger")
dl_headers = ["#","Feature/Decision","Why built","Pain solved","KPI improved","Trade-offs","Alt architecture considered","Confidence"]
dl = [
 ["1","Practo Ray (2008)","Digitize clinic records","Paper records","Supply lock-in","Sales-heavy GTM","Build vs buy -> built","🟢"],
 ["2","Practo.com (2013)","Free discovery","Opaque discovery","Liquidity","Monetization delayed","Listing-only vs full -> full","🟢"],
 ["3","Series A (Sequoia 2012)","Scale Ray","Capital need","Coverage","Dilution","Bootstrap vs raise -> raise","🟢"],
 ["4","Marketplace (consumer)","Book appointments","Access gap","Bookings","Complexity","B2B-only vs B2C -> B2C","🟢"],
 ["5","Series C (Tencent 90M)","Aggressive scale","Capital","Geo expansion","Burn","Modest vs bold -> bold","🟢"],
 ["6","Acq FitHo (2015)","Preventive wedge","Wellness gap","Breadth","Integration","Build vs buy -> buy","🟢"],
 ["7","Acq Genii (2015)","Eng capacity","Talent gap","Velocity","Culture","Hire vs buy -> buy","🟢"],
 ["8","Acq Insta (12M,2015)","Hospital HMS","Enterprise gap","B2B ARR","Integration debt","Build vs buy -> buy","🟢"],
 ["9","Acq Qikwell (2015)","Hospital scheduling","Hospital booking","Appointments","Integration","Build vs buy -> buy","🟢"],
 ["10","Acq Enlightiks (2015)","Analytics","Insight gap","Efficiency","Integration","Build vs buy -> buy","🟢"],
 ["11","Series D (Tencent 55M,2017)","Global scale","Capital","Intl","Burn","Raise vs profit -> raise","🟢"],
 ["12","Instant Consult","24/7 monetize","Access gap","GMV/consults","Quality backlash","Scheduled vs instant -> instant","🟢"],
 ["13","Practo Reach","Monetize listings","Revenue","ARPU","Trust risk","Subscription vs ads -> ads","🟢"],
 ["14","Practo Plus","Recurring rev","Churn","MRR","Abuse caps","One-off vs sub -> sub","🟢"],
 ["15","Series D (AIA 32M,2020)","Survive winter","Capital","Runway","50% down val","Raise vs fold -> raise","🟢"],
 ["16","Profitability pivot (2022+)","Survive/IPO","Burn","EBITDA","Growth slow","Grow vs profit -> profit","🟢"],
 ["17","Layoff 41 eng (2023)","Cost discipline","Margin","EBITDA","Velocity","Hire vs cut -> cut","🟢"],
 ["18","Co-founder elevation (Nihalani)","Retain talent","Key-person","Retention","--","--","🟢"],
 ["19","Practo Assured (2025)","Quality curation","Trust gap","Trust","Curation cost","Open vs curated -> curated","🟢"],
 ["20","Agentic AI (2025)","Defensible moat","Differentiation","Engagement","Reg unknown","Feature vs system -> system","🟢"],
 ["21","UAE launch (2025)","Diaspora GMV","India saturation","GMV","Execution","US-first vs UAE -> UAE","🟢"],
 ["22","US launch (2025)","Big-market GMV","IPO story","GMV","Execution/reg","India-only vs US -> US","🟢"],
 ["23","PROMs publishing","Outcomes cred","Trust/proof","Credibility","Disclosure risk","Hide vs show -> show","🟢"],
 ["24","CK Mishra board (2026)","Governance/reg","IPO/reg","Governance","--","--","🟢"],
 ["25","Leadership build-out (2026)","IPO-grade org","Founder dependency","Exec","Cost","Founder-led vs pro -> pro","🟢"],
 ["26","VP AI hire (2026)","AI depth","Talent","AI","Cost","--","🟢"],
 ["27","CPTO hire (2026)","Tech unification","Tech debt","Velocity","Cost","--","🟢"],
 ["28","20k AI/day scale","Engagement","Support cost","Interactions","Quality risk","Human vs AI -> AI","🟢"],
 ["29","No public API","Lock-in","Ecosystem gap","Lock-in","Dev gap","Open vs closed -> closed","🟢"],
 ["30","Phone OTP auth","Low friction","Friction","Signup","Security","Email vs OTP -> OTP","🟢"],
 ["31","Provider-scoped records","Clinic control","Data gap","Embed","No longitudinal","Patient vs provider -> provider","🟢"],
 ["32","Free doctor listing","Supply CAC","Liquidity","Supply","Spam","Paid vs free -> free","🟢"],
 ["33","Health-content SEO","Funnel","Awareness","Traffic","Content cost","Paid vs SEO -> SEO","🟢"],
 ["34","Surgery coordination","Secondary care","High-ACU gap","GMV","Ops","--","🟢"],
 ["35","Medicine ordering","GMV","Fulfilment","GMV","Logistics","Build vs partner -> partner","🟢"],
 ["36","Lab booking","GMV","Diagnostics","GMV","Ops","--","🟢"],
 ["37","Tata AIA tie-up","B2B2C","Insurer gap","Rev","Dependency","--","🟢"],
 ["38","ISO 27001","Compliance","Trust","Sales","Cost","--","🟢"],
 ["39","HIPAA server claim","Compliance","US story","Sales","Relevance","--","🟢"],
 ["40","City-by-city GTM","Land-expand","Coverage","Coverage","Slow","Blitz vs city -> city","🟢"],
]
write_table(ws2, dl_headers, dl, [4,30,24,22,18,30,30,12])

# =====================================================================
# SHEET 3: Risk Register
# =====================================================================
ws3 = wb.create_sheet("Risk Register")
rr_headers = ["#","Risk","Category","Likelihood","Impact","Evidence","Practo mitigation","Ovexis lesson","Confidence"]
rr = [
 ["1","Trust erosion from opaque 'instant consult' pricing","Business/Trust","High","High","Trustpilot/Reddit 2026","Prime guarantee","Transparent pricing + doctor choice","🟢"],
 ["2","Confirmation not synced to provider","Operational","Med","High","Play/Trustpilot","--","Reliability-first booking","🟢"],
 ["3","Refund delays / accountability gaps","Business","Med","High","Trustpilot","--","Refund SLA + concierge","🟢"],
 ["4","DPDP 'right to erasure' gap","Regulatory","Med","High","Reddit 2026","Claims compliance","Portable/erasable by design","🟢"],
 ["5","AI liability (undisclosed guardrails)","AI/Clinical","Med","High","No public guardrails","--","Disclosed eval + HITL","🔴"],
 ["6","Clinician-quality / clinical-harm incident","Clinical","Med","High","Instant-consult complaints","Vetting","Quality monitoring + EBM","🟢"],
 ["7","Thin take-rate (~6-7%)","Economic","Med","Med","GMV vs rev ratio","Care-Nav margin +30% CAGR","Diversify B2B2C","🟡"],
 ["8","Tech debt from 5 acquisitions","Technical","High","Med","History","Lean eng post-layoff","Modern architecture from start","🟡"],
 ["9","Platform disintermediation (Apple/Google)","Distribution","Med","High","OS health graphs","Neutral aggregator","Open API + integration","🟡"],
 ["10","New longitudinal/AI entrants (Ovexis-type)","Strategic","Med","High","Category gap","Agentic AI narrative","Own longitudinal + dev moat","🟡"],
 ["11","Regulatory (HIPAA outside US / US state)","Regulatory","Med","High","US expansion","HIPAA server claim","BAA + US Core FHIR","🟡"],
 ["12","Profitability thinness (15cr EBITDA)","Economic","Med","Med","FY25 letter","Cost discipline","Diversified revenue","🟢"],
 ["13","Doctor-app UX neglect","Product","High","Med","App Store reviews","--","First-class clinician UX","🟢"],
 ["14","International execution risk (US/UAE)","Strategic","Med","High","New markets","Diaspora focus","Phased GTM","🟢"],
 ["15","Sales-heavy post-purchase B2B","Business","Med","Med","Trustpilot (clinic)","--","Outcome-based B2B","🟢"],
 ["16","Data lock-in backlash","Trust/Reg","Med","Med","Reddit 2026","--","Portability","🟢"],
 ["17","No-show friction","Operational","Med","Med","Confirm complaints","Reminders","Proactive no-show prevention","🟢"],
 ["18","Brand vs 'scam' narrative","Trust","Med","High","Multiple 2026","--","Radical transparency","🟢"],
]
write_table(ws3, rr_headers, rr, [4,34,14,11,9,30,28,30,12])

# =====================================================================
# SHEET 4: Evidence Register
# =====================================================================
ws4 = wb.create_sheet("Evidence Register")
ev_headers = ["#","Claim","Source","Evidence (excerpt/type)","Screenshot/Artifact","Confidence","Observed vs Inferred"]
ev = [
 ["1","Founded 2008 by Shashank ND & Abhinav Lal","Financial Express / Forbes India","Founder story; NIT Surathkal; father's surgery insight","Web article","🟢","Observed"],
 ["2","Name = 'practice automation'","Financial Express","Quote from Abhinav Lal","Web article","🟢","Observed"],
 ["3","Practo Ray launched 2009 (rebuild Jan 2009)","Forbes India / FE","Live 2009; first clients dentists","Web article","🟢","Observed"],
 ["4","Sequoia Series A $4M (2012)","LiveMint / StartupTalky","Funding rounds","Web article","🟢","Observed"],
 ["5","Series C $90M Tencent-led (2015)","LiveMint / ET","Investor list","Web article","🟢","Observed"],
 ["6","Acquisitions FitHo/Genii/Insta/Qikwell/Enlightiks","Business Standard / LiveMint / ET","Deal reports; Insta $12M","Web article","🟢","Observed"],
 ["7","Series D $55M (2017), val $600-650M","LiveMint","Investor + valuation","Web article","🟢","Observed"],
 ["8","Series D $32M AIA (2020), val ~$310M","StartupTalky / TimesNow","Investor + 50% down","Web article","🟢","Observed"],
 ["9","FY24 rev +22%, GMV 3500cr, loss 17cr","Business Standard","FY results","Web article","🟢","Observed"],
 ["10","FY25 EBITDA +15cr, rev 234cr, GMV 3500cr","Entrackr (annual letter)","Exclusive","Web article","🟢","Observed"],
 ["11","US pilot 50-60 customers; 50M patients/640 cities/5L doctors","Entrackr","Annual letter","Web article","🟢","Observed"],
 ["12","UAE launch May 2025; 50k MAU; 100cr run-rate","Business Standard / FE","Launch report","Web article","🟢","Observed"],
 ["13","US GMV $100M; agentic AI; 20k AI/day; 700k doctors/2400 cities","CIOL / CIO&Leader (May 2026)","Press release","Web article","🟢","Observed"],
 ["14","AI layers: Clinical/Consultation/Care Companion","CPTO Srijesh Kumar quote","Quote","Web article","🟢","Observed"],
 ["15","40M structured data points power AI","FY25 annual letter","Letter","Web article","🟢","Observed"],
 ["16","Leadership: Singh/Biswas/Chopra/George/Srijesh","AngelOne / Wiretel / IndianStartupNews","Hiring news","Web article","🟢","Observed"],
 ["17","CK Mishra independent director (Mar 2026)","Business Standard / IndianStartupNews","Board news","Web article","🟢","Observed"],
 ["18","Security: HIPAA/256-bit/2FA/access zones/ISO27001","Practo /company/security + blog","Security page","Web page","🟢","Observed"],
 ["19","Pricing: Plus 399/mo; 2999-5999/yr; caps 5/day,15/mo","Help center / Express Healthcare","FAQ","Web page","🟢","Observed"],
 ["20","Prime: 15-min wait, 500 guarantee, free","Help center","FAQ","Web page","🟢","Observed"],
 ["21","Insta HMS $25-40/user/mo; 1250+ centers; 22 countries","Capterra / Techjockey","Product page","Web page","🟢","Observed"],
 ["22","Ray premium 1500-5000/mo","Aidukan / comparison","Pricing","Web page","🟢","Observed"],
 ["23","No public API / dev program","Rapidevelopers (2026)","Explicit 'doesn't exist'","Web article","🟢","Observed"],
 ["24","Customer complaints (scam/confirm/refund/data)","Trustpilot / Play / App Store / Reddit 2026","Reviews","Web pages","🟢","Observed"],
 ["25","Competitor mkt share (~22% MediBuddy)","BusinessModelCanvasTemplate","Analysis","Web article","🟡","Inferred"],
 ["26","Tech stack / cloud provider","None public","--","--","🔴","Inferred"],
 ["27","Public patents","None surfaced","No patent search run","--","🔴","Unverified"],
 ["28","Valuation $900M-1.1B (2025-26)","BusinessModelCanvasTemplate (chatter)","Estimate","Web article","🔴","Speculation"],
 ["29","Employees ~400 (Dec 2024) vs 1500 (2017/2023)","TheCompanyCheck / startupwiki","Conflicting","Web article","🟢","Observed(conflict)"],
 ["30","Revenue ~17.76M (TheCompanyCheck) vs 234cr (Entrackr)","Mixed","Conflicting scales","Web article","🟡","Inferred"],
]
write_table(ws4, ev_headers, ev, [4,40,26,34,18,12,16])

# =====================================================================
# SHEET 5: Moat Score
# =====================================================================
ws5 = wb.create_sheet("Moat Score")
mo_headers = ["Moat","Strength","Class","Notes","Confidence"]
mo = [
 ["Marketplace / network effects","Doctors<->patients liquidity (India)","Strong","Hard to replicate locally","🟢"],
 ["Distribution (brand+app)","16M downloads, 700k doctors","Strong","App-store dominance","🟢"],
 ["Clinical moat","None (navigation only)","Weak","No diagnostic IP","🟡"],
 ["AI moat","40M data pts, 20k AI/day, agentic","Medium","Unproven durability; no patents","🟡"],
 ["Data moat","Provider-scoped records","Medium","Weaker than it looks","🟡"],
 ["Brand moat","Trusted but wobbling","Medium","'Scam' narrative risk","🟡"],
 ["Developer moat","None (no API)","Weak","Gap = Ovexis opener","🟢"],
 ["Regulatory moat","ISO27001/HIPAA/CK Mishra","Medium","Governance strength","🟢"],
 ["Switching costs","Ray/Insta embed clinics","Medium-Strong","B2B lock-in","🟢"],
 ["Trust moat","'Verified' but eroded","Medium","Pricing complaints","🟡"],
 ["Future moat","Agentic AI + PROMs + payor","Emerging","Watch this","🟡"],
]
write_table(ws5, mo_headers, mo, [26,34,14,34,12])

# =====================================================================
# SHEET 6: Metrics / Financials
# =====================================================================
ws6 = wb.create_sheet("Metrics")
mt_headers = ["Metric","Value","Year/Date","Source","Confidence"]
mt = [
 ["Founded","2008","--","FE/Forbes","🟢"],
 ["Total raised","~$231-250M / 13 rounds","2008-2022","Dealroom/Tracxn","🟢"],
 ["Valuation (peak)","$600-650M","2017","LiveMint","🟢"],
 ["Valuation (2020)","~$310M (-50%)","2020","StartupTalky","🟢"],
 ["Valuation (2022)","~$418M","2022","Tracxn","🟡"],
 ["Valuation (est 2025-26)","$900M-1.1B","2026","chatter","🔴"],
 ["FY22 loss","₹162 cr","FY22","Entrackr","🟢"],
 ["FY23 revenue","₹204.4 cr (-3.2%)","FY23","FE","🟢"],
 ["FY24 revenue","+22% ; GMV ₹3,500 cr ; loss ₹17 cr","FY24","BS","🟢"],
 ["FY25 revenue","₹234 cr","FY25","Entrackr","🟢"],
 ["FY25 EBITDA","+₹15 cr (first full-year)","FY25","Entrackr","🟢"],
 ["FY25 GMV","~₹3,500 cr (steady)","FY25","Entrackr","🟢"],
 ["Contribution margin","40%->46%","FY24->FY25","Entrackr","🟢"],
 ["Care-Nav gross margin CAGR","+30% (3 yrs)","2022-25","Entrackr","🟢"],
 ["Patients served","50M+","FY25","Entrackr","🟢"],
 ["Cities","640+ (claimed) / 2,400 (2026)","2025/2026","Entrackr/CIOL","🟢"],
 ["Doctors/providers","5 lakh (FY25) / 700k (2026)","2025/2026","Entrackr/CIOL","🟢"],
 ["US GMV","$100M","May 2026","CIOL","🟢"],
 ["UAE MAU","50,000 (weeks)","May 2025","BS","🟢"],
 ["AI interactions","20,000+/day","2026","CIOL","🟢"],
 ["Structured data points","40M","FY25","Entrackr","🟢"],
 ["App downloads","16M ; 4.43* / 273k ratings","2026","AppBrain","🟢"],
 ["Employees","401 (Dec 2024) / 1500 (claims)","mixed","TheCompanyCheck","🟢"],
 ["PROMs recovery","78% tele / 80% physical (3 wks)","FY25","Letter","🟢"],
 ["Implied take-rate","~6.7% (rev/GMV)","FY25","Derived","🟡"],
]
write_table(ws6, mt_headers, mt, [22,40,16,22,12])

# =====================================================================
# SHEET 7: References
# =====================================================================
ws7 = wb.create_sheet("References")
rf_headers = ["#","Title / Source","URL","Type","Date"]
refs = [
 ["1","Practo official website","https://www.practo.com/","Primary","2026-07-25"],
 ["2","Practo Security page","https://www.practo.com/company/security","Primary","2026"],
 ["3","Practo Prime FAQ","https://help.practo.com/practo-prime/faqs-for-practo-prime-patients/","Primary","2026"],
 ["4","Practo Plus FAQ","https://help.practo.com/practo-plus/faqs-for-practo-plus/","Primary","2026"],
 ["5","Practo data privacy blog","https://blog.practo.com/data-privacy-security-practo/","Primary","2018-03"],
 ["6","LiveMint - Series D $55M","https://www.livemint.com/Companies/...","Secondary","2017-01"],
 ["7","Economic Times - Qikwell acq","https://economictimes.indiatimes.com/...","Secondary","2015-09"],
 ["8","Business Standard - acq spree","https://www.business-standard.com/...","Secondary","2015-10"],
 ["9","Business Standard - CK Mishra board","https://www.business-standard.com/...","Secondary","2026-03"],
 ["10","Entrackr - FY25 EBITDA/US pilot","https://entrackr.com/...","Secondary","2025-08"],
 ["11","CIOL - US $100M GMV / agentic AI","https://www.ciol.com/...","Secondary","2026-05"],
 ["12","CIO&Leader - agentic AI scale","https://www.cioandleader.com/...","Secondary","2026-05"],
 ["13","AngelOne - leadership hires","https://www.angelone.in/news/ipos/...","Secondary","2026-04"],
 ["14","Wiretel - leadership bench","https://wiretel.in/...","Secondary","2026-04"],
 ["15","IndianStartupNews - CK Mishra","https://indianstartupnews.com/...","Secondary","2026-03"],
 ["16","Financial Express - founder journey","https://www.financialexpress.com/...","Secondary","2025-06"],
 ["17","Forbes India - founders 30u30","https://www.forbesindia.com/...","Secondary","2015-02"],
 ["18","StartupTalky - success story","https://startuptalky.com/practo-success-story/","Secondary","2021"],
 ["19","Trustpilot - Practo reviews","https://www.trustpilot.com/review/practo.com","Secondary","2025-26"],
 ["20","Google Play - Practo app","https://play.google.com/store/apps/details?id=com.practo.fabric","Secondary","2026"],
 ["21","App Store - Practo Pro","https://apps.apple.com/in/app/practo-pro-for-doctors/id592116111","Secondary","2026"],
 ["22","Reddit r/india - Practo scam","https://www.reddit.com/r/india/...","Secondary","2026-01"],
 ["23","Rapidevelopers - no Practo API","https://www.rapidevelopers.com/bolt-ai-integrations/practo","Secondary","2026-04"],
 ["24","Capterra - Insta HMS","https://www.capterra.com/p/130501/Insta-HMS/","Secondary","2026"],
 ["25","Techjockey - Insta HMS","https://www.techjockey.com/detail/insta-hms","Secondary","2019"],
 ["26","Dealroom - Practo funding","https://app.dealroom.co/companies/practo","Secondary","2026"],
 ["27","Tracxn - Practo funding","https://tracxn.com/d/companies/practo/...","Secondary","2026"],
 ["28","TheCompanyCheck - Practo","https://www.thecompanycheck.com/company/b/practo/...","Secondary","2026"],
 ["29","BusinessModelCanvasTemplate - history/comp","https://businessmodelcanvastemplate.com/...","Secondary","2026"],
 ["30","Express Healthcare - Practo Plus","https://www.expresshealthcare.in/...","Secondary","2020-03"],
]
write_table(ws7, rf_headers, refs, [4,40,52,14,12])

# save
out = "/home/user/practo_feature_inventory.xlsx"
wb.save(out)
print("WROTE", out, "sheets:", wb.sheetnames)
