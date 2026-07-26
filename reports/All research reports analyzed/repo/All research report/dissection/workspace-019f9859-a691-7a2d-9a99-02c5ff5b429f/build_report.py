from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

out=Path('/home/user')
# Feature inventory
features=[
('Ambient encounter capture','Capture visit conversation for downstream drafting','Official clinician page','High','High','H','H','M','3-5','4-8','P0','Clinician workflow','Improve','Strong','High'),
('Speaker diarized transcription','Separate speakers in >50 languages','Official API docs','High','High','H','M','M','2-4','3-6','P0','Scribe/API','Copy','Medium','High'),
('Live clinical insights','Surface complaint/DDx/history/next steps during visit','Official comparisons/resources','High','High','H','H','H','4-7','6-12','P0','CDS','Reinvent','Strong','Medium'),
('Tiered differential diagnosis','Most likely / expanded / cannot miss','Official API docs','High','High','M','H','M','3-6','4-9','P0','Clinical reasoning','Improve','Strong','High'),
('Evidence-grounded clinical Q&A','Guideline/literature answer with citations','Official API docs','High','High','H','H','H','5-8','8-15','P0','Clinical reasoning','Improve','Strong','High'),
('Assessment & plan','Problem-based Dx/Tx next steps','Official API docs','High','High','M','H','H','3-6','5-10','P0','Documentation','Improve','Medium','High'),
('Note generator','Generate H&P, progress, clinic notes and more','Official clinician page','High','High','M','M','M','3-5','4-8','P0','Documentation','Copy','Medium','High'),
('60+ document types','Specialty and admin documents','Official clinician page','Medium','High','M','M','M','2-4','3-8','P1','Documentation','Improve','Medium','High'),
('Patient handouts/instructions','Plain-language clinical output','Official API docs','High','Medium','M','H','M','2-4','3-6','P1','Patient','Improve','Medium','High'),
('Coding suggestions','E/M, ICD-10-CM, CPT with justification','Official clinician page','High','High','M','H','H','4-7','6-12','P1','Revenue cycle','Improve','Medium','High'),
('SMART on FHIR launch','In-EHR OAuth context-aware app','Official EHR page','High','High','H','H','H','4-8','6-15','P0','Integration','Improve','Strong','High'),
('EHR context read','Authorized patient/encounter context','Official EHR guide','High','High','H','H','H','4-8','6-15','P0','Integration','Reinvent','Strong','High'),
('EHR write-back','Return reviewed output to chart destination','Official EHR guide says scope confirm','High','High','H','H','H','5-9','9-18','P0','Integration','Reinvent','Strong','Medium'),
('Longitudinal patient context','Persist history across encounters','Official comparison pages','High','High','H','H','H','5-8','8-15','P0','Longitudinal','Reinvent','Strong','Medium'),
('Patient health overview','Unify records, labs, wearables, medications','Official patient page','High','High','H','H','H','5-9','9-18','P0','Consumer longitudinal','Reinvent','Future','High'),
('Records summary','Summarize connected data','Official patient page','High','Medium','H','H','H','4-7','6-12','P0','Consumer longitudinal','Improve','Strong','High'),
('Visit prep','Prepare patient for physician visit','Official patient page','High','Medium','M','H','M','2-4','4-8','P1','Patient engagement','Improve','Medium','High'),
('Doctor message draft','Draft communication','Official patient page','Medium','Medium','L','M','M','1-2','2-4','P2','Patient engagement','Copy','Weak','High'),
('Wearable trends','Analyze sleep/recovery/BP and signals','Official patient page','High','High','H','H','H','5-8','8-16','P0','Consumer longitudinal','Reinvent','Future','High'),
('Medical record connection','Connect external provider/lab data','Official patient-page mockup','High','High','H','H','H','5-9','9-18','P0','Data ingestion','Reinvent','Strong','Medium'),
('Lab connection','Connect lab results','Official patient-page mockup','High','High','M','H','H','3-6','5-10','P0','Data ingestion','Improve','Medium','Medium'),
('Goal tracking','Turn overview into health goals','Official patient page','Medium','High','M','H','M','3-5','4-8','P1','Retention','Improve','Medium','High'),
('Timeline','Chronological care and data timeline','Official patient page','High','High','M','H','H','4-7','6-12','P0','Longitudinal','Reinvent','Strong','High'),
('Developer Messages API','Clinical AI API, JSON/SSE','Official API docs','High','High','M','H','H','4-7','5-10','P0','Platform','Copy','Strong','High'),
('Developer Scribing API','Audio jobs/direct upload and notes','Official API docs','High','High','H','M','H','4-7','6-12','P1','Platform','Improve','Medium','High'),
('Evidence reference objects','Typed literature/drug/billing citations','Official API docs','High','High','M','H','H','4-7','5-10','P0','Trust/API','Reinvent','Strong','High'),
('SSE progress events','Reveal retrieval/generation status','Official API docs','Medium','Medium','M','L','M','2-3','2-5','P2','Developer UX','Copy','Weak','High'),
('API key management','Provision/revoke secrets','Official API docs','High','High','M','M','H','2-4','3-6','P0','Security','Copy','Medium','High'),
('Click-through BAA','Enable production PHI API path','Official API docs','High','High','M','H','H','3-5','5-8','P0','Compliance','Improve','Strong','High'),
('API usage metering','Tokens and minimum subscription','Official API docs','Medium','High','M','L','M','2-3','2-4','P1','Monetization','Copy','Medium','High'),
('Template prompting','User-specified structured outputs','Official API docs','Medium','Medium','L','M','L','1-2','1-3','P1','Developer UX','Improve','Weak','High'),
('Deep reasoning mode','Higher-depth complex case analysis','Official pricing/compare pages','High','High','H','H','H','5-8','8-15','P1','Clinical reasoning','Reinvent','Medium','Medium'),
('Customization','Clinician templates/style','Official pricing pages','High','High','M','M','M','2-4','3-6','P1','Workflow','Improve','Medium','Medium'),
('Role/organization controls','Manage teams and permissions','Not publicly evidenced','High','High','M','H','H','4-7','7-12','P0','Admin','Reinvent','Unknown','Low'),
('Audit log','Trace PHI/output/writeback','Not publicly evidenced','High','High','M','H','H','4-7','6-12','P0','Security','Reinvent','Strong','Low'),
('Patient consent management','Manage record/wearable sharing','Privacy policy supports user authorization; UI unknown','High','High','H','H','H','5-8','8-15','P0','Consent','Reinvent','Strong','Medium'),
('Safety escalation','Urgent symptom triage','API capability claim','High','High','M','H','H','4-8','8-16','P0','Safety','Reinvent','Strong','High'),
('FDA label retrieval','Drug information grounding','Official clinician/API page','High','High','M','H','H','3-5','5-9','P0','Evidence','Copy','Medium','High'),
('Bias evaluation','Demographic-axis testing','Official API docs (vendor claim)','High','High','H','H','H','5-8','8-15','P0','Safety','Reinvent','Strong','Medium'),
('Security red-teaming','Test abuse/security vulnerabilities','Official API docs (vendor claim)','High','High','H','M','H','3-5','5-8','P0','Security','Improve','Medium','Medium'),
]
wb=Workbook(); ws=wb.active; ws.title='Feature Inventory'
headers=['Feature','Purpose','Evidence','User Value','Business Value','Engineering Complexity','Clinical Complexity','Infrastructure Complexity','Estimated Team','Estimated Months','Priority','Category','Action','Moat','Confidence']
ws.append(headers)
for r in features: ws.append(r)
for c in ws[1]: c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor='172554');c.alignment=Alignment(wrap_text=True)
for row in ws.iter_rows(min_row=2):
 for c in row: c.alignment=Alignment(wrap_text=True,vertical='top')
ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
for i,w in enumerate([28,38,38,14,16,17,17,19,14,16,10,20,14,14,12],1): ws.column_dimensions[get_column_letter(i)].width=w
# evidence
es=wb.create_sheet('Evidence Register'); es.append(['ID','Claim / observation','Label','Source','URL','Evidence excerpt / method','Confidence','Observed vs inferred','Date checked'])
evidence=[
('E01','Founded in 2021 by Dereck Paul, MD and Graham Ramsey','🟢 Confirmed','Glass blog','https://blog.glass.health/yc/','Founders state they founded Glass in 2021.','High','Observed','2026-07-25'),
('E02','$5M seed led by Initialized announced Sep 2023; YC and Breyer involvement','🟢 Confirmed','Glass LinkedIn / Finsmes','https://www.finsmes.com/2023/09/glass-health-raises-5m-in-funding.html','Company announcement and independent funding report.','High','Observed','2026-07-25'),
('E03','Two developer model versions: Glass 5.0 and 5.5','🟢 Confirmed','Glass API documentation','https://glass.health/api-documentation','Model section specifies version strings.','High','Observed','2026-07-25'),
('E04','Messages API POST endpoint, X-Api-Key auth, SSE option','🟢 Confirmed','Glass API documentation','https://glass.health/api-documentation','Technical reference exposes endpoint/schema.','High','Observed','2026-07-25'),
('E05','Current clinician price tiers $0/$18/$81 and higher Max plan','🟢 Confirmed','Glass pricing','https://glass.health/pricing/clinicians','Pricing fetched 25 July 2026; subject to change.','High','Observed','2026-07-25'),
('E06','Supports Epic, eClinicalWorks, athenahealth and Elation on Max; exact workflow must be confirmed','🟢 Confirmed','Glass EHR guide','https://glass.health/resources/clinical-ai-api-ehr-integration','Explicitly says confirm exact read/writeback/implementation.','High','Observed','2026-07-25'),
('E07','Consumer service permits imports/uploads/connections of external records, devices and more','🟢 Confirmed','Glass privacy policy','https://glass.health/privacy','Outside Information section lists these categories.','High','Observed','2026-07-25'),
('E08','Consumer account data is not necessarily HIPAA PHI in Glass hands absent BAA/covered relationship','🟢 Confirmed','Glass privacy policy','https://glass.health/privacy','Policy explicitly draws this boundary.','High','Observed','2026-07-25'),
('E09','RAG over PubMed-indexed literature and guidelines','🟢 Confirmed','Glass API documentation','https://glass.health/api-documentation','Vendor docs describe retrieval process; performance remains vendor claim.','Medium','Observed vendor claim','2026-07-25'),
('E10','Underlying frontier model/provider, data store, cloud, database, SOC 2, FDA clearance, formal validation paper unknown','🟢 Confirmed absence of public evidence','Public-site review','https://glass.health/','No public evidence located in reviewed official material; absence is not proof of absence.','Medium','Observed absence','2026-07-25'),
('E11','Consumer longitudinal expansion is strategic pivot/broadened wedge','🟡 Strong Inference','Official patient and homepage','https://glass.health/for-patients','New patient product uses records/labs/wearables and shares clinical AI positioning.','Medium','Inferred','2026-07-25'),
('E12','No evidence Glass uses a knowledge graph','🟢 Confirmed absence of public evidence','API/site review','https://glass.health/api-documentation','Docs disclose agentic retrieval, not a graph. Do not infer graph use.','High','Observed absence','2026-07-25'),
]
for r in evidence: es.append(r)
for c in es[1]: c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor='172554');c.alignment=Alignment(wrap_text=True)
for row in es.iter_rows(min_row=2):
 for c in row:c.alignment=Alignment(wrap_text=True,vertical='top')
es.freeze_panes='A2'; es.auto_filter.ref=es.dimensions
for i,w in enumerate([10,46,20,26,48,48,14,20,15],1):es.column_dimensions[get_column_letter(i)].width=w
# roadmap
rs=wb.create_sheet('Roadmap Reconstruction');rs.append(['Era','Likely product state','Evidence / rationale','Label','Confidence'])
for r in [
('MVP (2021–2022)','Medical knowledge-management notebook + clinician learning; early experimental clinical AI.','Founders’ 2023 post says original mission/new software; launch March 2022.','🟢 Confirmed','High'),
('V2 (2023–2025)','LLM-powered DDx and clinical-plan drafting; RAG / doctor-supervised positioning.','2023 funding coverage and Glass statements.','🟢 Confirmed','High'),
('Current (2026)','Three-sided platform: clinician point-of-care suite, consumer longitudinal agent, paid Developer API; scribing, EHR, billing/coding surface.','Current official public pages/API docs.','🟢 Confirmed','High'),
('Next likely 12m','Deepen data connectors, write-back, patient/clinician shared workflow, enterprise governance and clinical evaluation.','Necessary complements to launched consumer + clinician + API surfaces.','🟡 Strong Inference','Medium'),
('Not established','Exact engineering headcount, technical debt, model provider, data architecture, or acquisition plan.','No reliable public disclosure reviewed.','🟢 Confirmed unknown','High')]:rs.append(r)
for c in rs[1]: c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor='172554')
for col,w in zip(range(1,6),[22,55,55,22,14]):rs.column_dimensions[get_column_letter(col)].width=w
wb.save(out/'glass_health_feature_inventory.xlsx')
